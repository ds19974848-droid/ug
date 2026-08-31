import requests
import xlrd
import pdfplumber
import json
import re
import os
import tempfile
import base64
import time
import zipfile
from html import unescape
from datetime import datetime
from io import BytesIO
from pathlib import Path, PurePosixPath
from urllib.parse import parse_qs, unquote, urljoin, urlparse

from bs4 import BeautifulSoup
from openpyxl import load_workbook

API_HANDLERS = {}


def _extract_period(*values):
    text = " ".join(str(value or "") for value in values)
    match = re.search(r"(20\d{2})\D{0,3}(\d{1,2})(?:\D|$)", text)
    if not match:
        return ""
    month = int(match.group(2))
    return f"{match.group(1)}-{month:02d}" if 1 <= month <= 12 else ""

def register(domain_pattern):
    def wrapper(func):
        API_HANDLERS[domain_pattern] = func
        return func
    return wrapper


def has_api_handler(url: str) -> bool:
    return any(pattern in (url or "") for pattern in API_HANDLERS)

@register("ciac.zjw.sh.gov.cn")
def shanghai_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """Shanghai official construction cost info API."""
    base = "https://ciac.zjw.sh.gov.cn/JGBXMGCZJInterWeb/interWeb"
    headers = {"User-Agent": "Mozilla/5.0", "Content-Type": "application/x-www-form-urlencoded"}
    try:
        if progress_callback:
            progress_callback(5, "上海: 获取栏目树...")
        r = requests.post(base + "/hyxx/getHyxxDhcd", data={"bmCode": "003002"}, headers=headers, timeout=30)
        data = r.json()
        items = data.get("data", [])
        zdId = items[0].get("id") if items else None
        if not zdId:
            return {"prices": [], "count": 0, "error": "无法获取栏目ID: " + r.text[:200]}
        
        if progress_callback:
            progress_callback(15, "上海: 获取发布列表...")
        r2 = requests.post(base + "/hyxx/getHyxxList",
                          data={"zdId": zdId, "pageNum": "1", "pageSize": "10"},
                          headers=headers, timeout=30)
        d2 = r2.json()
        plist = d2.get("rows") or d2.get("result") or (d2.get("data") or {}).get("list") or []
        if not plist:
            return {"prices": [], "count": 0, "error": "发布列表为空"}
        
        target = plist[0]
        if period_filter:
            for p in plist:
                t = (p.get("title") or "") + (p.get("fbrq") or "")
                if period_filter in t:
                    target = p
                    break
        
        wjid = target.get("wjid")
        period_text = _extract_period(target.get("title"), target.get("fbrq"))
        if not wjid:
            return {"prices": [], "count": 0, "error": "无法获取文件ID"}
        
        if progress_callback:
            progress_callback(30, f"上海: 下载{period_text}信息价...")
        r3 = requests.post(base + "/currently/bdFileDownload",
                          data={"id": wjid}, headers=headers, timeout=120)
        
        filedata = None
        ct = r3.headers.get("Content-Type", "")
        if "json" in ct or r3.text.strip().startswith("{"):
            d3 = r3.json()
            b64 = d3.get("data", "") or d3.get("base64", "") or ""
            if b64 and isinstance(b64, str):
                filedata = base64.b64decode(b64)
            elif isinstance(d3.get("data"), dict):
                b64 = d3["data"].get("base64", "") or d3["data"].get("data", "")
                if b64 and isinstance(b64, str):
                    filedata = base64.b64decode(b64)
            else:
                return {"prices": [], "count": 0, "error": "返回JSON但无法找到base64内容"}
        else:
            filedata = r3.content
        
        if not filedata or len(filedata) < 200:
            return {"prices": [], "count": 0, "error": f"文件过小 {len(filedata)} bytes"}
        
        if progress_callback:
            progress_callback(55, "上海: 解析XLS文件...")
        
        tmp = tempfile.NamedTemporaryFile(suffix=".xls", delete=False)
        tmp_path = tmp.name
        tmp.write(filedata)
        tmp.close()
        
        prices = []
        wb = xlrd.open_workbook(tmp_path, formatting_info=True)
        total_sheets = len(wb.sheet_names())
        
        for si, sheet_name in enumerate(wb.sheet_names()):
            sh = wb.sheet_by_name(sheet_name)
            for ri in range(2, sh.nrows):
                try:
                    code = str(sh.cell_value(ri, 1)).strip() if sh.ncols > 1 else ""
                    suffix = str(sh.cell_value(ri, 2)).strip() if sh.ncols > 2 else ""
                    name = str(sh.cell_value(ri, 3)).strip() if sh.ncols > 3 else ""
                    spec = str(sh.cell_value(ri, 4)).strip() if sh.ncols > 4 else ""
                    unit = str(sh.cell_value(ri, 5)).strip() if sh.ncols > 5 else ""
                    
                    price = None
                    raw_price = ""
                    for pi in [6, 7, 8, 9]:
                        if sh.ncols > pi:
                            val = str(sh.cell_value(ri, pi)).strip()
                            if val and val != "" and val != "-":
                                raw_price = val
                                m = re.match(r"^([\d,.]+)\s*-\s*([\d,.]+)$", val)
                                if m:
                                    price = (float(m.group(1)) + float(m.group(2))) / 2
                                else:
                                    try:
                                        price = float(val)
                                    except ValueError:
                                        continue
                                break
                    
                    if not name or price is None or price <= 0:
                        continue
                    
                    prices.append({
                        "code": code + suffix,
                        "name": name,
                        "spec": spec,
                        "unit": unit,
                        "price": price,
                        "raw_price": raw_price,
                        "category": sheet_name,
                        "period": period_text or "2026-07",
                    })
                except Exception:
                    continue
            
            if progress_callback and total_sheets > 0:
                progress_callback(55 + int(35 * (si + 1) / total_sheets), f"上海: 解析{sheet_name}...")
        
        os.unlink(tmp_path)
        
        if progress_callback:
            progress_callback(95, f"上海: 解析完成, {len(prices)}条")
        
        return {
            "prices": prices,
            "count": len(prices),
            "period": period_text or "2026-07",
            "document": {
                "file_name": f"{period_text or 'latest'}_上海市人工材料机械信息价.xls",
                "url": f"{base}/currently/bdFileDownload?id={wjid}",
                "content": filedata,
            },
        }
        
    except Exception as e:
        import traceback
        return {"prices": [], "count": 0, "error": str(e), "traceback": traceback.format_exc()}


@register("zjj.sz.gov.cn/szzjxx")
def shenzhen_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """深圳住建局官方造价信息查询系统公开接口。"""
    base = "https://zjj.sz.gov.cn/szzjxx/priceinfo/pc"
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://zjj.sz.gov.cn/szzjxx/web/pc/index",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    }
    try:
        if progress_callback:
            progress_callback(5, "深圳: 获取发布期数...")
        response = requests.post(base + "/currentyear", data={"year": str(datetime.now().year)}, headers=headers, timeout=30)
        periods = response.json() if response.text.strip() else []
        if not periods:
            return {"prices": [], "count": 0, "error": "未获取到深圳发布期数"}
        target = periods[0]
        if period_filter:
            for period in periods:
                if period_filter in (period.get("yearMonth") or "") or period_filter in (period.get("periodName") or ""):
                    target = period
                    break
        period_id = target.get("id")
        if not period_id:
            return {"prices": [], "count": 0, "error": "无法获取深圳期数ID"}
        period_text = _extract_period(target.get("periodName")) or target.get("yearMonth") or ""

        if progress_callback:
            progress_callback(12, f"深圳: 解析{target.get('periodName')}分类...")
        category_names = {}
        try:
            tree_response = requests.post(base + "/getcategorytreelist", data={"periodid": period_id}, headers=headers, timeout=30)
            tree = tree_response.json() if tree_response.text.strip().startswith("[") else []
        except Exception:
            tree = []

        def _walk_categories(nodes):
            for node in nodes or []:
                node_id = str(node.get("id") or "")
                if node_id:
                    category_names[node_id] = node.get("name", "")
                _walk_categories(node.get("children") or [])

        _walk_categories(tree)

        raw_rows = []
        page = 1
        while True:
            if progress_callback:
                progress_callback(min(80, 15 + page * 4), f"深圳: 获取第{page}页...")
            payload = requests.post(base + "/all", data={
                "key": "", "categoryIds": "", "periodId": period_id,
                "rows": "1000", "page": str(page), "order": "asc", "sort": "sequencenum",
            }, headers=headers, timeout=60).json()
            rows = payload.get("rows") or []
            total = int(payload.get("total") or 0)
            if not rows:
                break
            raw_rows.extend(rows)
            if total and len(raw_rows) >= total:
                break
            if len(rows) < 1000:
                break
            page += 1

        prices = []
        for row in raw_rows:
            name = str(row.get("mc") or "").strip()
            spec = str(row.get("gg") or "").strip()
            unit = str(row.get("dw") or "").strip()
            raw_price = row.get("djSq")
            if raw_price in (None, ""):
                continue
            raw_text = str(raw_price).strip()
            price = None
            range_match = re.match(r"^([\d,.]+)\s*-\s*([\d,.]+)$", raw_text)
            if range_match:
                price = (float(range_match.group(1).replace(",", "")) + float(range_match.group(2).replace(",", ""))) / 2
            else:
                try:
                    price = float(raw_text.replace(",", ""))
                except ValueError:
                    continue
            if not name or price <= 0:
                continue
            prices.append({
                "code": str(row.get("jgCode") or "").strip(),
                "name": name,
                "spec": spec,
                "unit": unit,
                "price": price,
                "raw_price": raw_text,
                "category": category_names.get(str(row.get("categoryid") or ""), ""),
                "period": period_text,
            })

        snapshot = json.dumps({
            "source": "深圳市住房和建设局 深圳市建设工程造价信息查询系统",
            "period": target,
            "category_tree": tree,
            "rows": raw_rows,
        }, ensure_ascii=False).encode("utf-8")
        if progress_callback:
            progress_callback(95, f"深圳: 解析完成, {len(prices)}条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period_text,
            "document": {
                "file_name": f"{period_text or 'latest'}_深圳价格信息_官方接口快照.json",
                "url": f"https://zjj.sz.gov.cn/szzjxx/web/pc/index?periodId={period_id}",
                "content": snapshot,
            },
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


@register("www.xjzj.com")
def urumqi_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """新疆造价信息网乌鲁木齐综合价格信息 XLSX 适配器。"""
    base = "https://www.xjzj.com"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126 Safari/537.36",
        "Referer": base + "/",
    }
    try:
        if progress_callback:
            progress_callback(5, "乌鲁木齐: 查找最新综合价格信息...")
        response = requests.get(base + "/", headers=headers, timeout=30)
        response.raise_for_status()
        response.encoding = response.apparent_encoding or response.encoding or "utf-8"
        soup = BeautifulSoup(response.text, "lxml")
        candidates = []
        for anchor in soup.find_all("a", href=True):
            text = anchor.get_text(" ", strip=True)
            if "乌鲁木齐市" not in text or "价格信息" not in text:
                continue
            period = _extract_period(text, anchor.get("href", ""))
            candidates.append((period, text, urljoin(base + "/", anchor["href"])))
        if not candidates:
            return {"prices": [], "count": 0, "error": "新疆官网首页没有找到乌鲁木齐价格信息栏目"}
        candidates.sort(key=lambda item: item[0], reverse=True)
        target = next((item for item in candidates if period_filter and item[0] == period_filter), candidates[0])
        period_text, title, detail_url = target

        if progress_callback:
            progress_callback(20, f"乌鲁木齐: 打开{period_text or '最新'}发布页...")
        detail = requests.get(detail_url, headers=headers, timeout=30)
        detail.raise_for_status()
        detail.encoding = detail.apparent_encoding or detail.encoding or "utf-8"
        file_matches = re.findall(r"LookFile\(\s*['\"]([^'\"]+\.xlsx?)['\"]", detail.text, re.I)
        file_path = next((path for path in file_matches if "乌鲁木齐市" in unescape(path)), "")
        if not file_path:
            return {"prices": [], "count": 0, "error": "乌鲁木齐发布页没有找到 XLS/XLSX 价格附件"}
        file_url = urljoin(base + "/", unescape(file_path))

        if progress_callback:
            progress_callback(35, f"乌鲁木齐: 下载{period_text or '最新'} XLSX...")
        download = requests.get(file_url, headers=headers, timeout=120)
        download.raise_for_status()
        content = download.content
        if len(content) < 500:
            return {"prices": [], "count": 0, "error": f"乌鲁木齐 XLSX 文件过小: {len(content)} bytes"}

        if progress_callback:
            progress_callback(55, "乌鲁木齐: 解析材料价格...")
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        prices = []
        try:
            for worksheet in workbook.worksheets:
                rows = worksheet.iter_rows(min_col=1, max_col=5, values_only=True)
                header = None
                category = ""
                for row in rows:
                    cells = ["" if value is None else str(value).strip() for value in row]
                    compact = [value.replace("\n", "").replace(" ", "") for value in cells]
                    if header is None:
                        name_index = next((i for i, value in enumerate(compact) if "材料名称" in value), None)
                        unit_index = next((i for i, value in enumerate(compact) if "单位" in value), None)
                        price_index = next((i for i, value in enumerate(compact) if "除税" in value and "信息价" in value), None)
                        if name_index is not None and unit_index is not None and price_index is not None:
                            header = (name_index, unit_index, price_index)
                        continue
                    name_index, unit_index, price_index = header
                    name = cells[name_index]
                    unit = cells[unit_index]
                    raw_price = cells[price_index]
                    if cells[0] and not name and not unit:
                        category = cells[0]
                        continue
                    try:
                        price = float(raw_price.replace(",", ""))
                    except (TypeError, ValueError):
                        continue
                    if not name or not unit or price <= 0:
                        continue
                    prices.append({
                        "code": cells[0] if cells[0].isdigit() else "",
                        "name": name,
                        "spec": "",
                        "unit": unit,
                        "price": price,
                        "raw_price": raw_price,
                        "category": category or worksheet.title,
                        "period": period_text,
                    })
        finally:
            workbook.close()

        if not prices:
            return {"prices": [], "count": 0, "error": "乌鲁木齐 XLSX 已下载，但没有识别出价格行"}
        if progress_callback:
            progress_callback(95, f"乌鲁木齐: 解析完成, {len(prices)}条")
        file_name = unescape(file_path.rsplit("/", 1)[-1])
        return {
            "prices": prices,
            "count": len(prices),
            "period": period_text,
            "document": {"file_name": file_name, "url": file_url, "content": content},
            "message": f"已从新疆造价信息网解析乌鲁木齐 {period_text} 综合价格 {len(prices)} 条",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


OFFICIAL_BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36"
    ),
}


def _official_get(url, *, referer="", timeout=60, params=None):
    headers = dict(OFFICIAL_BROWSER_HEADERS)
    if referer:
        headers["Referer"] = referer
    response = requests.get(url, headers=headers, params=params, timeout=timeout)
    response.raise_for_status()
    return response


def _response_text(response):
    response.encoding = (
        getattr(response, "apparent_encoding", None)
        or getattr(response, "encoding", None)
        or "utf-8"
    )
    return response.text


def _clean_table_cell(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def _positive_number(value):
    text = _clean_table_cell(value).replace(",", "")
    if not re.fullmatch(r"\d+(?:\.\d+)?", text):
        return None
    number = float(text)
    return number if number > 0 else None


def _price_range_midpoint(value):
    text = _clean_table_cell(value).replace(",", "")
    match = re.fullmatch(r"(\d+(?:\.\d+)?)\s*[～~-]\s*(\d+(?:\.\d+)?)", text)
    if not match:
        return _positive_number(text)
    low, high = (float(number) for number in match.groups())
    if low <= 0 or high <= 0 or high < low:
        return None
    return (low + high) / 2


def _latest_monthly_article(html_text, base_url, required_terms, period_filter=None):
    candidates = []
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.find_all("a", href=True):
        title = _clean_table_cell(anchor.get_text(" ", strip=True))
        if not title or not all(term in title for term in required_terms):
            continue
        period = _extract_period(title, anchor.get("href", ""))
        if period:
            candidates.append((period, title, urljoin(base_url, anchor["href"])))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    if period_filter:
        exact = next((item for item in candidates if item[0] == period_filter), None)
        if exact:
            return exact
    return candidates[0]


def _split_pdf_cell(value):
    """Keep line breaks in PDF cells so merged multi-material rows can be expanded."""
    if value is None:
        return []
    return [part.strip() for part in str(value).replace("\r", "").split("\n") if part.strip()]


def _parse_beijing_pdf(content, period):
    prices = []
    with pdfplumber.open(BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                table_text = " ".join(
                    _clean_table_cell(cell)
                    for row in (table or [])[:3]
                    for cell in (row or [])
                )
                if "工程造价" not in table_text or "信息价" not in table_text:
                    continue
                if "市场" in table_text or "参考价" in table_text:
                    continue
                for raw_row in table or []:
                    if len(raw_row or []) < 6:
                        continue
                    code_values = _split_pdf_cell(raw_row[0])
                    name_values = _split_pdf_cell(raw_row[1])
                    spec_values = _split_pdf_cell(raw_row[2])
                    unit_values = _split_pdf_cell(raw_row[3])
                    price_values = _split_pdf_cell(raw_row[4])
                    if not code_values or not name_values or not unit_values or not price_values:
                        continue
                    if code_values[0].replace(" ", "") in {"代号", "代码"}:
                        continue
                    count = min(
                        len(code_values), len(name_values), len(spec_values),
                        len(unit_values), len(price_values),
                    )
                    for index in range(count):
                        code = code_values[index].replace(" ", "")
                        price = _positive_number(price_values[index])
                        if not re.fullmatch(r"[0-9A-Za-z-]+", code):
                            continue
                        if not name_values[index] or not unit_values[index] or price is None:
                            continue
                        prices.append({
                            "code": code,
                            "name": name_values[index],
                            "spec": spec_values[index],
                            "unit": unit_values[index],
                            "price": price,
                            "raw_price": price_values[index],
                            "category": "北京市工程造价信息价（含税）",
                            "period": period,
                        })
    return prices


_CHONGQING_WATERMARK_CHARS = set(
    "\u91cd\u5e86\u5e02\u4f4f\u623f\u548c\u57ce\u4e61"
    "\u5efa\u8bbe\u5de5\u7a0b\u9020\u4ef7\u603b\u7ad9"
)
_CHONGQING_ORDINALS = {
    "\u4e00": 1, "\u4e8c": 2, "\u4e09": 3, "\u56db": 4,
    "\u4e94": 5, "\u516d": 6, "\u4e03": 7, "\u516b": 8,
    "\u4e5d": 9, "\u5341": 10, "\u5341\u4e00": 11, "\u5341\u4e8c": 12,
}


def _clean_chongqing_cell(value):
    """Remove the official PDF watermark characters mixed into table cells."""
    text = _clean_table_cell(value)
    if not text:
        return text
    text = "".join(ch for ch in text if ch not in _CHONGQING_WATERMARK_CHARS)
    return " ".join(text.split())


def _chongqing_period(text):
    match = re.search(
        r"(20\d{2})\u5e74\u7b2c([\u4e00\u4e8c\u4e09\u56db\u4e94\u516d\u4e03\u516b\u4e5d\u5341]+)\u671f",
        text or "",
    )
    if not match:
        return ""
    month = _CHONGQING_ORDINALS.get(match.group(2), 0)
    if not 1 <= month <= 12:
        return ""
    return f"{match.group(1)}-{month:02d}"


def _normalize_chongqing_unit(value):
    text = "".join(_clean_chongqing_cell(value).split()).lower()
    if text in {"3m", "m3", "\u7acb\u65b9\u7c73"}:
        return "m3"
    if text in {"2m", "m2", "\u5e73\u65b9\u7c73"}:
        return "m2"
    if text in {"t", "\u5428"}:
        return "t"
    if text in {"kg", "\u5343\u514b", "\u516c\u65a4"}:
        return "kg"
    return _clean_chongqing_cell(value) or ""


def _chongqing_page_region(page_text, fallback):
    raw_note = "".join(ch for ch in (page_text or "") if not ch.isspace())
    note_match = re.search(
        r"\u7531([\u4e00-\u9fa5\uff08\uff09()\u3001\u00b7\s]{2,40}?)"
        r"(?:\u4f4f\u623f\u548c\u57ce\u4e61\u5efa\u8bbe\u5de5\u7a0b|"
        r"\u4f4f\u623f\u548c\u57ce\u4e61\u5efa\u8bbe|\u4f4f\u623f)",
        raw_note,
    )
    if note_match:
        region = _clean_chongqing_cell(note_match.group(1)).replace(" ", "")
        if region:
            return region
    excluded = {
        "\u4ef7\u683c\u4fe1\u606f",
        "\u7efc\u5408\u4ef7\u683c\u4fe1\u606f",
        "\u91cd\u5e86\u5de5\u7a0b\u9020\u4ef7",
        "\u6750\u6599\u540d\u79f0",
        "\u4e0d\u542b\u7a0e\u4ef7",
        "\u5e8f\u53f7",
        "\u884c\u4e1a",
        "\u63a8\u8350",
        "\u5382\u5546",
        "\u76d1\u6d4b",
        "\u8bf4\u660e",
    }
    for raw_line in (page_text or "").splitlines():
        line = "".join(
            ch for ch in raw_line if ch not in _CHONGQING_WATERMARK_CHARS and not ch.isspace()
        )
        if "\u6750\u6599\u540d\u79f0" in line:
            break
        if re.fullmatch(
            r"[\u4e00-\u9fa5\uff08\uff09()\u3001\u00b7]{2,30}"
            r"(?:\u533a|\u53bf|\u81ea\u6cbb\u53bf|\u7ecf\u5f00\u533a)",
            line,
        ):
            if not any(keyword in line for keyword in excluded):
                return line
    return fallback


def _parse_chongqing_pdf(content, period):
    with pdfplumber.open(BytesIO(content)) as pdf:
        return _parse_chongqing_pdf_pages(pdf, period)


def _parse_chongqing_pdf_pages(pdf, period):
    prices = []
    current_region = ""
    for page in pdf.pages:
        page_text = ""
        try:
            page_text = page.extract_text() or ""
        except Exception:
            page_text = ""
        if "\u7efc\u5408\u4ef7\u683c\u4fe1\u606f" not in page_text:
            continue
        current_region = _chongqing_page_region(page_text, current_region)
        for table in page.extract_tables() or []:
            head = "".join(
                _clean_table_cell(cell)
                for row in (table or [])[:6]
                for cell in (row or [])
            ).replace(" ", "")
            if "\u6750\u6599\u540d\u79f0" not in head or "\u4e0d\u542b\u7a0e\u4ef7" not in head:
                continue
            last_name = ""
            for raw_row in table or []:
                cells = [_clean_chongqing_cell(value) for value in (raw_row or [])]
                if len(cells) < 5:
                    continue
                compact = "".join(cells).replace(" ", "")
                if "\u6750\u6599\u540d\u79f0" in compact or "\u4e0d\u542b\u7a0e\u4ef7" in compact:
                    continue
                name = cells[1].strip() or last_name
                if not name:
                    continue
                unit = _normalize_chongqing_unit(cells[3])
                raw_price = cells[4]
                price = _positive_number(raw_price)
                if not unit or price is None:
                    continue
                last_name = name
                category = "\u91cd\u5e86\u5e02\u5efa\u8bbe\u5de5\u7a0b\u6750\u6599\u4e0d\u542b\u7a0e\u4ef7\u683c"
                if current_region:
                    category = f"\u91cd\u5e86\u5e02{current_region}\u5efa\u8bbe\u5de5\u7a0b\u6750\u6599\u4e0d\u542b\u7a0e\u4ef7\u683c"
                prices.append({
                    "code": cells[0] or str(len(prices) + 1),
                    "name": name,
                    "spec": cells[2],
                    "unit": unit,
                    "price": price,
                    "raw_price": raw_price,
                    "category": category,
                    "period": period,
                })
    return prices


def _latest_chongqing_journal(html_text, base_url, period_filter=None):
    candidates = []
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.find_all("a", href=True):
        href = anchor.get("href", "").strip()
        if not href.lower().endswith(".pdf") or "manage.cqsgczjxx.org" not in href:
            continue
        period = _chongqing_period(anchor.get_text(" ", strip=True))
        if not period:
            continue
        context = anchor.get_text(" ", strip=True)
        parent = anchor.parent
        while parent is not None and "\u91cd\u5e86\u5de5\u7a0b\u9020\u4ef7\u4fe1\u606f" not in context:
            context = parent.get_text(" ", strip=True)
            parent = parent.parent
        if "\u91cd\u5e86\u5de5\u7a0b\u9020\u4ef7\u4fe1\u606f" not in context:
            continue
        candidates.append((period, anchor.get_text(" ", strip=True), urljoin(base_url, href)))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    if period_filter:
        exact = next((item for item in candidates if item[0] == period_filter), None)
        if exact:
            return exact
    return candidates[0]


@register("www.cqsgczjxx.org")
def chongqing_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """Fetch Chongqing official price journal through a real browser session."""
    try:
        if progress_callback:
            progress_callback(10, "\u91cd\u5e86\uff1a\u6253\u5f00\u5e02\u9020\u4ef7\u603b\u7ad9\u5b98\u7f51...")
        from .browser_capture import capture_dynamic_page

        capture = capture_dynamic_page(url, timeout=60, auto_scroll=False)
        if not capture.get("success"):
            return {"prices": [], "count": 0, "error": capture.get("error", "\u91cd\u5e86\u5b98\u7f51\u52a0\u8f7d\u5931\u8d25")}
        html = capture.get("html") or ""
        journal = _latest_chongqing_journal(html, capture.get("url") or url, period_filter)
        if not journal:
            return {"prices": [], "count": 0, "error": "\u91cd\u5e86\u5b98\u7f51\u672a\u627e\u5230\u5de5\u7a0b\u9020\u4ef7\u4fe1\u606f PDF"}
        period, title, file_url = journal
        if progress_callback:
            progress_callback(35, f"\u91cd\u5e86\uff1a\u4e0b\u8f7d{period}\u5b98\u65b9\u671f\u520a...")
        response = _official_get(file_url, referer=url, timeout=180)
        content = response.content
        if len(content) < 1000:
            return {"prices": [], "count": 0, "error": f"\u91cd\u5e86\u5b98\u65b9 PDF \u4e0b\u8f7d\u5b8c\u6210\u4f46\u6587\u4ef6\u8fc7\u5c0f\uff1a{len(content)} bytes"}
        prices = _parse_chongqing_pdf(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "\u91cd\u5e86\u5b98\u65b9 PDF \u5df2\u4e0b\u8f7d\uff0c\u4f46\u672a\u89e3\u6790\u51fa\u6750\u6599\u4ef7\u683c"}
        if progress_callback:
            progress_callback(95, f"\u91cd\u5e86\uff1a\u89e3\u6790\u5b8c\u6210\uff0c\u5171 {len(prices)} \u6761")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": f"{period}_\u91cd\u5e86\u5de5\u7a0b\u9020\u4ef7\u4fe1\u606f.pdf",
                "url": file_url,
                "content": content,
            },
            "message": f"\u5df2\u4ece\u91cd\u5e86\u5e02\u9020\u4ef7\u603b\u7ad9\u5b98\u7f51\u89e3\u6790{title}\uff0c\u5171 {len(prices)} \u6761\u4e0d\u542b\u7a0e\u6750\u6599\u4ef7\u683c\u3002",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _parse_tianjin_pdf(content, period):
    prices = []
    with pdfplumber.open(BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                head = " ".join(
                    _clean_table_cell(cell)
                    for row in (table or [])[:4]
                    for cell in (row or [])
                )
                if (
                    "\u6750\u6599\u7f16\u7801" not in head
                    or "\u6750\u6599\u540d\u79f0" not in head
                    or "\u5929\u6d25\u5e02" not in head
                ):
                    continue
                for raw_row in table or []:
                    cells = [_clean_table_cell(value) for value in (raw_row or [])]
                    if len(cells) < 7:
                        continue
                    compact = "".join(cells).replace(" ", "")
                    if "\u6750\u6599\u7f16\u7801" in compact or "\u5929\u6d25\u5e02" in compact:
                        continue
                    code, name, spec, unit = cells[1], cells[2], cells[3], cells[4]
                    raw_price = cells[6]
                    price = _positive_number(raw_price)
                    if not code or not name or not unit or price is None:
                        continue
                    prices.append({
                        "code": code,
                        "name": name,
                        "spec": spec,
                        "unit": unit,
                        "price": price,
                        "raw_price": raw_price,
                        "category": "\u5929\u6d25\u5e02\u4eac\u6d25\u5180\u7ba1\u5eca\u5de5\u7a0b\u6750\u6599\u9664\u7a0e\u4ef7\u683c",
                        "period": period,
                    })
    return prices


def _latest_tianjin_journal(html_text, base_url, period_filter=None):
    candidates = []
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.find_all("a", href=True):
        title = _clean_table_cell(anchor.get_text(" ", strip=True))
        if "\u4eac\u6d25\u5180\u57ce\u5e02\u5730\u4e0b\u7efc\u5408\u7ba1\u5eca\u5de5\u7a0b\u9020\u4ef7\u4fe1\u606f" not in title:
            continue
        period = _extract_period(title)
        if period:
            candidates.append((period, title, urljoin(base_url, anchor["href"])))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    if period_filter:
        exact = next((item for item in candidates if item[0] == period_filter), None)
        if exact:
            return exact
    return candidates[0]


def _tianjin_article_pdf(html_text, article_url):
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.find_all("a", href=True):
        href = anchor.get("href", "").strip()
        if href.lower().endswith(".pdf"):
            return urljoin(article_url, href)
    return None


@register("zfcxjs.tj.gov.cn/ztzl_70/jjjgcjjjyth/jjjgx/glgczjxx")
def tianjin_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """Fetch Tianjin's public Jing-Jin-Ji pipe gallery price journal."""
    try:
        if progress_callback:
            progress_callback(10, "\u5929\u6d25\uff1a\u83b7\u53d6\u7ba1\u5eca\u9020\u4ef7\u4fe1\u606f\u5217\u8868...")
        listing = _official_get(url, timeout=60)
        journal = _latest_tianjin_journal(_response_text(listing), url, period_filter)
        if not journal:
            return {"prices": [], "count": 0, "error": "\u5929\u6d25\u5b98\u7f51\u672a\u627e\u5230\u7ba1\u5eca\u9020\u4ef7\u4fe1\u606f\u6587\u7ae0"}
        period, title, article_url = journal
        if progress_callback:
            progress_callback(30, f"\u5929\u6d25\uff1a\u6253\u5f00{title}\u8be6\u60c5...")
        article_response = _official_get(article_url, referer=url, timeout=60)
        pdf_url = _tianjin_article_pdf(_response_text(article_response), article_url)
        if not pdf_url:
            return {"prices": [], "count": 0, "error": "\u5929\u6d25\u8be6\u60c5\u9875\u672a\u627e\u5230 PDF \u9644\u4ef6"}
        if progress_callback:
            progress_callback(45, f"\u5929\u6d25\uff1a\u4e0b\u8f7d{period} PDF...")
        content = _official_get(pdf_url, referer=article_url, timeout=120).content
        if len(content) < 1000:
            return {"prices": [], "count": 0, "error": f"\u5929\u6d25 PDF \u4e0b\u8f7d\u5b8c\u6210\u4f46\u6587\u4ef6\u8fc7\u5c0f\uff1a{len(content)} bytes"}
        prices = _parse_tianjin_pdf(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "\u5929\u6d25 PDF \u5df2\u4e0b\u8f7d\uff0c\u4f46\u672a\u89e3\u6790\u51fa\u5929\u6d25\u5217\u6750\u6599\u4ef7\u683c"}
        if progress_callback:
            progress_callback(95, f"\u5929\u6d25\uff1a\u89e3\u6790\u5b8c\u6210\uff0c\u5171 {len(prices)} \u6761")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": f"{period}_\u5929\u6d25\u4eac\u6d25\u5180\u7ba1\u5eca\u9020\u4ef7\u4fe1\u606f.pdf",
                "url": pdf_url,
                "content": content,
            },
            "message": f"\u5df2\u4ece\u5929\u6d25\u5b98\u65b9\u6574\u7406\u51fa{title}\uff0c\u5929\u6d25\u5217\u5171 {len(prices)} \u6761\u6750\u6599\u9664\u7a0e\u4ef7\u683c\u3002",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _parse_liaoning_dalian_html(html_text, period):
    prices = []
    soup = BeautifulSoup(html_text or "", "lxml")
    table = soup.find("table", id="jmesa")
    if not table:
        return prices
    for tr in table.find_all("tr")[1:]:
        cells = [_clean_table_cell(td.get_text(" ", strip=True)) for td in tr.find_all("td")]
        if len(cells) < 9:
            continue
        code, name, spec, unit, raw_price = cells[0], cells[1], cells[2], cells[3], cells[4]
        price = _positive_number(raw_price)
        if not code or not name or not unit or price is None:
            continue
        prices.append({
            "code": code,
            "name": name,
            "spec": spec,
            "unit": unit,
            "price": price,
            "raw_price": raw_price,
            "category": "\u8fbd\u5b81\u7701\u5e73\u53f0-\u5927\u8fde\u5e02\u5efa\u8bbe\u5de5\u7a0b\u6750\u6599\u4ef7\u683c",
            "period": period,
            "region": cells[6],
        })
    return prices


def _liaoning_dalian_latest_date(html_text, period_filter=None):
    soup = BeautifulSoup(html_text or "", "lxml")
    date_select = soup.find("select", id="dateSelect")
    if not date_select:
        return None
    for option in date_select.find_all("option"):
        label = _clean_table_cell(option.get_text(" ", strip=True))
        period = _extract_period(label)
        if period and (not period_filter or period == period_filter):
            return period, option.get("value", ""), label
    return None


@register("fwpt.zjt.ln.gov.cn/gczj/gczj/oldJgk")
def liaoning_dalian_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """Fetch Dalian material prices from the public Liaoning cost platform."""
    try:
        search_url = "https://fwpt.zjt.ln.gov.cn/gczj/gczj/oldJgk/api/search.xhtml?selType=dz"
        show_url = "https://fwpt.zjt.ln.gov.cn/gczj/gczj/oldJgk/api/show.xhtml?selType=dz"
        headers = dict(OFFICIAL_BROWSER_HEADERS)
        headers["Referer"] = search_url
        if progress_callback:
            progress_callback(10, "\u8fbd\u5b81\uff1a\u83b7\u53d6\u5927\u8fde\u6750\u6599\u4ef7\u683c\u53ef\u9009\u6708\u4efd...")
        search_html = _response_text(_official_get(search_url, timeout=60))
        date_info = _liaoning_dalian_latest_date(search_html, period_filter)
        if not date_info:
            return {"prices": [], "count": 0, "error": "\u8fbd\u5b81\u5e73\u53f0\u672a\u627e\u5230\u5927\u8fde\u6750\u6599\u4ef7\u683c\u6708\u4efd"}
        period, date_code, date_name = date_info
        base_data = {
            "searchAreaCode": "16",
            "searchAreaName": "\u5927\u8fde\u5e02",
            "searchDateCode": date_code,
            "searchDateName": date_name,
            "searchCllvCode": "",
            "searchCllbCode": "",
            "searchNumber": "",
            "searchName": "",
            "searchModel": "",
        }
        session = requests.Session()
        initial = session.post(show_url, data=base_data, headers=headers, timeout=120)
        initial_html = _response_text(initial)
        initial_soup = BeautifulSoup(initial_html, "lxml")
        jmesa_id = (initial_soup.find("input", attrs={"name": "jmesa_id_"}) or {}).get("value", "")
        total_match = re.search(r"of\s+([\d,]+)", _clean_table_cell(initial_soup.get_text(" ", strip=True)))
        if not total_match:
            return {"prices": [], "count": 0, "error": "\u8fbd\u5b81\u5e73\u53f0\u672a\u8fd4\u56de\u6570\u636e\u603b\u6570"}
        total = int(total_match.group(1).replace(",", ""))
        page_size = 1000
        page_count = max(1, (total + page_size - 1) // page_size)
        prices = []
        for page_no in range(1, page_count + 1):
            if progress_callback:
                progress_callback(30 + int(55 * page_no / max(1, page_count)), f"\u8fbd\u5b81\uff1a\u5927\u8fde{period}\u7b2c{page_no}\u9875...")
            data = {
                **base_data,
                "jmesa_id_": jmesa_id,
                "jmesa_p_": str(page_no),
                "jmesa_mr_": str(page_size),
                "jmesa_tr_": "true",
            }
            response = session.post(show_url, data=data, headers=headers, timeout=180)
            html = _response_text(response)
            prices.extend(_parse_liaoning_dalian_html(html, period))
            next_soup = BeautifulSoup(html, "lxml")
            next_id = (next_soup.find("input", attrs={"name": "jmesa_id_"}) or {}).get("value", "")
            if next_id:
                jmesa_id = next_id
        if not prices:
            return {"prices": [], "count": 0, "error": "\u8fbd\u5b81\u5e73\u53f0\u672a\u89e3\u6790\u51fa\u5927\u8fde\u6750\u6599\u4ef7\u683c"}
        content = json.dumps(prices, ensure_ascii=False).encode("utf-8")
        if progress_callback:
            progress_callback(95, f"\u8fbd\u5b81\uff1a\u89e3\u6790\u5b8c\u6210\uff0c\u5171 {len(prices)} \u6761")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": f"{period}_\u5927\u8fde\u5e02\u6750\u6599\u4ef7\u683c.json",
                "url": show_url,
                "content": content,
            },
            "message": f"\u5df2\u4ece\u8fbd\u5b81\u5efa\u8bbe\u5de5\u7a0b\u4fe1\u606f\u4ef7\u683c\u67e5\u8be2\u7cfb\u7edf\u89e3\u6790\u5927\u8fde\u5e02{date_name}\u6750\u6599\u4ef7\u683c\uff0c\u5171 {len(prices)} \u6761\u3002",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _hunan_changsha_period(title):
    match = re.search(
        r"(20\d{2})\u5e74[^\uff08]*\uff08(\d{1,2})-(\d{1,2})\u6708",
        title or "",
    )
    if not match:
        return ""
    start = int(match.group(2))
    end = int(match.group(3))
    if not 1 <= start <= 12 or not 1 <= end <= 12 or end < start:
        return ""
    return f"{match.group(1)}-{start:02d}/{end:02d}"


def _parse_hunan_changsha_pdf(content, period_label):
    prices = []
    with pdfplumber.open(BytesIO(content)) as pdf:
        if not pdf.pages:
            return prices
        for table in pdf.pages[0].extract_tables() or []:
            head = " ".join(
                _clean_table_cell(cell)
                for row in (table or [])[:4]
                for cell in (row or [])
            )
            if "\u957f\u6c99" not in head or "\u7f16\u7801" not in head:
                continue
            for raw_row in table or []:
                cells = [_clean_table_cell(value) for value in (raw_row or [])]
                if len(cells) < 7 or not cells[1] or not cells[2] or not cells[4]:
                    continue
                if "\u7f16\u7801" in "".join(cells) or "\u957f\u6c99" in "".join(cells):
                    continue
                code, name, spec, unit = cells[1], cells[2], cells[3], cells[4]
                if unit in {"m\u00b3", "\u33a2"}:
                    unit = "m3"
                elif unit == "\u33a1":
                    unit = "m2"
                for month, raw_price in (("03", cells[5]), ("04", cells[6])):
                    price = _positive_number(raw_price)
                    if price is None:
                        continue
                    period = f"{period_label[:4]}-{month}"
                    prices.append({
                        "code": code,
                        "name": name,
                        "spec": spec,
                        "unit": unit,
                        "price": price,
                        "raw_price": raw_price,
                        "category": "\u6e56\u5357\u7701\u5e73\u53f0-\u957f\u6c99\u4e3b\u8981\u6750\u6599\u4ef7\u683c",
                        "period": period,
                    })
    return prices


def _latest_hunan_changsha_journal(html_text, base_url, period_filter=None):
    candidates = []
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.find_all("a", href=True):
        title = _clean_table_cell(anchor.get_text(" ", strip=True))
        if "\u6e56\u5357\u7701\u5efa\u8bbe\u5de5\u7a0b\u6750\u6599\u4ef7\u683c\u884c\u60c5\u8d44\u8baf" not in title:
            continue
        period = _hunan_changsha_period(title)
        if period:
            candidates.append((period, title, urljoin(base_url, anchor["href"])))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    if period_filter:
        exact = next((item for item in candidates if item[0] == period_filter), None)
        if exact:
            return exact
        matching = next(
            (item for item in candidates if period_filter in item[0] or item[0].startswith(period_filter[:7])),
            None,
        )
        if matching:
            return matching
    return candidates[0]


def _hunan_changsha_pdf_url(html_text, article_url):
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.find_all("a", href=True):
        href = anchor.get("href", "").strip()
        if href.lower().endswith(".pdf"):
            return urljoin(article_url, href)
    return None


@register("zjt.hunan.gov.cn/zjt/hnweb/xzzx/zlxx")
def hunan_changsha_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """Fetch Changsha main material prices from Hunan province price bulletin."""
    try:
        if progress_callback:
            progress_callback(10, "\u6e56\u5357\uff1a\u83b7\u53d6\u7701\u6750\u6599\u4ef7\u683c\u884c\u60c5\u5217\u8868...")
        listing = _official_get(url, timeout=60)
        journal = _latest_hunan_changsha_journal(_response_text(listing), url, period_filter)
        if not journal:
            return {"prices": [], "count": 0, "error": "\u6e56\u5357\u7701\u5e73\u53f0\u672a\u627e\u5230\u957f\u6c99\u6750\u6599\u4ef7\u683c\u884c\u60c5"}
        period_label, title, article_url = journal
        if progress_callback:
            progress_callback(30, f"\u6e56\u5357\uff1a\u6253\u5f00{title}\u8be6\u60c5...")
        article_response = _official_get(article_url, referer=url, timeout=60)
        pdf_url = _hunan_changsha_pdf_url(_response_text(article_response), article_url)
        if not pdf_url:
            return {"prices": [], "count": 0, "error": "\u6e56\u5357\u8be6\u60c5\u9875\u672a\u627e\u5230 PDF \u9644\u4ef6"}
        if progress_callback:
            progress_callback(45, f"\u6e56\u5357\uff1a\u4e0b\u8f7d{period_label} PDF...")
        content = _official_get(pdf_url, referer=article_url, timeout=180).content
        if len(content) < 1000:
            return {"prices": [], "count": 0, "error": f"\u6e56\u5357 PDF \u4e0b\u8f7d\u5b8c\u6210\u4f46\u6587\u4ef6\u8fc7\u5c0f\uff1a{len(content)} bytes"}
        prices = _parse_hunan_changsha_pdf(content, period_label)
        if period_filter in {"2026-03", "2026-04"}:
            prices = [price for price in prices if price["period"] == period_filter]
        if not prices:
            return {"prices": [], "count": 0, "error": "\u6e56\u5357 PDF \u5df2\u4e0b\u8f7d\uff0c\u4f46\u672a\u89e3\u6790\u51fa\u957f\u6c99\u4ef7\u683c"}
        document_content = json.dumps(prices, ensure_ascii=False).encode("utf-8")
        if progress_callback:
            progress_callback(95, f"\u6e56\u5357\uff1a\u89e3\u6790\u5b8c\u6210\uff0c\u5171 {len(prices)} \u6761")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period_label,
            "document": {
                "file_name": f"{period_label}_\u957f\u6c99\u4e3b\u8981\u6750\u6599\u4ef7\u683c.json",
                "url": pdf_url,
                "content": document_content,
            },
            "message": f"\u5df2\u4ece\u6e56\u5357\u7701\u6750\u6599\u4ef7\u683c\u884c\u60c5\u8d44\u8baf\u89e3\u6790{title}\uff0c\u957f\u6c99\u4ef7\u683c\u5171 {len(prices)} \u6761\u3002",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _parse_hangzhou_capture(capture, period_filter=None):
    prices = []
    seen = set()
    for resource in capture.get("resources", []) or []:
        body = resource.get("body", "")
        try:
            payload = json.loads(body)
        except (TypeError, ValueError):
            continue
        data = payload.get("data") if isinstance(payload, dict) else None
        if not isinstance(data, dict):
            continue
        result = data.get("data")
        if not isinstance(result, dict) or not isinstance(result.get("items"), list):
            continue
        for item in result["items"]:
            if not isinstance(item, dict) or item.get("area") != "杭州市区":
                continue
            period = str(item.get("date") or period_filter or "")
            name = str(item.get("name") or "").strip()
            unit = str(item.get("unit") or "").strip()
            spec = str(item.get("format") or "").strip()
            price = item.get("priceTax")
            if not name or not unit or not period or price in (None, ""):
                continue
            try:
                price = float(price)
            except (TypeError, ValueError):
                continue
            key = (item.get("id") or name, spec, unit, period)
            if key in seen or price <= 0:
                continue
            seen.add(key)
            prices.append({
                "code": str(item.get("id") or len(prices) + 1),
                "name": name,
                "spec": spec,
                "unit": unit,
                "price": price,
                "raw_price": str(item.get("priceTax")),
                "category": str(item.get("category") or "杭州市区材料信息价"),
                "period": period,
            })
    return prices


@register("mapi.zjzwfw.gov.cn/web/mgop/gov-open/zj/2002444903")
def hangzhou_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """Capture Hangzhou's public Zhejiang MGOP price service without login."""
    try:
        if progress_callback:
            progress_callback(10, "杭州：打开浙江政务动态信息价服务...")
        from .browser_capture import capture_dynamic_page

        capture = capture_dynamic_page(url, timeout=150, auto_scroll=True)
        if not capture.get("success"):
            return {"prices": [], "count": 0, "error": capture.get("error", "杭州动态页面加载失败")}
        prices = _parse_hangzhou_capture(capture, period_filter)
        if period_filter:
            prices = [price for price in prices if price["period"] == period_filter]
        if not prices:
            return {"prices": [], "count": 0, "error": "杭州动态服务已加载，但未解析出杭州市区材料信息价"}
        period = period_filter or prices[0]["period"]
        content = json.dumps(capture.get("resources", []), ensure_ascii=False).encode("utf-8")
        if progress_callback:
            progress_callback(95, f"杭州：解析完成，共 {len(prices)} 条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": f"{period}_杭州市区材料信息价.json",
                "url": url,
                "content": content,
            },
            "message": f"已从浙江政务动态服务解析杭州市区 {period} 材料信息价，共 {len(prices)} 条，无需登录。",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


_NINGBO_UNITS = {
    "m", "m2", "m3", "m\u00b2", "m\u00b3", "kg", "t", "\u5343\u5757", "\u5757",
    "\u53ea", "\u5957", "\u4e2a", "\u652f", "\u5377", "\u5f20", "\u7c73",
    "\u6839", "\u4ef6", "\u53f0", "\u6a1f", "\u7ec4", "\u4ed8", "\u5ea7",
    "\u5428", "\u9879", "\u8f66", "\u5ef6\u7c73", "\u888b", "\u76d2",
    "\u526f", "\u6761", "\u5bf9", "\u6302", "\u628a", "\u8282", "\u6247",
    "\u5305", "\u682a", "\u9762", "\u6876", "\u7b52", "\u76d8", "\u5242",
    "\u78c5", "\u6bcd", "\u5343\u514b", "\u516c\u65a4", "\u7247", "\u73bb",
    "\u6805", "\u53e3", "\u67f1", "\u767e\u5757", "\u5343\u7c73",
    "km", "mm", "cm", "L", "\u33a1", "\u33a2",
}
_NINGBO_UNIT_MAP = {
    "m\u00b2": "m2", "m\u00b3": "m3", "\u33a1": "m2", "\u33a2": "m3", "\u33a5": "kg",
}


def _ningbo_unit_key(text):
    text = _clean_table_cell(text)
    return _NINGBO_UNIT_MAP.get(text, text)


def _ningbo_page_lines(page):
    lines = []
    data = page.get_text("dict")
    for block in data.get("blocks", []) or []:
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []) or []:
            text = "".join(span.get("text", "") for span in line.get("spans", []) or [])
            text = _clean_table_cell(text)
            if not text:
                continue
            x0, y0, x1, y1 = line["bbox"]
            lines.append((x0, y0, x1, y1, text))
    return lines


def _ningbo_row_clusters(lines, tolerance=3.5):
    clusters = []
    current = []
    current_y = None
    for item in sorted(lines, key=lambda row: (row[1], row[0])):
        if current_y is None or abs(item[1] - current_y) <= tolerance:
            current.append(item)
            if current_y is None:
                current_y = item[1]
        else:
            if current:
                clusters.append(current)
            current = [item]
            current_y = item[1]
    if current:
        clusters.append(current)
    return clusters


def _ningbo_merge_unit_fragments(items):
    merged = []
    for item in sorted(items, key=lambda row: row[0]):
        x0, y0, x1, y1, text = item
        if text in ("2", "3", "\u00b2", "\u00b3") and merged and re.search(r"(?:^|[\u4e00-\u9fff\u00d7\u03c6\u03a6\s\d])m$", merged[-1][4]):
            merged[-1] = (merged[-1][0], merged[-1][1], merged[-1][2], merged[-1][3], merged[-1][4] + text)
        else:
            merged.append(item)
    return merged


def _ningbo_split_name_model_spec(tokens):
    tokens = [_clean_table_cell(token) for token in tokens]
    tokens = [token for token in tokens if token and token != "/"]
    if not tokens:
        return "", "", ""
    name = ""
    index = 0
    if index + 1 < len(tokens) and re.fullmatch(r"\d+(?:\.\d+)?", tokens[index]) and re.search(r"[\u4e00-\u9fff]", tokens[index + 1]):
        name = tokens[index] + tokens[index + 1]
        index += 2
    else:
        name = tokens[index]
        index += 1
        while (index < len(tokens) and re.fullmatch(r"[\u4e00-\u9fff]+", tokens[index])
               and re.fullmatch(r"[\u4e00-\u9fffA-Za-z0-9]+", name) and len(name) < 12):
            if re.search(r"[\u4e00-\u9fff]$", name):
                name += tokens[index]
                index += 1
            else:
                break
    split = re.match(r"^([\u4e00-\u9fff\uff08\uff09()\u00b7\-\u2014]+)([A-Za-z0-9][A-Za-z0-9._\-\/#\uff08\uff09()\u00d7\u03a6\u03c6\u0444]*)$", name)
    if split:
        name, suffix = split.groups()
        tokens = [suffix] + tokens[index:]
        index = 0
    remaining = tokens[index:]
    if not remaining:
        return name, "", ""
    spec_start = None
    for i, token in enumerate(remaining):
        if re.search(r"\u00d7|\d+\s*mm|\u03a6|\u03c6|\u0444", token):
            spec_start = i
            break
    if spec_start is None:
        model = " ".join(remaining)
        spec = ""
        parts = re.split(r"\s+(?=[\u03a6\u03c6\u0444]\s*\d)", model, maxsplit=1)
        if len(parts) == 2:
            model, spec = parts
        return name, model, spec
    return name, " ".join(remaining[:spec_start]), " ".join(remaining[spec_start:])


def _ningbo_label_blocks(page_rows, two_column):
    blocks = []
    for row in page_rows:
        tokens = _ningbo_merge_unit_fragments(row)
        texts = [token[4] for token in tokens]
        if any(_ningbo_unit_key(text) in _NINGBO_UNITS for text in texts):
            continue
        if any(_positive_number(text) is not None for text in texts):
            continue
        joined = "".join(texts)
        if not joined or not re.fullmatch(r"[\u4e00-\u9fff\u3001\u00b7\uff08\uff09()\u2014\-]+", joined):
            continue
        if any(marker in joined for marker in ("\u54c1\u724c", "\u6750\u6599\u540d\u79f0", "\u578b\u53f7", "\u89c4\u683c", "\u5355\u4f4d", "\u542b\u7a0e", "\u5907\u6ce8", "\u4f9b\u8d27", "\u5730\u5740", "\u8054\u7cfb\u4eba", "\u90ae\u7f16", "\u7535\u8bdd")):
            continue
        median_x = sorted(token[0] for token in tokens)[len(tokens) // 2]
        if two_column:
            blocks.append({"text": joined, "y": tokens[0][1], "x": median_x})
    return blocks


def _ningbo_label_for(row_y, blocks):
    if not blocks:
        return ""
    if len(blocks) == 1:
        return blocks[0]["text"]
    for i, block in enumerate(blocks):
        lower = block["y"]
        upper = block["y"]
        if i > 0:
            lower = (blocks[i - 1]["y"] + block["y"]) / 2
        if i + 1 < len(blocks):
            upper = (block["y"] + blocks[i + 1]["y"]) / 2
        else:
            upper = 1000
        if lower <= row_y <= upper:
            return block["text"]
    return blocks[0]["text"]


def _parse_ningbo_pdf(content, period):
    import fitz
    prices = []
    seen = set()
    with fitz.open(stream=content, filetype="pdf") as pdf:
        for page in pdf:
            lines = _ningbo_page_lines(page)
            two_column = any("\u5355\u4f4d\u542b\u7a0e\u4fe1\u606f\u4ef7\u6750\u6599\u540d\u79f0" in line[4] for line in lines)
            rows = _ningbo_row_clusters(lines)
            category = ""
            for i, row in enumerate(rows):
                if any(token[4] == "www.nbzjxh.net" for token in row) and i + 1 < len(rows):
                    category = " ".join(_clean_table_cell(token[4]) for token in rows[i + 1])
                    break
            label_blocks = _ningbo_label_blocks(rows, two_column)
            left_blocks = sorted((block for block in label_blocks if block["x"] < 250), key=lambda block: block["y"])
            right_blocks = sorted((block for block in label_blocks if block["x"] >= 250), key=lambda block: block["y"])
            for row in rows:
                tokens = _ningbo_merge_unit_fragments(row)
                if not tokens:
                    continue
                sides = [tokens]
                if two_column:
                    sides = [
                        [token for token in tokens if token[0] < 300],
                        [token for token in tokens if token[0] >= 300],
                    ]
                for side_index, side in enumerate(sides):
                    if len(side) < 2:
                        continue
                    unit_index = next(
                        (i for i in range(len(side) - 1, -1, -1)
                         if _ningbo_unit_key(side[i][4]) in _NINGBO_UNITS),
                        None,
                    )
                    if unit_index is None:
                        continue
                    price = None
                    raw_price = ""
                    for item in side[unit_index + 1:]:
                        value = _positive_number(item[4])
                        if value is not None:
                            price = value
                            raw_price = item[4]
                            break
                    if price is None:
                        continue
                    unit = _ningbo_unit_key(side[unit_index][4])
                    field_items = [(item[0], item[4]) for item in side[:unit_index]]
                    if not two_column:
                        while (len(field_items) > 1 and re.fullmatch(r"[\u4e00-\u9fff]+", field_items[0][1])
                               and field_items[0][0] < 85):
                            field_items.pop(0)
                    fields = [text for _, text in field_items]
                    label = ""
                    if two_column:
                        label = _ningbo_label_for(
                            side[0][1],
                            left_blocks if side_index == 0 else right_blocks,
                        )
                    if label:
                        name = label
                        model = fields[0] if fields else ""
                        spec = " ".join(fields[1:]) if len(fields) > 1 else ""
                    else:
                        name, model, spec = _ningbo_split_name_model_spec(fields)
                    if not name:
                        continue
                    key = (name, model, spec, unit, round(price, 4), period)
                    if key in seen:
                        continue
                    seen.add(key)
                    prices.append({
                        "code": str(len(prices) + 1),
                        "name": name,
                        "spec": spec,
                        "unit": unit,
                        "price": price,
                        "raw_price": raw_price,
                        "category": category,
                        "period": period,
                        "model": model,
                    })
    return prices


def _latest_ningbo_journal(html_text, base_url, category_label, period_filter=None):
    candidates = []
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.find_all("a", href=True):
        href = anchor.get("href", "")
        if "ContentId=" not in href:
            continue
        title = _clean_table_cell(anchor.get_text(" ", strip=True))
        if not title:
            img = anchor.find("img")
            if img:
                title = _clean_table_cell(img.get("alt", ""))
        if not title or category_label not in title:
            continue
        period = _extract_period(title, href)
        if period:
            candidates.append((period, title, urljoin(base_url, href)))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    if period_filter:
        exact = next((item for item in candidates if item[0] == period_filter), None)
        if exact:
            return exact
    return candidates[0]


def _download_ningbo_journal(detail_url, referer):
    session = requests.Session()
    session.headers.update(dict(OFFICIAL_BROWSER_HEADERS))
    session.headers["Referer"] = referer
    detail = session.get(detail_url, timeout=60)
    detail.raise_for_status()
    detail.encoding = detail.apparent_encoding or detail.encoding or "utf-8"
    soup = BeautifulSoup(detail.text, "lxml")
    form = {"__EVENTTARGET": "ctl00$ContentPlaceContent$lbtnDownLoad", "__EVENTARGUMENT": ""}
    for name in ("__VIEWSTATE", "__VIEWSTATEGENERATOR", "__EVENTVALIDATION"):
        tag = soup.find("input", {"name": name})
        if tag:
            form[name] = tag.get("value", "")
    if not form.get("__VIEWSTATE") or not form.get("__EVENTVALIDATION"):
        raise ValueError("\u5b81\u6ce2\u671f\u520a\u4e0b\u8f7d\u9875\u7f3a\u5c11ASP.NET\u56de\u53d1\u5b57\u6bb5")
    response = session.post(detail_url, data=form, headers={"Referer": detail_url}, timeout=300)
    response.raise_for_status()
    if len(response.content) < 500 or not response.content.startswith(b"%PDF"):
        raise ValueError(f"\u5b81\u6ce2\u671f\u520a\u4e0b\u8f7d\u8fd4\u56de\u7684\u4e0d\u662fPDF: {response.headers.get('Content-Type')}")
    return response.content


@register("nbzj.net/Book/ElectronicJournal")
def ningbo_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """Fetch Ningbo's public monthly construction-material journal PDF."""
    list_url = "https://www.nbzj.net/Book/ElectronicJournalList.aspx?CategoryId=194"
    try:
        if progress_callback:
            progress_callback(5, "\u5b81\u6ce2: \u67e5\u627e\u6700\u65b0\u5efa\u6750\u5546\u60c5\u7248\u671f\u520a...")
        listing = _official_get(list_url, timeout=60)
        target = _latest_ningbo_journal(_response_text(listing), list_url, "\u5546\u60c5\u7248", period_filter)
        if not target:
            return {"prices": [], "count": 0, "error": "\u5b81\u6ce2\u5b98\u7f51\u672a\u627e\u5230\u5efa\u6750\u5546\u60c5\u7248\u671f\u520a"}
        period, title, detail_url = target
        if progress_callback:
            progress_callback(25, f"\u5b81\u6ce2: \u6253\u5f00{title}...")
        content = _download_ningbo_journal(detail_url, list_url)
        if progress_callback:
            progress_callback(60, "\u5b81\u6ce2: \u89e3\u6790\u5b98\u65b9PDF\u4ef7\u683c\u8868...")
        prices = _parse_ningbo_pdf(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "\u5b81\u6ce2\u5b98\u65b9PDF\u5df2\u4e0b\u8f7d\uff0c\u4f46\u672a\u89e3\u6790\u51fa\u6750\u6599\u4ef7\u683c"}
        if progress_callback:
            progress_callback(95, f"\u5b81\u6ce2: \u89e3\u6790\u5b8c\u6210\uff0c\u5171 {len(prices)} \u6761")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": f"{period}_\u5b81\u6ce2\u5e02\u5efa\u6750\u5546\u60c5\u7248.pdf",
                "url": detail_url,
                "content": content,
            },
            "message": f"\u5df2\u4ece\u5b81\u6ce2\u5e02\u5efa\u8bbe\u5de5\u7a0b\u9020\u4ef7\u7ba1\u7406\u534f\u4f1a\u89e3\u6790 {title}\uff0c\u5171 {len(prices)} \u6761\u542b\u7a0e\u4fe1\u606f\u4ef7\uff0c\u65e0\u9700\u767b\u5f55\u3002",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _latest_beijing_pdf(html_text, base_url, period_filter=None):
    candidates = []
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.find_all("a", href=True):
        title = _clean_table_cell(anchor.get_text(" ", strip=True))
        if "北京工程造价信息" not in title or "市场参考价" in title or "厂家参考" in title:
            continue
        period = _extract_period(title, anchor.get("href", ""))
        if period and anchor["href"].lower().endswith(".pdf"):
            candidates.append((period, title, urljoin(base_url, anchor["href"])))
    candidates.sort(key=lambda item: item[0], reverse=True)
    if period_filter:
        return next((item for item in candidates if item[0] == period_filter), None)
    return candidates[0] if candidates else None


@register("zjw.beijing.gov.cn/bjjs/gczj14/zjxx")
def beijing_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """Fetch Beijing's public monthly construction-cost information PDF."""
    list_url = "https://zjw.beijing.gov.cn/bjjs/gczj14/zjxx/index.shtml"
    try:
        if progress_callback:
            progress_callback(5, "北京：查找最新工程造价信息 PDF...")
        listing = _official_get(list_url, timeout=45)
        target = _latest_beijing_pdf(_response_text(listing), list_url, period_filter)
        if not target:
            return {"prices": [], "count": 0, "error": "北京官网未找到工程造价信息 PDF"}
        period, title, file_url = target
        if progress_callback:
            progress_callback(35, f"北京：下载 {period} 官方 PDF...")
        content = _download_official_file(file_url, list_url)
        prices = _parse_beijing_pdf(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "北京官方 PDF 已下载，但未解析出含税信息价"}
        if progress_callback:
            progress_callback(95, f"北京：解析完成，共 {len(prices)} 条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": _download_name(file_url, f"{period}_北京工程造价信息.pdf"),
                "url": file_url,
                "content": content,
            },
            "message": f"已从北京市住建委解析 {title}，共 {len(prices)} 条含税信息价。",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _parse_kunming_price_html(html_text, period_filter=None):
    prices = []
    soup = BeautifulSoup(html_text or "", "lxml")
    for row in soup.select("div.body"):
        cells = [
            _clean_table_cell(cell.get_text(" ", strip=True))
            for cell in row.find_all("div", recursive=False)
        ]
        if len(cells) < 6:
            continue
        name, spec, unit, raw_price, period, region = cells[:6]
        price = _positive_number(raw_price)
        if not name or not unit or not period or region != "昆明市" or price is None:
            continue
        if period_filter and period != period_filter:
            continue
        prices.append({
            "code": str(len(prices) + 1),
            "name": name,
            "spec": spec,
            "unit": unit,
            "price": price,
            "raw_price": raw_price,
            "category": "主材综合除税价",
            "period": period,
            "region": region,
        })
    return prices


@register("www.ynbzde.com")
def kunming_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """云南省工程建设科技与标准定额管理网公开昆明主材综合除税价。"""
    list_url = "https://www.ynbzde.com/catlist.html?catid=32"
    base_url = "https://www.ynbzde.com/cgprice"
    try:
        if progress_callback:
            progress_callback(5, "昆明: 查找最新主材综合价格月份...")
        listing = _official_get(list_url, timeout=60)
        soup = BeautifulSoup(_response_text(listing), "lxml")
        periods = []
        for anchor in soup.find_all("a", href=True):
            title = _clean_table_cell(anchor.get_text(" ", strip=True))
            if "主材综合价格信息" not in title:
                continue
            period = _extract_period(title, anchor.get("href", ""))
            if period and period not in periods:
                periods.append(period)
        periods.sort(reverse=True)
        if period_filter:
            periods = [period_filter] + [period for period in periods if period != period_filter]
        if not periods:
            return {"prices": [], "count": 0, "error": "昆明官网未找到主材综合价格月份"}

        prices = []
        selected_period = ""
        selected_page_url = ""
        for period in periods:
            year, month = period.split("-", 1)
            page_url = f"{base_url}?search%5Byear%5D={year}&search%5Bmonth%5D={int(month)}"
            if progress_callback:
                progress_callback(20, f"昆明: 查询 {period} 主材综合价格...")
            response = _official_get(
                base_url,
                referer=list_url,
                timeout=60,
                params={"search[year]": year, "search[month]": int(month)},
            )
            parsed = _parse_kunming_price_html(_response_text(response), period)
            if parsed:
                prices = parsed
                selected_period = period
                selected_page_url = page_url
                break
        if not prices:
            return {"prices": [], "count": 0, "error": "昆明主材综合价格页面没有解析出昆明市材料价格"}
        if progress_callback:
            progress_callback(95, f"昆明: 解析完成，共 {len(prices)} 条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": selected_period,
            "document": {
                "file_name": f"{selected_period}_昆明主材综合价格信息.html",
                "url": selected_page_url,
                "content": _response_text(response).encode("utf-8"),
            },
            "message": (
                f"已从云南省工程建设科技与标准定额管理网解析昆明 {selected_period} 主材综合除税价，"
                f"共 {len(prices)} 条，无需登录。"
            ),
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _latest_quarterly_article(html_text, base_url, required_terms, period_filter=None):
    candidates = []
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.find_all("a", href=True):
        title = _clean_table_cell(anchor.get_text(" ", strip=True))
        if not title or not all(term in title for term in required_terms):
            continue
        period = _extract_quarter_period(title, anchor.get("href", ""))
        if period:
            candidates.append((period, title, urljoin(base_url, anchor["href"])))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    requested = period_filter or ""
    month_match = re.fullmatch(r"(20\d{2})-(\d{2})", requested)
    if month_match:
        quarter = (int(month_match.group(2)) - 1) // 3 + 1
        requested = f"{month_match.group(1)}-Q{quarter}"
    return next((item for item in candidates if item[0] == requested), candidates[0])


def _attachment_url(html_text, article_url, extensions):
    extensions = tuple(extension.lower() for extension in extensions)
    soup = BeautifulSoup(html_text or "", "lxml")
    candidates = []
    for tag in soup.find_all(True):
        for attribute in ("href", "src", "data"):
            raw_url = tag.get(attribute)
            if not raw_url:
                continue
            full_url = urljoin(article_url, unescape(raw_url))
            nested_file = parse_qs(urlparse(full_url).query).get("file", [])
            if nested_file:
                candidates.extend(unquote(value) for value in nested_file)
            candidates.append(full_url)
    for candidate in candidates:
        path = urlparse(candidate).path.lower()
        if path.endswith(extensions):
            return candidate
    return ""


def _download_official_file(file_url, referer):
    response = _official_get(file_url, referer=referer, timeout=180)
    content = response.content
    if len(content) < 500:
        raise ValueError(f"官方附件过小: {len(content)} bytes")
    return content


def _download_name(file_url, fallback):
    return unquote(PurePosixPath(urlparse(file_url).path).name) or fallback


# 造价HOME页面与附件使用不同子域名。只允许造价HOME自有域名、
# 造价HOME明确使用的对象存储域名，以及页面明确指向的政府附件域名。
# 不开放任意跳转，避免把“公开检索”变成任意网址下载器。
_ZAOJIAHOME_HOSTS = {
    "wlg.zaojiahome.com",
    "file.zaojiahome.com",
    "pan.zaojiahome.com",
    "zaojiahome.oss-cn-shanghai.aliyuncs.com",
}
_ZAOJIAHOME_OFFICIAL_ATTACHMENT_SUFFIXES = (
    ".gov.cn",
    "zjjcmspublic.oss-cn-hangzhou-zwynet-d01-a.internet.cloud.zj.gov.cn",
)
_ZAOJIAHOME_PROVINCES = (
    "上海", "江苏", "浙江", "安徽", "福建", "江西", "山东",
    "广东", "广西", "海南", "北京", "天津", "河北", "山西",
    "内蒙古", "河南", "湖北", "湖南", "重庆", "四川", "贵州",
    "云南", "西藏", "陕西", "甘肃", "青海", "宁夏", "新疆",
    "辽宁", "吉林", "黑龙江",
)


def _zaojiahome_is_allowed_attachment_host(host):
    host = (host or "").lower().rstrip(".")
    return host in _ZAOJIAHOME_HOSTS or any(
        host.endswith(suffix) for suffix in _ZAOJIAHOME_OFFICIAL_ATTACHMENT_SUFFIXES
    )


def _zaojiahome_get(url, *, referer="", timeout=60):
    parsed = urlparse(url)
    host = (parsed.hostname or "").lower().rstrip(".")
    if host != "wlg.zaojiahome.com" and not _zaojiahome_is_allowed_attachment_host(host):
        raise ValueError(f"造价HOME附件地址不在允许的来源范围内: {url}")
    headers = dict(OFFICIAL_BROWSER_HEADERS)
    if referer:
        headers["Referer"] = referer
    try:
        from urllib3.exceptions import InsecureRequestWarning
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", InsecureRequestWarning)
            response = requests.get(
                url,
                headers=headers,
                timeout=timeout,
                verify=False,
                allow_redirects=True,
            )
    except ImportError:
        response = requests.get(
            url,
            headers=headers,
            timeout=timeout,
            verify=False,
            allow_redirects=True,
        )
    final_host = (urlparse(response.url).hostname or "").lower().rstrip(".")
    if not _zaojiahome_is_allowed_attachment_host(final_host):
        if final_host.endswith("123pan.cn"):
            raise ValueError("该期附件跳转到需要登录或会员权限的网盘，请先完成授权后再抓取")
        raise ValueError(f"造价HOME请求跳转到了不允许的域名: {response.url}")
    response.raise_for_status()
    return response


def _zaojiahome_periods(period_filter=""):
    requested = _extract_period(period_filter)
    if requested:
        return [requested]
    now = datetime.now()
    periods = []
    year, month = now.year, now.month
    for _ in range(18):
        periods.append(f"{year}-{month:02d}")
        month -= 1
        if month == 0:
            year -= 1
            month = 12
    return periods


def _zaojiahome_page_url(province, city, period=""):
    """Return the public city information-price table, never an attachment URL.

    The site exposes the month links from Home/GetXinXiJia. The adapter
    selects the requested month from those links before downloading anything.
    """
    params = {
        "shengFen": province,
        "chengShi": city,
    }
    query = "&".join(
        f"{key}={requests.utils.quote(str(value), safe='')}"
        for key, value in params.items()
    )
    return f"https://wlg.zaojiahome.com/Home/GetXinXiJia?{query}"


def _zaojiahome_province_url(province):
    return (
        "https://wlg.zaojiahome.com/home/index/?shengfen="
        + requests.utils.quote(str(province or ""), safe="")
    )


def _zaojiahome_catalog_city(province, city):
    """Return the site's original city parameter and a city-specific period page.

    The trailing spaces in data-city are significant on some deployments. Keep the
    raw value for the request, while all database/display values remain trimmed.
    """
    province = (province or "").strip()
    city = (city or "").strip().removesuffix("市")
    home_url = _zaojiahome_province_url(province)
    home = _zaojiahome_get(home_url, timeout=60)
    soup = BeautifulSoup(home.text or "", "lxml")
    raw_city = next(
        (str(button.get("data-city") or "") for button in soup.select("[data-city]")
         if str(button.get("data-city") or "").strip().removesuffix("市") == city),
        city,
    )
    page_url = _zaojiahome_page_url(province, raw_city)
    page = _zaojiahome_get(page_url, referer=home_url, timeout=60)
    return raw_city, page_url, page


def _zaojiahome_link_period(anchor, page_url):
    href = urljoin(page_url, unescape(str(anchor.get("href") or "").strip()))
    parsed = urlparse(href)
    query = parse_qs(parsed.query)
    year = (query.get("nianFen") or [""])[0]
    month = (query.get("yueFen") or [""])[0]
    if year and month and str(month).isdigit():
        return f"{year}-{int(month):02d}"
    match = re.search(r"(20\d{2})[-_/](\d{1,2})(?:\D|$)", unquote(href))
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}"
    row = anchor.find_parent("tr")
    row_text = row.get_text(" ", strip=True) if row else ""
    year_match = re.search(r"(20\d{2})\s*年?", row_text)
    month_match = re.search(r"(\d{1,2})\s*月", anchor.get_text(" ", strip=True))
    if year_match and month_match and 1 <= int(month_match.group(1)) <= 12:
        return f"{year_match.group(1)}-{int(month_match.group(1)):02d}"
    return _extract_period(row_text)


def _zaojiahome_period_links(html_text, page_url):
    links = []
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.select("a.xxj-link[href]"):
        href = unescape(str(anchor.get("href") or "").strip())
        if not href or href.lower().startswith(("javascript:", "mailto:", "#")):
            continue
        period = _zaojiahome_link_period(anchor, page_url)
        if period:
            links.append((period, urljoin(page_url, href), anchor.get_text(" ", strip=True)))
    unique = {}
    for item in links:
        unique.setdefault((item[0], item[1]), item)
    return sorted(unique.values(), key=lambda item: item[0], reverse=True)


def _zaojiahome_attachments(html_text, page_url, period, city):
    soup = BeautifulSoup(html_text or "", "lxml")
    candidates = []
    for anchor in soup.find_all("a", href=True):
        href = unescape(str(anchor.get("href") or "").strip())
        if not href or href.lower().startswith(("javascript:", "mailto:", "#")):
            continue
        file_url = urljoin(page_url, href)
        parsed = urlparse(file_url)
        host = (parsed.hostname or "").lower().rstrip(".")
        extension = Path(parsed.path).suffix.lower()
        is_direct_file = extension in {".xls", ".xlsx", ".pdf", ".zip"}
        is_public_pan_file = host == "pan.zaojiahome.com" and parsed.path.strip("/")
        if not _zaojiahome_is_allowed_attachment_host(host) or not (is_direct_file or is_public_pan_file):
            continue
        text = _clean_table_cell(anchor.get_text(" ", strip=True))
        blob = f"{text} {unquote(file_url)}".lower()
        # The common navigation contains unrelated software ZIP downloads.
        # They are not part of the selected city/month price publication.
        if any(marker in blob for marker in (
            "zaojiahome_v3", "/ak/ak/", "免费下载安装",
        )):
            continue
        score = 0
        if extension in {".xls", ".xlsx"}:
            score += 100
        if "excel" in blob or "xls" in blob:
            score += 30
        if city and city in text:
            score += 20
        if any(token in blob for token in (
            "人工", "材料", "施工机具", "机械设备", "原材料", "信息价",
            "价格", "建材", "工料机",
        )):
            score += 25
        if period.replace("-", "") in blob.replace("-", "") or f"{period.split('-')[0]}-{int(period.split('-')[1])}" in blob:
            score += 20
        candidates.append((score, extension, file_url, text))
    return sorted(candidates, key=lambda item: item[0], reverse=True)


def _zaojiahome_price(value):
    text = _clean_table_cell(value).replace(",", "")
    if text in {"", "-", "--", "暂无", "无"}:
        return None
    match = re.fullmatch(r"\d+(?:\.\d+)?(?:\s*[-~—–至]\s*\d+(?:\.\d+)?)?", text)
    if not match:
        return None
    numbers = [float(item) for item in re.findall(r"\d+(?:\.\d+)?", text)]
    price = sum(numbers) / len(numbers) if numbers else 0
    return price if price > 0 else None


def _zaojiahome_tax_rate(value):
    text = _clean_table_cell(value).replace("%", "")
    try:
        number = float(text)
    except (TypeError, ValueError):
        return 0.0
    return number * 100 if 0 < number < 1 else number


def _zaojiahome_header_mapping(cells):
    mapping = {}
    no_tax_candidates = []
    tax_candidates = []
    generic_candidates = []
    for index, value in enumerate(cells):
        header = re.sub(r"\s+", "", str(value or "")).lower()
        if not header:
            continue
        if "name" not in mapping and any(token in header for token in ("材料名称", "名称", "品名", "品种名称")):
            mapping["name"] = index
        if "spec" not in mapping and any(token in header for token in ("规格型号", "规格", "型号", "材质规格")):
            mapping["spec"] = index
        if "unit" not in mapping and any(token in header for token in ("计量单位", "单位")):
            mapping["unit"] = index
        if any(token in header for token in ("除税", "不含税", "税前", "不含规费")):
            no_tax_candidates.append(index)
        elif "含税" in header:
            tax_candidates.append(index)
        elif any(token in header for token in ("单价", "价格", "综合价", "信息价")):
            generic_candidates.append(index)
        elif re.fullmatch(r"20\d{2}年?\d{1,2}月?|\d{1,2}月", header):
            # Some city tables use the publication month itself as the price column.
            generic_candidates.append(index)
        if "rate" not in mapping and "税率" in header:
            mapping["rate"] = index
    if no_tax_candidates:
        mapping["no_tax"] = no_tax_candidates[0]
        mapping["no_tax_candidates"] = no_tax_candidates
    if tax_candidates:
        mapping["tax"] = tax_candidates[0]
        mapping["tax_candidates"] = tax_candidates
    if generic_candidates:
        mapping["price"] = generic_candidates[0]
        mapping["price_candidates"] = generic_candidates
    return mapping


def _zaojiahome_first_price(cells, indexes):
    for index in indexes or []:
        if index < len(cells):
            value = _zaojiahome_price(cells[index])
            if value is not None:
                return value, index
    return None, None


def _zaojiahome_unit(value):
    unit = _clean_table_cell(value).replace(" ", "")
    aliases = {
        "m²": "m2", "㎡": "m2", "平方米": "m2",
        "m³": "m3", "立方米": "m3", "吨": "t",
        "千克": "kg", "米": "m",
    }
    unit = aliases.get(unit, unit)
    allowed = {
        "m", "m2", "m3", "t", "kg", "g", "L", "l",
        "个", "只", "套", "台", "件", "块", "根", "樘",
        "组", "副", "座", "孔", "工日", "台班", "千匹",
        "千块", "百块", "千米", "延米", "瓶", "卷", "箱",
        "盏", "批", "项", "km",
    }
    return unit if unit in allowed else ""


def _zaojiahome_rows_from_xls(content):
    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    try:
        return [
            (worksheet.name, [[_clean_table_cell(value) for value in worksheet.row_values(i)] for i in range(worksheet.nrows)])
            for worksheet in workbook.sheets()
        ]
    finally:
        workbook.release_resources()


def _zaojiahome_rows_from_xlsx(content):
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        return [
            (worksheet.title, [[_clean_table_cell(value) for value in row] for row in worksheet.iter_rows(values_only=True)])
            for worksheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _parse_zaojiahome_xls(content, period, extension=".xls"):
    sheets = _zaojiahome_rows_from_xlsx(content) if extension == ".xlsx" else _zaojiahome_rows_from_xls(content)
    prices = []
    for sheet_name, rows in sheets:
        header = None
        category = ""
        for cells in rows:
            compact = [value.replace(" ", "") for value in cells]
            if header is None:
                candidate = _zaojiahome_header_mapping(compact)
                if {"name", "unit"} <= set(candidate) and any(key in candidate for key in ("no_tax", "tax", "price")):
                    header = candidate
                continue
            if not any(cells):
                continue
            name_index = header["name"]
            spec_index = header.get("spec", -1)
            unit_index = header["unit"]
            rate_index = header.get("rate", -1)
            name = cells[name_index] if name_index < len(cells) else ""
            unit = _zaojiahome_unit(cells[unit_index] if unit_index < len(cells) else "")
            no_tax, no_tax_index = _zaojiahome_first_price(cells, header.get("no_tax_candidates", []))
            tax, tax_index = _zaojiahome_first_price(cells, header.get("tax_candidates", []))
            generic, generic_index = _zaojiahome_first_price(cells, header.get("price_candidates", []))
            if not name or not unit:
                first = next((value for value in cells if value), "")
                if first and ("类" in first or first.startswith(("一、", "二、", "三、", "四、", "五、", "六、", "七、", "八、", "九、"))):
                    category = first
                continue
            price = no_tax or tax or generic
            selected_index = no_tax_index if no_tax is not None else tax_index if tax is not None else generic_index
            if price is None:
                continue
            raw_value = cells[selected_index] if selected_index is not None and selected_index < len(cells) else price
            code = cells[0] if cells else ""
            if re.fullmatch(r"\d+\.0+", code):
                code = code.split(".", 1)[0]
            prices.append({
                "code": code,
                "name": name,
                "spec": cells[spec_index] if spec_index >= 0 and spec_index < len(cells) else "",
                "unit": unit,
                "price": price,
                "raw_price": f"原值:{raw_value};除税:{no_tax if no_tax is not None else ''};含税:{tax if tax is not None else ''}",
                "tax_price": tax,
                "tax_rate": _zaojiahome_tax_rate(cells[rate_index]) if rate_index >= 0 and rate_index < len(cells) else 0,
                "category": category or sheet_name,
                "period": period,
                "price_basis": "tax_exclusive" if no_tax is not None else "tax_inclusive" if tax is not None else "as_published",
            })
    return prices


def _parse_zaojiahome_ocr_pdf(content, period, progress_callback=None):
    """Parse scanned public price PDFs with the bundled OCR engine.

    The parser only accepts pages whose OCR contains a recognizable price-table
    header and rows with a known unit plus a numeric price. This keeps article
    pages and the journal contents page out of the price library.
    """
    try:
        import fitz
        import cv2
        import numpy as np
        from rapidocr_onnxruntime import RapidOCR
    except Exception:
        return []

    def center(box):
        return (
            sum(float(point[0]) for point in box) / 4,
            sum(float(point[1]) for point in box) / 4,
        )

    def token_text(value):
        return re.sub(r"\s+", "", _clean_table_cell(value))

    def number(value):
        text = token_text(value).replace(",", "")
        if not re.fullmatch(r"\d+(?:\.\d+)?", text):
            return None
        parsed = float(text)
        return parsed if parsed > 0 else None

    prices = []
    seen = set()
    try:
        ocr = RapidOCR()
        with fitz.open(stream=content, filetype="pdf") as pdf:
            total_pages = max(1, pdf.page_count)
            for page_index, page in enumerate(pdf):
                # Text PDFs have already been handled by the normal parser.
                if (page.get_text() or "").strip():
                    continue
                pixmap = page.get_pixmap(matrix=fitz.Matrix(1.25, 1.25), alpha=False)
                image = np.frombuffer(pixmap.samples, dtype=np.uint8).reshape(
                    pixmap.height, pixmap.width, pixmap.n,
                )
                image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
                result, _ = ocr(image)
                tokens = []
                for item in result or []:
                    if len(item) < 2:
                        continue
                    box, text = item[0], token_text(item[1])
                    if not text:
                        continue
                    x, y = center(box)
                    tokens.append({"x": x, "y": y, "text": text})
                if len(tokens) < 8:
                    continue

                header_tokens = [
                    item for item in tokens
                    if any(word in item["text"] for word in (
                        "材料名称", "名称", "品名", "规格", "型号", "计量", "单位",
                        "含税", "除税", "税前", "价格", "单价",
                    ))
                ]
                header_y_values = [item["y"] for item in header_tokens]
                if len(header_tokens) < 4 or not header_y_values:
                    continue
                header_y = sorted(header_y_values)[len(header_y_values) // 2]
                nearby_headers = [
                    item for item in header_tokens if abs(item["y"] - header_y) <= 24
                ]

                def header_x(*words):
                    matches = [item["x"] for item in nearby_headers if any(word in item["text"] for word in words)]
                    return sum(matches) / len(matches) if matches else None

                name_x = header_x("材料名称", "名称", "品名")
                spec_x = header_x("规格", "型号")
                unit_x = header_x("计量", "单位")
                tax_x = header_x("含税")
                no_tax_x = header_x("除税", "税前")
                price_x = header_x("价格", "单价")
                if name_x is None or unit_x is None or (tax_x is None and no_tax_x is None and price_x is None):
                    continue
                if spec_x is None:
                    spec_x = name_x + 145
                if price_x is None:
                    price_x = tax_x or no_tax_x

                serial_x = min((item["x"] for item in tokens if item["y"] > header_y + 18), default=0)
                serials = []
                for item in tokens:
                    if item["y"] <= header_y + 18:
                        continue
                    text = item["text"].replace(".", "")
                    if not re.fullmatch(r"\d{1,4}", text):
                        continue
                    if abs(item["x"] - serial_x) <= 55:
                        serials.append(item)
                serials.sort(key=lambda item: item["y"])
                if not serials:
                    continue

                for serial_index, serial in enumerate(serials):
                    lower = serial["y"] - 14
                    upper = (
                        (serial["y"] + serials[serial_index + 1]["y"]) / 2
                        if serial_index + 1 < len(serials)
                        else float("inf")
                    )
                    row = [item for item in tokens if lower <= item["y"] < upper]
                    unit_items = [
                        item for item in row
                        if _zaojiahome_unit(item["text"])
                        and abs(item["x"] - unit_x) <= 75
                    ]
                    if not unit_items:
                        continue
                    unit = _zaojiahome_unit(min(unit_items, key=lambda item: abs(item["x"] - unit_x))["text"])
                    if not unit:
                        continue

                    def closest_number(anchor_x, minimum=0):
                        if anchor_x is None:
                            return None, None
                        candidates = []
                        for item in row:
                            if item["x"] < minimum or abs(item["x"] - anchor_x) > 65:
                                continue
                            value = number(item["text"])
                            if value is not None:
                                candidates.append((abs(item["x"] - anchor_x), value, item["text"]))
                        if not candidates:
                            return None, None
                        _, value, raw = min(candidates, key=lambda item: item[0])
                        return value, raw

                    tax, tax_raw = closest_number(tax_x, unit_x + 20)
                    no_tax, no_tax_raw = closest_number(no_tax_x, unit_x + 20)
                    generic, generic_raw = closest_number(price_x, unit_x + 20)
                    price = no_tax or tax or generic
                    if price is None:
                        continue

                    name_items = [
                        item for item in row
                        if name_x - 70 <= item["x"] < spec_x - 30
                        and not re.fullmatch(r"\d+(?:\.\d+)?", item["text"])
                    ]
                    name = " ".join(item["text"] for item in sorted(name_items, key=lambda item: (item["y"], item["x"])))
                    if len(name) < 2 or name in {"材料名称", "名称", "价格"}:
                        continue
                    spec_items = [
                        item for item in row
                        if spec_x - 55 <= item["x"] < unit_x - 45
                    ]
                    spec = " ".join(item["text"] for item in sorted(spec_items, key=lambda item: (item["y"], item["x"])))
                    key = (name, spec, unit, price, period)
                    if key in seen:
                        continue
                    seen.add(key)
                    prices.append({
                        "name": name,
                        "spec": spec,
                        "unit": unit,
                        "price": price,
                        "raw_price": no_tax_raw or tax_raw or generic_raw or str(price),
                        "tax_price": tax,
                        "category": "OCR公开表格",
                        "period": period,
                        "price_basis": "tax_exclusive" if no_tax is not None else "tax_inclusive" if tax is not None else "as_published",
                    })
                if progress_callback and page_index % 5 == 0:
                    progress_callback(
                        45 + int(45 * (page_index + 1) / total_pages),
                        f"公开市场参考价：OCR扫描页 {page_index + 1}/{total_pages}...",
                    )
    except Exception:
        return prices
    return prices


def _parse_zaojiahome_pdf(content, period, progress_callback=None):
    """Parse text/table PDFs, then OCR scanned public documents."""
    text_unit_pattern = re.compile(
        r"(?<![A-Za-z0-9/])(m2|m3|m²|m³|㎡|t|kg|吨|千克|米|个|只|套|台|件|块|根|工日|台班)(?![A-Za-z0-9/])",
        re.I,
    )

    def find_text_unit(line):
        """Use the last unit token before the final numeric price.

        Chinese unit characters can also be part of a material name, e.g.
        ``混凝土砌块 块 8.00`` or ``预制构件 件 49.00``.
        """
        number_matches = list(re.finditer(r"(?<![A-Za-z])\d+(?:,\d{3})*(?:\.\d+)?", line))
        if not number_matches:
            return None
        price_start = number_matches[-1].start()
        candidates = [match for match in text_unit_pattern.finditer(line) if match.start() < price_start]
        return candidates[-1] if candidates else None

    def plausible_text_row(name, unit):
        compact_name = re.sub(r"\s+", "", str(name or "")).strip("-:：;；,，")
        if not compact_name or len(compact_name) < 2:
            return False
        if compact_name in {"附件", "附", "件", "说明", "备注", "表"}:
            return False
        if any(marker in compact_name for marker in (
            "详见附", "相关材料税前", "价格情况", "建设工程部分实物量",
            "表中每", "本页", "本期", "报价说明", "编制说明",
        )):
            return False
        if re.match(r"^\d+[.、]", compact_name):
            return False
        # Section headings and footer prose are often split by PDF extraction
        # around a unit/number and otherwise look like a valid price row.
        if re.match(r"^[一二三四五六七八九十]+[、.]", compact_name):
            return False
        if unit == "件" and compact_name in {"附", "附件", "构", "构件"}:
            return False
        return True

    prices = []
    seen = set()
    document_is_price_index = False
    with pdfplumber.open(BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                mapping = None
                for row in table:
                    cells = [_clean_table_cell(value) for value in (row or [])]
                    if mapping is None:
                        candidate = _zaojiahome_header_mapping(cells)
                        if {"name", "unit"} <= set(candidate) and any(key in candidate for key in ("no_tax", "tax", "price")):
                            mapping = candidate
                        continue
                    if not cells:
                        continue
                    name = cells[mapping["name"]] if mapping["name"] < len(cells) else ""
                    unit = _zaojiahome_unit(cells[mapping["unit"]] if mapping["unit"] < len(cells) else "")
                    no_tax, no_tax_index = _zaojiahome_first_price(cells, mapping.get("no_tax_candidates", []))
                    tax, tax_index = _zaojiahome_first_price(cells, mapping.get("tax_candidates", []))
                    generic, generic_index = _zaojiahome_first_price(cells, mapping.get("price_candidates", []))
                    price = no_tax or tax or generic
                    selected_index = no_tax_index if no_tax is not None else tax_index if tax is not None else generic_index
                    basis = "tax_exclusive" if no_tax is not None else "tax_inclusive" if tax is not None else "as_published"
                    if not name or not unit or price is None:
                        continue
                    key = (name, cells[mapping.get("spec", -1)] if mapping.get("spec", -1) >= 0 else "", unit, price)
                    if key in seen:
                        continue
                    seen.add(key)
                    prices.append({
                        "name": name, "spec": key[1], "unit": unit, "price": price,
                        "raw_price": cells[selected_index] if selected_index is not None and selected_index < len(cells) else str(price),
                        "tax_price": tax,
                        "category": "PDF公开表格", "period": period, "price_basis": basis,
                    })
            text = page.extract_text() or ""
            if not text:
                continue
            # Index publications are not unit prices. Their rows commonly end
            # with a tax-rate number, which the generic text parser would
            # otherwise mistake for a price.
            if "价格指数" in text or "指数（%）" in text or "指数(%)" in text:
                document_is_price_index = True
            if document_is_price_index:
                continue
            for line in text.splitlines():
                line = _clean_table_cell(line)
                if not line or any(marker in line for marker in ("目录", "说明", "通知", "地址")):
                    continue
                unit_match = find_text_unit(line)
                if not unit_match:
                    continue
                numbers = re.findall(r"(?<![A-Za-z])\d+(?:,\d{3})*(?:\.\d+)?", line)
                if not numbers:
                    continue
                price = _zaojiahome_price(numbers[-1])
                name = line[:unit_match.start()].strip(" :-")
                if (
                    not plausible_text_row(name, _zaojiahome_unit(unit_match.group(0)))
                    or price is None
                    or len(name) > 160
                ):
                    continue
                unit = _zaojiahome_unit(unit_match.group(0))
                if not unit:
                    continue
                key = (name, "", unit, price)
                if key in seen:
                    continue
                seen.add(key)
                prices.append({
                    "name": name, "spec": "", "unit": unit, "price": price,
                    "raw_price": numbers[-1], "tax_price": None, "category": "PDF公开文本",
                    "period": period, "price_basis": "as_published",
                })
    if prices:
        return prices
    return _parse_zaojiahome_ocr_pdf(content, period, progress_callback=progress_callback)


@register("wlg.zaojiahome.com")
def zaojiahome_public_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """造价HOME全国公开入口：按省份、城市和期数定位公开附件并解析。"""
    try:
        query = parse_qs(urlparse(url or "").query)
        city = (query.get("chengShi") or [""])[0].strip()
        province = (query.get("shengFen") or query.get("shengfen") or [""])[0].strip()
        if not city and session_db is not None and region_id:
            try:
                from .models import Region
                region = session_db.query(Region).filter(Region.id == region_id).first()
                city = (region.name if region else "").strip()
                province = (region.province if region else "").strip()
            except Exception:
                city = ""
        city = city.removesuffix("市")
        if not province:
            try:
                from .official_sources_config import get_zaojiahome_province
                province = get_zaojiahome_province(city)
            except Exception:
                province = ""
        if not province or province not in _ZAOJIAHOME_PROVINCES:
            return {"prices": [], "count": 0, "error": f"公开市场参考价无法确定{city}所属省份，请先选择正确的信息来源地区"}
        if not city:
            return {"prices": [], "count": 0, "error": "公开市场参考价需要先选择对应城市"}

        raw_city, city_page_url, city_page = _zaojiahome_catalog_city(province, city)
        period_links = _zaojiahome_period_links(city_page.text, city_page_url)
        target_period = _extract_period(period_filter) if period_filter else ""
        selected_link = next((item for item in period_links if item[0] == target_period), None) if target_period else (period_links[0] if period_links else None)
        if selected_link is None:
            return {"prices": [], "count": 0, "error": f"公开市场参考价未找到{city}对应的期数"}
        period, period_url, _ = selected_link
        if progress_callback:
            progress_callback(15, f"公开市场参考价：已定位{province}{city} {period}期数...")
        period_page = city_page if period_url.rstrip("/") == city_page_url.rstrip("/") else _zaojiahome_get(period_url, referer=city_page_url, timeout=60)
        parsed_period_links = _zaojiahome_attachments(period_page.text, period_url, period, city)
        direct = urlparse(period_url).hostname or ""
        extension = Path(urlparse(period_url).path).suffix.lower()
        file_url = period_url if direct != "wlg.zaojiahome.com" and _zaojiahome_is_allowed_attachment_host(direct) else ""
        download_candidates = []
        if file_url:
            download_candidates = [(1000, extension, file_url, "")]
        else:
            download_candidates = [
                item for item in parsed_period_links
                if item[1] in {".xls", ".xlsx", ".pdf"}
            ]
        if not download_candidates:
            return {"prices": [], "count": 0, "error": f"公开市场参考价已定位{province}{city} {period}，但该期没有可用公开数据"}
        if progress_callback:
            progress_callback(35, f"公开市场参考价：读取{period}公开数据...")
        prices = []
        first_content = b""
        first_extension = ""
        first_file_url = ""
        # A detail page may publish separate labor/material/machinery files.
        # Parse each relevant public file and deduplicate records below.
        for _, candidate_extension, candidate_url, _ in download_candidates[:4]:
            try:
                file_response = _zaojiahome_get(candidate_url, referer=period_url, timeout=180)
                content = file_response.content
                if len(content) < 500:
                    continue
                content_type = (file_response.headers.get("Content-Type") or "").lower()
                actual_extension = candidate_extension
                if actual_extension not in {".xls", ".xlsx", ".pdf"}:
                    actual_extension = (
                        ".pdf" if "pdf" in content_type or content.startswith(b"%PDF")
                        else ".xlsx" if content.startswith(b"PK") else ".xls"
                    )
                parsed_prices = (
                    _parse_zaojiahome_xls(content, period, actual_extension)
                    if actual_extension in {".xls", ".xlsx"}
                    else _parse_zaojiahome_pdf(content, period, progress_callback=progress_callback)
                )
                if parsed_prices:
                    prices.extend(parsed_prices)
                    if not first_content:
                        first_content = content
                        first_extension = actual_extension
                        first_file_url = candidate_url
            except Exception:
                # One malformed/publicly restricted file must not discard other
                # valid price files on the same period page.
                continue
        file_url = first_file_url or file_url
        extension = first_extension or extension
        content = first_content
        unique_prices = []
        seen_price_keys = set()
        for price in prices:
            key = (
                str(price.get("name") or "").strip(),
                str(price.get("spec") or "").strip(),
                str(price.get("unit") or "").strip(),
                str(price.get("price") or ""),
                str(price.get("period") or period),
            )
            if key in seen_price_keys:
                continue
            seen_price_keys.add(key)
            unique_prices.append(price)
        prices = unique_prices
        if not prices:
            return {
                "prices": [],
                "count": 0,
                "error": f"公开市场参考价已定位{province}{city} {period}，文件可以公开下载，但当前未解析出有效的名称、单位和价格字段；可能是扫描件或表格版式暂未适配，不是会员限制",
            }
        if progress_callback:
            progress_callback(95, f"公开市场参考价：解析完成，共{len(prices)}条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": _download_name(file_url, f"{period}_{city}公开市场参考价{extension}"),
                # 只留期数页面作为来源凭据，不把实际附件下载地址写入数据库。
                "url": _zaojiahome_page_url(province, city, period),
                "content": content,
            },
            "message": (
                f"已解析{province}{city}{period}公开市场参考价格{len(prices)}条；"
                "来源性质：公开市场参考（非政府官方），已优先采用除税单价。"
            ),
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


_GUANGZHOU_CODE_RE = re.compile(r"^\d{18}$")
_GUANGZHOU_NUMBER_RE = re.compile(r"^\d+(?:\.\d+)?$")
_GUANGZHOU_RANGE_RE = re.compile(
    r"^(\d+(?:\.\d+)?)\s*[-~\uFF5E\u2014\u2013]\s*(\d+(?:\.\d+)?)$"
)
_GUANGZHOU_UNITS = {
    "t", "m", "m2", "m\u00b2", "m3", "m\u00b3", "kg", "mm", "cm",
    "\u5957", "\u4e2a", "\u4ef6", "\u5757", "\u7c73", "\u5428", "\u5343\u5757",
    "\u767e\u5757", "\u5343\u7c73", "\u7ec4", "\u53f0", "\u53ea", "\u6839",
    "\u6a1f", "\u6247", "\u33a1", "\u5e73\u65b9\u7c73", "\u7acb\u65b9\u7c73",
    "\u516c\u65a4", "\u74f6", "\u5377", "\u7bb1", "\u76cf", "\u6279", "\u9879",
    "km", "L",
}


def _guangzhou_number(value):
    text = _clean_table_cell(value).replace(",", "")
    if _GUANGZHOU_NUMBER_RE.fullmatch(text):
        number = float(text)
        return number if number > 0 else None
    match = _GUANGZHOU_RANGE_RE.fullmatch(text)
    if not match:
        return None
    low, high = (float(item) for item in match.groups())
    if low <= 0 or high < low:
        return None
    return (low + high) / 2


def _guangzhou_is_unit(value):
    return _clean_table_cell(value).replace(" ", "") in _GUANGZHOU_UNITS


def _guangzhou_page_category(page_text):
    lines = [line.strip() for line in (page_text or "").splitlines() if line.strip()]
    candidates = [
        line for line in lines[:4]
        if "\u4ef7\u683c" in line
        and "\u6750\u6599\u7f16\u7801" not in line
        and not line.startswith("\u7a0e\u524d\u7efc\u5408\u4ef7\u683c")
        and len(line) < 100
    ]
    return candidates[-1] if candidates else (lines[0] if lines else "")


def _guangzhou_price_from_row(segment, price_index):
    if price_index is None or price_index >= len(segment):
        return None, ""
    raw_price = segment[price_index]
    price = _guangzhou_number(raw_price)
    if price is None:
        return None, ""
    # When both a range and mean are printed, the official table's mean wins.
    if _GUANGZHOU_RANGE_RE.fullmatch(raw_price):
        for candidate in segment[price_index + 1:]:
            value = _guangzhou_number(candidate)
            if value is not None:
                return value, candidate
            if candidate and candidate not in ("-", "\u2014"):
                break
    if (
        price_index + 2 < len(segment)
        and segment[price_index + 1] in ("-", "\u2014")
        and _guangzhou_number(segment[price_index + 2]) is not None
    ):
        raw_price = f"{raw_price}-{segment[price_index + 2]}"
        return _guangzhou_number(raw_price), raw_price
    return price, raw_price


def _parse_guangzhou_pdf(content, period):
    """Parse Guangzhou's public pre-tax comprehensive-price PDF."""
    prices = []
    with pdfplumber.open(BytesIO(content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            category = _guangzhou_page_category(page_text)
            first_lines = [line.strip() for line in page_text.splitlines() if line.strip()]
            default_unit = "m" if any("\u5143/m" in line for line in first_lines[:4]) else ""
            for table in page.extract_tables() or []:
                rows = [
                    [_clean_table_cell(value) for value in (row or [])]
                    for row in table
                ]
                header = next(
                    (
                        row for row in rows[:8]
                        if any("\u6750\u6599\u7f16\u7801" in cell for cell in row)
                    ),
                    [],
                )
                if not header:
                    continue
                name_columns = [
                    index for index, cell in enumerate(header)
                    if any(
                        marker in cell.replace(" ", "")
                        for marker in (
                            "\u6750\u6599\u540d\u79f0", "\u9879\u76ee\u540d\u79f0", "\u4ea7\u54c1\u540d\u79f0"
                        )
                    )
                ]
                unit_columns = [
                    index for index, cell in enumerate(header)
                    if "\u5355\u4f4d" in cell.replace(" ", "")
                ]
                price_columns = [
                    index for index, cell in enumerate(header)
                    if "\u7a0e\u524d\u7efc\u5408" in cell.replace(" ", "")
                    and "\u4ef7\u683c" in cell.replace(" ", "")
                ]
                previous_names = {}
                previous_units = {}
                for row in rows:
                    code_indexes = [
                        index for index, cell in enumerate(row)
                        if _GUANGZHOU_CODE_RE.fullmatch(cell)
                    ]
                    for code_number, code_index in enumerate(code_indexes):
                        end = (
                            code_indexes[code_number + 1]
                            if code_number + 1 < len(code_indexes)
                            else len(row)
                        )
                        segment = row[code_index:end]
                        name_column = next(
                            (
                                column for column in name_columns
                                if code_index <= column < end
                            ),
                            None,
                        )
                        unit_column = next(
                            (
                                column for column in unit_columns
                                if code_index <= column < end
                            ),
                            None,
                        )
                        price_column = next(
                            (
                                column for column in price_columns
                                if code_index <= column < end
                            ),
                            None,
                        )
                        price_index = (
                            price_column - code_index
                            if price_column is not None else None
                        )
                        price, raw_price = _guangzhou_price_from_row(segment, price_index)

                        # A few PDF rows shift cells because of a merged note column.
                        # Recover those rows from the actual unit cell instead of dropping them.
                        if price is None:
                            actual_unit_index = next(
                                (
                                    index for index, value in enumerate(segment[1:], 1)
                                    if _guangzhou_is_unit(value)
                                ),
                                None,
                            )
                            if actual_unit_index is not None:
                                for index in range(actual_unit_index + 1, len(segment)):
                                    price, raw_price = _guangzhou_price_from_row(segment, index)
                                    if price is not None:
                                        unit_column = code_index + actual_unit_index
                                        price_index = index
                                        break
                        if price is None:
                            continue

                        name = (
                            row[name_column]
                            if name_column is not None and name_column < len(row)
                            else ""
                        )
                        if name:
                            previous_names[code_index] = name
                        else:
                            name = previous_names.get(code_index, "")
                        if not name:
                            name = category

                        unit = (
                            row[unit_column]
                            if unit_column is not None and unit_column < len(row)
                            else ""
                        )
                        if unit:
                            previous_units[code_index] = unit
                        else:
                            unit = previous_units.get(code_index, default_unit)
                        if not unit:
                            unit = next(
                                (
                                    value for value in segment[1:]
                                    if _guangzhou_is_unit(value)
                                ),
                                "",
                            )
                        if not unit:
                            continue

                        excluded = {
                            column - code_index
                            for column in (name_column, unit_column)
                            if column is not None
                        }
                        price_index = price_index or len(segment)
                        spec = " ".join(
                            value for index, value in enumerate(segment[1:price_index], 1)
                            if index not in excluded and value
                        )
                        prices.append(
                            {
                                "code": segment[0],
                                "name": name,
                                "spec": spec,
                                "unit": unit,
                                "price": price,
                                "raw_price": raw_price,
                                "category": category,
                                "period": period,
                            }
                        )
    return prices


@register("zfcj.gz.gov.cn")
def guangzhou_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """Fetch Guangzhou's latest official material-price PDF without login."""
    list_url = "https://zfcj.gz.gov.cn/zwgk/zsdwwj/"
    try:
        if progress_callback:
            progress_callback(5, "\u5e7f\u5dde: \u67e5\u627e\u6700\u65b0\u4eba\u5de5\u3001\u6750\u6599\u3001\u65bd\u5de5\u673a\u5177\u4ef7\u683c\u4fe1\u606f...")
        listing = _official_get(list_url, timeout=45)
        target = _latest_monthly_article(
            _response_text(listing),
            list_url,
            ("\u4eba\u5de5", "\u6750\u6599", "\u65bd\u5de5\u673a\u5177", "\u4ef7\u683c\u4fe1\u606f"),
            period_filter,
        )
        if not target:
            return {"prices": [], "count": 0, "error": "\u5e7f\u5dde\u5b98\u65b9\u901a\u77e5\u680f\u672a\u627e\u5230\u5f53\u671f\u4ef7\u683c\u4fe1\u606f"}
        period, title, article_url = target
        if progress_callback:
            progress_callback(25, f"\u5e7f\u5dde: \u4e0b\u8f7d {period} \u4ef7\u683c\u4fe1\u606f PDF...")
        article = _official_get(article_url, referer=list_url, timeout=45)
        file_url = _attachment_url(_response_text(article), article_url, (".pdf",))
        if not file_url:
            return {"prices": [], "count": 0, "error": "\u5e7f\u5dde\u5b98\u65b9\u901a\u77e5\u672a\u627e\u5230 PDF \u9644\u4ef6"}
        content = _download_official_file(file_url, article_url)
        if progress_callback:
            progress_callback(55, "\u5e7f\u5dde: \u89e3\u6790\u53cc\u680f\u8868\u683c\u3001\u533a\u95f4\u503c\u548c\u5408\u5e76\u5355\u5143...")
        prices = _parse_guangzhou_pdf(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "\u5e7f\u5dde PDF \u5df2\u4e0b\u8f7d\uff0c\u4f46\u6ca1\u6709\u89e3\u6790\u51fa\u6750\u6599\u4ef7\u683c"}
        if progress_callback:
            progress_callback(95, f"\u5e7f\u5dde: \u89e3\u6790\u5b8c\u6210\uff0c\u5171 {len(prices)} \u6761")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": _download_name(file_url, f"{period}_\u5e7f\u5dde\u5e02\u5efa\u8bbe\u5de5\u7a0b\u6750\u6599\u4ef7\u683c\u4fe1\u606f.pdf"),
                "url": file_url,
                "content": content,
            },
            "message": (
                f"\u5df2\u4ece\u5e7f\u5dde\u5e02\u4f4f\u5efa\u5c40\u89e3\u6790 {title}\uff0c"
                f"\u5171 {len(prices)} \u6761\u7a0e\u524d\u7efc\u5408\u4ef7\u683c\uff0c\u65e0\u9700\u767b\u5f55\u3002"
            ),
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _parse_shijiazhuang_pdf(content, period):
    prices = []
    category = "建安材料"
    previous_name = ""
    with pdfplumber.open(BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for raw_row in table or []:
                    cells = [_clean_table_cell(value) for value in (raw_row or [])]
                    cells = [
                        value.replace("尧", "、").replace("渊", "（").replace("冤", "）").replace("耀", "～")
                        for value in cells
                    ]
                    if len(cells) < 5:
                        continue
                    sequence = cells[0]
                    if not re.fullmatch(r"\d+", sequence):
                        if sequence and "序号" not in sequence and not any(cells[1:]):
                            category = sequence
                        continue
                    name = cells[1].replace(" ", "") or previous_name
                    if name:
                        previous_name = name
                    price = _positive_number(cells[4])
                    if not name or price is None:
                        continue
                    prices.append({
                        "code": sequence,
                        "name": name,
                        "spec": cells[2],
                        "unit": cells[3],
                        "price": price,
                        "raw_price": cells[4],
                        "category": category,
                        "period": period,
                    })
    return prices


@register("zjj.sjz.gov.cn")
def shijiazhuang_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """石家庄住建局公开建安材料信息价 PDF。"""
    list_url = "https://zjj.sjz.gov.cn/columns/38884fcc-b2c6-46a2-83e5-e9a2a178d559/index.html"
    try:
        if progress_callback:
            progress_callback(5, "石家庄: 查找最新建安材料信息价...")
        listing = _official_get(list_url, timeout=30)
        target = _latest_monthly_article(
            _response_text(listing), list_url, ("工程造价信息", "建安材料"), period_filter
        )
        if not target:
            return {"prices": [], "count": 0, "error": "石家庄官网未找到建安材料信息价发布页"}
        period, title, article_url = target
        if progress_callback:
            progress_callback(20, f"石家庄: 打开 {period} 发布页...")
        article = _official_get(article_url, referer=list_url, timeout=30)
        file_url = _attachment_url(_response_text(article), article_url, (".pdf",))
        if not file_url:
            return {"prices": [], "count": 0, "error": "石家庄发布页没有找到 PDF 附件"}
        if progress_callback:
            progress_callback(35, f"石家庄: 下载 {period} PDF...")
        content = _download_official_file(file_url, article_url)
        if progress_callback:
            progress_callback(55, "石家庄: 解析 PDF 表格...")
        prices = _parse_shijiazhuang_pdf(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "石家庄 PDF 已下载，但没有解析出价格行"}
        if progress_callback:
            progress_callback(95, f"石家庄: 解析完成，共 {len(prices)} 条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": _download_name(file_url, f"{period}_石家庄建安材料.pdf"),
                "url": file_url,
                "content": content,
            },
            "message": f"已从石家庄市住建局解析 {title}，共 {len(prices)} 条。",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _extract_quarter_period(*values):
    text = " ".join(str(value or "") for value in values)
    match = re.search(r"(20\d{2})\D{0,4}第?([一二三四1-4])季度", text)
    if not match:
        return ""
    quarter = {"一": "1", "二": "2", "三": "3", "四": "4"}.get(match.group(2), match.group(2))
    return f"{match.group(1)}-Q{quarter}"


def _jilin_latest_article(period_filter=None):
    search_url = "http://infogate.jl.gov.cn/govsearch/jsonp/zf_jd_list.jsp"
    params = {
        "page": "1",
        "lb": "139891",
        "channelId": "139891",
        "searchColumn": "all",
        "searchColumnYear": "all",
        "searchYear": "all",
        "SType": "1",
        "callback": "result",
    }
    response = _official_get(
        search_url,
        referer="http://xxgk.jl.gov.cn/zcbm/fgw_98022/xxgkmlqy/",
        timeout=30,
        params=params,
    )
    match = re.search(r"result\s*\((\{.*\})\)\s*;?", _response_text(response), re.S)
    if not match:
        return None
    payload = json.loads(match.group(1))
    candidates = []
    for item in payload.get("data") or []:
        title = _clean_table_cell(item.get("title"))
        if "建设工程价格信息" not in title or "季度" not in title:
            continue
        period = _extract_quarter_period(title)
        article_url = item.get("puburl") or ""
        if period and article_url:
            candidates.append((period, title, article_url))
    candidates.sort(key=lambda item: item[0], reverse=True)
    if not candidates:
        return None
    requested = period_filter or ""
    month_match = re.fullmatch(r"(20\d{2})-(\d{2})", requested)
    if month_match:
        quarter = (int(month_match.group(2)) - 1) // 3 + 1
        requested = f"{month_match.group(1)}-Q{quarter}"
    return next((item for item in candidates if item[0] == requested), candidates[0])


def _parse_changchun_zip(content, period):
    with zipfile.ZipFile(BytesIO(content)) as archive:
        entry = next((
            info for info in archive.infolist()
            if re.split(r"[\\/]", info.filename)[-1].startswith("长春市")
            and info.filename.lower().endswith(".xlsx")
        ), None)
        if entry is None:
            raise ValueError("吉林省 ZIP 中没有找到长春市 XLSX")
        workbook_content = archive.read(entry)

    workbook = load_workbook(BytesIO(workbook_content), read_only=True, data_only=True)
    prices = []
    category = "建设工程价格信息"
    try:
        for worksheet in workbook.worksheets:
            for raw_row in worksheet.iter_rows(min_col=1, max_col=6, values_only=True):
                cells = [_clean_table_cell(value) for value in raw_row]
                if cells[0] and not cells[1] and not cells[3] and _positive_number(cells[5]) is None:
                    category = cells[0]
                    continue
                price = _positive_number(cells[5])
                if not cells[1] or not cells[3] or price is None:
                    continue
                prices.append({
                    "code": cells[0],
                    "name": cells[1],
                    "spec": cells[2],
                    "unit": cells[3],
                    "price": price,
                    "raw_price": cells[5],
                    "category": category,
                    "period": period,
                })
    finally:
        workbook.close()
    return prices


@register("xxgk.jl.gov.cn")
def changchun_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """吉林省住建厅季度 ZIP 中的长春市价格信息 XLSX。"""
    try:
        if progress_callback:
            progress_callback(5, "长春: 查找最新季度价格信息...")
        target = _jilin_latest_article(period_filter)
        if not target:
            return {"prices": [], "count": 0, "error": "吉林省官网未找到季度建设工程价格信息"}
        period, title, article_url = target
        if progress_callback:
            progress_callback(20, f"长春: 打开 {period} 发布页...")
        article = _official_get(article_url, timeout=30)
        file_url = _attachment_url(_response_text(article), article_url, (".zip",))
        if not file_url:
            return {"prices": [], "count": 0, "error": "吉林省发布页没有找到价格信息 ZIP"}
        if progress_callback:
            progress_callback(35, f"长春: 下载 {period} ZIP...")
        content = _download_official_file(file_url, article_url)
        if progress_callback:
            progress_callback(55, "长春: 解析长春市 XLSX...")
        prices = _parse_changchun_zip(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "长春市 XLSX 已下载，但没有解析出含税价格"}
        if progress_callback:
            progress_callback(95, f"长春: 解析完成，共 {len(prices)} 条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": _download_name(file_url, f"{period}_吉林省价格信息.zip"),
                "url": file_url,
                "content": content,
            },
            "message": f"已从吉林省住建厅 {title} 中解析长春市含税价格 {len(prices)} 条。",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _qingdao_page_category(page):
    text = _clean_table_cell(page.extract_text() or "")
    for keyword in ("钢材", "预拌混凝土", "门窗", "保温", "防水", "安装材料"):
        if keyword in text:
            return keyword
    return "建设工程材料价格"


def _parse_qingdao_pdf(content, period):
    prices = []
    with pdfplumber.open(BytesIO(content)) as pdf:
        for page in pdf.pages:
            category = _qingdao_page_category(page)
            for table in page.extract_tables() or []:
                for raw_row in table or []:
                    cells = [_clean_table_cell(value) for value in (raw_row or [])]
                    if len(cells) < 5 or not re.fullmatch(r"\d+", cells[0]):
                        continue
                    price = _positive_number(cells[4])
                    if not cells[1] or not cells[3] or price is None:
                        continue
                    prices.append({
                        "code": cells[0],
                        "name": cells[1],
                        "spec": cells[2],
                        "unit": cells[3],
                        "price": price,
                        "raw_price": cells[4],
                        "category": category,
                        "period": period,
                    })
    return prices


@register("sjw.qingdao.gov.cn")
def qingdao_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """青岛住建局公开建设工程材料价格 PDF。"""
    list_url = "http://sjw.qingdao.gov.cn/cxjsj13/cxjs_95/cxjsj_gczjxx13/"
    try:
        if progress_callback:
            progress_callback(5, "青岛: 查找最新材料价格...")
        listing = _official_get(list_url, timeout=30)
        target = _latest_monthly_article(
            _response_text(listing), list_url, ("建设工程材料价格",), period_filter
        )
        if not target:
            return {"prices": [], "count": 0, "error": "青岛官网未找到建设工程材料价格发布页"}
        period, title, article_url = target
        if progress_callback:
            progress_callback(20, f"青岛: 打开 {period} 发布页...")
        article = _official_get(article_url, referer=list_url, timeout=30)
        file_url = _attachment_url(_response_text(article), article_url, (".pdf",))
        if not file_url:
            return {"prices": [], "count": 0, "error": "青岛发布页没有找到 PDF 附件"}
        if progress_callback:
            progress_callback(35, f"青岛: 下载 {period} PDF...")
        content = _download_official_file(file_url, article_url)
        if progress_callback:
            progress_callback(55, "青岛: 解析 PDF 表格...")
        prices = _parse_qingdao_pdf(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "青岛 PDF 已下载，但没有解析出材料价格"}
        if progress_callback:
            progress_callback(95, f"青岛: 解析完成，共 {len(prices)} 条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": _download_name(file_url, f"{period}_青岛材料价格.pdf"),
                "url": file_url,
                "content": content,
            },
            "message": f"已从青岛市住建局解析 {title}，共 {len(prices)} 条。",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _parse_mianyang_xls(content, period):
    workbook = xlrd.open_workbook(file_contents=content)
    prices = []
    for worksheet in workbook.sheets():
        header = None
        for row_index in range(worksheet.nrows):
            cells = [_clean_table_cell(value) for value in worksheet.row_values(row_index)]
            compact = [value.replace(" ", "") for value in cells]
            if header is None:
                name_index = next((i for i, value in enumerate(compact) if "材料名称" in value), None)
                spec_index = next((i for i, value in enumerate(compact) if "规格型号" in value), None)
                unit_index = next((i for i, value in enumerate(compact) if value == "单位"), None)
                price_index = next((i for i, value in enumerate(compact) if "不含税信息价" in value), None)
                if None not in (name_index, spec_index, unit_index, price_index):
                    header = (name_index, spec_index, unit_index, price_index)
                continue
            name_index, spec_index, unit_index, price_index = header
            if max(header) >= len(cells):
                continue
            price = _positive_number(cells[price_index])
            if not cells[name_index] or not cells[unit_index] or price is None:
                continue
            code = cells[0] if cells else ""
            prices.append({
                "code": code,
                "name": cells[name_index],
                "spec": cells[spec_index],
                "unit": cells[unit_index],
                "price": price,
                "raw_price": cells[price_index],
                "category": worksheet.name,
                "period": period,
            })
    return prices


@register("zjw.my.gov.cn")
def mianyang_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """绵阳市住建委公开市区材料价格 XLS。"""
    list_url = "https://zjw.my.gov.cn/myszjj/c101133/list.shtml"
    try:
        if progress_callback:
            progress_callback(5, "绵阳: 查找最新材料价格...")
        listing = _official_get(list_url, timeout=30)
        target = _latest_monthly_article(
            _response_text(listing), list_url, ("材料价格信息",), period_filter
        )
        if not target:
            return {"prices": [], "count": 0, "error": "绵阳官网未找到市区材料价格发布页"}
        period, title, article_url = target
        if progress_callback:
            progress_callback(20, f"绵阳: 打开 {period} 发布页...")
        article = _official_get(article_url, referer=list_url, timeout=30)
        file_url = _attachment_url(_response_text(article), article_url, (".xls", ".xlsx"))
        if not file_url:
            return {"prices": [], "count": 0, "error": "绵阳发布页没有找到 XLS 附件"}
        if progress_callback:
            progress_callback(35, f"绵阳: 下载 {period} XLS...")
        content = _download_official_file(file_url, article_url)
        if progress_callback:
            progress_callback(55, "绵阳: 解析材料价格 XLS...")
        prices = _parse_mianyang_xls(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "绵阳 XLS 已下载，但没有解析出材料价格"}
        if progress_callback:
            progress_callback(95, f"绵阳: 解析完成，共 {len(prices)} 条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": _download_name(file_url, f"{period}_绵阳材料价格.xls"),
                "url": file_url,
                "content": content,
            },
            "message": f"已从绵阳市住建委解析 {title}，共 {len(prices)} 条。",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _parse_zhengzhou_pdf(content, period):
    prices = []
    with pdfplumber.open(BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for raw_row in table or []:
                    cells = [_clean_table_cell(value) for value in (raw_row or [])]
                    if len(cells) < 6 or not re.fullmatch(r"\d+", cells[0]):
                        continue
                    price = _positive_number(cells[4])
                    if not cells[1] or not cells[3] or price is None:
                        continue
                    prices.append({
                        "code": cells[0],
                        "name": cells[1],
                        "spec": cells[2],
                        "unit": cells[3],
                        "price": price,
                        "raw_price": cells[4],
                        "category": "建设工程主要材料含税价格",
                        "period": period,
                    })
    return prices


@register("zzjsj.zhengzhou.gov.cn")
def zhengzhou_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """郑州市城乡建设局公开主要材料价格信息 PDF。"""
    list_url = "https://zzjsj.zhengzhou.gov.cn/zjxx/index.jhtml"
    try:
        if progress_callback:
            progress_callback(5, "郑州: 查找最新月度主要材料价格...")
        listing = _official_get(list_url, timeout=30)
        target = _latest_monthly_article(
            _response_text(listing), list_url, ("月份", "主要材料价格信息"), period_filter
        )
        if not target:
            return {"prices": [], "count": 0, "error": "郑州官网未找到月度主要材料价格发布页"}
        period, title, article_url = target
        if progress_callback:
            progress_callback(20, f"郑州: 打开 {period} 发布页...")
        article = _official_get(article_url, referer=list_url, timeout=30)
        file_url = _attachment_url(_response_text(article), article_url, (".pdf",))
        if not file_url:
            return {"prices": [], "count": 0, "error": "郑州发布页没有找到 PDF 附件"}
        if progress_callback:
            progress_callback(35, f"郑州: 下载 {period} PDF...")
        content = _download_official_file(file_url, article_url)
        if progress_callback:
            progress_callback(55, "郑州: 解析含税材料价格...")
        prices = _parse_zhengzhou_pdf(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "郑州 PDF 已下载，但没有解析出含税价格"}
        if progress_callback:
            progress_callback(95, f"郑州: 解析完成，共 {len(prices)} 条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": _download_name(file_url, f"{period}_郑州主要材料价格.pdf"),
                "url": file_url,
                "content": content,
            },
            "message": f"已从郑州市城乡建设局解析 {title}，共 {len(prices)} 条含税价格。",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _parse_foshan_green_pdf(content, period):
    prices = []
    previous_name = ""
    name_prefix = ""
    with pdfplumber.open(BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for raw_row in table or []:
                    cells = [_clean_table_cell(value) for value in (raw_row or [])]
                    if len(cells) < 8:
                        continue

                    name = cells[1].replace("材料名称", "").replace(" ", "").strip()
                    raw_price = cells[7]
                    price = _price_range_midpoint(raw_price)
                    unit = cells[3]

                    if name and not unit and price is None:
                        name_prefix = name
                        continue
                    if price is None or not unit:
                        continue
                    if name_prefix and name:
                        name = name_prefix + name
                        name_prefix = ""
                    elif not name:
                        name = previous_name
                    if not name:
                        continue
                    previous_name = name
                    prices.append({
                        "code": cells[0] or str(len(prices) + 1),
                        "name": name,
                        "spec": cells[2],
                        "unit": unit,
                        "price": price,
                        "raw_price": raw_price,
                        "category": "绿色建材市场价格（税前季度价）",
                        "period": period,
                    })
    return prices


@register("fszj.foshan.gov.cn")
def foshan_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """佛山市建设工程造价服务中心公开绿色建材市场价格 PDF。"""
    list_url = (
        "http://fszj.foshan.gov.cn/ywxt/jsgczjfwzx/zwzt_1110045/"
        "jjyjgl/jgxx/scjg/index.html"
    )
    try:
        if progress_callback:
            progress_callback(5, "佛山: 查找最新季度绿色建材市场价格...")
        listing = _official_get(list_url, timeout=30)
        target = _latest_quarterly_article(
            _response_text(listing), list_url, ("绿色建材市场价格",), period_filter
        )
        if not target:
            return {"prices": [], "count": 0, "error": "佛山官网未找到绿色建材市场价格发布页"}
        period, title, article_url = target
        if progress_callback:
            progress_callback(20, f"佛山: 打开 {period} 发布页...")
        article = _official_get(article_url, referer=list_url, timeout=30)
        file_url = _attachment_url(_response_text(article), article_url, (".pdf",))
        if not file_url:
            return {"prices": [], "count": 0, "error": "佛山发布页没有找到 PDF 附件"}
        if progress_callback:
            progress_callback(35, f"佛山: 下载 {period} 绿色建材 PDF...")
        content = _download_official_file(file_url, article_url)
        if progress_callback:
            progress_callback(55, "佛山: 解析税前季度区间价...")
        prices = _parse_foshan_green_pdf(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "佛山 PDF 已下载，但没有解析出绿色建材价格"}
        if progress_callback:
            progress_callback(95, f"佛山: 解析完成，共 {len(prices)} 条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "scope": "绿色建材市场价格",
            "document": {
                "file_name": _download_name(file_url, f"{period}_佛山绿色建材市场价格.pdf"),
                "url": file_url,
                "content": content,
            },
            "message": (
                f"已从佛山市建设工程造价服务中心解析 {title}，共 {len(prices)} 条。"
                "当前仅覆盖绿色建材，价格采用官方税前季度区间的中点并保留原始区间。"
            ),
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _split_combined_name_spec(value):
    text = _clean_table_cell(value)
    text = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])", "", text)
    keyword_match = re.match(r"^(.+?)(胸径|地径|冠幅|高度)(.+)$", text)
    if keyword_match:
        return keyword_match.group(1).strip(), "".join(keyword_match.groups()[1:]).strip()
    compact_match = re.match(r"^([\u4e00-\u9fff（）()、]+?)([A-Z]{2,}\d.+)$", text)
    if compact_match:
        return compact_match.group(1).strip(), compact_match.group(2).strip()
    spec_match = re.search(
        r"\s+(?=(?:[A-Za-z]{1,12}[A-Za-z0-9./+_-]*\d|DN\s*\d|[Φφ∅]\s*\d|"
        r"\d+(?:\.\d+)?(?:mm|cm|m|×|\*)))",
        text,
    )
    if spec_match:
        return text[:spec_match.start()].strip(), text[spec_match.end():].strip()
    return text, ""


def _hohhot_period(value):
    text = _clean_table_cell(value)
    match = re.search(r"(20\d{2}).{0,12}?(\d{1,2})\s*[-—至]\s*(\d{1,2})\s*月份", text)
    if not match:
        return _extract_period(text)
    end_month = int(match.group(3))
    return f"{match.group(1)}-{end_month:02d}" if 1 <= end_month <= 12 else ""


def _latest_hohhot_article(html_text, base_url, period_filter=None):
    candidates = []
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.find_all("a", href=True):
        title = _clean_table_cell(anchor.get_text(" ", strip=True))
        if "建设工程造价信息" not in title:
            continue
        period = _hohhot_period(title)
        if period:
            candidates.append((period, title, urljoin(base_url, anchor["href"])))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    if period_filter:
        exact = next((item for item in candidates if item[0] == period_filter), None)
        if exact:
            return exact
    return candidates[0]


def _parse_hohhot_pdf(content, period):
    prices = []
    seen = set()
    category = "建设工程材料市场价格"
    with pdfplumber.open(BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for raw_row in table or []:
                    cells = [_clean_table_cell(value) for value in (raw_row or [])]
                    if len(cells) < 5:
                        continue
                    if cells[0] and not any(cells[1:]) and not cells[0].startswith(("注：", "编者按")):
                        category = cells[0]
                        continue
                    if not re.fullmatch(r"\d{8,9}", cells[0]):
                        continue
                    price = _positive_number(cells[3])
                    if not cells[1] or not cells[2] or price is None:
                        continue
                    name, spec = _split_combined_name_spec(cells[1])
                    unique_key = (cells[0], name, spec, cells[2], price)
                    if unique_key in seen:
                        continue
                    seen.add(unique_key)
                    prices.append({
                        "code": cells[0],
                        "name": name,
                        "spec": spec,
                        "unit": cells[2],
                        "price": price,
                        "raw_price": cells[3],
                        "category": category,
                        "period": period,
                    })
    return prices


@register("zfcxjsj.huhhot.gov.cn")
def hohhot_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """呼和浩特市住建局公开建设工程造价信息 PDF。"""
    list_url = "http://zfcxjsj.huhhot.gov.cn/bsfw_91/xzzx/zjxx/"
    try:
        if progress_callback:
            progress_callback(5, "呼和浩特: 查找最新建设工程造价信息...")
        listing = _official_get(list_url, timeout=30)
        target = _latest_hohhot_article(_response_text(listing), list_url, period_filter)
        if not target:
            return {"prices": [], "count": 0, "error": "呼和浩特官网未找到建设工程造价信息发布页"}
        period, title, article_url = target
        if progress_callback:
            progress_callback(20, f"呼和浩特: 打开截至 {period} 的发布页...")
        article = _official_get(article_url, referer=list_url, timeout=30)
        file_url = _attachment_url(_response_text(article), article_url, (".pdf",))
        if not file_url:
            return {"prices": [], "count": 0, "error": "呼和浩特发布页没有找到 PDF 附件"}
        if progress_callback:
            progress_callback(35, f"呼和浩特: 下载截至 {period} 的 PDF...")
        content = _download_official_file(file_url, article_url)
        if progress_callback:
            progress_callback(55, "呼和浩特: 解析材料含税价格...")
        prices = _parse_hohhot_pdf(content, period)
        if not prices:
            return {"prices": [], "count": 0, "error": "呼和浩特 PDF 已下载，但没有解析出材料含税价格"}
        if progress_callback:
            progress_callback(95, f"呼和浩特: 解析完成，共 {len(prices)} 条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": _download_name(file_url, f"{period}_呼和浩特建设工程造价信息.pdf"),
                "url": file_url,
                "content": content,
            },
            "message": f"已从呼和浩特市住建局解析 {title}，共 {len(prices)} 条材料含税价格。",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _parse_sichuan_price_rows(html_text):
    rows = []
    for row_html in re.findall(r"<tr[^>]*>(.*?)</tr>", html_text or "", re.I | re.S):
        cells = []
        for cell_html in re.findall(r"<td[^>]*>(.*?)</td>", row_html, re.I | re.S):
            value = re.sub(r"<[^>]+>", "", cell_html)
            value = unescape(value).replace("\xa0", " ").strip()
            cells.append(value)
        if len(cells) >= 7 and any(cells):
            rows.append({
                "code": cells[0],
                "name": cells[1],
                "spec": cells[2],
                "unit": cells[3],
                "raw_price": cells[4],
                "period": cells[5],
                "district": cells[6],
            })
    return rows


@register("202.61.90.35:8037")
def sichuan_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """四川省工程造价信息网成都材料价格查询接口。"""
    base = "http://202.61.90.35:8037"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126 Safari/537.36",
        "Referer": base + "/jgxx.htm?code=5101",
        "X-Requested-With": "XMLHttpRequest",
    }
    if source_id:
        try:
            from .source_credentials import source_auth_headers
            headers.update(source_auth_headers(source_id))
        except Exception:
            pass

    def request_json(method, endpoint, *, params=None, data=None, timeout=45):
        last_error = None
        for attempt in range(3):
            try:
                response = requests.request(
                    method,
                    base + endpoint,
                    params=params,
                    data=data,
                    headers=headers,
                    timeout=timeout,
                )
                response.raise_for_status()
                return response.json()
            except Exception as error:
                last_error = error
                if attempt < 2:
                    time.sleep(attempt + 1)
        raise last_error

    try:
        if progress_callback:
            progress_callback(5, "成都: 连接四川省工程造价信息网...")
        years_payload = request_json("GET", "/cl/jgxx.ashx", params={"Type": "HQCLJGNFXX"})
        years = [str(item.get("id") or "") for item in years_payload if item.get("id")]
        if not years:
            return {"prices": [], "count": 0, "error": "四川官网没有返回材料价格年份列表"}

        requested_year = ""
        requested_month = ""
        if period_filter:
            match = re.match(r"^(20\d{2})(?:-(\d{2}))?$", period_filter)
            if match:
                requested_year, requested_month = match.groups()
        if requested_year and requested_year not in years:
            return {
                "prices": [],
                "count": 0,
                "error": (
                    f"四川官网价格查询接口没有 {period_filter} 数据；"
                    f"该接口当前公开的最新年份为 {years[0]} 年"
                ),
            }
        year = requested_year or years[0]
        month = str(int(requested_month)) if requested_month else "0"

        if progress_callback:
            progress_callback(15, f"成都: 查询{year}年材料目录...")
        query = {
            "nf": year,
            "yf": month,
            "lb": "0",
            "zonecode": "5101",
            "code": "510100",
            "mc": "",
            "ggxh": "",
        }
        first_page = request_json(
            "POST",
            "/cl/jgxx.ashx",
            params={"Type": "BDJGXX", "start": 0, "limit": 1000},
            data=query,
            timeout=60,
        )
        total = int(first_page.get("total") or 0)
        rows = _parse_sichuan_price_rows(first_page.get("html", ""))
        start = len(rows)
        while start < total:
            if progress_callback:
                progress_callback(min(75, 20 + int(50 * start / max(total, 1))), f"成都: 获取材料目录 {start}/{total}...")
            page = request_json(
                "POST",
                "/cl/jgxx.ashx",
                params={"Type": "BDJGXX", "start": start, "limit": 1000},
                data=query,
                timeout=60,
            )
            page_rows = _parse_sichuan_price_rows(page.get("html", ""))
            if not page_rows:
                break
            rows.extend(page_rows)
            start += len(page_rows)

        if not rows:
            return {"prices": [], "count": 0, "error": f"四川官网没有返回成都 {year} 年材料目录"}

        available_periods = sorted({row["period"] for row in rows if re.fullmatch(r"20\d{4}", row["period"])})
        target_period_raw = f"{year}{int(month):02d}" if month != "0" else (available_periods[-1] if available_periods else "")
        selected_rows = [row for row in rows if not target_period_raw or row["period"] == target_period_raw]
        if not selected_rows:
            selected_rows = rows
        period_text = (
            f"{target_period_raw[:4]}-{target_period_raw[4:]}"
            if len(target_period_raw) == 6 else year
        )

        prices = []
        restricted_count = 0
        for row in selected_rows:
            raw_price = row["raw_price"].replace(",", "").strip()
            try:
                price = float(raw_price)
            except ValueError:
                restricted_count += 1
                continue
            if not row["name"] or price <= 0:
                continue
            prices.append({
                "code": row["code"],
                "name": row["name"],
                "spec": row["spec"],
                "unit": row["unit"],
                "price": price,
                "raw_price": row["raw_price"],
                "category": f"四川材料信息价/{row['district']}",
                "period": period_text,
            })

        snapshot = json.dumps({
            "source": "四川省工程造价信息网",
            "source_url": base + "/jgxx.htm?code=5101",
            "query": query,
            "available_years": years,
            "catalog_total": total,
            "selected_period": period_text,
            "selected_rows": selected_rows,
        }, ensure_ascii=False, indent=2).encode("utf-8")
        document = {
            "file_name": f"{period_text}_成都材料价格_四川官网接口快照.json",
            "url": base + "/jgxx.htm?code=5101",
            "content": snapshot,
        }

        if prices:
            if progress_callback:
                progress_callback(95, f"成都: 解析完成，共 {len(prices)} 条价格")
            return {
                "prices": prices,
                "count": len(prices),
                "period": period_text,
                "catalog_count": total,
                "document": document,
            }

        message = (
            f"已从四川官网查到成都材料目录 {total:,} 条，{period_text} 共 {len(selected_rows):,} 条；"
            "官网接口把具体单价返回为“会员查看”。网站和目录均可访问，缺少的是价格查看授权，"
            "不是网址错误或抓取失败。若浏览器已获授权，可在“登录授权”中填写该站 Cookie 后重试。"
        )
        if progress_callback:
            progress_callback(95, f"成都: 查到 {total} 条目录，价格受官网权限限制")
        return {
            "prices": [],
            "count": 0,
            "period": period_text,
            "catalog_count": total,
            "access_restricted": restricted_count > 0,
            "message": message,
            "document": document,
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


def _latest_jiangsu_article(html_text, base_url, city_label, period_filter=None):
    candidates = []
    soup = BeautifulSoup(html_text or "", "lxml")
    for anchor in soup.find_all("a", href=True):
        title = _clean_table_cell(anchor.get_text(" ", strip=True))
        if city_label not in title:
            continue
        if not any(term in title for term in ("信息价", "材料", "造价信息")):
            continue
        period = _extract_period(title, anchor.get("href", ""))
        if period:
            candidates.append((period, title, urljoin(base_url, anchor["href"])))
    candidates.sort(key=lambda item: item[0], reverse=True)
    if period_filter:
        exact = next((item for item in candidates if item[0] == period_filter), None)
        if exact:
            return exact
    return candidates[0] if candidates else None


def _parse_jiangsu_xuzhou_xls(content, period):
    workbook = xlrd.open_workbook(file_contents=content)
    prices = []
    category = ""
    for worksheet in workbook.sheets():
        for row_index in range(worksheet.nrows):
            cells = [_clean_table_cell(value) for value in worksheet.row_values(row_index)]
            if len(cells) < 6:
                continue
            compact = [value.replace(" ", "") for value in cells]
            if "材料编码" in compact and "含税单价" in compact:
                continue
            code, name, spec, unit, raw_price = cells[1:6]
            price = _positive_number(raw_price)
            if not code and (name or spec) and not unit and price is None:
                category = name or spec
                continue
            if not name or not unit or price is None:
                continue
            prices.append({
                "code": code,
                "name": name,
                "spec": spec,
                "unit": unit,
                "price": price,
                "raw_price": raw_price,
                "category": category or worksheet.name,
                "period": period,
            })
    return prices


def _parse_jiangsu_lianyungang_pdf(content, period):
    prices = []
    category = ""
    with pdfplumber.open(BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for raw_row in table or []:
                    cells = [_clean_table_cell(value) for value in (raw_row or [])]
                    if len(cells) < 6:
                        continue
                    compact = [value.replace(" ", "") for value in cells]
                    if "材料编码" in compact and "含税单价" in compact:
                        continue
                    code, name, unit, spec, raw_price = cells[:5]
                    price = _positive_number(raw_price)
                    if code and not name and not unit and price is None:
                        category = code
                        continue
                    if not name or not unit or price is None:
                        continue
                    prices.append({
                        "code": code or str(len(prices) + 1),
                        "name": name,
                        "spec": spec,
                        "unit": unit,
                        "price": price,
                        "raw_price": raw_price,
                        "category": category or "建筑工程材料",
                        "period": period,
                    })
    return prices


@register("49.77.204.6:10081/continue/xydt/001004/")
def jiangsu_city_gov_fetch(
    url,
    session_db,
    task_id=None,
    region_id=None,
    period_filter=None,
    progress_callback=None,
    source_id=None,
):
    """江苏省造价管理信息系统公开的徐州/连云港月度材料价格附件。"""
    is_xuzhou = "001004003" in (url or "")
    city = "徐州" if is_xuzhou else "连云港"
    city_label = f"{city}市"
    list_url = url or (
        "http://49.77.204.6:10081/continue/xydt/001004/001004003/secondPageThird.html"
        if is_xuzhou else
        "http://49.77.204.6:10081/continue/xydt/001004/001004007/secondPageThird.html"
    )
    try:
        if progress_callback:
            progress_callback(5, f"{city}: 查找最新月度材料价格公告...")
        listing = _official_get(list_url, timeout=45)
        target = _latest_jiangsu_article(_response_text(listing), list_url, city_label, period_filter)
        if not target:
            return {"prices": [], "count": 0, "error": f"{city}官方栏目未找到月度材料价格公告"}
        period, title, article_url = target
        article = _official_get(article_url, referer=list_url, timeout=45)
        extensions = (".xls", ".xlsx") if is_xuzhou else (".pdf",)
        file_url = _attachment_url(_response_text(article), article_url, extensions)
        if not file_url:
            return {"prices": [], "count": 0, "error": f"{city}公告未找到价格附件"}
        if progress_callback:
            progress_callback(35, f"{city}: 下载 {period} 官方附件...")
        content = _download_official_file(file_url, article_url)
        if progress_callback:
            progress_callback(60, f"{city}: 解析官方附件...")
        prices = (
            _parse_jiangsu_xuzhou_xls(content, period)
            if is_xuzhou else
            _parse_jiangsu_lianyungang_pdf(content, period)
        )
        if not prices:
            return {"prices": [], "count": 0, "error": f"{city}官方附件已下载，但未解析出材料价格"}
        if progress_callback:
            progress_callback(95, f"{city}: 解析完成，共 {len(prices)} 条")
        return {
            "prices": prices,
            "count": len(prices),
            "period": period,
            "document": {
                "file_name": _download_name(file_url, f"{period}_{city}材料价格"),
                "url": file_url,
                "content": content,
            },
            "message": f"已从江苏省造价管理信息系统解析{title}，共 {len(prices)} 条。",
        }
    except Exception as error:
        import traceback
        return {"prices": [], "count": 0, "error": str(error), "traceback": traceback.format_exc()}


_ZAOJIAHOME_SEARCH_CACHE = {}


def search_zaojiahome_market_reference(
    city,
    period="",
    keywords=None,
    unit="",
    limit=20,
    province="",
):
    """Return structured public market-price evidence for a verified city.

    The returned URL is the public period page, never the hidden attachment URL.
    """
    city = (city or "").strip().removesuffix("市")
    if not province:
        try:
            from .official_sources_config import get_zaojiahome_province
            province = get_zaojiahome_province(city)
        except Exception:
            province = ""
    province = (province or "").strip()
    if not city or not province or province not in _ZAOJIAHOME_PROVINCES:
        return []
    normalized_period = _extract_period(period) if period else ""
    key = (
        province,
        city,
        normalized_period,
        tuple(str(value).strip() for value in (keywords or []) if str(value).strip()),
        unit or "",
    )
    cached = _ZAOJIAHOME_SEARCH_CACHE.get(key)
    if cached and time.monotonic() - cached[0] < 30 * 60:
        return [dict(value) for value in cached[1]]
    page_url = _zaojiahome_page_url(province, city, normalized_period)
    result = zaojiahome_public_fetch(page_url, None, period_filter=normalized_period)
    if not result.get("prices"):
        return []
    terms = [str(value).strip() for value in (keywords or []) if str(value).strip()]
    compact_terms = [re.sub(r"\s+", "", value) for value in terms if len(value.strip()) >= 2]
    target_unit = re.sub(r"\s+", "", str(unit or "")).lower()
    records = []
    for price in result["prices"]:
        name = str(price.get("name") or "").strip()
        spec = str(price.get("spec") or "").strip()
        material_text = re.sub(r"\s+", "", f"{name}{spec}")
        matched = sum(1 for term in compact_terms if term in material_text or material_text in term)
        unit_match = bool(
            target_unit
            and target_unit == re.sub(r"\s+", "", str(price.get("unit") or "")).lower()
        )
        if compact_terms and not matched:
            continue
        score = matched * 20 + (10 if unit_match else 0)
        records.append({
            "query": " ".join(terms[:4]),
            "title": f"公开市场参考价：{name} {spec}".strip(),
            "url": "",
            "host": "公开来源",
            "snippet": (
                f"{name} {spec}，单位{price.get('unit') or ''}，除税单价{price.get('price')}; "
                f"含税单价{price.get('tax_price') or ''}，期数{price.get('period') or normalized_period}，"
                f"类别{price.get('category') or ''}"
            ),
            "excerpt": "公开结构化记录；价格来自对应城市和期数的公开数据。",
            "official": False,
            "trusted": False,
            "source_type": "market_reference",
            "market_price": price.get("price"),
            "market_tax_price": price.get("tax_price"),
            "material_name": name,
            "spec": spec,
            "unit": price.get("unit") or "",
            "period": price.get("period") or normalized_period,
            "province": province,
            "price_basis": price.get("price_basis") or "tax_exclusive",
            "relevance_score": score,
            "evidence_score": score + 20,
        })
    records.sort(key=lambda item: (-item["relevance_score"], item["material_name"]))
    records = records[: max(1, int(limit))]
    _ZAOJIAHOME_SEARCH_CACHE[key] = (time.monotonic(), [dict(value) for value in records])
    return records


def fetch_api_source(url, session_db, task_id=None, region_id=None, **kwargs):
    """根据 URL 匹配已注册的专用适配器。

    如果适配器返回 access_restricted=True，说明官网需要登录授权。
    此时调用方应提示用户使用"浏览器登录"获取 Cookie 后重试。
    """
    for pattern, handler in API_HANDLERS.items():
        if pattern in url:
            return handler(url, session_db, task_id=task_id, region_id=region_id, **kwargs)
    return {"prices": [], "count": 0, "error": f"没有匹配的API处理器: {url}"}


def fetch_with_browser_cookies(url, session_db, task_id=None, region_id=None, source_id=None, **kwargs):
    """使用浏览器登录后保存的 Cookie 请求官网接口。

    适用于 browser_login 策略的来源。需要 source_id 以读取已保存的 Cookie。
    如果 Cookie 无效或未配置，返回 access_required=True 提示用户先登录。
    """
    if not source_id:
        return {"prices": [], "count": 0, "access_required": True,
                "error": "浏览器登录来源需要提供 source_id 以读取 Cookie"}

    from .browser_login import get_source_cookie_header, has_valid_cookies

    if not has_valid_cookies(source_id):
        return {"prices": [], "count": 0, "access_required": True,
                "error": "该来源需要浏览器登录授权。请在来源配置中点击「浏览器登录」按钮，在打开的浏览器中完成登录。"}

    cookie_header = get_source_cookie_header(source_id)

    # 尝试用 Cookie 访问官网首页，验证 Cookie 是否仍然有效
    import requests as req
    try:
        test_resp = req.get(url, headers={"User-Agent": "Mozilla/5.0", "Cookie": cookie_header},
                           timeout=15, allow_redirects=True)
        if test_resp.status_code in (401, 403):
            return {"prices": [], "count": 0, "access_required": True,
                    "error": "浏览器登录 Cookie 已过期或被拒绝。请重新点击「浏览器登录」获取新的 Cookie。"}
    except Exception:
        pass  # 网络问题时继续尝试

    # Cookie 有效 — 但通用 browser_login 源没有专用 API 适配器时，
    # 不能自动知道怎么提取价格数据。返回提示让用户在浏览器中手动下载附件上传。
    # 
    # 如果该源同时有专用 API 适配器（如 202.61.90.35:8037 同时匹配 api 和 browser_login），
    # 则 fetch_api_source 已经会被调用并自动使用 source_auth_headers(Cookie)。
    # 这个函数只是兜底：browser_login 源已有 Cookie 但无专用适配器时，告知用户现状。
    return {
        "prices": [],
        "count": 0,
        "access_required": False,
        "cookie_valid": True,
        "message": (
            "浏览器登录 Cookie 有效。该来源尚未配置专用价格提取适配器。"
            "请在浏览器中查找并下载官方信息价文件（XLS/PDF），"
            "然后通过「原始文件管理」上传；或使用「API嗅探」发现接口后配置适配器。"
        ),
    }
