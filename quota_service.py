"""定额库服务层：Excel/JSON 导入、导出、查询与 CRUD。"""
from __future__ import annotations

# Bump this whenever matching rules, quantity conversion, candidate safety
# checks, or the AI matching contract changes. Persisted project results and
# row-level AI caches are rejected when their version is older.
QUOTA_MATCH_ALGORITHM_VERSION = "quota-match-20260831-01"

import hashlib
import json
import re
import unicodedata
from pathlib import Path
from typing import Iterable

from rapidfuzz import fuzz
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, selectinload

from .db import write_audit
from .bill_list_extractor import invalid_engineering_name_reason, looks_like_engineering_code
from .excel_compat import read_excel_sheets
from .models import (
    CostItem, Material, MaterialPrice, Project, ProjectListData,
    QuotaComposition, QuotaItem, Region, RegionStandard,
)


QUOTA_DETAIL_CATEGORIES = ("材料费", "辅材费", "人工费", "主材费", "机械费", "专业分包")
QUOTA_EXCEL_HEADERS = (
    "类别",
    "定额编码",
    "定额名称",
    "单位",
    "除税单价",
    "含税单价",
)

DEFAULT_MAJORS = ["装修", "土建", "安装", "市政", "园林", "其他"]
QUOTA_MAJOR_ALIASES = {
    "土建": "建筑",
    "建筑工程": "建筑",
    "房建": "建筑",
    "装修": "装饰",
    # Project imports commonly use these business labels for the formal
    # decoration quota major. Without this mapping, selecting "工装" creates
    # an empty candidate index and makes every row appear unmatched.
    "工装": "装饰",
    "精装": "装饰",
    "精装修": "装饰",
    "装潢": "装饰",
    "装饰装修": "装饰",
    "建筑装饰": "装饰",
    "市政工程": "市政",
    "安装工程": "安装",
    "园林绿化": "园林",
}
PRIMARY_MATCH_THRESHOLD = 40
SUPPLEMENTAL_MATCH_OFFSET = 8
LOW_CONFIDENCE_THRESHOLD = 75
AI_COMPOSITION_CATEGORIES = {
    "人工费", "材料费", "辅材费", "主材费", "机械费", "专业分包",
}
AI_COMPOSITION_CATEGORY_ALIASES = {
    "人工": "人工费",
    "人工费": "人工费",
    "材料": "材料费",
    "材料费": "材料费",
    "辅材": "辅材费",
    "辅材费": "辅材费",
    "主材": "主材费",
    "主材费": "主材费",
    "机械": "机械费",
    "机械费": "机械费",
    "专业分包": "专业分包",
    "分包": "专业分包",
}

UNIT_ALIASES = {
    "㎡": "m2", "m²": "m2", "平方米": "m2", "平方": "m2",
    "m³": "m3", "立方米": "m3", "立方": "m3",
    "米": "m", "延米": "m", "吨": "t", "千克": "kg",
    "km": "km", "公里": "km", "千米": "km",
    "1km": "km", "1公里": "km", "1千米": "km",
}
UNIT_DIMENSIONS = {
    "m2": "area", "m3": "volume", "m": "length", "km": "length",
    "t": "mass", "kg": "mass",
    "座": "count", "套": "count", "个": "count", "只": "count",
    "组": "count", "台": "count", "项": "count", "块": "count",
    "件": "count",
}
COUNT_UNITS = {"座", "套", "个", "只", "组", "台", "项", "块", "件"}
LOCATION_GROUPS = {
    "道路": ("道路", "路床", "路基", "路面", "人行道", "车行道", "水稳", "沥青", "侧石", "缘石"),
    "楼地面": ("楼地面", "地面", "地坪", "地砖", "地毯", "踢脚"),
    "墙面": ("墙面", "墙体", "隔墙", "墙砖", "墙裙"),
    "顶棚": ("顶棚", "天棚", "吊顶"),
    "门窗": ("门窗", "门套", "窗套", "门框", "窗框"),
    "管道": ("管道", "水管", "风管", "排水管", "给水管", "雨水管"),
    "电气": ("电缆", "电线", "配电", "桥架", "灯具", "开关", "插座"),
    "园林": ("绿化", "苗木", "乔木", "灌木", "草坪", "种植"),
}
CRAFT_TERMS = (
    "拆除", "开挖", "回填", "运输", "浇筑", "砌筑", "铺贴", "安装", "制作",
    "涂刷", "喷涂", "抹灰", "找平", "防水", "保温", "打磨", "美缝", "切缝",
    "焊接", "植筋", "绑扎", "支模", "养护", "清理", "碾压", "铺设", "铺装",
    "铺砌", "正铺", "斜铺", "擦缝", "勾缝", "嵌缝", "防滑", "刻纹",
    "缩缝", "伸缝", "灌缝", "填缝", "锯缝", "封缝", "密封", "刻槽", "拉毛", "拼花",
)
MATERIAL_TERMS = (
    "混凝土", "钢筋", "砂浆", "水泥", "碎石", "沥青", "瓷砖", "地砖", "墙砖",
    "石材", "地毯", "木地板", "乳胶漆", "腻子", "涂料", "玻璃", "铝合金",
    "不锈钢", "钢板", "钢材", "铁件", "预埋件", "钢构件", "钢盖板",
    "电缆", "电线", "pvc", "pe管", "砖", "木材",
    "塑料膜", "塑料薄膜", "油膏", "密封膏", "密封胶", "玛蹄脂", "聚乙烯", "防滑条",
    "美缝剂", "勾缝剂", "填缝剂",
    "铝板", "铝塑板", "方钢", "方钢管", "镀锌方钢", "大理石", "花岗岩",
    "无纺土工布", "土工布",
)
MATERIAL_COMPONENT_CATEGORIES = {"材料费", "辅材费", "主材费"}

# Material-price fallback is allowed only when the BOQ names a material or a
# concrete material object. Process-only rows such as excavation and hauling
# must never receive an unrelated information-price row merely because the
# units happen to match.
MATERIAL_TERMS = (*MATERIAL_TERMS, "\u782a\u77f3", "\u7802\u783e\u77f3", "\u7802\u77f3", "\u7ea7\u914d\u7802\u77f3", "\u6728\u65b9", "\u94f8\u94c1", "\u4e95\u76d6")

# These are construction objects/material systems, rather than descriptive
# modifiers such as colour, thickness, size, or "综合考虑". When the BOQ
# names one of them, a quota must carry the same evidence in its name or
# composition. This prevents long feature text from making "aluminium wall"
# look like an unrelated stone-wall operation.
CORE_OBJECT_TERMS = (
    "铝板", "铝塑板", "景墙", "景观墙", "方钢", "方钢管", "镀锌方钢",
    "大理石", "花岗岩", "石材", "混凝土", "砼", "沥青混凝土", "砖墙", "砌块墙",
    "玻璃幕墙", "幕墙", "预埋铁件", "预埋件", "钢盖板", "电缆", "电线",
    "给水管", "排水管", "雨水管", "风管", "路床", "路面", "人行道",
    "无纺土工布", "土工布",
)
CORE_MATERIAL_FAMILIES = {
    "铝板体系": ("铝板", "铝塑板"),
    "石材体系": ("大理石", "花岗岩", "石材"),
    "瓷砖体系": ("瓷砖", "地砖", "墙砖", "面砖", "陶瓷砖", "大板砖"),
    "混凝土体系": ("混凝土", "砼"),
    "沥青体系": ("沥青混凝土",),
    "砖墙体系": ("砖墙", "砌块墙"),
    "玻璃幕墙体系": ("玻璃幕墙",),
}
CORE_OBJECT_GROUPS = {
    "\u5bb6\u5177\u684c\u6905": ("\u684c\u6905", "\u4f11\u95f2\u684c", "\u4f11\u95f2\u6905", "\u5ea7\u6905", "\u6210\u54c1\u5bb6\u5177", "\u5bb6\u5177", "\u9910\u684c", "\u6905\u5b50"),
    "\u95e8\u7a97": ("\u95e8\u7a97", "\u5e73\u5f00\u95e8", "\u63a8\u62c9\u95e8", "\u9632\u706b\u95e8", "\u6728\u95e8", "\u94dd\u5408\u91d1\u95e8", "\u7a97", "\u95e8"),
    "\u8e22\u811a\u7ebf": ("\u8e22\u811a\u7ebf", "\u8e22\u811a"),
    "\u680f\u6746\u6276\u624b": ("\u680f\u6746", "\u6276\u624b", "\u6807\u6746"),
    "\u536b\u751f\u6d01\u5177": ("\u5750\u4fbf\u5668", "\u5c0f\u4fbf\u5668", "\u6d17\u624b\u76c6", "\u6d01\u5177"),
    "\u706f\u5177": ("\u706f\u5177", "\u706f\u76d8", "\u5438\u9876\u706f", "\u7b52\u706f"),
    "铝板": ("铝板", "铝塑板"),
    "景墙": ("景墙", "景观墙"),
    "方钢骨架": ("方钢", "方钢管", "镀锌方钢", "钢骨架", "钢框架"),
    "大理石": ("大理石",),
    "石材": ("石材", "花岗岩"),
    "石英砖": ("石英砖",),
    "地砖": ("地砖", "瓷砖地面"),
    "墙砖": ("墙砖", "瓷砖墙面", "面砖墙面"),
    "混凝土": ("混凝土", "砼"),
    "沥青混凝土": ("沥青混凝土",),
    "砖墙": ("砖墙", "砌块墙"),
    "玻璃幕墙": ("玻璃幕墙",),
    "幕墙": ("幕墙",),
    "预埋件": ("预埋铁件", "预埋件"),
    "钢盖板": ("钢盖板",),
    "电缆": ("电缆",),
    "电线": ("电线",),
    "给水管": ("给水管",),
    "排水管": ("排水管",),
    "雨水管": ("雨水管",),
    "风管": ("风管",),
    "路床": ("路床", "路基", "道路基层"),
    "路面": ("路面", "道路面层", "混凝土面层", "沥青面层"),
    "人行道": ("人行道",),
    "土工布": ("无纺土工布", "土工布"),
}
MATERIAL_EQUIVALENT_GROUPS = {
    "塑料薄膜": ("塑料膜", "塑料薄膜", "聚乙烯膜", "聚乙烯薄膜", "聚乙烯塑料薄膜"),
    "密封填缝材料": (
        "油膏", "密封膏", "密封胶", "嵌缝膏", "填缝膏", "玛蹄脂",
        "沥青玛蹄脂", "沥青密封膏", "沥青嵌缝料", "沥青密封材料",
    ),
    "瓷砖填缝材料": ("美缝剂", "勾缝剂", "填缝剂", "瓷砖填缝剂"),
}
# A quota composition may legitimately decompose a named engineering material
# into its constituent resources. These are allowed only for the matching
# material system; they do not make unrelated materials interchangeable.
MATERIAL_DECOMPOSITION_GROUPS = {
    "混凝土": ("混凝土", "水泥", "砂", "碎石", "石子", "骨料", "水", "外加剂"),
    "砂浆": ("砂浆", "水泥", "砂", "水", "外加剂"),
    "砌筑": ("砖", "砌块", "水泥", "砂浆", "砂", "水"),
    "路基": ("土", "砂砾石", "碎石", "级配碎石", "石屑", "填料"),
    "涂料": ("腻子", "乳胶漆", "底漆", "面漆", "涂料", "砂纸"),
    "钢构件": ("钢材", "钢板", "型钢", "方钢", "螺栓", "焊条", "焊丝", "防锈漆"),
}
OBJECT_CONFLICT_GROUPS = {
    "预埋铁件": ("预埋铁件", "预埋件", "铁件", "埋件"),
    "钢盖板": ("钢盖板", "盖板", "钢篦子", "篦子"),
    "幕墙埋件": ("幕墙埋件", "幕墙预埋件", "幕墙"),
}
MATERIAL_REPLACEMENT_PRIORITY = {
    "塑料薄膜": ("塑料薄膜", "塑料膜", "聚乙烯薄膜", "聚乙烯塑料薄膜"),
    "密封填缝材料": ("密封膏", "油膏", "沥青玛蹄脂", "密封胶", "嵌缝膏", "填缝膏"),
    "瓷砖填缝材料": ("美缝剂", "瓷砖填缝剂", "勾缝剂", "填缝剂"),
}
CRAFT_EQUIVALENT_GROUPS = {
    "路面防滑处理": ("防滑", "防滑条", "刻纹", "刻槽", "拉毛"),
    "接缝填封": ("灌缝", "填缝", "嵌缝", "封缝", "密封"),
    "路面接缝": ("缩缝", "伸缝", "切缝", "锯缝"),
    # BOQ feature wording and quota work content use different names for the
    # same layer operation. Keep these as controlled aliases so a valid floor
    # or wall quota is not rejected for lacking the exact phrase "找平" or
    # "铺贴".
    "找平结合层": (
        "找平", "找平层", "结合层", "抹平", "垫层", "粘结层", "粘接层",
        "砂浆粘结", "砂浆粘接", "砂浆结合", "铺砂浆",
    ),
    "铺贴铺装": ("铺贴", "铺设", "铺筑", "铺装", "粘贴", "正铺", "斜铺"),
    "缝隙处理": ("擦缝", "美缝", "勾缝", "灌缝", "填缝", "嵌缝"),
    "基层清理": ("基层处理", "清理基层", "清扫基层", "清理"),
    # Exported BOQs and regional quota books use different names for the same
    # earthwork operation. Treat these as controlled process aliases, while
    # keeping soil/rock, scope and transport conflicts hard.
    "土方开挖": ("开挖", "挖土", "挖方", "挖装", "挖槽", "挖坑", "基坑", "沟槽"),
    "土方回填": ("回填", "填土", "夯填"),
    "土方压实": ("压实", "碾压", "夯实"),
}
STRONG_CRAFT_TERMS = (
    "碾压", "整形", "压实", "找平", "摊铺", "铣刨",
    "铺筑", "铺装", "浇筑", "切缝", "锯缝", "刻纹", "灌缝",
    "填缝", "养护", "养生", "防滑", "拉毛",
)

TILE_OBJECT_TERMS = (
    "瓷砖", "地砖", "墙砖", "面砖", "陶瓷砖", "陶瓷板", "大板砖",
)
STONE_APPEARANCE_TILE_TERMS = (
    "仿石材", "仿大理石", "大理石瓷砖", "通体大理石瓷砖", "仿花岗岩",
)
PAVING_ASSEMBLY_TERMS = (
    "铺贴", "铺设", "铺装", "粘贴", "地砖", "墙砖", "面砖",
    "结合层", "粘结层", "粘接层",
)
SURFACE_TREATMENT_ONLY_TERMS = (
    "结晶", "晶面", "打磨结晶", "抛光处理", "翻新处理",
)
MECHANICAL_CRAFT_RULES = (
    ("土方开挖", (
        "挖一般土方", "挖土方", "挖装土", "挖装槽坑土方", "挖装沟槽土方",
        "机械挖槽坑土方", "机械挖沟槽土方", "挖方", "开挖", "基坑", "沟槽",
    )),
    ("土方运输", ("土方运输", "运土", "外运", "弃置", "倒运", "运距", "装运")),
    ("压实碾压", ("碾压", "压实", "压路机", "打夯机", "夯实")),
    ("钻孔打桩", ("钻孔", "旋挖", "冲击钻", "打桩", "桩机", "沉桩", "成孔")),
    ("起重吊装", ("吊装", "起重", "汽车吊", "塔吊", "垂直运输", "提升")),
    ("机械破除", ("破碎", "铣刨", "机械拆除", "风镐", "破碎锤")),
    ("机械切割", ("锯缝机", "切缝机", "切割机", "路面切缝")),
    ("混凝土机械施工", ("泵送", "泵车", "混凝土运输", "混凝土泵", "搅拌站")),
    ("路面机械施工", ("摊铺", "路拌", "拌和", "洒布车", "摊铺机")),
)
MANUAL_OVERRIDE_TERMS = ("人工挖", "人工开挖", "人工运输", "人工搬运", "人工夯实", "手工")
# These phrases mean that a machine quota already includes the small amount of
# manual coordination required to finish the same machine operation. They do
# not replace an independent manual operation stated by the BOQ.
LABOR_COORDINATION_TERMS = (
    "人工配合", "人工修整", "人工清底", "人工捡底", "人工修坡",
    "清底", "拍底", "修整边坡", "修整坡面", "配合施工", "机下余土",
)
FULL_MANUAL_EXCAVATION_TERMS = ("人工挖土方", "人工挖土", "人工开挖")
MECHANICAL_COMPONENT_CATEGORIES = {"机械费"}
COST_CATEGORY_GROUPS = {
    "人工费": {"人工费"},
    "材料费": MATERIAL_COMPONENT_CATEGORIES,
    "机械费": {"机械费"},
}
ROAD_SCOPE_GROUPS = {
    "路床": ("路床", "路基", "道路基层", "路床整形"),
    "路面": ("路面", "道路面层", "混凝土面层", "沥青面层"),
    "路肩": ("路肩", "硬路肩"),
    "人行道": ("人行道", "人行步道"),
    "车行道": ("车行道", "机动车道"),
}
SPECIAL_SCOPE_TERMS = (
    "洞门", "隧道", "桥墩", "桥台", "箱涵", "涵洞", "烟囱", "水塔",
    "铁路", "地铁", "管廊", "码头", "船闸",
    # These are construction-location constraints, not descriptive keywords.
    # A quota for a named exterior/interior scope must never be selected when
    # the BOQ does not state that scope.
    "外墙", "内墙", "幕墙", "室外", "室内", "楼地面", "屋面", "地下室", "地下工程",
    "楼梯", "踏步", "台阶", "拼花", "斜铺", "错铺", "盲道", "波打线",
)

EARTHWORK_SOIL_TERMS = (
    "土方", "土壤", "一般土", "一类土", "二类土", "三类土", "四类土",
    "槽坑土", "挖土", "土方开挖", "土方回填", "土方运输",
)
EARTHWORK_ROCK_TERMS = (
    "石方", "石渣", "岩石开挖", "岩石", "软岩", "较软岩", "较硬岩",
    "坚硬岩", "中风化岩", "破碎岩",
)
EARTHWORK_EXCAVATION_TERMS = (
    "挖一般土方", "挖土方", "挖装土", "挖装槽坑土方", "挖装沟槽土方",
    "机械挖槽坑土方", "机械挖沟槽土方", "挖方", "开挖", "基坑", "沟槽",
)
EARTHWORK_TRANSPORT_TERMS = (
    "土方运输", "运土", "外运", "外弃", "弃置", "倒运", "运距", "装运",
    "自卸汽车运土",
)


def _text(value) -> str:
    if value is None:
        return ""
    # Excel exports sometimes store carriage returns as the literal
    # ``_x000D_`` token. It is formatting noise, not matching evidence.
    return re.sub(r"(?i)_?x000d_?", "\n", str(value)).strip()


def _period_sort_key(value: str) -> tuple[int, int, int, str]:
    """Sort monthly/quarterly price periods newest first without trusting labels."""
    text = _text(value)
    match = re.search(r"(20\d{2})[-/]?(\d{1,2})", text)
    if match:
        return (int(match.group(1)), int(match.group(2)), 0, text)
    year = re.search(r"(20\d{2})", text)
    return (int(year.group(1)) if year else 0, 0, 0, text)


def _information_region_matches(selected_region: str, region_name: str, province: str = "") -> bool:
    """Match only the user-selected information-price region.

    Project city/province are intentionally absent from this helper. A user
    may price a project in one place using an information-price source from
    another place, so the two concepts must remain independent.
    """
    selected = _text(selected_region)
    if not selected:
        return False
    selected_key = _match_text(selected)
    name_key = _match_text(region_name)
    province_key = _match_text(province)
    return bool(
        selected_key
        and (
            selected_key == name_key
            or selected_key in name_key
            or name_key in selected_key
            or selected_key == province_key
            or selected_key in province_key
            or province_key in selected_key
        )
    )


def latest_price_period(
    session: Session,
    project: Project | None = None,
    region_override: str = "",
) -> tuple[str, str]:
    """Return the newest confirmed period for the selected information region.

    ``region_override`` is deliberately the only regional selector in this
    workflow. Project location fields describe the job background, but must not
    silently change the information-price city or month.
    """
    query = (
        session.query(MaterialPrice.period, MaterialPrice.region_id, Region.name, Region.province)
        .join(Region, MaterialPrice.region_id == Region.id)
        .filter(
            MaterialPrice.is_confirmed.is_(True),
            MaterialPrice.is_withdrawn.is_(False),
            MaterialPrice.is_anomaly.is_(False),
            MaterialPrice.price > 0,
            MaterialPrice.period.isnot(None),
        )
    )
    rows = query.limit(20000).all()
    if not rows:
        return "", "none"
    selected_region = _text(region_override)
    if not selected_region:
        periods = [_text(row[0]) for row in rows if _text(row[0])]
        return max(periods, key=_period_sort_key), "national" if periods else "none"

    # A selected city/province is matched against the Region record only. The
    # project Region.id is intentionally not consulted here.
    scopes = (
        ("selected", lambda row: _information_region_matches(selected_region, row[2], row[3])),
    )
    for scope, predicate in scopes:
        periods = [_text(row[0]) for row in rows if predicate(row) and _text(row[0])]
        if periods:
            return max(periods, key=_period_sort_key), scope
    # The selected region may have no confirmed row yet. Use the newest known
    # period only as a neutral AI/search context; do not relabel another
    # region's price as selected-region information price.
    periods = [_text(row[0]) for row in rows if _text(row[0])]
    return (max(periods, key=_period_sort_key), "national") if periods else ("", "none")


def infer_quota_major(filepath: str | Path, fallback: str = "装修") -> str:
    """Suggest a major from a quota workbook filename.

    Files such as ``电气成本定额.xlsx`` and ``轨道交通成本定额.xlsx`` are
    handled without maintaining a fixed list of disciplines.  The UI still
    asks the user to confirm the suggestion before importing.
    """
    stem = Path(filepath).stem.strip()
    for marker in ("成本定额库", "成本定额", "定额库", "定额"):
        index = stem.find(marker)
        if index > 0:
            suggested = stem[:index].strip(" _-（）()")
            if suggested:
                return suggested
    return _text(fallback)


def _match_text(value) -> str:
    text = _text(value)
    text = re.sub(r"(?i)_?x000d_?", "", text)
    text = re.sub(r"\d+(?:[.*×x]\d+)+", "", text)
    return re.sub(r"\s+", "", text)


def canonical_quota_major(major: str = "") -> str:
    """Resolve UI/import aliases to the major names stored by the quota library."""
    value = _text(major)
    if not value:
        return ""
    return QUOTA_MAJOR_ALIASES.get(value, value)


def quota_major_variants(major: str = "") -> tuple[str, ...]:
    """Return the selected major and its controlled business aliases.

    Imported books may store the same decoration scope as ``工装`` or
    ``装饰``.  Canonicalising only one side made a selected ``工装`` scope
    silently query the other label and lose valid rows.  These aliases remain
    inside the same controlled discipline family; they do not enable an
    unrestricted cross-major search.
    """
    value = _text(major)
    if not value:
        return ()
    canonical = canonical_quota_major(value)
    variants = {value, canonical}
    variants.update(
        alias for alias, target in QUOTA_MAJOR_ALIASES.items()
        if target == canonical
    )
    return tuple(sorted(item for item in variants if item))


def _quota_detail_text(category: str = "", code: str = "", name: str = "", unit: str = "") -> str:
    """Build the searchable text for one library detail row.

    Quota library matching deliberately excludes the legacy work-content fields.
    The BOQ feature/work-content text remains the target evidence supplied by the
    caller; this helper defines the library-side evidence boundary.
    """
    return " ".join(
        value for value in (
            _match_text(category),
            _match_text(code),
            _match_text(name),
            _match_text(unit),
        )
        if value
    )


def _size_specs(value) -> set[tuple[float, ...]]:
    """提取 600x600、1000*1000 等明确多维规格，统一符号后用于硬校验。"""
    text = _text(value).lower().replace("＊", "*").replace("×", "x")
    result = set()
    for match in re.finditer(r"(?<!\d)(\d+(?:\.\d+)?(?:\s*[x*]\s*\d+(?:\.\d+)?){1,2})(?!\d)", text):
        numbers = tuple(float(number) for number in re.split(r"\s*[x*]\s*", match.group(1)))
        result.add(numbers)
    return result


def _tile_face_sizes(value) -> set[tuple[float, float]]:
    """Extract tile face dimensions such as 600*1200 or 800mm×800mm."""
    text = unicodedata.normalize("NFKC", _text(value)).lower()
    text = text.replace("＊", "*").replace("×", "x")
    result = set()
    pattern = re.compile(
        r"(?<!\d)(\d+(?:\.\d+)?)\s*(mm|cm|m)?\s*[x*]\s*"
        r"(\d+(?:\.\d+)?)\s*(mm|cm|m)?",
    )

    def millimetres(amount: str, unit: str) -> float:
        number = float(amount)
        if unit == "cm":
            return number * 10
        if unit == "m":
            return number * 1000
        return number

    for match in pattern.finditer(text):
        width = millimetres(match.group(1), match.group(2) or "mm")
        length = millimetres(match.group(3), match.group(4) or match.group(2) or "mm")
        result.add(tuple(sorted((width, length))))
    return result


def _explicit_tile_plan_size_conflict(target_text: str, candidate_text: str) -> str:
    """Reject a quota that declares a different exact tile face size.

    Parent quota names such as ``600x600地砖`` are product/labour identities,
    not adjustable layer dimensions. Generic range quotas without an exact
    face size remain eligible for controlled matching and AI review.
    """
    target_materials = _material_evidence_terms(target_text)
    candidate_materials = _material_evidence_terms(candidate_text)
    tile_terms = {"瓷砖", "地砖", "墙砖", "面砖"}
    if not target_materials.intersection(tile_terms):
        return ""
    if not candidate_materials.intersection(tile_terms):
        return ""
    target_sizes = _tile_face_sizes(target_text)
    candidate_sizes = _tile_face_sizes(candidate_text)
    if target_sizes and candidate_sizes and not target_sizes.intersection(candidate_sizes):
        format_sizes = lambda values: "/".join(
            "x".join(f"{number:g}" for number in value)
            for value in sorted(values)
        )
        return (
            f"地砖/瓷砖明确长宽规格冲突（清单:{format_sizes(target_sizes)}；"
            f"定额:{format_sizes(candidate_sizes)}）"
        )
    return ""


def _legacy_technical_specs(value) -> set[str]:
    """Extract explicit grades/diameters/thicknesses that AI must not invent."""
    text = _text(value).lower().replace("φ", "Φ").replace("ø", "Φ")
    patterns = (
        r"(?<![a-z0-9])c\d{2,3}(?![a-z0-9])",
        r"(?<![a-z0-9])m\d+(?:\.\d+)?(?![a-z0-9])",
        r"(?<![a-z0-9])dn\s*\d+(?![a-z0-9])",
        r"Φ\s*\d+(?:\.\d+)?",
        r"\d+(?:\.\d+)?\s*(?:mm|cm)\s*厚",
        r"厚\s*\d+(?:\.\d+)?\s*(?:mm|cm)",
        r"(?<![a-z0-9])(?:hrb|hpb)\s*\d+(?![a-z0-9])",
        r"(?<![a-z0-9])q\s*\d{3}(?![a-z0-9])",
    )
    result = set()
    for pattern in patterns:
        for match in re.finditer(pattern, text):
            value = re.sub(r"\s+", "", match.group(0))
            # m2/m3 are measurement units, not mortar grades M2/M3.
            if value.lower() in {"m2", "m3"}:
                continue
            result.add(value)
    return result


def _technical_specs(value) -> set[str]:
    """Extract canonical specification families used by matching and quantity review."""
    text = unicodedata.normalize("NFKC", _text(value)).lower()
    text = text.replace("蠁", "φ").replace("酶", "φ")
    text = re.sub(r"[\s\u3000]+", "", text)
    # Composition fields are often serialized as ``m2 20厚`` or
    # ``m3 1:2.5``. Removing spaces would turn these into ``m220厚`` and
    # ``m31:2.5``, creating false thickness/ratio values. Preserve the unit
    # boundary before extracting specifications.
    text = re.sub(r"(?i)(m[23])(?=\d)", r"\1|", text)
    result: set[str] = set()

    def add(family: str, value: str):
        result.add(f"{family}:{value.upper()}")

    for match in re.finditer(r"(?<![a-z0-9])c[-]?(\d{2,3})(?![a-z0-9])", text):
        add("concrete_grade", f"C{match.group(1)}")
    for match in re.finditer(r"(?<![a-z0-9])mu[-]?(\d+(?:\.\d+)?)(?![a-z0-9])", text):
        add("masonry_grade", f"MU{match.group(1)}")
    for match in re.finditer(r"(?<![a-z0-9])m[-]?(\d+(?:\.\d+)?)(?![a-z0-9])", text):
        if match.group(1) not in {"2", "3"}:
            add("mortar_grade", f"M{match.group(1)}")
    for match in re.finditer(r"(?<!\d)(\d+(?:\.\d+)?)\s*[:：]\s*(\d+(?:\.\d+)?)(?!\d)", text):
        add("mix_ratio", f"R{match.group(1)}:{match.group(2)}")
    for match in re.finditer(r"(?<![a-z0-9])(?:dn|de)[-]?(\d+(?:\.\d+)?)(?![a-z0-9])", text):
        add("diameter", f"DN{match.group(1)}")
    for match in re.finditer(r"(?<![a-z0-9])(?:hrb|hpb)[-]?(\d{3,4}[a-z]?)(?![a-z0-9])", text):
        add("rebar_grade", match.group(0).replace("-", "").upper())
    for match in re.finditer(r"(?<![a-z0-9])q[-]?(\d{3}[a-z]?)(?![a-z0-9])", text):
        add("steel_grade", f"Q{match.group(1)}")
    for match in re.finditer(r"(?:φ|ø|\u03c6)(\d+(?:\.\d+)?)", text):
        add("diameter", f"D{match.group(1)}")
    for match in re.finditer(r"(?<![a-z0-9])pn[-]?(\d+(?:\.\d+)?)(?![a-z0-9])", text):
        add("pressure_grade", f"PN{match.group(1)}")
    for match in re.finditer(r"(\d+(?:\.\d+)?)mpa", text):
        add("pressure_grade", f"{match.group(1)}MPA")
    for match in re.finditer(r"(\d+(?:\.\d+)?)(kv|v)(?![a-z])", text):
        add("voltage_grade", f"{match.group(1)}{match.group(2).upper()}")
    for match in re.finditer(r"(?<![a-z0-9])(?:pe|hdpe)[-]?(\d{2,3})(?![a-z0-9])", text):
        add("material_grade", match.group(0).replace("-", "").upper())
    for match in re.finditer(r"(?<![a-z0-9])(?:pvc-u|upvc|ppr)(?![a-z0-9])", text):
        add("material_grade", match.group(0).replace("-", "").upper())

    dimension_labels = {
        "厚": "thickness", "厚度": "thickness",
        "深": "depth", "深度": "depth",
        "宽": "width", "宽度": "width", "寬": "width",
        "高": "height", "高度": "height",
        "长": "length", "长度": "length",
    }

    def add_dimension(label: str, amount: str, unit: str):
        number = float(amount)
        if unit == "cm":
            number *= 10
        elif unit == "m":
            number *= 1000
        family = dimension_labels[label]
        prefix = {
            "thickness": "T", "depth": "D", "width": "W",
            "height": "H", "length": "L",
        }[family]
        add(family, f"{prefix}{number:g}MM")

    labels = "厚度|厚|深度|深|宽度|宽|寬|高度|高|长度|长"
    for match in re.finditer(rf"({labels})[:：]?(\d+(?:\.\d+)?)(mm|cm|m)(?![a-z])", text):
        add_dimension(match.group(1), match.group(2), match.group(3))
    # Do not merge a concrete grade with a following thickness after spaces
    # are removed: ``C25 ... 15cm厚`` must yield T150MM, never T2515MM.
    for match in re.finditer(rf"(?<![a-z0-9])(\d+(?:\.\d+)?)(mm|cm|m)({labels})", text):
        # ``宽35cm厚3mm`` means W350 + T3.  After whitespace removal the
        # substring ``35cm厚`` also looks like a suffix-style thickness.  Do
        # not reinterpret a dimension that already has an explicit prefix.
        prefix = text[max(0, match.start() - 3):match.start()]
        if any(prefix.endswith(label) for label in dimension_labels):
            continue
        add_dimension(match.group(3), match.group(1), match.group(2))
    for match in re.finditer(
        r"(?:管径|桩径|直径|公称直径)[:：]?(\d+(?:\.\d+)?)(mm|cm)?",
        text,
    ):
        number = float(match.group(1))
        if match.group(2) == "cm":
            number *= 10
        add("diameter", f"DN{number:g}")
    # Cost descriptions commonly omit ``mm`` and continue immediately with a
    # mix ratio, for example ``30厚1:3干硬性水泥砂浆``. The digit after ``厚``
    # belongs to the following ratio and must not suppress the 30mm thickness.
    for match in re.finditer(r"(?<!\d)(\d+(?:\.\d+)?)厚", text):
        add("thickness", f"T{float(match.group(1)):g}MM")

    # Chinese descriptions sometimes omit the C prefix: 混凝土强度等级:25.
    # Requiring the explicit strength label is important. Treating every
    # number following ``砂浆`` as a concrete grade previously converted
    # ``30厚1:3砂浆`` and even later ``68%相关度`` into false C30/C68 specs.
    marker = r"(?:混凝土|砼)"
    pattern = rf"{marker}[^0-9a-z]{{0,12}}(?:强度等级|标号)[^0-9a-z]{{0,8}}(\d{{2,3}})(?![0-9a-z])"
    for match in re.finditer(pattern, text):
        add("concrete_grade", f"C{match.group(1)}")
    for match in re.finditer(
        r"砂浆[^0-9a-z]{0,12}(?:强度等级|标号)[^0-9a-z]{0,8}(\d+(?:\.\d+)?)(?![0-9a-z])",
        text,
    ):
        add("mortar_grade", f"M{match.group(1)}")
    return result


SPECIFICATION_FAMILY_LABELS = {
    "concrete_grade": "混凝土强度等级",
    "mortar_grade": "砂浆强度等级",
    "masonry_grade": "砌体强度等级",
    "diameter": "管径/直径",
    "rebar_grade": "钢筋等级",
    "steel_grade": "钢材牌号",
    "material_grade": "材料型号",
    "pressure_grade": "压力等级",
    "voltage_grade": "电压等级",
    "thickness": "厚度",
    "depth": "深度",
    "width": "宽度",
    "height": "高度",
    "length": "长度",
    "mix_ratio": "配合比",
}
CONVERTIBLE_SPECIFICATION_FAMILIES = {"thickness"}
LAYERED_SPECIFICATION_FAMILIES = {
    # One BOQ feature block commonly contains several layers, each with its
    # own thickness, mortar grade or mix ratio. These families are checked by
    # layer evidence instead of comparing the whole block as one specification.
    "thickness",
    "mortar_grade",
    "mix_ratio",
}
STRICT_SPECIFICATION_FAMILIES = {
    "concrete_grade",
    "masonry_grade",
    "diameter",
    "rebar_grade",
    "steel_grade",
    "material_grade",
    "pressure_grade",
    "voltage_grade",
}


def _specifications_by_family(value) -> dict[str, set[str]]:
    result: dict[str, set[str]] = {}
    for token in _technical_specs(value):
        family, _, specification = token.partition(":")
        if family and specification:
            result.setdefault(family, set()).add(specification)
    return result


def _spec_numeric_mm(token: str) -> float | None:
    match = re.fullmatch(r"[TDWHL](\d+(?:\.\d+)?)MM", _text(token).upper())
    return float(match.group(1)) if match else None


def specification_relation(
    target_text: str,
    candidate_text: str,
    *,
    allow_quantity_conversion: bool = True,
    require_target_specs: bool = True,
    allow_missing_convertible_specs: bool = False,
    allow_missing_layered_specs: bool = False,
    allow_layered_mix_substitution: bool = False,
) -> dict:
    """Compare each specification family independently.

    Strength, grade, diameter, model and pressure conflicts are never made
    equivalent by text similarity. A single explicit layer-thickness change
    can be converted by ratio; the actual component adjustment is performed
    separately and remains auditable.
    """
    target = _specifications_by_family(target_text)
    candidate = _specifications_by_family(candidate_text)
    errors: list[str] = []
    warnings: list[str] = []
    conversions: list[dict] = []
    for family, target_values in target.items():
        label = SPECIFICATION_FAMILY_LABELS.get(family, family)
        candidate_values = candidate.get(family) or set()
        if not candidate_values:
            missing_layer_can_be_reviewed = (
                family in LAYERED_SPECIFICATION_FAMILIES
                and len(target_values) > 1
            )
            missing_dimension_can_be_converted = (
                allow_missing_convertible_specs
                and family in CONVERTIBLE_SPECIFICATION_FAMILIES
            )
            missing_layered_spec_can_be_reviewed = (
                allow_missing_layered_specs
                and family in LAYERED_SPECIFICATION_FAMILIES
            )
            if require_target_specs and not (
                missing_layer_can_be_reviewed
                or missing_dimension_can_be_converted
                or missing_layered_spec_can_be_reviewed
            ):
                errors.append(
                    f"定额候选缺少清单明确{label}：{'/'.join(sorted(target_values))}"
                )
            elif family in LAYERED_SPECIFICATION_FAMILIES:
                if missing_layered_spec_can_be_reviewed:
                    warnings.append(
                        f"定额候选未注明清单明确{label}，按清单规格保留候选并复核含量"
                    )
                else:
                    warnings.append(
                        f"定额候选未完整注明清单{label}，将按匹配层次保留并对相关含量进行复核"
                    )
            continue
        if target_values == candidate_values:
            continue

        shared_values = target_values.intersection(candidate_values)
        if shared_values:
            # A quota may describe only one of several BOQ layers. Shared
            # evidence is therefore compatible; the unmatched values remain a
            # visible review warning instead of rejecting a valid whole quota.
            if (
                family in LAYERED_SPECIFICATION_FAMILIES
                and (
                    len(target_values) > 1
                    or len(candidate_values) > 1
                )
            ):
                missing_target = target_values.difference(candidate_values)
                extra_candidate = candidate_values.difference(target_values)
                detail = []
                if missing_target:
                    detail.append("清单层次未全部覆盖：" + "/".join(sorted(missing_target)))
                if extra_candidate:
                    detail.append("定额含其他层次：" + "/".join(sorted(extra_candidate)))
                warnings.append(
                    f"{label}存在层次对应关系（已匹配：{'/'.join(sorted(shared_values))}）；"
                    + "；".join(detail)
                    + "，按具体组成明细复核含量"
                )
            continue

        # Explicit strength, grade, diameter and model specifications are
        # identity constraints. They remain hard conflicts even when another
        # layer in the same feature block has a different specification.
        if family in STRICT_SPECIFICATION_FAMILIES:
            errors.append(
                f"明确{label}冲突（清单:{'/'.join(sorted(target_values))}；"
                f"定额:{'/'.join(sorted(candidate_values))}）"
            )
            continue

        # Thickness/mortar/mix-ratio values can belong to different layers.
        # When a block contains multiple values there is no sound basis for
        # rejecting the quota as a whole; leave the layer mapping auditable.
        if family == "mix_ratio" and allow_layered_mix_substitution:
            warnings.append(
                f"{label}不同（清单:{'/'.join(sorted(target_values))}；"
                f"定额:{'/'.join(sorted(candidate_values))}），按同类砂浆层作为可调整替补"
            )
            continue
        if family in LAYERED_SPECIFICATION_FAMILIES and (
            len(target_values) > 1 or len(candidate_values) > 1
        ):
            warnings.append(
                f"{label}未能在多层特征中一一对应（清单:{'/'.join(sorted(target_values))}；"
                f"定额:{'/'.join(sorted(candidate_values))}），保留候选并按组成层次复核"
            )
            continue

        if (
            allow_quantity_conversion
            and family in CONVERTIBLE_SPECIFICATION_FAMILIES
            and len(target_values) == len(candidate_values) == 1
        ):
            target_token = next(iter(target_values))
            candidate_token = next(iter(candidate_values))
            target_value = _spec_numeric_mm(target_token)
            candidate_value = _spec_numeric_mm(candidate_token)
            if target_value and candidate_value:
                ratio = target_value / candidate_value
                if 0.05 <= ratio <= 20:
                    conversions.append({
                        "family": family,
                        "label": label,
                        "target": target_token,
                        "candidate": candidate_token,
                        "target_value": target_value,
                        "candidate_value": candidate_value,
                        "ratio": ratio,
                        "formula": (
                            f"{label}换算系数={target_value:g}mm÷{candidate_value:g}mm"
                            f"={ratio:.6g}"
                        ),
                    })
                    warnings.append(
                        f"{label}不同，必须按{target_value:g}/{candidate_value:g}换算受影响材料含量"
                    )
                    continue
        errors.append(
            f"明确{label}冲突（清单:{'/'.join(sorted(target_values))}；"
            f"定额:{'/'.join(sorted(candidate_values))}）"
        )
    target_sizes = _size_specs(target_text)
    candidate_sizes = _size_specs(candidate_text)
    if target_sizes:
        if not candidate_sizes:
            # A generic quota without a stored plan dimension can remain a
            # review candidate, but it must not be treated as dimensionally
            # identical. Explicit size-to-size conflicts remain hard errors.
            warnings.append("定额候选未注明清单明确的长宽尺寸，材料含量和适用范围需人工确认")
        elif target_sizes.intersection(candidate_sizes):
            target_label = "/".join("x".join(f"{number:g}" for number in value) for value in sorted(target_sizes))
            candidate_label = "/".join("x".join(f"{number:g}" for number in value) for value in sorted(candidate_sizes))
            warnings.append(
                f"长宽尺寸存在层次对应（清单:{target_label}；定额:{candidate_label}），"
                "未对应层次按具体材料组成复核"
            )
        elif len(target_sizes) > 1 or len(candidate_sizes) > 1:
            target_label = "/".join("x".join(f"{number:g}" for number in value) for value in sorted(target_sizes))
            candidate_label = "/".join("x".join(f"{number:g}" for number in value) for value in sorted(candidate_sizes))
            warnings.append(
                f"多层长宽尺寸未能一一对应（清单:{target_label}；定额:{candidate_label}），"
                "按对应材料组成复核"
            )
        elif target_sizes != candidate_sizes:
            target_label = "/".join("x".join(f"{number:g}" for number in value) for value in sorted(target_sizes))
            candidate_label = "/".join("x".join(f"{number:g}" for number in value) for value in sorted(candidate_sizes))
            errors.append(f"规格冲突：明确长宽尺寸冲突（清单:{target_label}；定额:{candidate_label}）")
    return {
        "target": target,
        "candidate": candidate,
        "errors": list(dict.fromkeys(errors)),
        "warnings": list(dict.fromkeys(warnings)),
        "conversions": conversions,
        "compatible": not errors,
    }


def apply_specification_quantity_conversion(
    composition: dict,
    target_text: str,
    candidate_text: str,
) -> dict:
    """Apply only an auditable, material-only specification quantity ratio.

    A quota with a different explicit layer thickness can be reused as a
    controlled substitute. Material consumption follows the thickness ratio;
    labor, machinery, subcontract and fees retain their quota quantities until
    a separate construction-method review confirms otherwise.
    """
    result = dict(composition or {})
    relation = specification_relation(
        target_text,
        candidate_text,
        allow_quantity_conversion=True,
        require_target_specs=True,
    )
    if relation["errors"] or not relation["conversions"]:
        return result
    category = _text(result.get("cat") or result.get("category"))
    if category not in MATERIAL_COMPONENT_CATEGORIES:
        result.setdefault("specConversionFactor", 1.0)
        result["specConversionWarning"] = (
            "清单与定额存在可换算规格差异；人工、机械等非材料分量未按厚度同比调整，需复核"
        )
        return result
    component_text = " ".join(
        _text(result.get(key))
        for key in ("name", "feature", "code", "unit")
        if _text(result.get(key))
    )
    target_materials = _clause_material_terms(target_text)
    candidate_materials = _clause_material_terms(candidate_text)
    component_materials = _clause_material_terms(component_text)
    if target_materials and candidate_materials and not (
        target_materials.intersection(candidate_materials)
        or target_materials.intersection(component_materials)
    ):
        return result
    factor = 1.0
    formulas = []
    for conversion in relation["conversions"]:
        factor *= float(conversion["ratio"])
        formulas.append(conversion["formula"])
    result["specConversionFactor"] = round(factor, 8)
    result["specConversionBasis"] = "；".join(formulas)
    existing_conversion = _text(result.get("unitConversion") or result.get("unit_conversion"))
    conversion_note = (
        f"规格换算：材料单位含量×{factor:.8g}；{result['specConversionBasis']}；"
        "仅调整受规格影响的材料含量，人工费和机械费不自动同比调整"
    )
    result["unitConversion"] = "；".join(
        value for value in (existing_conversion, conversion_note) if value
    )
    result["note"] = "；".join(
        value for value in (_text(result.get("note")), conversion_note, "需人工复核施工工艺和材料范围")
        if value
    )
    return result


def _technical_evidence(*values) -> str:
    """Index only normalized hidden legacy specification evidence."""
    return " ".join(sorted({token for value in values for token in _technical_specs(value)}))


def _canonical_unit(value) -> str:
    unit = _match_text(value).lower().replace("²", "2").replace("³", "3")
    # Imported BOQs often describe a selectable count unit as ``个/组`` or
    # ``台(套)``.  It is still a count dimension; treating the whole string as
    # an unknown unit caused valid AI components to be rejected.
    count_parts = {part for part in re.split(r"[/、或()（）]+", unit) if part}
    if count_parts and len(count_parts) > 1 and count_parts.issubset(COUNT_UNITS):
        # Prefer the common generic counter while retaining ordinary single
        # count units (台/套/组) for display and formula parsing.
        return "个" if "个" in count_parts else next(
            part for part in re.split(r"[/、或()（）]+", unit) if part
        )
    return UNIT_ALIASES.get(unit, unit)


def _text_terms(text: str, terms: Iterable[str]) -> set[str]:
    normalized = _match_text(text).lower()
    return {term for term in terms if term.lower() in normalized}


def _semantic_groups(text: str, groups: dict[str, Iterable[str]]) -> set[str]:
    normalized = _match_text(text).lower()
    return {
        label
        for label, terms in groups.items()
        if any(term.lower() in normalized for term in terms)
    }


def _stone_word_is_tile_appearance(text: str) -> bool:
    """Whether a stone word describes a ceramic tile finish, not stone."""
    normalized = _match_text(text).lower()
    return (
        any(term in normalized for term in TILE_OBJECT_TERMS)
        and any(term in normalized for term in STONE_APPEARANCE_TILE_TERMS)
    )


def _core_material_family_groups(text: str) -> set[str]:
    """Classify the actual material system instead of decorative wording."""
    groups = _semantic_groups(text, CORE_MATERIAL_FAMILIES)
    if _stone_word_is_tile_appearance(text):
        groups.discard("石材体系")
        groups.add("瓷砖体系")
    return groups


def _specific_terms(terms: Iterable[str]) -> set[str]:
    values = set(terms)
    return {
        term for term in values
        if not any(term != other and term in other for other in values)
    }


MATERIAL_TERM_ALIASES = {
    "砼": "混凝土",
    "商品砼": "混凝土",
    "普通商品砼": "混凝土",
    "水泥砼": "混凝土",
    "水泥混凝土": "混凝土",
}


def _canonical_material_text(value) -> str:
    """Normalize common BOQ material aliases before evidence matching."""
    normalized = _match_text(value).lower()
    for alias, canonical in sorted(
        MATERIAL_TERM_ALIASES.items(), key=lambda pair: len(pair[0]), reverse=True
    ):
        normalized = normalized.replace(alias, canonical)
    return normalized


def _material_evidence_terms(text: str) -> set[str]:
    """Return material evidence with aliases collapsed to one material system."""
    normalized = _canonical_material_text(text)
    terms = _specific_terms(
        term for term in MATERIAL_TERMS
        if term.lower() in normalized
    )
    if _stone_word_is_tile_appearance(text):
        terms.difference_update({"石材", "大理石", "花岗岩"})
    return terms


def _clause_material_terms(text: str) -> set[str]:
    terms = _material_evidence_terms(text)
    normalized = _match_text(text)
    if (
        "钢筋混凝土结构层" in normalized
        and any(term in normalized for term in TILE_OBJECT_TERMS)
        and not any(term in normalized for term in ("结构层浇筑", "现浇结构层", "新建结构层"))
    ):
        # In floor/wall finish schedules this phrase normally identifies the
        # existing substrate. Price the stated treatment above it, not a new
        # reinforced-concrete slab unless construction is explicitly required.
        terms.difference_update({"混凝土", "钢筋"})
    # 道路“防滑条”通常描述拉毛/刻纹工艺，不按楼梯用 L 型防滑条主材处理。
    if "道路" in _location_groups(text) and "防滑条" in terms:
        terms.remove("防滑条")
    return terms


def _canonical_cost_category_group(value: str) -> str:
    """Return a stable evidence group for category aliases."""
    value = _text(value)
    if value in MATERIAL_COMPONENT_CATEGORIES or value == "材料费":
        return "材料费"
    if value == "人工费":
        return "人工费"
    if value == "机械费":
        return "机械费"
    return value


def _material_match_allowed(target_text: str, candidate_text: str) -> tuple[bool, str]:
    """Allow only the same material system or a known constituent material."""
    explicit_conflict = _explicit_material_conflict(target_text, candidate_text)
    if explicit_conflict:
        return False, explicit_conflict
    target_materials = _clause_material_terms(target_text)
    target_core_objects = _core_object_groups(target_text)
    if not target_materials:
        candidate_core_objects = _core_object_groups(candidate_text)
        if target_core_objects.intersection(candidate_core_objects):
            return True, "清单核心工程对象与组成材料体系一致"
        return True, ""
    candidate_materials = _clause_material_terms(candidate_text)
    if target_materials.intersection(candidate_materials):
        return True, ""
    candidate_core_objects = _core_object_groups(candidate_text)
    if target_core_objects.intersection(candidate_core_objects):
        return True, "清单核心工程对象与组成材料体系一致"
    # A broad library label such as “石材” can safely serve a more specific
    # BOQ material such as “大理石” when it does not assert a competing stone
    # type. The BOQ material remains the controlling identity for pricing.
    target_families = _core_material_family_groups(target_text)
    candidate_families = _core_material_family_groups(candidate_text)
    broad_candidate = candidate_materials.intersection({"石材"})
    specific_candidate = candidate_materials.intersection({"大理石", "花岗岩"})
    if target_families.intersection(candidate_families) and broad_candidate and not specific_candidate:
        return True, "候选为同一材料体系的通用石材定额，按清单具体品种计价"
    for group_terms in MATERIAL_EQUIVALENT_GROUPS.values():
        target_group = target_materials.intersection(group_terms)
        candidate_group = candidate_materials.intersection(group_terms)
        if target_group and candidate_group:
            return True, ""
    for family, component_terms in MATERIAL_DECOMPOSITION_GROUPS.items():
        if not target_materials.intersection({family, *component_terms}):
            continue
        if target_materials.intersection({family, *component_terms}) and candidate_materials.intersection(component_terms):
            return True, f"允许{family}按组成明细分解"
    if not candidate_materials:
        return False, "定额候选缺少清单明确材料对象"
    return False, (
        "明确材料对象冲突（清单："
        + "/".join(sorted(target_materials))
        + "；定额："
        + "/".join(sorted(candidate_materials))
        + "）"
    )


def _earthwork_process_profile(text: str) -> tuple[bool, bool]:
    """Identify excavation and hauling as separate earthwork processes."""
    profile = _earthwork_profile(text)
    excavation = bool(profile["excavation"])
    transport = bool(profile["transport"])
    return excavation, transport


def _earthwork_profile(text: str) -> dict:
    """Extract earthwork object and process evidence without fuzzy guessing.

    Soil/rock identity is treated as an object constraint. An explicit ``土方``
    or soil class wins over incidental rock words in a broad soil description
    such as ``土方类别综合，含软岩``.
    """
    normalized = _match_text(text)
    has_soil = any(term in normalized for term in EARTHWORK_SOIL_TERMS)
    has_rock = any(term in normalized for term in EARTHWORK_ROCK_TERMS)
    explicit_rock = any(term in normalized for term in ("石方", "石渣", "岩石开挖", "破碎岩"))
    if has_soil:
        material = "土方"
    elif explicit_rock or has_rock:
        material = "石方"
    else:
        material = ""

    excavation = any(term in normalized for term in EARTHWORK_EXCAVATION_TERMS)
    transport = any(term in normalized for term in EARTHWORK_TRANSPORT_TERMS)
    loading = any(term in normalized for term in ("挖装", "装车", "装运"))
    if any(term in normalized for term in ("槽坑", "基坑", "沟槽")):
        scope = "槽坑"
    elif "一般土方" in normalized or "一般土" in normalized:
        scope = "一般"
    else:
        scope = ""
    return {
        "material": material,
        "excavation": excavation,
        "transport": transport,
        "loading": loading,
        "scope": scope,
        "terms": [
            *([material] if material else []),
            *(["挖装" if loading else ""] if loading else []),
            *([scope] if scope else []),
        ],
    }


def _earthwork_logic_conflict(target_text: str, candidate_text: str) -> str:
    """Reject soil/rock or excavation/hauling process substitutions."""
    target = _earthwork_profile(target_text)
    candidate = _earthwork_profile(candidate_text)
    if target["material"] and candidate["material"] and target["material"] != candidate["material"]:
        return (
            f"土石方对象冲突（清单：{target['material']}；"
            f"定额/组成：{candidate['material']}）"
        )
    if target["excavation"] and candidate["transport"] and not candidate["excavation"]:
        return "清单要求土方开挖，候选仅为土方运输/外运工序"
    if target["transport"] and candidate["excavation"] and not candidate["transport"]:
        return "清单要求土方运输/外运，候选仅为土方开挖工序"
    if target["scope"] and candidate["scope"] and target["scope"] != candidate["scope"]:
        return (
            f"土方开挖范围冲突（清单：{target['scope']}；"
            f"定额/组成：{candidate['scope']}）"
        )
    return ""


def _integrated_transport_scope(text: str) -> bool:
    """Return true when excavation already includes hauling or disposal."""
    _, transport = _earthwork_process_profile(text)
    normalized = _match_text(text)
    return transport and any(term in normalized for term in (
        "\u542b\u573a\u5185\u5916", "\u542b\u5012\u8fd0", "\u542b\u5916\u5f03",
        "\u8fd0\u8ddd\u7efc\u5408\u8003\u8651", "\u5916\u5f03\u8fd0\u8ddd\u7efc\u5408\u8003\u8651",
    ))


def _standalone_transport_scope(text: str) -> bool:
    """Return true for a transport-only quota or BOQ row."""
    excavation, transport = _earthwork_process_profile(text)
    normalized = _match_text(text)
    return transport and not excavation and any(term in normalized for term in (
        "\u571f\u65b9\u8fd0\u8f93", "\u8fd0\u571f", "\u81ea\u5378\u6c7d\u8f66", "\u88c5\u8fd0",
    ))


def _cjk_bigrams(value) -> set[str]:
    """Extract Chinese word fragments used to anchor AI output to the BOQ object."""
    chars = re.sub(r"[^\u4e00-\u9fff]", "", _match_text(value))
    return {chars[index:index + 2] for index in range(len(chars) - 1)}


def _location_groups(text: str) -> set[str]:
    normalized = _match_text(text)
    return {
        group
        for group, terms in LOCATION_GROUPS.items()
        if any(term in normalized for term in terms)
    }


def _strong_craft_terms(text: str) -> set[str]:
    return _text_terms(text, STRONG_CRAFT_TERMS)


def _road_scope_groups(text: str) -> set[str]:
    return _semantic_groups(text, ROAD_SCOPE_GROUPS)


def _core_object_groups(text: str) -> set[str]:
    """Return named engineering objects/material systems.

    These groups intentionally ignore modifiers such as colour, thickness,
    dimensions, grade, and generic words like "installation". They are used
    as hard evidence for candidate selection, so a visually similar but
    different object cannot win on fuzzy text length alone.
    """
    normalized = _match_text(text)
    groups = {
        group
        for group, terms in CORE_OBJECT_GROUPS.items()
        if any(term in normalized for term in terms)
    }
    if _stone_word_is_tile_appearance(text):
        groups.difference_update({"石材", "大理石"})
        if any(term in normalized for term in ("墙砖", "瓷砖墙面", "面砖墙面")):
            groups.add("墙砖")
        else:
            groups.add("地砖")
    if (
        "钢筋混凝土结构层" in normalized
        and any(term in normalized for term in TILE_OBJECT_TERMS)
        and not any(term in normalized for term in ("结构层浇筑", "现浇结构层", "新建结构层"))
    ):
        groups.discard("混凝土")
    return groups


def _core_object_conflict(target_text: str, candidate_text: str) -> str:
    target_groups = _core_object_groups(target_text)
    candidate_groups = _core_object_groups(candidate_text)
    if not target_groups or not candidate_groups:
        return ""
    # Related aliases are grouped above; unrelated named systems must not be
    # substituted just because they share a location or a generic craft word.
    target_material = _core_material_family_groups(target_text)
    candidate_material = _core_material_family_groups(candidate_text)
    if target_material and candidate_material:
        if not target_material.intersection(candidate_material) or candidate_material.difference(target_material):
            return (
                "核心工程对象/材料体系冲突（清单："
                + "/".join(sorted(target_groups))
                + "；定额/组成："
                + "/".join(sorted(candidate_groups))
                + "）"
            )
        # A matching material may be stored as a component-only quota while
        # the BOQ name also carries an assembly/scope word such as 景墙 or
        # 骨架. Missing those ancillary words is handled by clause coverage;
        # it is not a material conflict and must not prevent a valid
        # supplemental quota from being selected.
        return ""
    missing_groups = target_groups.difference(candidate_groups)
    if not missing_groups:
        return ""
    return (
        "候选定额/组成缺少清单核心工程对象："
        + "/".join(sorted(missing_groups))
    )


def extract_quota_key_terms(item_name: str = "", feature: str = "") -> dict:
    """提取清单中高辨识度的工艺和道路部位词，供本地匹配与 AI 共同约束。"""
    text = f"{_text(item_name)} {_text(feature)}".strip()
    crafts = _strong_craft_terms(text)
    road_groups = _road_scope_groups(text)
    return {
        "crafts": sorted(crafts),
        "road_scope_groups": sorted(road_groups),
        "strict": bool(crafts or road_groups) and len(crafts) <= 2,
        "required": [
            *[f"工艺:{value}" for value in sorted(crafts)],
            *[f"部位:{value}" for value in sorted(road_groups)],
        ],
    }


def extract_material_requirements(item_name: str = "", feature: str = "", unit: str = "") -> list[dict]:
    """Split BOQ material evidence into auditable matching fields.

    The original clause is retained. Matching consumers can use the material
    object and explicit specifications as hard evidence, while qualifiers such
    as colour, finish, location, and construction method remain constraints.
    """
    requirements = []
    for clause in boq_cost_clauses(item_name, feature):
        text = _text(clause)
        material_terms = _clause_material_terms(text)
        object_groups = _core_object_groups(text)
        if not material_terms and not object_groups:
            continue
        material_candidates = set(material_terms) | set(object_groups)
        material_name = max(
            material_candidates,
            key=lambda value: (len(_match_text(value)), len(value)),
            default="",
        )
        specifications = sorted(
            set(_technical_specs(text)) | {
                f"尺寸:{'x'.join(str(number).rstrip('0').rstrip('.') for number in size)}"
                for size in _size_specs(text)
            }
        )
        normalized = _match_text(text)
        material_normalized = _match_text(material_name)
        qualifier_text = normalized.replace(material_normalized, "") if material_normalized else normalized
        qualifiers = re.sub(
            r"(?:材料|品种|规格型号|规格|型号|厚度|材质|颜色|综合考虑|：|:|、)",
            "",
            qualifier_text,
        )
        requirements.append({
            "source_clause": text,
            "material_name": material_name,
            "material_terms": sorted(set(material_terms) | ({material_name} if material_name else set())),
            "core_object_groups": sorted(object_groups),
            "specifications": specifications,
            "qualifiers": qualifiers,
            "crafts": sorted(_strong_craft_terms(text)),
            "locations": sorted(_location_groups(text)),
            "unit": _text(unit),
        })
    return requirements


def extract_resource_requirements(item_name: str = "", feature: str = "", unit: str = "") -> dict:
    """Build the same structured evidence for material, labour, and machinery."""
    text = f"{_text(item_name)} {_text(feature)}"
    material_requirements = extract_material_requirements(item_name, feature, unit)
    labor_terms = sorted(_text_terms(text, (
        "人工", "安装", "制作", "运输", "清理", "修整", "配合施工",
    )))
    machine_requirement = infer_mechanical_requirement(item_name, feature)
    return {
        "materials": material_requirements,
        "labor": {
            "required": bool(labor_terms or material_requirements),
            "terms": labor_terms,
        },
        "machinery": machine_requirement,
    }


def infer_mechanical_requirement(item_name: str = "", feature: str = "") -> dict:
    """Infer whether a BOQ item requires a mechanical cost component.

    This is a construction-method guard, not a price estimate. Explicit manual
    wording wins over broad craft terms, while an explicit machine name keeps
    the requirement active.
    """
    text = _match_text(f"{_text(item_name)} {_text(feature)}")
    if not text:
        return {"required": False, "profiles": [], "terms": [], "reason": ""}
    matched = []
    terms = []
    for profile, profile_terms in MECHANICAL_CRAFT_RULES:
        hits = [term for term in profile_terms if term in text]
        if hits:
            matched.append(profile)
            terms.extend(hits)
    # Context guards for terms that frequently occur in material/process
    # descriptions without meaning machine construction.
    compaction_context = (
        "土", "路", "基层", "地基", "回填", "夯", "碾", "场地", "密实度",
    )
    if "压实碾压" in matched and "压实" in terms and not any(
        value in text for value in compaction_context
    ):
        matched.remove("压实碾压")
        terms = [term for term in terms if term != "压实"]
    if "起重吊装" in matched and "提升" in terms and "泵" in text:
        # “提升泵” is a pump type, not a lifting operation.
        matched.remove("起重吊装")
        terms = [term for term in terms if term != "提升"]
    explicit_machine = any(
        term in text
        for term in ("机械", "挖掘机", "压路机", "打夯机", "钻机", "桩机", "吊车", "泵车", "摊铺机", "洒布车")
    )
    manual_override = any(term in text for term in MANUAL_OVERRIDE_TERMS)
    required = bool(matched or explicit_machine)
    if manual_override and not explicit_machine:
        required = False
    unique_terms = list(dict.fromkeys(terms))
    reason = (
        f"识别机械工艺：{'、'.join(dict.fromkeys(matched))}（{'、'.join(unique_terms[:5])}）"
        if required else ""
    )
    return {
        "required": required,
        "profiles": list(dict.fromkeys(matched)),
        "terms": unique_terms,
        "reason": reason,
        "manual_override": manual_override,
    }


def labor_coordination_evidence(item_name: str = "", feature: str = "") -> list[str]:
    """Return explicit evidence that a quota includes manual coordination.

    This is intentionally narrower than merely seeing an ``人工费`` category:
    a full manual excavation quota must not be treated as the manual part of a
    machine excavation quota. The evidence is taken from the quota name and
    work content, not inferred from price.
    """
    text = _match_text(f"{_text(item_name)} {_text(feature)}")
    return list(dict.fromkeys(
        term for term in LABOR_COORDINATION_TERMS if term in text
    ))


def quota_covers_labor_coordination(item_name: str = "", feature: str = "") -> bool:
    """Whether a quota's work content already includes manual coordination."""
    return bool(labor_coordination_evidence(item_name, feature))


def is_full_manual_excavation_quota(item_name: str = "", feature: str = "") -> bool:
    """Identify a standalone manual excavation operation.

    The guard only applies when a machine quota already covers coordination;
    standalone manual excavation remains a valid match for manual-only BOQs.
    """
    text = _match_text(f"{_text(item_name)} {_text(feature)}")
    return any(term in text for term in FULL_MANUAL_EXCAVATION_TERMS)


def mechanical_dedup_keys(item_name: str = "", feature: str = "") -> set[str]:
    """Build a construction-aware identity for mechanical quota de-duplication.

    The same broad profile can contain different scopes, equipment sizes,
    loading modes, or haul distances. Those variants must remain separate;
    only the same profile with the same operating constraints is deduplicated.
    """
    text = _match_text(f"{_text(item_name)} {_text(feature)}")
    requirement = infer_mechanical_requirement(item_name, feature)
    if not requirement.get("profiles"):
        return set()
    earthwork = _earthwork_profile(text)
    distance_tokens = tuple(sorted(set(re.findall(
        r"\d+(?:\.\d+)?(?:km|公里|m|米)", text, flags=re.IGNORECASE
    ))))
    equipment_tokens = tuple(sorted(set(re.findall(
        r"(?:\d+(?:\.\d+)?\s*(?:m3|m³|吨|t|m|米)(?:以内|以上|以下)?|履带式|轮胎式|自卸汽车|挖掘机|压路机|夯实机|钻机)",
        text,
        flags=re.IGNORECASE,
    ))))
    signature = (
        earthwork.get("material", ""),
        earthwork.get("scope", ""),
        "loading" if earthwork.get("loading") else "",
        "transport" if earthwork.get("transport") else "",
        distance_tokens,
        equipment_tokens,
    )
    return {
        f"{profile}|{signature}"
        for profile in requirement.get("profiles") or []
    }


def required_cost_categories(
    item_name: str = "",
    feature: str = "",
    source_costs: dict | None = None,
) -> dict:
    """Derive cost-category requirements before a quota can be accepted.

    This deliberately distinguishes an absent source value from a zero value.
    A source cost or a clearly stated construction/material term makes that
    category mandatory. Management and profit are always checked separately by
    the UI fee calculation and are never silently discarded.
    """
    source_costs = source_costs or {}
    text = f"{_text(item_name)} {_text(feature)}"
    normalized = _match_text(text)
    mechanical = infer_mechanical_requirement(item_name, feature)
    has_craft = bool(_text_terms(text, CRAFT_TERMS))
    explicit_material = bool(_clause_material_terms(text))
    material_only = explicit_material and not has_craft and not mechanical["required"]
    requirements = {
        "人工费": _number(source_costs.get("labor")) > 0 or has_craft or not material_only,
        "材料费": _number(source_costs.get("material")) > 0 or explicit_material,
        "机械费": _number(source_costs.get("machinery")) > 0 or mechanical["required"],
        "管理费": True,
        "利润": True,
    }
    reasons = {
        "人工费": "原表有人工费或清单存在施工工艺" if requirements["人工费"] else "未识别到人工施工依据",
        "材料费": "原表有材料费或清单明确材料/规格" if requirements["材料费"] else "清单未明确材料组成",
        "机械费": mechanical["reason"] or (
            "原表有机械费" if _number(source_costs.get("machinery")) > 0 else "未识别到机械施工依据"
        ),
        "管理费": "按取费设置生成",
        "利润": "按取费设置生成",
    }
    return {
        "required": requirements,
        "reasons": reasons,
        "mechanical": mechanical,
        "source_costs": {
            key: _number(value)
            for key, value in source_costs.items()
        },
    }


def cost_category_coverage(
    item_name: str = "",
    feature: str = "",
    compositions: Iterable[dict] = (),
    source_costs: dict | None = None,
    fee_details: Iterable[dict] = (),
) -> dict:
    """Return an auditable five-category coverage result for one BOQ row."""
    requirements = required_cost_categories(item_name, feature, source_costs)
    present = {"人工费": False, "材料费": False, "机械费": False, "管理费": False, "利润": False}
    reference_only = set()
    actual_categories = set()
    for detail in [*(compositions or ()), *(fee_details or ())]:
        if not isinstance(detail, dict) or not detail.get("included", True):
            continue
        category = _text(detail.get("cat") or detail.get("category"))
        if not category:
            continue
        actual_categories.add(category)
        if detail.get("sourceCostReferenceCategory"):
            reference_only.add(category)
        if category == "人工费":
            present["人工费"] = True
        elif category in MATERIAL_COMPONENT_CATEGORIES:
            present["材料费"] = True
        elif category == "机械费":
            present["机械费"] = True
        elif category in {"管理费", "利润"}:
            present[category] = True
    missing = [
        category
        for category, required in requirements["required"].items()
        if required and not present[category]
    ]
    return {
        "required": dict(requirements["required"]),
        "present": present,
        "missing": missing,
        "reference_only": sorted(reference_only),
        "actual_categories": sorted(actual_categories),
        "reasons": dict(requirements["reasons"]),
        "mechanical": dict(requirements["mechanical"]),
        "complete": not missing and not reference_only,
    }


def _category_candidate_evidence(
    category: str,
    item_name: str,
    feature: str,
    entry: dict,
    source_costs: dict | None = None,
) -> tuple[bool, float, list[str]]:
    """Check category-specific evidence before a supplemental quota is used."""
    source_costs = source_costs or {}
    if category in MATERIAL_COMPONENT_CATEGORIES:
        category = "材料费"
    target_text = f"{_text(item_name)} {_text(feature)}"
    candidate_text = _text(entry.get("raw_logic_text"))
    target_mechanical = infer_mechanical_requirement(item_name, feature)
    candidate_mechanical = infer_mechanical_requirement(
        entry.get("item").name if entry.get("item") is not None else "",
        candidate_text,
    )
    target_anchors = _cjk_bigrams(target_text)
    candidate_anchors = _cjk_bigrams(candidate_text)
    shared_anchors = target_anchors.intersection(candidate_anchors)
    shared_materials = _clause_material_terms(target_text).intersection(
        _clause_material_terms(candidate_text)
    )
    shared_crafts = _text_terms(target_text, CRAFT_TERMS).intersection(
        _text_terms(candidate_text, CRAFT_TERMS)
    )
    shared_locations = _location_groups(target_text).intersection(
        _location_groups(candidate_text)
    )
    reasons = []
    if category == "机械费":
        if target_mechanical["required"]:
            if not candidate_mechanical["required"]:
                return False, 0.0, ["候选没有对应机械施工依据"]
            target_profiles = set(target_mechanical.get("profiles") or [])
            candidate_profiles = set(candidate_mechanical.get("profiles") or [])
            if target_profiles and candidate_profiles and not target_profiles.intersection(candidate_profiles):
                return False, 0.0, ["候选机械工艺与清单机械工艺不一致"]
            reasons.append(candidate_mechanical.get("reason") or "机械工艺一致")
        elif _number(source_costs.get("machinery")) > 0:
            # Some exported BOQs carry a machine cost column while their
            # feature text only describes the object/process (for example a
            # brick wall). Do not discard that evidence merely because the
            # text does not contain the word "mechanical". The candidate is
            # already restricted to a mechanical composition; it still needs
            # an object, material, craft, or location connection to the BOQ.
            object_evidence = (
                shared_anchors
                or shared_materials
                or shared_crafts
                or shared_locations
            )
            if not object_evidence:
                return False, 0.0, ["原表虽有机械费，但候选缺少与清单对象相符的依据"]
            reasons.append("原表有机械费，候选为相关机械组成")
        else:
            return False, 0.0, ["清单未提供机械费或机械施工依据"]
    elif category == "材料费":
        if not (shared_materials or _technical_specs(target_text).intersection(_technical_specs(candidate_text))):
            return False, 0.0, ["候选材料或规格与清单不一致"]
        reasons.append("材料/规格证据一致")
    elif category == "人工费":
        if not (shared_anchors or shared_crafts or shared_locations or shared_materials):
            return False, 0.0, ["候选人工工序缺少清单对象或工艺依据"]
        reasons.append("人工工序与清单对象相关")
    else:
        if not (shared_anchors or shared_crafts or shared_locations or shared_materials):
            return False, 0.0, ["候选缺少清单对象依据"]
    evidence_bonus = min(
        12.0,
        len(shared_anchors) * 2.0
        + len(shared_materials) * 4.0
        + len(shared_crafts) * 3.0
        + len(shared_locations) * 3.0,
    )
    return True, evidence_bonus, reasons


def _unit_logic(target_unit: str, candidate_unit: str) -> tuple[bool, float, str]:
    target = _canonical_unit(target_unit)
    candidate = _canonical_unit(candidate_unit)
    if not target or not candidate:
        return True, 0.0, ""
    if target == candidate:
        return True, 6.0, "单位一致"
    target_dimension = UNIT_DIMENSIONS.get(target)
    candidate_dimension = UNIT_DIMENSIONS.get(candidate)
    if target_dimension and candidate_dimension and target_dimension != candidate_dimension:
        return False, 0.0, "单位量纲冲突"
    return True, -4.0, "单位不同"


def _unit_conversion_basis_is_actionable(
    conversion: str,
    *,
    category: str = "",
    component_name: str = "",
    boq_unit: str = "",
    component_unit: str = "",
) -> bool:
    """Accept only a conversion note that can be checked by a reviewer.

    A note such as ``已换算`` is not enough to turn m3 into m2. The text must
    contain a conversion action and either a physical basis (thickness,
    density, weight, etc.), an explicit per-unit/formula expression, or an
    identifiable labour conversion. This keeps AI estimates from silently
    treating every different unit as one unit of consumption.
    """
    text = _match_text(conversion).lower()
    if not text:
        return False
    # An explicit positive consumption ratio is already independently
    # checkable, even when the sentence omits words such as ``换算``.  Typical
    # AI output is ``0.03m3/m2`` or ``每米2个、每个0.5kg，共1kg``.
    if _declared_conversion_quantity(conversion, boq_unit, component_unit) is not None:
        return True
    action_terms = ("换算", "折算", "折合", "计量")
    physical_terms = (
        "厚度", "深度", "长度", "面积", "体积", "重量", "质量", "密度",
        "直径", "宽度", "高度", "每", "按", "数量",
    )
    formula_terms = ("=", "÷", "/", "×", "*", "每")
    has_physical_basis = any(term in text for term in physical_terms)
    has_formula = any(term in text for term in formula_terms)
    has_action = any(term in text for term in action_terms)
    if has_action and has_physical_basis and (has_formula or "按" in text):
        return True

    # Common quota wording gives labour conversion as ``安装人工按块数换算``
    # without repeating the labour unit. The category/name anchors make this
    # exception specific instead of accepting any vague conversion sentence.
    category_text = _match_text(category).lower()
    component_text = _match_text(component_name).lower()
    is_labor = "人工" in category_text or "人工" in component_text or "工日" in component_text
    if is_labor and "人工" in text and "按" in text:
        return True

    # If both units are written in the formula, the expression is reviewable
    # even when the physical basis uses project-specific wording.
    unit_texts = {
        _canonical_unit(boq_unit),
        _canonical_unit(component_unit),
        _match_text(boq_unit).lower(),
        _match_text(component_unit).lower(),
    } - {""}
    explicit_unit_formula = (
        len(unit_texts) >= 2
        and sum(value in text for value in unit_texts) >= 2
        and has_formula
        and has_physical_basis
    )
    return explicit_unit_formula


def _declared_conversion_quantity(
    conversion: str,
    boq_unit: str,
    component_unit: str,
) -> float | None:
    """Read an explicit per-BOQ-unit consumption from an AI formula."""
    text = unicodedata.normalize("NFKC", _text(conversion)).lower().replace("³", "3").replace("²", "2")
    boq_canonical = _canonical_unit(boq_unit)
    component_canonical = _canonical_unit(component_unit)
    boq = re.escape(boq_canonical)
    component = re.escape(component_canonical)
    if not text or not boq or not component:
        return None
    # Accept explicit natural-language ratios such as ``每米管道含管件0.3个``
    # and ``每公里电杆数量=20根``. Only numeric, unit-bearing expressions
    # are accepted; vague market assumptions remain review-required.
    unit_words = {
        "m": r"(?:m|米|延米|米长)",
        "km": r"(?:km|公里|千米)",
        "m2": r"(?:m2|平方米|平米)",
        "m3": r"(?:m3|立方米|方)",
        "t": r"(?:t|吨)",
        "kg": r"(?:kg|千克|公斤)",
        "个": r"(?:个|件|只)",
        "根": r"(?:根)",
        "套": r"(?:套)",
    }
    boq_word = unit_words.get(boq_canonical, boq)
    component_word = unit_words.get(component_canonical, component)
    number = r"(\d+(?:\.\d+)?)"

    # A component can be expressed through an intermediate count unit:
    # ``每公里按20根电杆，每根1套拉线`` means 20套/km. Resolve the
    # intermediate ratio before the direct ratio so ``1套/根`` is not
    # incorrectly returned as the final per-km quantity.
    relation_units = {
        "个": "(?:个|件|只)",
        "根": "(?:根)",
        "套": "(?:套)",
        "组": "(?:组|付)",
        "台": "(?:台|台班)",
        "块": "(?:块|片)",
    }
    for base_canonical, base_word in relation_units.items():
        relation_pattern = (
            rf"每\s*(?:1\s*)?{base_word}[^\n,，。;；]{{0,40}}?"
            rf"(?:1|一)\s*{component_word}"
        )
        if not re.search(relation_pattern, text):
            continue
        base_pattern = (
            rf"(?:每\s*)?(?:1\s*)?{boq_word}[^\n,，。;；]{{0,140}}?"
            rf"{number}\s*{base_word}"
        )
        base_match = re.search(base_pattern, text)
        if base_match:
            value = float(base_match.group(1))
            if value > 0:
                return value

    natural_patterns = (
        rf"每\s*(?:1\s*)?{boq_word}[^\n,，。;；]{{0,140}}?{number}\s*{component_word}",
        rf"(?:每\s*)?(?:1\s*)?{boq_word}[^\n,，。;；]{{0,140}}?{number}\s*{component_word}",
        rf"{number}\s*{component_word}\s*/\s*(?:1\s*)?{boq_word}",
    )
    for pattern in natural_patterns:
        match = re.search(pattern, text)
        if match:
            value = float(match.group(1))
            if value > 0:
                return value

    # Equation form: 1km / 40m/根 = 25根/km. The right-hand ratio is
    # already the required consumption per BOQ unit.
    equation_pattern = (
        rf"{number}\s*{boq_word}[^=\n]{{0,120}}?=\s*{number}\s*{component_word}\s*/\s*{boq_word}"
    )
    match = re.search(equation_pattern, text)
    if match:
        value = float(match.group(2))
        if value > 0:
            return value

    unit_patterns = {
        "m2": r"(?:m2|㎡|平方米|平方)",
        "m3": r"(?:m3|立方米|立方)",
        "m": r"(?:m(?![m23])|米|延米)",
        "kg": r"(?:kg|千克|公斤)",
        "t": r"(?:t|吨)",
    }
    boq_pattern = unit_patterns.get(boq_canonical, boq)
    component_pattern = unit_patterns.get(component_canonical, component)
    patterns = (
        rf"每\s*1?\s*{boq}[^0-9]{{0,20}}(?:用量|含量|体积|重量)?\s*[=:：为]?\s*(\d+(?:\.\d+)?)\s*{component}",
        rf"(\d+(?:\.\d+)?)\s*{component}\s*/\s*{boq}",
        rf"(?:含量|用量|单位含量)\s*[=:：为]?\s*(\d+(?:\.\d+)?)\s*{component_pattern}\s*/\s*{boq_pattern}",
        rf"(\d+(?:\.\d+)?)\s*{component_pattern}\s*/\s*{boq_pattern}",
        rf"每\s*1?\s*{boq_pattern}.{{0,80}}?(?:共|合计|折合|用量|含量|重量|体积)\s*[=:：为]?\s*(\d+(?:\.\d+)?)\s*{component_pattern}",
    )
    values = []
    for pattern in patterns:
        values.extend(float(match.group(1)) for match in re.finditer(pattern, text))
    values = [value for value in values if value > 0]
    return min(values) if values else None


def _derived_material_supported_by_work(
    boq_text: str,
    component_text: str,
    component_materials: set[str],
) -> str:
    """Return the explicit work clause supporting a necessary derived material.

    Construction BOQs often name a process/object (垫层、砌井、找平层) rather
    than repeat every material.  This is weaker than an explicit material name,
    so it becomes a review warning instead of bypassing unrelated-material
    checks globally.
    """
    target = _match_text(boq_text)
    component = _match_text(component_text)
    rules = (
        (("垫层",), ("混凝土", "水泥", "砂浆", "碎石"), ("垫层",)),
        (("找平层", "找平"), ("砂浆", "水泥"), ("找平", "砂浆")),
        (("砌筑", "砌体", "砖墙"), ("砖", "砂浆", "水泥"), ("砌", "井体", "墙体")),
        (("雨水井", "雨水口", "进水井", "检查井"), ("砖", "混凝土", "砂浆", "水泥"), ("井体", "井圈", "垫层", "砌")),
        (("基础",), ("混凝土", "碎石", "砂浆", "水泥"), ("基础", "垫层")),
        (("固定配件", "固定件", "固定"), ("钢材", "方钢", "方钢管", "镀锌方钢"), ("龙骨", "骨架", "固定")),
    )
    for target_terms, allowed_materials, component_terms in rules:
        if (
            any(term in target for term in target_terms)
            and component_materials.intersection(allowed_materials)
            and any(term in component for term in component_terms)
        ):
            return next(term for term in target_terms if term in target)
    return ""


def _infer_unit_conversion_basis(
    source_text: str,
    boq_unit: str,
    component_unit: str,
) -> str:
    """Derive only simple physical conversions from explicit BOQ data.

    The conversion is intentionally limited to area/volume and an explicit
    thickness or depth. It never treats a raw quantity as proof of a missing
    dimension, and it does not guess density, weight, loss, or workmanship.
    """
    boq = _canonical_unit(boq_unit)
    component = _canonical_unit(component_unit)
    if not source_text or boq == component:
        return ""
    if {boq, component} != {"m2", "m3"}:
        return ""
    explicit_thickness = re.search(
        r"(?:厚度|厚|深度|深|压实厚度)\s*[:：]?\s*(\d+(?:\.\d+)?)\s*(mm|cm|m)",
        _text(source_text),
        flags=re.IGNORECASE,
    )
    if not explicit_thickness:
        explicit_thickness = re.search(
            r"(\d+(?:\.\d+)?)\s*(mm|cm|m)\s*(?:厚度|厚|深度|深)",
            _text(source_text),
            flags=re.IGNORECASE,
        )
    if explicit_thickness:
        amount = float(explicit_thickness.group(1))
        unit = explicit_thickness.group(2).lower()
        thickness_mm = amount * (10 if unit == "cm" else 1000 if unit == "m" else 1)
        thickness_m = thickness_mm / 1000
        if boq == "m2" and component == "m3":
            return f"单位换算：每1m2按明确厚度{thickness_mm:g}mm计算体积={thickness_m:g}m3"
        reciprocal = 1 / thickness_m
        return f"单位换算：每1m3按明确厚度{thickness_mm:g}mm折算面积={reciprocal:g}m2"
    dimensions = _specifications_by_family(source_text)
    thickness_tokens = dimensions.get("thickness") or dimensions.get("depth") or set()
    if len(thickness_tokens) != 1:
        return ""
    thickness_mm = _spec_numeric_mm(next(iter(thickness_tokens)))
    if not thickness_mm or thickness_mm <= 0:
        return ""
    thickness_m = thickness_mm / 1000
    if boq == "m2" and component == "m3":
        return (
            f"单位换算：每1m2按明确厚度{thickness_mm:g}mm（{thickness_m:g}m）"
            f"计算体积，1m2×{thickness_m:g}m={thickness_m:g}m3"
        )
    reciprocal = 1 / thickness_m
    return (
        f"单位换算：每1m3按明确厚度{thickness_mm:g}mm（{thickness_m:g}m）"
        f"折算面积，1m3÷{thickness_m:g}m={reciprocal:g}m2"
    )


def _infer_area_volume_quantity(
    source_text: str,
    boq_unit: str,
    component_unit: str,
) -> float | None:
    """Return the per-BOQ-unit m2/m3 quantity for one explicit thickness."""
    boq = _canonical_unit(boq_unit)
    component = _canonical_unit(component_unit)
    if {boq, component} == {"m2", "m3"}:
        explicit_thickness = re.search(
            r"(?:厚度|厚|深度|深|压实厚度)\s*[:：]?\s*(\d+(?:\.\d+)?)\s*(mm|cm|m)",
            _text(source_text),
            flags=re.IGNORECASE,
        )
        if not explicit_thickness:
            explicit_thickness = re.search(
                r"(\d+(?:\.\d+)?)\s*(mm|cm|m)\s*(?:厚度|厚|深度|深)",
                _text(source_text),
                flags=re.IGNORECASE,
            )
        if explicit_thickness:
            amount = float(explicit_thickness.group(1))
            unit = explicit_thickness.group(2).lower()
            thickness_mm = amount * (10 if unit == "cm" else 1000 if unit == "m" else 1)
            thickness_m = thickness_mm / 1000
            return thickness_m if boq == "m2" and component == "m3" else 1 / thickness_m
        dimensions = _specifications_by_family(source_text)
        thickness_tokens = dimensions.get("thickness") or dimensions.get("depth") or set()
        if len(thickness_tokens) != 1:
            return None
        thickness_mm = _spec_numeric_mm(next(iter(thickness_tokens)))
        if not thickness_mm or thickness_mm <= 0:
            return None
        thickness_m = thickness_mm / 1000
        return thickness_m if boq == "m2" and component == "m3" else 1 / thickness_m
    if {boq, component} == {"t", "kg"}:
        return 1000.0 if boq == "t" and component == "kg" else 0.001
    return None


def _component_conversion_source(
    boq_name: str,
    boq_feature: str,
    component_name: str,
    component_feature: str,
    boq_unit: str,
    component_unit: str,
) -> str:
    """Select the BOQ clause that actually describes one component.

    A BOQ may contain several layers, such as an 18mm tile and a 30mm mortar
    bed. Inferring m3/m2 from the complete row is ambiguous, while the mortar
    component has an unambiguous 30mm source clause. Prefer the component's
    own feature, then rank individual BOQ clauses by material, craft and text
    relevance. Returning the full row is a final non-destructive fallback.
    """
    component_text = " ".join(
        value for value in (_text(component_name), _text(component_feature)) if value
    )
    if _infer_unit_conversion_basis(component_text, boq_unit, component_unit):
        return component_text

    component_materials = _clause_material_terms(component_text)
    component_crafts = _text_terms(component_text, CRAFT_TERMS)
    ranked: list[tuple[float, str]] = []
    for clause in split_boq_work_items(boq_name, boq_feature):
        if not _infer_unit_conversion_basis(clause, boq_unit, component_unit):
            continue
        clause_materials = _clause_material_terms(clause)
        clause_crafts = _text_terms(clause, CRAFT_TERMS)
        material_overlap = component_materials.intersection(clause_materials)
        craft_overlap = component_crafts.intersection(clause_crafts)
        relation = fuzz.WRatio(_match_text(component_text), _match_text(clause))
        score = relation + len(material_overlap) * 25 + len(craft_overlap) * 15
        if component_materials and clause_materials and not material_overlap and relation < 55:
            continue
        ranked.append((score, clause))
    if ranked:
        ranked.sort(key=lambda value: (value[0], len(value[1])), reverse=True)
        return ranked[0][1]
    return f"{_text(boq_name)} {_text(boq_feature)}".strip()


def _component_requires_physical_conversion(
    category: str,
    boq_unit: str,
    component_unit: str,
) -> bool:
    """Whether a different component unit needs a physical formula.

    Labour, machinery, and subcontract amounts are ordinary quota
    consumptions such as 0.08 工日/块 or 0.02 台班/m3. Their basis is the
    quantity itself. Material amounts with a different dimension, such as
    m3/m2 or kg/t, require a formula or an explicit dimension-derived ratio.
    """
    if category not in MATERIAL_COMPONENT_CATEGORIES:
        return False
    target = _canonical_unit(boq_unit)
    component = _canonical_unit(component_unit)
    if not target or not component or target == component:
        return False
    return True


def apply_explicit_unit_conversion(
    composition: dict,
    target_text: str,
    boq_unit: str,
    *,
    strict: bool = False,
) -> dict:
    """Attach and apply a deterministic material-unit conversion.

    This is shared by local quota matches and AI-generated compositions. The
    source quantity is retained in ``sourceQty`` by the caller; any correction
    is recorded in ``unitConversion`` and ``note`` so exports show why the
    displayed quantity changed.
    """
    result = dict(composition or {})
    category = _text(result.get("cat") or result.get("category"))
    component_unit = _text(result.get("unit"))
    requires = _component_requires_physical_conversion(category, boq_unit, component_unit)
    result["physicalUnitConversionRequired"] = requires
    result["requiresUnitConversion"] = bool(requires and strict)
    if not requires:
        return result
    existing = _text(result.get("unitConversion") or result.get("unit_conversion"))
    formula = existing or _infer_unit_conversion_basis(
        target_text,
        boq_unit,
        component_unit,
    )
    if formula:
        result["unitConversion"] = formula
    elif requires:
        result["unitConversionReview"] = (
            f"组成单位{component_unit}与清单单位{boq_unit}存在量纲差异，"
            "当前沿用本地定额单位含量，需核对定额适用单位"
        )
    factor = _infer_area_volume_quantity(target_text, boq_unit, component_unit)
    if factor is None or category not in MATERIAL_COMPONENT_CATEGORIES:
        return result
    raw_qty = _number(result.get("qty"), 0.0)
    if raw_qty > 0 and abs(raw_qty - factor) > max(0.000001, factor * 0.05):
        result["qty"] = round(factor, 8)
        correction = f"含量校正：按明确尺寸换算为{factor:g}{component_unit}/{boq_unit}"
        result["unitQuantityCorrection"] = correction
        result["note"] = "；".join(
            value for value in (_text(result.get("note")), correction) if value
        )
    return result


def _primary_process_conflict(target_text: str, candidate_text: str) -> str:
    """Reject a finishing-only quota for an installation/paving BOQ item."""
    target = _match_text(target_text)
    candidate = _match_text(candidate_text)
    target_requires_assembly = any(term in target for term in PAVING_ASSEMBLY_TERMS)
    candidate_is_surface_treatment = any(
        term in candidate for term in SURFACE_TREATMENT_ONLY_TERMS
    )
    candidate_has_assembly = any(term in candidate for term in PAVING_ASSEMBLY_TERMS)
    if target_requires_assembly and candidate_is_surface_treatment and not candidate_has_assembly:
        return "清单要求铺贴/安装成层，候选仅为结晶、打磨等表面处理工序"
    return ""


def _explicit_material_conflict(target_text: str, candidate_text: str) -> str:
    """Reject named material systems that are not interchangeable by similarity."""
    target = _canonical_material_text(target_text)
    candidate = _canonical_material_text(candidate_text)
    dry_mortar = "\u5e72\u786c\u6027\u6c34\u6ce5\u7802\u6d46"
    expansion_mortar = "\u81a8\u80c0\u6c34\u6ce5\u7802\u6d46"
    quartz_tile = "\u77f3\u82f1\u7816"
    face_tile = "\u9762\u7816"
    ceramic_tile_terms = (
        "\u74f7\u7816", "\u9676\u74f7\u7816", "\u9676\u74f7\u677f", "\u901a\u4f53\u5927\u7406\u77f3\u74f7\u7816",
        "\u5927\u7406\u77f3\u74f7\u7816", "\u4eff\u5927\u7406\u77f3\u74f7\u7816",
    )
    concrete = "\u6df7\u51dd\u571f"
    asphalt_concrete = "\u6ca5\u9752\u6df7\u51dd\u571f"
    if concrete in target and asphalt_concrete in candidate:
        return "清单要求普通/水泥混凝土，候选为沥青混凝土，材料体系不一致"
    if asphalt_concrete in target and concrete in candidate and asphalt_concrete not in candidate:
        return "清单要求沥青混凝土，候选为普通混凝土，材料体系不一致"
    if dry_mortar in target and expansion_mortar in candidate:
        return "清单要求干硬性水泥砂浆，候选为膨胀水泥砂浆，材料体系不一致"
    if "\u7d20\u6c34\u6ce5\u6d46" in target and expansion_mortar in candidate:
        return "清单要求素水泥浆，候选为膨胀水泥砂浆，材料体系不一致"
    if quartz_tile in target and quartz_tile not in candidate and any(
        term in candidate for term in ceramic_tile_terms + (face_tile,)
    ):
        return "清单明确石英砖，候选为其他陶瓷/面砖品种，材料体系不一致"
    # “大理石”默认指天然石材。候选明确写成“大理石瓷砖/通体大理石瓷砖”
    # 时不能仅因共享“大理石”三个字就放行；反向选择则允许，因为清单已
    # 明确要求的是陶瓷砖体系。
    engineered_marble_terms = tuple(
        term for term in ceramic_tile_terms if "大理石" in term
    )
    target_is_natural_marble = (
        "大理石" in target
        and not any(term in target for term in engineered_marble_terms)
    )
    candidate_is_engineered_marble = (
        "大理石" in candidate
        and any(term in candidate for term in engineered_marble_terms)
    )
    if target_is_natural_marble and candidate_is_engineered_marble:
        return "清单要求天然大理石，候选为大理石瓷砖，材料品种不一致"
    target_groups = {
        group for group, terms in OBJECT_CONFLICT_GROUPS.items()
        if any(term in target for term in terms)
    }
    candidate_groups = {
        group for group, terms in OBJECT_CONFLICT_GROUPS.items()
        if any(term in candidate for term in terms)
    }
    if target_groups and candidate_groups and not target_groups.intersection(candidate_groups):
        return (
            "工程对象不一致（清单："
            + "/".join(sorted(target_groups))
            + "；定额/组成："
            + "/".join(sorted(candidate_groups))
            + "）"
        )
    if "幕墙埋件" in candidate_groups and "幕墙埋件" not in target_groups:
        return "候选为幕墙埋件，但清单未明确幕墙部位"
    return ""


def _unsupported_special_scope(target_text: str, candidate_text: str) -> str:
    """Return a hard conflict for a scope asserted only by the candidate."""
    target = _match_text(target_text)
    candidate = _match_text(candidate_text)
    scopes = []
    for term in SPECIAL_SCOPE_TERMS:
        if term not in candidate or term in target:
            continue
        # Generic paving BOQs may omit the floor word, so do not reject a
        # generic floor quota unless the BOQ explicitly points to another
        # construction scope. Named exterior/interior scopes remain strict
        # because they change the construction method and quota family.
        if term == "楼地面" and not any(
            value in target for value in ("墙面", "墙体", "幕墙", "顶棚", "天棚", "吊顶", "屋面")
        ):
            continue
        scopes.append(term)
    scopes = sorted(scopes)
    return (
        "定额含清单未说明的专用部位:" + "/".join(scopes)
        if scopes else ""
    )


def quota_reference_scope_errors(boq_name: str, boq_feature: str, reference: dict) -> list[str]:
    """Validate a selected/generated quota and all of its components together."""
    target = f"{_text(boq_name)} {_text(boq_feature)}"
    errors = []
    quota_text = " ".join(
        _text(reference.get(key))
        for key in ("quota_code", "quota_name", "code", "name", "feature")
        if _text(reference.get(key))
    )
    process_conflict = _primary_process_conflict(target, quota_text)
    if process_conflict:
        errors.append(process_conflict)
    earthwork_conflict = _earthwork_logic_conflict(target, quota_text)
    if earthwork_conflict:
        errors.append(earthwork_conflict)
    conflict = _unsupported_special_scope(target, quota_text)
    if conflict:
        errors.append(conflict)
    for component in reference.get("compositions") or reference.get("components") or []:
        if not isinstance(component, dict):
            continue
        component_text = " ".join(
            _text(component.get(key))
            for key in ("code", "name", "feature", "note", "quotaName")
            if _text(component.get(key))
        )
        earthwork_conflict = _earthwork_logic_conflict(target, component_text)
        if earthwork_conflict:
            errors.append(
                f"组成“{_text(component.get('name')) or _text(component.get('code')) or '未命名'}”："
                f"{earthwork_conflict}"
            )
        material_conflict = _explicit_material_conflict(target, component_text)
        if material_conflict:
            errors.append(
                f"组成“{_text(component.get('name')) or _text(component.get('code')) or '未命名'}”："
                f"{material_conflict}"
            )
        component_category = _text(
            component.get("cat") or component.get("category")
        )
        if component_category in MATERIAL_COMPONENT_CATEGORIES:
            tile_size_conflict = _explicit_tile_plan_size_conflict(
                target,
                component_text,
            )
            if tile_size_conflict:
                errors.append(
                    f"组成“{_text(component.get('name')) or _text(component.get('code')) or '未命名'}”："
                    f"{tile_size_conflict}"
                )
            material_allowed, material_reason = _material_match_allowed(
                target,
                component_text,
            )
            if not material_allowed:
                errors.append(
                    f"组成“{_text(component.get('name')) or _text(component.get('code')) or '未命名'}”："
                    f"{material_reason}"
                )
        conflict = _unsupported_special_scope(target, component_text)
        if conflict:
            errors.append(
                f"组成“{_text(component.get('name')) or _text(component.get('code')) or '未命名'}”：{conflict}"
            )
    return list(dict.fromkeys(errors))


def _candidate_logic(
    target_text: str,
    candidate_text: str,
    target_unit: str,
    candidate_unit: str,
) -> tuple[bool, float, list[str]]:
    unit_ok, unit_score, unit_reason = _unit_logic(target_unit, candidate_unit)
    if not unit_ok:
        explicit_conversion = _infer_unit_conversion_basis(
            target_text,
            target_unit,
            candidate_unit,
        )
        if not explicit_conversion:
            return False, 0.0, [unit_reason]
        unit_ok = True
        unit_score = -2.0
        unit_reason = f"单位可按清单明确尺寸换算：{explicit_conversion}"
    process_conflict = _primary_process_conflict(target_text, candidate_text)
    if process_conflict:
        return False, 0.0, [process_conflict]
    material_conflict = _explicit_material_conflict(target_text, candidate_text)
    if material_conflict:
        return False, 0.0, [material_conflict]
    earthwork_conflict = _earthwork_logic_conflict(target_text, candidate_text)
    if earthwork_conflict:
        return False, 0.0, [earthwork_conflict]
    scope_conflict = _unsupported_special_scope(target_text, candidate_text)
    if scope_conflict:
        return False, 0.0, [scope_conflict]

    core_conflict = _core_object_conflict(target_text, candidate_text)
    if core_conflict:
        return False, 0.0, [core_conflict]
    target_core_groups = _core_object_groups(target_text)
    candidate_core_groups = _core_object_groups(candidate_text)
    if target_core_groups and not candidate_core_groups:
        return False, 0.0, [
            "候选定额/组成未提供清单明确的核心工程对象或材料体系"
        ]

    target_locations = _location_groups(target_text)
    candidate_locations = _location_groups(candidate_text)
    if target_locations and candidate_locations and not target_locations.intersection(candidate_locations):
        return False, 0.0, [
            f"工程部位冲突（清单:{'/'.join(sorted(target_locations))}；定额:{'/'.join(sorted(candidate_locations))}）"
        ]

    target_demolition = "拆除" in _match_text(target_text)
    candidate_demolition = "拆除" in _match_text(candidate_text)
    if target_demolition != candidate_demolition:
        return False, 0.0, ["拆除与新建/安装工序冲突"]

    target_key_terms = extract_quota_key_terms(target_text)
    candidate_key_terms = extract_quota_key_terms(candidate_text)
    target_crafts = set(target_key_terms["crafts"]) | _text_terms(
        target_text, STRONG_CRAFT_TERMS
    )
    candidate_crafts = set(candidate_key_terms["crafts"]) | _text_terms(
        candidate_text, STRONG_CRAFT_TERMS
    )
    target_road_groups = set(target_key_terms["road_scope_groups"])
    candidate_road_groups = set(candidate_key_terms["road_scope_groups"])
    if target_key_terms["strict"] and target_crafts:
        target_craft_groups = _semantic_groups(target_text, CRAFT_EQUIVALENT_GROUPS)
        candidate_craft_groups = _semantic_groups(candidate_text, CRAFT_EQUIVALENT_GROUPS)
        missing_crafts = {
            term for term in set(target_crafts)
            if term not in set(candidate_crafts)
            and not (
                _semantic_groups(term, CRAFT_EQUIVALENT_GROUPS)
                & candidate_craft_groups
            )
        }
        if missing_crafts:
            same_object_or_scope = bool(
                target_core_groups.intersection(candidate_core_groups)
                or target_locations.intersection(candidate_locations)
                or _core_material_family_groups(target_text).intersection(
                    _core_material_family_groups(candidate_text)
                )
            )
            if not same_object_or_scope:
                return False, 0.0, [
                    "缺少清单核心工艺词：" + "、".join(sorted(missing_crafts)),
                ]
            craft_review = (
                "候选未完整注明清单核心工艺词："
                + "、".join(sorted(missing_crafts))
                + "，已按同一工程对象/部位保留并需复核"
            )
        else:
            craft_review = ""
    else:
        craft_review = ""
    if target_road_groups and candidate_road_groups:
        conflicting_road_groups = target_road_groups - candidate_road_groups
        if conflicting_road_groups:
            return False, 0.0, [
                "关键道路部位冲突：" + "、".join(sorted(conflicting_road_groups)),
            ]
    target_normalized = _match_text(target_text).lower()
    candidate_normalized = _match_text(candidate_text).lower()
    reasons = [unit_reason] if unit_reason else []
    if craft_review:
        reasons.append(craft_review)
    material_ok, material_reason = _material_match_allowed(target_text, candidate_text)
    if not material_ok:
        return False, 0.0, [material_reason]
    target_materials = _clause_material_terms(target_text)
    candidate_materials = _clause_material_terms(candidate_text)
    if not target_materials and candidate_materials:
        derived_material_basis = _derived_material_supported_by_work(
            target_text,
            candidate_text,
            candidate_materials,
        )
        if not derived_material_basis:
            return False, 0.0, [
                "清单未明确材料对象，候选却包含未经工作内容推导的材料："
                + "/".join(sorted(candidate_materials))
            ]
    target_material_families = _core_material_family_groups(target_text)
    candidate_material_families = _core_material_family_groups(candidate_text)
    layered_mix_substitution = bool(
        "砂浆" in _match_text(target_text)
        and "砂浆" in _match_text(candidate_text)
        and (
            target_core_groups.intersection(candidate_core_groups)
            or target_material_families.intersection(candidate_material_families)
        )
        and any(
            term in _match_text(target_text)
            for term in ("找平", "结合层", "粘结层", "粘接层", "铺贴", "铺设", "铺装")
        )
        and any(
            term in _match_text(candidate_text)
            for term in ("找平", "结合层", "粘结层", "粘接层", "铺贴", "铺设", "铺装")
        )
    )
    specification = specification_relation(
        target_text,
        candidate_text,
        allow_quantity_conversion=True,
        require_target_specs=True,
        allow_missing_convertible_specs=bool(
            _infer_unit_conversion_basis(target_text, target_unit, candidate_unit)
        ),
        allow_missing_layered_specs=True,
        allow_layered_mix_substitution=layered_mix_substitution,
    )
    specification_penalty = 0.0
    if specification["errors"]:
        return False, 0.0, specification["errors"]
    if specification["warnings"]:
        # The candidate remains usable only as a controlled quantity
        # conversion. The caller applies the factor to affected materials;
        # labor and machinery are never scaled by this warning alone.
        reasons.extend(specification["warnings"])
        # A generic quota is usable when it has the same object and process,
        # but it must rank below a candidate that states the BOQ layer/spec.
        specification_penalty = min(6.0, 1.5 * len(specification["warnings"]))
    if (
        "水泥混凝土" in target_normalized
        and "沥青混凝土" in candidate_normalized
    ) or (
        "沥青混凝土" in target_normalized
        and "水泥混凝土" in candidate_normalized
    ):
        return False, 0.0, ["水泥混凝土与沥青混凝土材料体系冲突"]

    bonus = unit_score - specification_penalty
    if target_locations.intersection(candidate_locations):
        bonus += 6
        reasons.append(f"工程部位一致:{'/'.join(sorted(target_locations.intersection(candidate_locations)))}")
    if target_core_groups.intersection(candidate_core_groups):
        bonus += min(18, len(target_core_groups.intersection(candidate_core_groups)) * 9)
        reasons.append(
            "核心对象/材料体系一致："
            + "/".join(sorted(target_core_groups.intersection(candidate_core_groups)))
        )
    target_crafts = _text_terms(target_text, CRAFT_TERMS)
    candidate_crafts = _text_terms(candidate_text, CRAFT_TERMS)
    common_crafts = target_crafts.intersection(candidate_crafts)
    if common_crafts:
        bonus += min(8, len(common_crafts) * 3)
        reasons.append(f"工艺一致:{'/'.join(sorted(common_crafts))}")
    common_strong_crafts = set(target_key_terms["crafts"]).intersection(
        candidate_key_terms["crafts"]
    )
    if common_strong_crafts:
        bonus += min(12, len(common_strong_crafts) * 5)
        reasons.append("关键工艺一致：" + "、".join(sorted(common_strong_crafts)))
    common_road_scopes = set(target_key_terms["road_scope_groups"]).intersection(
        candidate_key_terms["road_scope_groups"]
    )
    if common_road_scopes:
        bonus += min(12, len(common_road_scopes) * 5)
        reasons.append("关键道路部位一致：" + "、".join(sorted(common_road_scopes)))

    target_materials = _clause_material_terms(target_text)
    candidate_materials = _clause_material_terms(candidate_text)
    common_materials = target_materials.intersection(candidate_materials)
    if common_materials:
        bonus += min(8, len(common_materials) * 3)
        reasons.append(f"材料一致:{'/'.join(sorted(common_materials))}")
    target_sizes = _size_specs(target_text)
    candidate_sizes = _size_specs(candidate_text)
    if target_sizes and target_sizes == candidate_sizes:
        bonus += 8
        reasons.append("明确规格尺寸一致")
    target_earthwork = _earthwork_profile(target_text)
    candidate_earthwork = _earthwork_profile(candidate_text)
    if target_earthwork["material"] and target_earthwork["material"] == candidate_earthwork["material"]:
        bonus += 12
        reasons.append(f"土石方对象一致:{target_earthwork['material']}")
    if target_earthwork["excavation"] and candidate_earthwork["excavation"]:
        bonus += 8
        reasons.append("开挖工艺一致")
    if target_earthwork["transport"] and candidate_earthwork["transport"]:
        bonus += 8
        reasons.append("运输/外运工艺一致")
    if target_earthwork["loading"] and candidate_earthwork["loading"]:
        bonus += 6
        reasons.append("挖装/装车工艺一致")
    if target_earthwork["scope"] and target_earthwork["scope"] == candidate_earthwork["scope"]:
        bonus += 8
        reasons.append(f"开挖范围一致:{target_earthwork['scope']}")
    return True, bonus, reasons


def _number(value, default=0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(str(value).replace(",", "").replace("%", ""))
    except (TypeError, ValueError):
        return default


def _percent(value) -> float:
    """把 '3%'、0.03 或 3 都归一为百分数 3.0 存储。"""
    if value is None or value == "":
        return 0.0
    try:
        result = float(str(value).replace(",", "").replace("%", ""))
    except (TypeError, ValueError):
        return 0.0
    if result != 0 and abs(result) <= 1:
        result = result * 100
    return result


def _source_key(kind: str, major: str, code: str, name: str, feature: str, unit: str) -> str:
    raw = "|".join([major or "", code or "", name or "", feature or "", unit or ""])
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]
    return f"{kind}|{major or '装修'}|{digest}"


def calculate_direct_fee(item: QuotaItem) -> float:
    """参考直接费：由全部组成明细重新汇总，不直接采用定额主表总价。"""
    total = sum(_composition_total(composition) for composition in item.compositions)
    return total if total else (item.no_tax_price or 0)


def _composition_total(composition: QuotaComposition, *, tax: bool = False) -> float:
    """优先使用来源合价；无合价时按含量、损耗率和单价计算。"""
    price = composition.tax_price if tax else composition.no_tax_price
    qty = float(composition.qty or 0)
    loss_factor = 1 + float(composition.loss_rate or 0) / 100
    if price not in (None, 0) and qty not in (None, 0):
        return qty * loss_factor * float(price)
    total = composition.tax_total if tax else composition.no_tax_total
    if total not in (None, 0):
        return float(total)
    return qty * loss_factor * float(price or 0)


def composition_cost_breakdown(composition, engineering_quantity: float | None = None) -> dict:
    """计算一条工料机组成的完整金额链，保留单位价和工程量金额两套口径。

    `composition` 可以是 ORM 的 QuotaComposition，也可以是 quota_reference_dict
    返回的字典。单位消耗先计损耗，再乘除税/含税单价；只有明确传入工程量时，
    才计算工程量合价，避免把单位综合单价误当成项目总价。
    """
    def value(*keys, default=0.0):
        if isinstance(composition, dict):
            for key in keys:
                current = composition.get(key)
                if current not in (None, ""):
                    return _number(current, default)
            return default
        for key in keys:
            current = getattr(composition, key, None)
            if current not in (None, ""):
                return _number(current, default)
        return default

    raw_qty = value("sourceQty", "qty")
    adjusted_qty = value("qty", default=raw_qty)
    quota_factor = value("quotaFactor", default=1.0)
    if quota_factor == 0 and isinstance(composition, dict) and "quotaFactor" not in composition:
        quota_factor = 1.0
    loss_rate = _percent(value("loss", "loss_rate"))
    specification_factor = value("specConversionFactor", "spec_conversion_factor", default=1.0)
    if specification_factor <= 0:
        specification_factor = 1.0
    no_tax_price = value("noTaxPrice", "no_tax_price")
    tax_price = value("taxPrice", "tax_price")
    source_no_tax_total = value("noTaxTotal", "no_tax_total")
    source_tax_total = value("taxTotal", "tax_total")
    loss_factor = 1 + loss_rate / 100
    factored_qty = adjusted_qty * quota_factor * specification_factor
    effective_qty = factored_qty * loss_factor
    force_formula = bool(composition.get("forceFormula")) if isinstance(composition, dict) else False
    use_no_tax_formula = force_formula or (
        adjusted_qty not in (None, 0) and no_tax_price not in (None, 0)
    )
    use_tax_formula = force_formula or (
        adjusted_qty not in (None, 0) and tax_price not in (None, 0)
    )
    unit_no_tax_total = (
        effective_qty * no_tax_price
        if use_no_tax_formula
        else source_no_tax_total * quota_factor * specification_factor
    )
    unit_tax_total = (
        effective_qty * tax_price
        if use_tax_formula
        else source_tax_total * quota_factor * specification_factor
    )
    component_unit = _text(
        composition.get("unit") if isinstance(composition, dict) else getattr(composition, "unit", "")
    )
    boq_unit = _text(
        composition.get("boqUnit") or composition.get("engineeringUnit")
        if isinstance(composition, dict)
        else ""
    )
    unit_conversion = _text(
        composition.get("unitConversion") or composition.get("unit_conversion")
        if isinstance(composition, dict)
        else ""
    )
    category = _text(
        composition.get("cat") or composition.get("category")
        if isinstance(composition, dict)
        else getattr(composition, "category", "")
    )
    physical_conversion_required = _component_requires_physical_conversion(
        category,
        boq_unit,
        component_unit,
    )
    conversion_is_actionable = bool(
        unit_conversion
        and _unit_conversion_basis_is_actionable(
            unit_conversion,
            category=category,
            component_name=_text(
                composition.get("name") if isinstance(composition, dict) else getattr(composition, "name", "")
            ),
            boq_unit=boq_unit,
            component_unit=component_unit,
        )
    )
    conversion_required = bool(
        physical_conversion_required
        and boq_unit
        and component_unit
        and _canonical_unit(boq_unit) != _canonical_unit(component_unit)
        and not conversion_is_actionable
    )
    quantity_valid = adjusted_qty > 0
    quantity_issue = "" if quantity_valid else "单位含量缺失或不大于0，禁止计价"
    result = {
        "baseQty": round(raw_qty, 8),
        "adjustedQty": round(adjusted_qty, 8),
        "quotaFactor": round(quota_factor, 8),
        "specConversionFactor": round(specification_factor, 8),
        "factoredQty": round(factored_qty, 8),
        "lossRate": round(loss_rate, 8),
        "effectiveQty": round(effective_qty, 8),
        "unitNoTaxTotal": round(unit_no_tax_total, 8),
        "unitTaxTotal": round(unit_tax_total, 8),
        "engineeringQuantity": None,
        "engineeringUnit": boq_unit,
        "componentUnit": component_unit,
        "totalQty": None,
        "unitConversion": unit_conversion,
        "quantityValid": quantity_valid,
        "quantityIssue": quantity_issue,
        "blockedByQuantity": not quantity_valid,
        "conversionBasisValid": conversion_is_actionable,
        "requiresUnitConversion": conversion_required,
        "physicalUnitConversionRequired": physical_conversion_required,
        "conversionRequired": conversion_required,
        "engineeringNoTaxTotal": None,
        "engineeringTaxTotal": None,
    }
    if conversion_required:
        # Never price incompatible units as if their raw quota quantity were
        # already expressed per BOQ unit. AI or a user must provide the
        # conversion formula before this component contributes to totals.
        result.update({
            "unitNoTaxTotal": None,
            "unitTaxTotal": None,
            "blockedByUnitConversion": True,
        })
    if not quantity_valid:
        # A stale source total must not survive a missing/zero unit quantity.
        # Every priced component has to be reproducible from quantity, factor,
        # loss and unit price.
        result.update({
            "unitNoTaxTotal": None,
            "unitTaxTotal": None,
            "totalQty": None,
            "engineeringNoTaxTotal": None,
            "engineeringTaxTotal": None,
        })
    if engineering_quantity not in (None, ""):
        quantity = _number(engineering_quantity)
        result["engineeringQuantity"] = round(quantity, 8)
        if not conversion_required and quantity_valid:
            result.update({
                "totalQty": round(effective_qty * quantity, 8),
                "engineeringNoTaxTotal": round(unit_no_tax_total * quantity, 8),
                "engineeringTaxTotal": round(unit_tax_total * quantity, 8),
            })
    return result


def quota_reference_dict(item: QuotaItem) -> dict:
    """由定额组成明细计算清单各费用和综合单价。"""
    labor = 0.0
    material = 0.0
    machinery = 0.0
    management = 0.0
    profit = 0.0
    subcontract = 0.0
    other = 0.0
    component_categories = set()
    compositions = []
    seen_compositions = set()
    for composition in item.compositions:
        value = _composition_total(composition)
        category = composition.category
        if category:
            component_categories.add(category)
        duplicate_key = (
            _text(category),
            _text(composition.code),
            _text(composition.name),
            round(_number(composition.qty), 8),
            round(_number(composition.loss_rate), 8),
            round(_number(composition.no_tax_price), 8),
            round(_number(composition.tax_rate), 8),
            round(_number(composition.tax_price), 8),
        )
        if duplicate_key in seen_compositions:
            continue
        seen_compositions.add(duplicate_key)
        if category == "人工费":
            labor += value
        elif category in ("材料费", "辅材费", "主材费"):
            material += value
        elif category == "机械费":
            machinery += value
        elif category == "管理费":
            management += value
        elif category == "利润":
            profit += value
        elif category == "专业分包":
            subcontract += value
        else:
            other += value
        detail = _composition_dict(composition)
        detail.update(composition_cost_breakdown(composition))
        detail.update({
            "calculatedNoTaxTotal": round(value, 4),
            "calculatedTaxTotal": round(_composition_total(composition, tax=True), 4),
            "quotaId": item.id,
            "quotaCode": item.code,
            "quotaName": item.name,
        })
        compositions.append(detail)
    component_total = labor + material + machinery + management + profit + subcontract + other
    return {
        "price": round(component_total, 4),
        "unit": item.unit,
        "source_type": "quota",
        "labor_unit_price": round(labor, 4),
        "material_unit_price": round(material, 4),
        "machinery_unit_price": round(machinery, 4),
        "management_unit_price": round(management, 4),
        "profit_unit_price": round(profit, 4),
        "subcontract_unit_price": round(subcontract, 4),
        "other_unit_price": round(other, 4),
        "unallocated_unit_price": 0.0,
        "component_categories": sorted(component_categories),
        "source": f"定额库 / {item.major} / {item.name}",
        "quota_id": item.id,
        "quota_code": item.code,
        "quota_name": item.name,
        "quota_feature": item.feature,
        "compositions": compositions,
    }


def deduplicate_quota_components(components: Iterable[dict]) -> list[dict]:
    """Deduplicate final human/material/machine rows by their cost identity."""
    result = []
    seen = set()
    for raw in components:
        component = dict(raw)
        key = (
            _text(component.get("cat") or component.get("category")),
            _text(component.get("name")),
            _text(component.get("unit")),
            round(_number(component.get("qty")), 8),
            round(_percent(component.get("loss") if "loss" in component else component.get("loss_rate")), 8),
            round(_number(component.get("noTaxPrice") if "noTaxPrice" in component else component.get("no_tax_price")), 8),
            round(_percent(component.get("taxRate") if "taxRate" in component else component.get("tax_rate")), 8),
            round(_number(component.get("taxPrice") if "taxPrice" in component else component.get("tax_price")), 8),
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(component)
    return result


def _quota_match_identity(match: dict) -> str:
    """Return a stable identity for one formal quota result."""
    quota_id = match.get("quota_id")
    if quota_id not in (None, ""):
        return f"quota:{quota_id}"
    return "generated:%s:%s:%s" % (
        _text(match.get("source_type")),
        _text(match.get("quota_code")),
        _text(match.get("quota_name")),
    )


def _quota_match_score(match: dict) -> float:
    """Rank formal matches without allowing labels or price to beat logic."""
    score = _number(match.get("score"))
    if score > 1.0:
        score /= 100.0
    evidence_bonus = {"green": 0.015, "yellow": 0.0, "red": -0.02}.get(
        _text(match.get("evidence_level")), 0.0
    )
    warning_text = " ".join(
        _text(value)
        for value in [
            *(match.get("logic_warnings") or []),
            *(match.get("reasons") or []),
        ]
    )
    hard_warning = any(token in warning_text for token in ("冲突", "乱匹配", "不符合", "禁止"))
    return score + evidence_bonus - (0.50 if hard_warning else 0.0)


def _quota_component_belongs_to_match(component: dict, match: dict) -> bool:
    """Identify components belonging to a formal or generated match."""
    quota_id = match.get("quota_id")
    if quota_id not in (None, ""):
        component_quota_id = component.get("quotaId") or component.get("quota_id")
        if component_quota_id not in (None, ""):
            return str(component_quota_id) == str(quota_id)
        # Legacy saved compositions may not have quotaId. Fall back to the
        # quota code/name so a discarded candidate cannot keep contributing
        # to the price after a result is reopened and normalized.
        return (
            _text(component.get("quotaCode") or component.get("quota_code"))
            == _text(match.get("quota_code"))
            and _text(component.get("quotaName") or component.get("quota_name"))
            == _text(match.get("quota_name"))
        )
    code = _text(match.get("quota_code"))
    name = _text(match.get("quota_name"))
    return (
        _text(component.get("quotaCode")) == code
        and _text(component.get("quotaName")) == name
    )


def normalize_quota_row_matches(row: dict) -> dict:
    """Keep exactly one formal quota and its complete cost composition.

    Matching functions may return several candidates while ranking. That is an
    internal operation only: a persisted row must have one winning quota.
    Component-only source references are retained as labelled evidence, but
    they are not additional quotas and are never used to select the winner.
    """
    raw_components = [dict(value) for value in row.get("compositions") or []]
    components = raw_components
    matches = []
    formal_matches = []
    evidence_matches = []
    seen_matches = set()
    for raw in row.get("quota_matches") or []:
        match = dict(raw)
        quota_id = match.get("quota_id")
        key = _quota_match_identity(match)
        if key in seen_matches:
            continue
        # Imported source-cost rows are evidence, not a quota candidate. Keep
        # them visible as labelled references, but never let them win or add a
        # second quota to the row.
        if match.get("source_type") == "imported_cost_reference":
            seen_matches.add(key)
            evidence_matches.append(match)
            continue
        seen_matches.add(key)
        # A component-only result is still only one candidate. It can win if
        # it is the best logical result, but it must never coexist with the
        # main quota as a second formal match.
        formal_matches.append(match)

    if formal_matches:
        winner_pool = [
            value for value in formal_matches
            if not value.get("component_only")
        ] or formal_matches
        winner = max(
            winner_pool,
            key=lambda value: (
                _quota_match_score(value),
                1 if value.get("quota_id") not in (None, "") else 0,
                len(value.get("matched_source_clauses") or []),
            ),
        )
        winner_key = _quota_match_identity(winner)
        rejected = row.setdefault("rejected_quota_matches", [])
        rejected_ids = {
            _quota_match_identity(value)
            for value in rejected
            if isinstance(value, dict)
        }
        for candidate in formal_matches:
            if _quota_match_identity(candidate) == winner_key:
                continue
            candidate_key = _quota_match_identity(candidate)
            if candidate_key not in rejected_ids:
                rejected.append({
                    **candidate,
                    "rejected_reason": "同一清单仅保留综合逻辑评分最高的一个定额，未参与计价",
                })
                rejected_ids.add(candidate_key)
            components = [
                value
                for value in components
                if not _quota_component_belongs_to_match(value, candidate)
            ]
        row["rejected_quota_matches"] = rejected[-20:]
        matches.extend([winner])
    # Remove duplicates only after losing candidates have been removed. This
    # preserves the winning quota's composition when two candidates happen to
    # contain an otherwise identical cost row.
    components = deduplicate_quota_components(components)
    row["compositions"] = components
    row["quota_matches"] = [*evidence_matches, *matches]
    winning_formal = next(
        (value for value in matches if value in formal_matches),
        None,
    )
    row["quota_ids"] = (
        [winning_formal.get("quota_id")]
        if winning_formal and winning_formal.get("quota_id") is not None
        else []
    )
    if winning_formal:
        winning_score = _number(winning_formal.get("score"))
        row["score"] = winning_score / 100.0 if winning_score > 1.0 else winning_score
        row["match_reasons"] = list(winning_formal.get("reasons") or row.get("match_reasons") or [])
        if winning_formal.get("evidence_level"):
            row["evidence_level"] = winning_formal["evidence_level"]
    if formal_matches and winner.get("quota_id") in (None, ""):
        # A generated/project-level result is the sole formal result. Remove
        # stale local identifiers and make its generated payload authoritative.
        row["quota_ids"] = []
        row["quota_id"] = None
        row["ai_generated"] = True
    elif formal_matches:
        row["quota_id"] = row["quota_ids"][0] if row["quota_ids"] else None
        if row.get("quota_id") is not None:
            row.pop("generated_quota", None)
            row["ai_generated"] = False
    else:
        # Evidence-only rows and empty rows must not retain an old formal ID.
        # This is important when an older saved result is normalized after a
        # failed correction or after its only quota has been removed.
        row["quota_id"] = None
        row["quota_ids"] = []
    return row


def _base_tax_rate(session: Session, project_id: int | None) -> float:
    """Return the active regional tax rate, falling back to a neutral value."""
    if project_id:
        project = session.get(Project, int(project_id))
        region_id = project.region_id if project is not None else None
        standard = (
            session.query(RegionStandard)
            .filter(
                RegionStandard.region_id == region_id,
                RegionStandard.is_active.is_(True),
                RegionStandard.tax_rate > 0,
            )
            .order_by(RegionStandard.created_at.desc())
            .first()
        )
        if standard is not None:
            return float(standard.tax_rate)
    return 0.0


def _linear_dimension_meters(value: str) -> float | None:
    """Extract an explicit linear dimension that can convert area to volume."""
    text = _text(value).lower()
    for match in re.finditer(r"(?<!\d)(\d+(?:\.\d+)?)\s*(mm|cm|m)(?!\d)", text):
        amount = float(match.group(1))
        unit = match.group(2)
        if unit == "mm":
            return amount / 1000
        if unit == "cm":
            return amount / 100
        if unit == "m":
            return amount
    return None


def _base_fallback_qty(
    boq_unit: str,
    price_unit: str,
    boq_text: str,
) -> tuple[float | None, str]:
    boq_canonical = _canonical_unit(boq_unit)
    price_canonical = _canonical_unit(price_unit)
    boq_dimension = UNIT_DIMENSIONS.get(boq_canonical, "other")
    price_dimension = UNIT_DIMENSIONS.get(price_canonical, "other")

    if boq_canonical and boq_canonical == price_canonical:
        return 1.0, f"基础库单位与清单单位一致（{boq_unit}），按1份清单工程量换算"
    if boq_dimension == "area" and price_dimension == "volume":
        thickness = _linear_dimension_meters(boq_text)
        if thickness is not None and thickness > 0:
            return round(thickness, 8), (
                f"清单单位{boq_unit}按{thickness}m厚度换算为{price_unit}，"
                "每1份清单工程量按该厚度计算"
            )
    if (
        boq_canonical in COUNT_UNITS
        and price_canonical in COUNT_UNITS
        and boq_dimension == price_dimension == "count"
    ):
        return 1.0, (
            f"清单计数单位{boq_unit}与基础库计数单位{price_unit}按1份对应，"
            "套件、井盖等配套范围需人工复核"
        )
    return None, ""


def _base_fallback_component(
    *,
    category: str,
    code: str,
    name: str,
    feature: str,
    unit: str,
    qty: float,
    no_tax_price: float,
    tax_rate: float,
    tax_price: float,
    basis: str,
    source_ids: list[int],
    evidence_level: str,
    note: str,
) -> dict:
    return {
        "cat": category,
        "code": code,
        "name": name,
        "feature": feature,
        "unit": unit,
        "qty": round(qty, 8),
        "loss": 0.0,
        "noTaxPrice": round(no_tax_price, 8),
        "taxRate": round(tax_rate * 100, 8),
        "taxPrice": round(tax_price, 8),
        "noTaxTotal": round(qty * no_tax_price, 8),
        "taxTotal": round(qty * tax_price, 8),
        "calculation_basis": basis,
        "calculationBasis": basis,
        "source_evidence_ids": [int(value) for value in source_ids],
        "sourceEvidenceIds": [int(value) for value in source_ids],
        "evidenceLevel": evidence_level,
        "note": note,
    }


def _fixed_base_fallback(
    session: Session,
    boq: dict,
    minimum_score: float = 45,
) -> dict | None:
    name = _text(boq.get("name"))
    unit = _text(boq.get("unit"))
    feature = _text(boq.get("feature"))
    if not name:
        return None

    candidates = (
        session.query(CostItem)
        .filter(
            CostItem.item_name == name,
            CostItem.comprehensive_price > 0,
        )
        .limit(100)
        .all()
    )
    if not candidates:
        candidates = (
            session.query(CostItem)
            .filter(
                CostItem.item_name.contains(name),
                CostItem.comprehensive_price > 0,
            )
            .limit(100)
            .all()
        )
    if not candidates:
        return None

    target_text = f"{name} {feature}".strip()
    target_sizes = _size_specs(target_text)
    target_tech = _technical_specs(target_text)
    best = None
    for item in candidates:
        if unit and _canonical_unit(unit) != _canonical_unit(item.unit):
            continue
        candidate_text = f"{item.item_name} {item.features}".strip()
        logic_ok, _, _ = _candidate_logic(
            target_text,
            candidate_text,
            unit,
            item.unit,
        )
        if not logic_ok:
            continue
        name_score = max(
            fuzz.ratio(_match_text(name), _match_text(item.item_name)),
            fuzz.WRatio(_match_text(name), _match_text(item.item_name)),
        )
        context_score = fuzz.WRatio(
            _match_text(target_text),
            _match_text(candidate_text),
        )
        if name_score < minimum_score and context_score < minimum_score:
            continue

        candidate_sizes = _size_specs(candidate_text)
        candidate_tech = _technical_specs(candidate_text)
        target_missing_specs = (
            target_sizes.difference(candidate_sizes)
            or target_tech.difference(candidate_tech)
        )
        candidate_extra_specs = (
            candidate_sizes.difference(target_sizes)
            or candidate_tech.difference(target_tech)
        )
        direct = (
            _number(item.labor_cost)
            + _number(item.material_cost)
            + _number(item.machinery_cost)
        )
        comprehensive = _number(item.comprehensive_price) or direct
        if comprehensive <= 0 or direct <= 0:
            continue

        score = (
            name_score * 0.62
            + context_score * 0.38
            + (0 if unit and _canonical_unit(unit) == _canonical_unit(item.unit) else -10)
            - (18 if target_missing_specs else 0)
            - (7 if candidate_extra_specs else 0)
        )
        score = min(max(score, 0), 100)
        evidence = "red"
        if not target_missing_specs:
            evidence = (
                "green"
                if not candidate_extra_specs and name_score >= 78 and context_score >= 55
                else "yellow"
            )
        best = {
            "item": item,
            "score": score,
            "name_score": name_score,
            "context_score": context_score,
            "evidence": evidence,
            "spec_mismatch": bool(target_missing_specs),
            "extra_specs": bool(candidate_extra_specs),
        }
        # Exact-name rows in the base library are intentionally preferred even
        # when their stored specification needs manual review.
        if not target_missing_specs:
            break

    if best is None or best["score"] < minimum_score:
        return None
    item = best["item"]
    components = []
    for category, value, code in (
        ("人工费", item.labor_cost, "BASE-FIXED-LABOR"),
        ("材料费", item.material_cost, "BASE-FIXED-MATERIAL"),
        ("机械费", item.machinery_cost, "BASE-FIXED-MACHINERY"),
    ):
        value = _number(value)
        if value <= 0:
            continue
        components.append(_base_fallback_component(
            category=category,
            code=code,
            name=f"{item.item_name} / {category}",
            feature=item.features,
            unit=item.unit,
            qty=1.0,
            no_tax_price=value,
            tax_rate=0.0,
            tax_price=value,
            basis=f"基础数据库固定清单: {item.data_source} / {item.item_name}",
            source_ids=[],
            evidence_level=best["evidence"],
            note="基础库价格需按项目地区、计价期和清单特征复核",
        ))
    if not components:
        return None

    generated = {
        "major": _text(boq.get("major")) or "其他",
        "code": "BASE-FIXED",
        "name": item.item_name,
        "feature": item.features,
        "unit": item.unit,
        "unit_conversion": (
            f"基础库单位{ item.unit }与清单单位{ unit }相同，按1份清单工程量采用"
        ),
        "assumptions": [
            f"来源：基础数据库 / 固定工程清单库 / {item.data_source}",
            "基础库综合价仅作为项目级参考，不含本项目规格调整或地区信息价联动",
        ],
        "category": "基础库换算",
        "confidence": round(best["score"] / 100, 4),
        "notes": (
            f"基础库固定清单匹配：名称相关度{best['name_score']:.0f}%，"
            f"内容相关度{best['context_score']:.0f}%"
        ),
        "components": components,
        "source_type": "base_fixed",
        "source": f"基础数据库 / 固定工程清单库 / {item.data_source}",
    }
    generated["no_tax_price"] = round(sum(comp["noTaxTotal"] for comp in components), 8)
    generated["tax_price"] = round(sum(comp["taxTotal"] for comp in components), 8)
    generated["evidence_level"] = best["evidence"]
    generated["evidence_summary"] = (
        f"基础数据库固定清单 {item.item_name}，"
        f"名称相关度{best['name_score']:.0f}%，"
        + ("规格与清单不完全一致，需人工复核" if best["spec_mismatch"] else "规格与单位一致")
    )
    return {
        "decision": "base",
        "source_type": "base_fixed",
        "generated_quota": generated,
        "evidence_level": best["evidence"],
        "score": best["score"] / 100,
        "actionable": bool(components),
        "auto_apply": best["evidence"] == "green",
        "summary": f"定额库无匹配，已从基础数据库固定清单调用：{item.item_name}",
        "errors": [],
        "warnings": (
            ["基础库规格与当前清单不完全一致，请确认是否按相似规格换算"]
            if best["spec_mismatch"]
            else []
        ),
    }


def _enterprise_reference_fallback(
    session: Session,
    boq: dict,
    minimum_score: float = 55,
) -> dict | None:
    """Use a compatible enterprise reference as a traceable project estimate.

    Enterprise rows are not government quotas.  They are considered only after
    the real quota matcher has failed, and only when object, craft, explicit
    specification and unit checks pass.  Their cost split is retained as
    project evidence rather than silently inserted into the formal quota table.
    """
    target_name = _text(boq.get("name"))
    target_feature = _text(boq.get("feature"))
    target_unit = _text(boq.get("unit"))
    if not target_name:
        return None
    target_text = f"{target_name} {target_feature}".strip()
    best = None
    rows = (
        session.query(ProjectListData)
        .filter(ProjectListData.item_name != "")
        .order_by(ProjectListData.updated_at.desc(), ProjectListData.id.desc())
        .limit(10000)
        .all()
    )
    for item in rows:
        # Every row stored by the "企业参考定额表" page is a project-level
        # reference, including generic BOQs with price columns. Older code
        # only accepted the two-sheet enterprise format and silently ignored
        # valid imported reference lists.
        candidate_text = f"{_text(item.item_name)} {_text(item.feature)}".strip()
        logic_ok, logic_bonus, logic_reasons = _candidate_logic(
            target_text, candidate_text, target_unit, _text(item.unit),
        )
        if not logic_ok:
            continue
        specification = specification_relation(
            target_text, candidate_text,
            allow_quantity_conversion=False, require_target_specs=True,
        )
        if specification.get("errors"):
            continue
        name_score = max(
            fuzz.ratio(_match_text(target_name), _match_text(item.item_name)),
            fuzz.WRatio(_match_text(target_name), _match_text(item.item_name)),
        )
        context_score = fuzz.WRatio(_match_text(target_text), _match_text(candidate_text))
        score = name_score * 0.62 + context_score * 0.38 + logic_bonus
        if _canonical_unit(target_unit) == _canonical_unit(item.unit):
            score += 6
        score = min(max(score, 0), 100)
        if score < max(55, minimum_score):
            continue
        current = (score, name_score, context_score, item, logic_reasons)
        if best is None or current[0] > best[0]:
            best = current
    if best is None:
        return None

    score, name_score, context_score, item, logic_reasons = best
    components = []
    for category, value, code in (
        ("人工费", item.labor_cost, "ENTERPRISE-LABOR"),
        ("材料费", item.material_cost, "ENTERPRISE-MATERIAL"),
        ("机械费", item.machinery_cost, "ENTERPRISE-MACHINERY"),
    ):
        amount = _number(value)
        if amount <= 0:
            continue
        components.append(_base_fallback_component(
            category=category, code=code,
            name=f"{item.item_name} / {category}", feature=item.feature,
            unit=item.unit, qty=1.0, no_tax_price=amount, tax_rate=0.0,
            tax_price=amount, source_ids=[], evidence_level="yellow",
            basis=(
                f"企业参考定额表：{item.source_project or item.source_file} / "
                f"{item.region or '地区未识别'} / {item.period or '期数未识别'}"
            ),
            note="企业参考案例，不是政府正式定额；已通过对象、工艺、规格和单位校验",
        ))
    if not components:
        return None
    source_label = " / ".join(value for value in (
        item.source_project, item.region, item.period,
    ) if _text(value)) or "来源信息不完整"
    generated = {
        "major": _text(boq.get("major")) or "其他",
        "code": item.item_code or "ENTERPRISE-REFERENCE",
        "name": item.item_name,
        "feature": item.feature or target_feature or target_name,
        "unit": item.unit or target_unit,
        "unit_conversion": f"企业参考与清单单位一致，按每1{target_unit or '清单单位'}的费用组成引用",
        "assumptions": [
            f"企业参考来源：{source_label}",
            "仅在正式定额无可靠候选时作为项目级参考，地区、计价期和费用口径需复核",
        ],
        "confidence": round(score / 100, 4),
        "components": components,
        "source_type": "enterprise_reference",
        "source": f"企业参考定额表 / {source_label}",
        "evidence_level": "yellow",
        "evidence_summary": (
            f"企业参考案例 {item.item_name}；名称相关度{name_score:.0f}%，"
            f"内容相关度{context_score:.0f}%；不是政府正式定额"
        ),
    }
    generated["no_tax_price"] = round(sum(value["noTaxTotal"] for value in components), 8)
    generated["tax_price"] = round(sum(value["taxTotal"] for value in components), 8)
    return {
        "decision": "base", "source_type": "enterprise_reference",
        "generated_quota": generated, "evidence_level": "yellow",
        "score": score / 100, "actionable": True,
        "auto_apply": score >= 78 and name_score >= 75,
        "summary": f"正式定额无可靠匹配，已引用企业参考定额表：{item.item_name}",
        "errors": [],
        "warnings": ["企业参考定额表仅作项目级组价证据，不等同政府正式定额"],
    }


def _material_base_fallback(
    session: Session,
    boq: dict,
    project_id: int | None,
    minimum_score: float = 45,
    include_unconfirmed: bool = False,
    region_override: str = "",
) -> dict | None:
    target_name = _match_text(boq.get("name"))
    target_feature = _match_text(boq.get("feature"))
    target_text = f"{target_name} {target_feature}".strip()
    if not target_name:
        return None

    target_materials = _clause_material_terms(target_text)
    target_core_objects = _core_object_groups(target_text)
    # A material price is not a generic price fallback. Excavation, hauling,
    # compaction, demolition, and similar process-only rows have no material
    # identity to price here. Returning no fallback is safer than turning a
    # coincidental text/bi-gram match into a 1,650 yuan/m3 stone row.
    if not target_materials and not target_core_objects:
        return None

    query = (
        session.query(MaterialPrice, Material, Region)
        .join(Material, MaterialPrice.material_id == Material.id)
        .join(Region, MaterialPrice.region_id == Region.id)
        .filter(
            MaterialPrice.is_withdrawn.is_(False),
            MaterialPrice.is_anomaly.is_(False),
            MaterialPrice.price > 0,
        )
    )
    if not include_unconfirmed:
        query = query.filter(MaterialPrice.is_confirmed.is_(True))

    project = session.get(Project, int(project_id)) if project_id else None
    pricing_date = _text(project.pricing_date) if project else ""
    price_year = _text(project.price_year) if project else ""
    selected_region = _text(region_override)
    target_period, period_scope = latest_price_period(session, project, selected_region)
    if not target_period:
        target_period = ""
        period_scope = "none"

    best = None
    for price, material, region in query.limit(5000).all():
        if selected_region and not _information_region_matches(
            selected_region, region.name, region.province
        ):
            continue
        material_text = _match_text(f"{material.name} {price.spec}")
        candidate_materials = _clause_material_terms(material_text)
        candidate_core_objects = _core_object_groups(material_text)
        if target_materials:
            material_allowed, _ = _material_match_allowed(target_text, material_text)
            if not material_allowed:
                continue
        elif not target_core_objects.intersection(candidate_core_objects):
            continue
        name_score = max(
            fuzz.ratio(target_name, _match_text(material.name)),
            fuzz.WRatio(target_name, _match_text(material.name)),
        )
        context_score = fuzz.WRatio(target_text, material_text)
        shared_bigrams = _cjk_bigrams(target_name) & _cjk_bigrams(material_text)
        if (
            name_score < minimum_score
            and context_score < minimum_score + 8
            and not shared_bigrams
        ):
            continue

        qty, conversion = _base_fallback_qty(
            boq.get("unit", ""),
            price.unit or material.default_unit,
            f"{boq.get('name', '')} {boq.get('feature', '')}",
        )
        if qty is None:
            continue

        region_score = 0
        region_match = "none"
        if selected_region and _information_region_matches(
            selected_region, region.name, region.province
        ):
            region_score = 30
            region_match = "selected"
        period_score = 0
        period_match = "none"
        if target_period and price.period == target_period:
            period_score = 15
            period_match = "exact"
        elif target_period and price.period[:4] == target_period[:4]:
            period_score = 7
            period_match = "year"
        unit_score = 8 if _canonical_unit(boq.get("unit", "")) == _canonical_unit(price.unit or material.default_unit) else 0
        confirmation_score = 20 if price.is_confirmed else -20
        score = (
            name_score * 0.58
            + context_score * 0.27
            + unit_score
            + region_score
            + period_score
            + confirmation_score
        )
        score = min(max(score, 0), 100)
        if score < minimum_score:
            continue
        candidate = {
            "price": price,
            "material": material,
            "region": region,
            "score": score,
            "name_score": name_score,
            "context_score": context_score,
            "qty": qty,
            "conversion": conversion,
            "region_match": region_match,
            "period_match": period_match,
            "unit_score": unit_score,
        }
        # Do not stop at the first confirmed row. The old early break made
        # database insertion order decide between unrelated or weakly related
        # materials, especially when several regions shared the same period.
        if best is None or candidate["score"] > best["score"]:
            best = candidate

    if best is None:
        return None
    price = best["price"]
    material = best["material"]
    region = best["region"]
    tax_rate = _base_tax_rate(session, project_id)
    basis = _text(price.price_basis or "as_published").lower()
    if "tax_inclusive" in basis or "含税" in basis:
        tax_price = _number(price.price)
        no_tax_price = tax_price / (1 + tax_rate) if tax_rate else tax_price
        tax_rate_value = tax_rate
    elif "tax_exclusive" in basis or "除税" in basis or "no_tax" in basis:
        no_tax_price = _number(price.price)
        tax_price = no_tax_price * (1 + tax_rate) if tax_rate else no_tax_price
        tax_rate_value = tax_rate
    else:
        no_tax_price = tax_price = _number(price.price)
        tax_rate_value = 0.0

    category = "主材费"
    if "机械" in material.category:
        category = "机械费"
    elif "人工" in material.category:
        category = "人工费"
    confirmed = bool(price.is_confirmed)
    evidence = "red"
    if confirmed and best["region_match"] == "selected":
        evidence = "green" if best["period_match"] == "exact" else "yellow"
    elif confirmed:
        evidence = "yellow"

    source_ids = [price.id] if confirmed else []
    component = _base_fallback_component(
        category=category,
        code=_text(price.source_key),
        name=material.name,
        feature=_text(price.spec),
        unit=price.unit or material.default_unit,
        qty=best["qty"],
        no_tax_price=no_tax_price,
        tax_rate=tax_rate_value,
        tax_price=tax_price,
        basis=(
            f"信息价/基础数据库：{region.name} {price.period} {material.name} {price.spec}；"
            f"价格口径：{price.price_basis or '原始发布'}；{best['conversion']}"
        ),
        source_ids=source_ids,
        evidence_level=evidence,
        note=(
            "已确认信息价换算"
            if confirmed
            else "信息价尚未确认，仅供人工复核，不得直接作为正式价格"
        ),
    )
    generated = {
        "major": _text(boq.get("major")) or material.category or "其他",
        "code": _text(price.source_key) or "BASE-INFO",
        "name": _text(boq.get("name")) or material.name,
        "feature": _text(boq.get("feature")) or material.name,
        "unit": boq.get("unit", "") or price.unit or material.default_unit,
        "unit_conversion": best["conversion"],
        "assumptions": [
            f"信息价证据：{region.name} / {price.period} / {material.name} / {price.spec}",
            "仅形成主材或机械信息价分量，安装人工、辅材和专业工序仍需定额或人工补充",
        ],
        "category": "基础库换算",
        "confidence": round(best["score"] / 100, 4),
        "notes": (
            f"名称相关度{best['name_score']:.0f}%，内容相关度{best['context_score']:.0f}%，"
            f"地区匹配：{best['region_match']}，计价期匹配：{best['period_match']}"
        ),
        "components": [component],
        "source_type": "base_material",
        "source": f"信息价与基础数据库 / {region.name} / {price.period}",
    }
    generated["no_tax_price"] = component["noTaxTotal"]
    generated["tax_price"] = component["taxTotal"]
    generated["evidence_level"] = evidence
    generated["evidence_summary"] = (
        f"{'1/1' if confirmed else '0/1'}个分量有已确认本地信息价证据；"
        f"来源：{region.name} {price.period} {material.name} {price.spec}"
    )
    return {
        "decision": "base",
        "source_type": "base_material",
        "generated_quota": generated,
        "evidence_level": evidence,
        "score": best["score"] / 100,
        "actionable": True,
        "auto_apply": evidence == "green",
        "summary": f"定额库无匹配，已调用信息价与基础数据库：{material.name} {price.spec}",
        "errors": [],
        "warnings": (
            ["信息价尚未确认，请先在信息价管理完成确认后使用"]
            if not confirmed
            else []
        ),
    }


def build_base_data_fallback(
    session: Session,
    boq: dict,
    project_id: int | None = None,
    *,
    major: str = "",
    minimum_score: float = 45,
    include_unconfirmed: bool = False,
    region_override: str = "",
) -> dict | None:
    """Fallback through enterprise cases, fixed BOQs and selected-region prices."""
    enterprise = _enterprise_reference_fallback(session, boq, max(55, minimum_score))
    if enterprise is not None:
        return enterprise
    fixed = _fixed_base_fallback(session, boq, minimum_score)
    if fixed is not None:
        return fixed
    return _material_base_fallback(
        session,
        boq,
        project_id,
        minimum_score=minimum_score,
        include_unconfirmed=include_unconfirmed,
        region_override=region_override,
    )


def apply_official_prices_to_compositions(
    session: Session,
    project_id: int | None,
    boq: dict,
    compositions: list[dict],
    region_override: str = "",
) -> dict:
    """Replace eligible component prices with confirmed regional information prices.

    This intentionally changes prices only. Quota quantities, loss, factors,
    categories and the composition selected by the quota matcher remain intact.
    A price is eligible only when material identity/specification and unit are
    compatible, and the source is confirmed and non-anomalous.
    """
    if not compositions:
        return {"applied": 0, "skipped": 0, "details": []}
    project = session.get(Project, int(project_id)) if project_id else None
    selected_region = _text(region_override)
    target_period, period_scope = latest_price_period(session, project, selected_region)
    if not target_period:
        target_period = ""
        period_scope = "none"
    tax_rate = _number(project.tax_rate if project else 0.09, 0.09)
    if tax_rate > 1:
        tax_rate /= 100

    prices = (
        session.query(MaterialPrice, Material, Region)
        .join(Material, MaterialPrice.material_id == Material.id)
        .join(Region, MaterialPrice.region_id == Region.id)
        .filter(
            MaterialPrice.is_confirmed.is_(True),
            MaterialPrice.is_withdrawn.is_(False),
            MaterialPrice.is_anomaly.is_(False),
            MaterialPrice.price > 0,
        )
        .limit(8000)
        .all()
    )
    applied = 0
    skipped = 0
    evidence_details = []
    for detail in compositions:
        # The user explicitly selected information-price priority. Generated
        # components may therefore be repriced too, provided identity, explicit
        # specification, unit, region and period checks below all pass.
        if detail.get("priceSourceMode") == "manual":
            skipped += 1
            continue
        category = _text(detail.get("cat") or detail.get("category"))
        if category in {"管理费", "利润"}:
            skipped += 1
            continue
        target_text = " ".join(
            _text(detail.get(key)) for key in ("name", "feature", "code") if _text(detail.get(key))
        )
        if not target_text:
            skipped += 1
            continue
        target_key = _match_text(target_text)
        target_specs = _technical_specs(target_text)
        target_unit = _canonical_unit(detail.get("unit"))
        best = None
        for price, material, region in prices:
            # A selected source region is a hard source boundary for replacing
            # component prices. Project location is never a fallback here.
            if selected_region and not _information_region_matches(
                selected_region, region.name, region.province
            ):
                continue
            material_text = " ".join(
                _text(value) for value in (
                    material.name, price.spec, material.category, price.unit or material.default_unit,
                ) if _text(value)
            )
            candidate_specs = _technical_specs(material_text)
            if target_specs:
                target_by_family = {}
                candidate_by_family = {}
                for spec in target_specs:
                    family, _, value = spec.partition(":")
                    target_by_family.setdefault(family, set()).add(value)
                for spec in candidate_specs:
                    family, _, value = spec.partition(":")
                    candidate_by_family.setdefault(family, set()).add(value)
                # An official price is a replacement only when every explicit
                # target family has exactly the same value. Intersection alone
                # incorrectly allowed C25+C30 to replace explicit C25.
                if any(
                    candidate_by_family.get(family, set()) != values
                    for family, values in target_by_family.items()
                ):
                    continue
            elif candidate_specs:
                # Do not invent an explicit grade/specification when the quota
                # side did not state one.
                continue
            price_unit = _canonical_unit(price.unit or material.default_unit)
            if target_unit and price_unit and target_unit != price_unit:
                continue
            material_category = _text(material.category)
            if category == "人工费" and material_category and "人工" not in material_category:
                continue
            if category == "机械费" and material_category and "机械" not in material_category:
                continue
            if category in {"材料费", "辅材费", "主材费"} and material_category and any(
                token in material_category for token in ("人工", "机械")
            ):
                continue
            name_score = max(
                fuzz.ratio(target_key, _match_text(material.name)),
                fuzz.WRatio(target_key, _match_text(material.name)),
            )
            context_score = fuzz.WRatio(target_key, _match_text(material_text))
            if name_score < 62 and context_score < 68:
                continue
            region_score = 0
            region_match = "none"
            if selected_region and _information_region_matches(
                selected_region, region.name, region.province
            ):
                region_score, region_match = 30, "selected"
            period_score = 0
            period_match = "none"
            if target_period and price.period == target_period:
                period_score, period_match = 15, "exact"
            elif target_period and price.period[:4] == target_period[:4]:
                period_score, period_match = 7, "year"
            unit_score = 10 if target_unit and target_unit == price_unit else 0
            score = name_score * 0.58 + context_score * 0.22 + region_score + period_score + unit_score
            candidate = {
                "price": price, "material": material, "region": region,
                "name_score": name_score, "context_score": context_score,
                "region_match": region_match, "period_match": period_match,
                "score": score + (
                    12 if _text(price.source_type or price.trust_level).lower() in {"official", "official_manual"}
                    else 4 if _text(price.source_type or price.trust_level).lower() == "market_reference"
                    else 0
                ),
            }
            if best is None or candidate["score"] > best["score"]:
                best = candidate
        if best is None or best["name_score"] < 62:
            skipped += 1
            continue

        price = best["price"]
        basis = _text(price.price_basis or "as_published").lower()
        published_price = _number(price.price)
        source_type = _text(price.source_type or price.trust_level).lower()
        if source_type == "market_reference" or _text(price.trust_level).lower() == "market_reference":
            source_mode = "market_reference"
            source_label = "公开市场参考数据"
        elif source_type in {"official", "official_manual"} or "official" in _text(price.trust_level).lower():
            source_mode = "official"
            source_label = "官方信息价"
        else:
            source_mode = "database"
            source_label = "已确认基础数据库价格"
        if "tax_inclusive" in basis or "含税" in basis:
            tax_price = published_price
            no_tax_price = published_price / (1 + tax_rate) if tax_rate else published_price
            basis_label = f"{source_label}含税，按项目税率换算除税价"
        elif "tax_exclusive" in basis or "除税" in basis or "no_tax" in basis:
            no_tax_price = published_price
            tax_price = published_price * (1 + tax_rate) if tax_rate else published_price
            basis_label = f"{source_label}除税，按项目税率换算含税价"
        else:
            no_tax_price = tax_price = published_price
            basis_label = f"{source_label}原始口径未明确，暂不换算税价"
        detail.setdefault("quotaNoTaxPrice", detail.get("noTaxPrice"))
        detail.setdefault("quotaTaxPrice", detail.get("taxPrice"))
        detail["noTaxPrice"] = round(no_tax_price, 8)
        detail["taxPrice"] = round(tax_price, 8)
        detail["taxRate"] = f"{tax_rate * 100:g}%" if tax_rate else ""
        detail["priceSourceMode"] = source_mode
        detail["officialPriceId"] = price.id
        detail["priceEvidence"] = (
            f"{source_label} / {best['region'].name} / {price.period} / "
            f"{best['material'].name} / {price.spec or '未注明规格'} / {price.unit or best['material'].default_unit}"
        )
        detail["calculationBasis"] = f"{detail['priceEvidence']}；{basis_label}"
        detail["sourceEvidence"] = [detail["priceEvidence"]]
        detail["evidenceLevel"] = (
            "green"
            if best["region_match"] in {"exact", "city"} and best["period_match"] == "exact" and "as_published" not in basis
            else "yellow"
        )
        detail["note"] = "；".join(filter(None, (
            _text(detail.get("note")),
            f"已按{source_label}更新单价，定额/AI单位含量和系数保持不变",
        )))
        applied += 1
        evidence_details.append(detail["priceEvidence"])
    return {"applied": applied, "skipped": skipped, "details": evidence_details}


def find_quota_reference(
    session: Session,
    item_name: str,
    feature: str = "",
    unit: str = "",
    major: str = "",
) -> tuple[QuotaItem, float] | None:
    """兼容接口：返回通过造价逻辑校验的最佳定额和置信度。"""
    result = find_quota_match_details(
        session,
        item_name,
        feature,
        unit,
        major,
    )
    if result is None:
        return None
    return result["item"], result["score"]


def find_quota_match_details(
    session: Session,
    item_name: str,
    feature: str = "",
    unit: str = "",
    major: str = "",
    candidate_index: list[dict] | None = None,
    item_code: str = "",
    minimum_score: float = PRIMARY_MATCH_THRESHOLD,
    reference_price: float | None = None,
    required_category: str | set[str] | None = None,
    source_costs: dict | None = None,
) -> dict | None:
    """综合名称、工作内容、单位、部位和工艺，给出可复核的定额候选。"""
    name = _text(item_name)
    if not name:
        return None
    name_key = _match_text(item_name)
    feature_key = _match_text(feature)
    code_key = _match_text(item_code).lower()
    target = f"{name_key} {feature_key}".strip()
    target_logic_text = f"{_text(item_name)} {_text(feature)}".strip()
    target_core_groups = _core_object_groups(target_logic_text)
    target_locations = _location_groups(target_logic_text)
    target_material_terms = _clause_material_terms(target_logic_text)
    index = candidate_index if candidate_index is not None else build_quota_match_index(session, major)
    if not index:
        return None

    rough_candidates = []
    for entry in index:
        candidate_core_groups = _core_object_groups(entry["raw_logic_text"])
        if target_core_groups:
            if not candidate_core_groups:
                continue
            if not target_core_groups.intersection(candidate_core_groups) and not _core_material_family_groups(
                target_logic_text
            ).intersection(_core_material_family_groups(entry["raw_logic_text"])):
                continue
        if required_category:
            allowed_categories = (
                {required_category}
                if isinstance(required_category, str)
                else set(required_category)
            )
            if not entry.get("categories", set()).intersection(allowed_categories):
                continue
            category_name = _canonical_cost_category_group(
                "材料费" if allowed_categories.intersection(MATERIAL_COMPONENT_CATEGORIES)
                else next(iter(sorted(allowed_categories)), "")
            )
            category_ok, category_bonus, category_reasons = _category_candidate_evidence(
                category_name,
                item_name,
                feature,
                entry,
                source_costs,
            )
            if not category_ok:
                continue
        else:
            category_bonus, category_reasons = 0.0, []
        item_name_score = max(
            fuzz.ratio(name_key, entry["name"]),
            fuzz.WRatio(name_key, entry["name"]),
        )
        code_score = fuzz.WRatio(code_key, entry["code"]) if code_key and entry["code"] else 0
        component_name_score = fuzz.WRatio(name_key, entry["component_names"])
        context_score = max(
            fuzz.WRatio(target, entry["main_text"]),
            fuzz.WRatio(target, entry["component_text"]),
        )
        rough_score = (
            max(item_name_score, component_name_score * 0.80, code_score * 0.90) * 0.72
            + context_score * 0.28
        )
        candidate_material_terms = _clause_material_terms(entry["raw_logic_text"])
        if target_material_terms.intersection(candidate_material_terms):
            # Exact material-object evidence outranks shared modifiers such as
            # colour, thickness, "综合考虑", or generic installation wording.
            rough_score += 18
        if target_core_groups.intersection(candidate_core_groups):
            rough_score += 24
        rough_candidates.append((rough_score, entry))
    rough_candidates.sort(key=lambda value: value[0], reverse=True)

    best = None
    best_score = 0.0
    best_reasons: list[str] = []
    best_name_score = 0.0
    best_context_score = 0.0
    best_component_name_score = 0.0
    best_reference = None
    best_source_type = "formal_quota"
    # Keep a wide deterministic shortlist.  A valid work-content candidate
    # can have a generic parent name while its hidden composition feature is
    # the strongest evidence; a 100-row rough cut could discard it first.
    for _, entry in rough_candidates[:300]:
        if required_category:
            allowed_categories = (
                {required_category}
                if isinstance(required_category, str)
                else set(required_category)
            )
            if not entry.get("categories", set()).intersection(allowed_categories):
                continue
            category_name = _canonical_cost_category_group(
                "材料费" if allowed_categories.intersection(MATERIAL_COMPONENT_CATEGORIES)
                else next(iter(sorted(allowed_categories)), "")
            )
            category_ok, category_bonus, category_reasons = _category_candidate_evidence(
                category_name,
                item_name,
                feature,
                entry,
                source_costs,
            )
            if not category_ok:
                continue
        else:
            category_bonus, category_reasons = 0.0, []
        candidate = entry["item"]
        exact_tile_size_conflict = _explicit_tile_plan_size_conflict(
            target_logic_text,
            f"{_text(candidate.name)} {_text(candidate.feature)}",
        )
        if exact_tile_size_conflict:
            continue
        # The rough-ranking loop has its own ``candidate_core_groups`` local.
        # Recompute it here for the current entry; reusing the last rough-loop
        # value caused valid core-object candidates to be discarded or scored
        # with another row's evidence.
        candidate_core_groups = _core_object_groups(entry["raw_logic_text"])
        candidate_name = entry["name"]
        candidate_main_text = entry["main_text"]
        item_name_score = max(
            fuzz.ratio(name_key, candidate_name),
            fuzz.WRatio(name_key, candidate_name),
        )
        code_score = fuzz.WRatio(code_key, entry["code"]) if code_key and entry["code"] else 0
        item_context_score = fuzz.WRatio(target, candidate_main_text)

        component_name_score = 0.0
        component_context_score = 0.0
        best_component_name = ""
        component_logic_texts = []
        for component_name, component_text, component_label in entry["components"]:
            current_component_name_score = max(
                fuzz.ratio(name_key, component_name),
                fuzz.WRatio(name_key, component_name),
            )
            if current_component_name_score > component_name_score:
                component_name_score = current_component_name_score
                best_component_name = component_label
            component_context_score = max(
                component_context_score,
                fuzz.WRatio(target, component_text),
            )
            component_logic_texts.append(component_text)

        logic_text = entry["logic_text"]
        logic_ok, logic_bonus, logic_reasons = _candidate_logic(
            target_logic_text,
            entry["raw_logic_text"],
            unit,
            candidate.unit,
        )
        if not logic_ok:
            continue

        # 名称承担主要判断，工作内容和组成项只增强证据，避免因常见材料词误套工序。
        primary_score = max(item_name_score, component_name_score * 0.72, code_score * 0.90)
        context_score = max(item_context_score, component_context_score * 0.86)
        score = primary_score * 0.70 + context_score * 0.30 + logic_bonus + category_bonus
        if entry.get("is_imported_cost_reference"):
            # Imported cost rows are useful price evidence, but are not formal
            # government quota items. Keep them as a last-resort candidate so
            # a formal quota with slightly different wording wins first.
            score -= 12
        if target_material_terms.intersection(_clause_material_terms(entry["raw_logic_text"])):
            score += 8
        price_deviation_rate = None
        price_reference = _number(reference_price)
        candidate_reference = entry.get("reference") or quota_reference_dict(candidate)
        candidate_price = _number(candidate_reference.get("price"))
        if price_reference > 0 and candidate_price > 0:
            price_deviation_rate = abs(candidate_price - price_reference) / price_reference
            if price_deviation_rate <= 0.50:
                # Deliberately capped: price can break a close tie, never fix
                # a wrong material, unit, specification, or craft.
                score += max(0.0, 4.0 * (1.0 - price_deviation_rate / 0.50))
        candidate_locations = _location_groups(entry["raw_logic_text"])
        if target_locations.intersection(candidate_locations):
            # A full BOQ feature block contains the object context. Prefer a
            # road/wall/pipe-specific detail over a generic material-only row.
            score += 10
            best_reasons_for_location = f"工程部位一致:{'/'.join(sorted(target_locations.intersection(candidate_locations)))}"
        else:
            best_reasons_for_location = ""
        if code_score >= 92:
            score += 6
        evidence_floor = max(35, min(float(minimum_score) - 5, 55))
        explicit_evidence = (
            item_name_score >= evidence_floor
            or (
                item_name_score >= max(30, evidence_floor - 10)
                and context_score >= max(55, float(minimum_score) + 10)
            )
            or (
                component_name_score >= max(68, float(minimum_score) + 18)
                and item_context_score >= max(40, float(minimum_score) - 5)
            )
            or (
                code_score >= max(75, float(minimum_score) + 25)
                and context_score >= max(35, float(minimum_score) - 15)
            )
        )
        # Short BOQ names often differ materially from quota titles, while
        # the object/material anchor and unit/process evidence are decisive.
        # Let that combination pass the evidence floor so a valid quota is not
        # discarded solely because its wording is longer or more formal.
        shared_core_groups = target_core_groups.intersection(candidate_core_groups)
        shared_material_terms = target_material_terms.intersection(
            _clause_material_terms(entry["raw_logic_text"])
        )
        if (
            not explicit_evidence
            and shared_core_groups
            and (shared_material_terms or context_score >= max(35, float(minimum_score) - 10))
            and (unit or not candidate.unit or _canonical_unit(unit) == _canonical_unit(candidate.unit))
        ):
            explicit_evidence = True
        if not explicit_evidence:
            continue
        if score > best_score:
            best = candidate
            best_score = score
            best_name_score = item_name_score
            best_context_score = context_score
            best_component_name_score = component_name_score
            best_reference = candidate_reference
            best_source_type = (
                "imported_cost_reference"
                if entry.get("is_imported_cost_reference")
                else "formal_quota"
            )
            best_reasons = [
                f"定额名称相关度 {item_name_score:.0f}%",
                f"清单名称及项目特征与库明细相关度 {context_score:.0f}%",
            ]
            if entry.get("is_imported_cost_reference"):
                best_reasons.append("导入表人材机仅作价格证据，正式定额优先")
            if price_deviation_rate is not None:
                best_reasons.append(
                    f"原表价格参考偏差 {price_deviation_rate:.0%}（仅用于合格候选择优）"
                )
            if best_reasons_for_location:
                best_reasons.append(best_reasons_for_location)
            if code_score >= 70:
                best_reasons.append(f"定额编码相关度 {code_score:.0f}%")
            if best_component_name and component_name_score >= 70:
                best_reasons.append(
                    f"组成明细相关:{best_component_name}（{component_name_score:.0f}%）"
                )
            best_reasons.extend(logic_reasons)
            best_reasons.extend(category_reasons)
    if best is None or best_score < float(minimum_score):
        return None
    return {
        "item": best,
        "score": min(best_score, 100) / 100,
        "name_score": round(best_name_score / 100, 4),
        "context_score": round(best_context_score / 100, 4),
        "component_name_score": round(best_component_name_score / 100, 4),
        "source_type": best_source_type,
        "reasons": best_reasons,
        "reference_price": _number(reference_price) if _number(reference_price) > 0 else None,
        "estimated_price": round(_number((best_reference or quota_reference_dict(best)).get("price")), 4),
    }


def build_quota_match_index(session: Session, major: str = "") -> list[dict]:
    """一次加载并归一化定额及组成文本，供批量匹配重复使用。"""
    query = session.query(QuotaItem).options(selectinload(QuotaItem.compositions))
    major_variants = quota_major_variants(major)
    if major_variants:
        query = query.filter(QuotaItem.major.in_(major_variants))
    # The previous 3000-row cap hid valid records and depended on insertion
    # order. Keep matching deterministic and stream rows in small batches.
    query = query.order_by(QuotaItem.major, QuotaItem.id).yield_per(500)
    result = []
    for item in query:
        name = _match_text(item.name)
        raw_main_text = " ".join(
            value for value in (
                _quota_detail_text("", item.code, item.name, item.unit),
                _text(item.feature),
                _text(item.notes),
            ) if value
        )
        main_text = _match_text(raw_main_text)
        components = []
        component_names = []
        component_texts = []
        raw_component_texts = []
        for composition in item.compositions:
            if composition.category not in QUOTA_DETAIL_CATEGORIES:
                continue
            component_name = _match_text(composition.name)
            component_text = _quota_detail_text(
                composition.category,
                composition.code,
                composition.name,
                composition.unit,
            )
            if _text(composition.feature):
                component_text = " ".join((component_text, _match_text(composition.feature)))
            raw_component_text = " ".join(
                value for value in (
                    _text(composition.category),
                    _text(composition.code),
                    _text(composition.name),
                    _text(composition.unit),
                    _text(composition.feature),
                    _text(composition.note),
                )
                if value
            )
            components.append((
                component_name,
                component_text,
                composition.name or composition.code or "",
            ))
            if component_name:
                component_names.append(component_name)
            if component_text:
                component_texts.append(component_text)
            if raw_component_text:
                raw_component_texts.append(raw_component_text)
        if not components:
            # Parent rows containing only legacy fee categories are not quota
            # detail candidates under the current five-category library.
            continue
        technical_evidence = _technical_evidence(
            item.feature,
            item.notes,
            *(value for composition in item.compositions for value in (
                composition.feature,
                composition.note,
            )),
        )
        raw_logic_text = " ".join(
            value for value in (raw_main_text, *raw_component_texts, technical_evidence) if value
        ).strip()
        reference = quota_reference_dict(item)
        is_imported_cost_reference = (
            _text(item.category) == "导入表人材机参考"
            or _text(item.major) == "导入表人材机"
        )
        result.append({
            "item": item,
            "code": _match_text(item.code).lower(),
            "name": name,
            "main_text": main_text,
            "components": components,
            "component_names": " ".join(component_names),
            "component_text": " ".join(component_texts),
            "categories": {
                _text(composition.category)
                for composition in item.compositions
                if _text(composition.category) in QUOTA_DETAIL_CATEGORIES
            },
            "reference": reference,
            "logic_text": " ".join([main_text, *component_texts]),
            "raw_logic_text": raw_logic_text,
            "technical_specs": sorted(_technical_specs(technical_evidence)),
            "is_imported_cost_reference": is_imported_cost_reference,
            "source_type": (
                "imported_cost_reference"
                if is_imported_cost_reference
                else "formal_quota"
            ),
        })
    return result


def rank_quota_ai_candidates(
    session: Session,
    item_name: str,
    feature: str = "",
    unit: str = "",
    major: str = "",
    item_code: str = "",
    limit: int = 12,
    source_costs: dict | None = None,
) -> list[dict]:
    """Return a broad candidate set with explicit logic warnings for AI review.

    Local automatic matching still uses strict rules. AI review is allowed to
    see candidates with unit or specification risks so it can distinguish
    "reuse this quota" from "generate a project-only conversion estimate".
    """
    name_key = _match_text(item_name)
    feature_key = _match_text(feature)
    code_key = _match_text(item_code).lower()
    target = f"{name_key} {feature_key}".strip()
    target_logic_text = f"{_text(item_name)} {_text(feature)}".strip()
    target_core_groups = _core_object_groups(target_logic_text)
    candidates = []
    requirements = required_cost_categories(item_name, feature, source_costs)
    for entry in build_quota_match_index(session, major):
        item = entry["item"]
        logic_ok, logic_bonus, logic_reasons = _candidate_logic(
            target_logic_text,
            entry["raw_logic_text"],
            unit,
            item.unit,
        )
        exact_tile_size_conflict = _explicit_tile_plan_size_conflict(
            target_logic_text,
            f"{_text(item.name)} {_text(item.feature)}",
        )
        if exact_tile_size_conflict:
            logic_ok = False
            logic_reasons = list(dict.fromkeys([
                *logic_reasons,
                exact_tile_size_conflict,
            ]))
        logic_warnings = [] if logic_ok else list(logic_reasons)
        logic_penalty = 0 if logic_ok else min(25, len(logic_warnings) * 8)
        candidate_core_groups = _core_object_groups(entry["raw_logic_text"])
        core_shared = target_core_groups.intersection(candidate_core_groups)
        core_conflict = _core_object_conflict(
            target_logic_text,
            entry["raw_logic_text"],
        )
        if core_conflict:
            logic_warnings.append(core_conflict)
            logic_penalty = min(60, logic_penalty + 28)
        name_score = max(
            fuzz.ratio(name_key, entry["name"]),
            fuzz.WRatio(name_key, entry["name"]),
        )
        code_score = fuzz.WRatio(code_key, entry["code"]) if code_key and entry["code"] else 0
        component_name_score = fuzz.WRatio(name_key, entry["component_names"])
        context_score = max(
            fuzz.WRatio(target, entry["main_text"]),
            fuzz.WRatio(target, entry["component_text"]),
        )
        local_score = (
            max(name_score, component_name_score * 0.72, code_score * 0.90) * 0.70
            + context_score * 0.30
            + logic_bonus
            - logic_penalty
        )
        is_imported_cost_reference = bool(entry.get("is_imported_cost_reference"))
        if is_imported_cost_reference:
            local_score -= 12
        if core_shared:
            local_score += min(24, len(core_shared) * 12)
        components = [
            {
                "category": composition.category,
                "code": composition.code,
                "name": composition.name,
                "feature": "",
                "unit": composition.unit,
                "qty": composition.qty,
                "loss_rate": composition.loss_rate,
                "no_tax_price": composition.no_tax_price,
                "tax_rate": composition.tax_rate,
                "tax_price": composition.tax_price,
                "no_tax_total": composition.no_tax_total,
                "tax_total": composition.tax_total,
                "note": composition.note,
            }
            for composition in item.compositions[:20]
        ]
        candidates.append({
            "item": item,
            "quota_id": item.id,
            "major": item.major,
            "code": item.code,
            "name": item.name,
            "feature": "",
            "unit": item.unit,
            "no_tax_price": item.no_tax_price,
            "tax_price": item.tax_price,
            "local_score": round(min(max(local_score, 0), 100) / 100, 4),
            "name_score": round(name_score / 100, 4),
            "context_score": round(context_score / 100, 4),
            "code_score": round(code_score / 100, 4),
            "logic_reasons": logic_reasons,
            "logic_allowed": logic_ok,
            "logic_warnings": logic_warnings,
            "core_object_groups": sorted(candidate_core_groups),
            "core_material_families": sorted(
                _core_material_family_groups(entry["raw_logic_text"])
            ),
            "cost_categories": sorted(entry.get("categories", set())),
            "source_type": entry.get("source_type", "formal_quota"),
            "is_imported_cost_reference": is_imported_cost_reference,
            "components": components,
        })
    # Recall first, score second: a safe lower-scoring candidate is more useful
    # to AI than a high-scoring candidate already rejected for an object,
    # process, scope or specification conflict. This prevents the old top-N
    # truncation from hiding every usable quota behind unsafe lookalikes.
    candidates.sort(
        key=lambda value: (
            bool(value.get("logic_allowed", True)),
            not bool(value.get("is_imported_cost_reference")),
            bool(value.get("core_object_groups")),
            value["local_score"],
            value["name_score"],
        ),
        reverse=True,
    )
    return candidates[:max(1, int(limit))]


def build_ai_quota_context(
    session: Session,
    project_id: int,
    boq: dict,
    *,
    major: str = "",
    candidate_limit: int = 24,
    price_limit: int = 50,
    source_costs: dict | None = None,
    region_override: str = "",
) -> dict:
    """Build bounded local evidence for one AI quota review/generation request."""
    project = session.query(Project).filter(Project.id == project_id).first() if project_id else None
    region = (
        session.query(Region).filter(Region.id == project.region_id).first()
        if project is not None and project.region_id
        else None
    )
    candidates = rank_quota_ai_candidates(
        session,
        boq.get("name", ""),
        boq.get("feature", ""),
        boq.get("unit", ""),
        major=major,
        item_code=boq.get("code", ""),
        source_costs=source_costs or boq.get("source_costs") or {
            "labor": boq.get("imported_labor"),
            "material": boq.get("imported_material"),
            "machinery": boq.get("imported_machinery"),
        },
        limit=candidate_limit,
    )
    usable_local_count = sum(
        1
        for candidate in candidates
        if candidate.get("logic_allowed", True)
        and _number(candidate.get("local_score")) >= 0.40
    )
    if major and usable_local_count < 5:
        broad_candidates = rank_quota_ai_candidates(
            session,
            boq.get("name", ""),
            boq.get("feature", ""),
            boq.get("unit", ""),
            major="",
            item_code=boq.get("code", ""),
            source_costs=source_costs or boq.get("source_costs") or {
                "labor": boq.get("imported_labor"),
                "material": boq.get("imported_material"),
                "machinery": boq.get("imported_machinery"),
            },
            limit=candidate_limit,
        )
        existing_ids = {candidate.get("quota_id") for candidate in candidates}
        selected_major = canonical_quota_major(major)
        for candidate in broad_candidates:
            if candidate.get("quota_id") in existing_ids:
                continue
            candidate = dict(candidate)
            candidate["cross_major_candidate"] = (
                canonical_quota_major(candidate.get("major")) != selected_major
            )
            if candidate["cross_major_candidate"]:
                candidate["logic_warnings"] = list(dict.fromkeys([
                    *(candidate.get("logic_warnings") or []),
                    f"跨专业候选：{candidate.get('major') or '未分类'}，仅供AI逻辑复核",
                ]))
            candidates.append(candidate)
            existing_ids.add(candidate.get("quota_id"))
            if len(candidates) >= candidate_limit:
                break
    safe_candidates = [
        {key: value for key, value in candidate.items() if key != "item"}
        for candidate in candidates
    ]

    selected_region = _text(region_override)
    # Project location is retained below as construction background only. It
    # must never become an implicit information-price region.
    effective_region = selected_region
    project_context = {
        "project_id": project.id if project else 0,
        "name": project.name if project else "",
        "project_type": project.project_type if project else "",
        "location": project.project_location if project else "",
        "province": project.pricing_province if project else "",
        "city": project.pricing_city if project else "",
        "district": project.pricing_district if project else "",
        "pricing_date": project.pricing_date if project else "",
        "price_year": project.price_year if project else "",
        "stage": project.stage if project else "",
        "specialty": project.specialty if project else "",
        "region_id": project.region_id if project else None,
        "region_name": region.name if region else "",
        "source_region": effective_region,
        "effective_region": effective_region,
        "source_region_selected": bool(selected_region),
        "tax_rate": project.tax_rate if project else 0.09,
        "pricing_priority": [
            "页面选择的信息来源地区同期官方信息价",
            "页面选择的信息来源地区公开市场参考价",
            "页面选择的信息来源地区最新可用期数",
            "联网检索到的可核验市场证据",
        ],
        "region_required": False,
    }
    target_text = " ".join(_text(value) for value in (
        boq.get("name"), boq.get("feature"), boq.get("code"),
    ) if _text(value))
    required_terms = extract_quota_key_terms(
        boq.get("name", ""),
        boq.get("feature", ""),
    )
    material_requirements = extract_material_requirements(
        boq.get("name", ""),
        boq.get("feature", ""),
        boq.get("unit", ""),
    )
    resource_requirements = extract_resource_requirements(
        boq.get("name", ""),
        boq.get("feature", ""),
        boq.get("unit", ""),
    )
    work_items = split_boq_work_items(boq.get("name", ""), boq.get("feature", ""))
    target_period, period_scope = latest_price_period(
        session, None, selected_region
    )
    project_context["effective_pricing_period"] = target_period
    project_context["effective_period_scope"] = period_scope
    # Historical project rows are evidence only. They help AI compare
    # finished project practice and local pricing, but never become formal
    # quota records automatically.
    from .project_list_data_service import find_project_list_references
    historical_references = []
    reference_region = selected_region
    for score, historical, name_score, context_score in find_project_list_references(
        session, boq.get("name", ""), boq.get("feature", ""),
        boq.get("unit", ""), region=reference_region, period=target_period, limit=12,
    ):
        historical_references.append({
            "id": historical.id,
            "score": round(score / 100, 4),
            "name_score": round(name_score / 100, 4),
            "context_score": round(context_score / 100, 4),
            "item_code": historical.item_code,
            "item_name": historical.item_name,
            "feature": historical.feature,
            "unit": historical.unit,
            "source_project": historical.source_project,
            "source_file": historical.source_file,
            "region": historical.region,
            "period": historical.period,
            "comprehensive_price": historical.comprehensive_price,
            "labor_cost": historical.labor_cost,
            "material_cost": historical.material_cost,
            "machinery_cost": historical.machinery_cost,
            "management_cost": historical.management_cost,
            "profit": historical.profit,
            "analysis": historical.analysis,
        })
    price_rows = (
        session.query(MaterialPrice, Material, Region)
        .join(Material, MaterialPrice.material_id == Material.id)
        .join(Region, MaterialPrice.region_id == Region.id)
        .filter(
            MaterialPrice.is_confirmed.is_(True),
            MaterialPrice.is_withdrawn.is_(False),
            MaterialPrice.is_anomaly.is_(False),
            MaterialPrice.price > 0,
        )
        .limit(5000)
        .all()
    )
    market_prices = []
    for price, material, price_region in price_rows:
        if selected_region and not _information_region_matches(
            selected_region, price_region.name, price_region.province,
        ):
            # A selected information-price city is a hard price-source scope.
            # Do not expose another city's price to AI and hope that a prompt
            # will prevent it from being selected.
            continue
        material_text = " ".join(_text(value) for value in (
            material.name, price.spec, material.category, material.description,
        ) if _text(value))
        relevance = fuzz.WRatio(_match_text(target_text), _match_text(material_text))
        region_score = 0
        region_match = "none"
        if effective_region and _information_region_matches(
            effective_region, price_region.name, price_region.province,
        ):
            region_score = 30
            region_match = "selected"
        # No project city/province fallback: an empty selector means neutral
        # latest-library context, while a selected region is the only regional
        # evidence allowed to rank a price.
        period_score = 0
        period_match = "none"
        if target_period and price.period == target_period:
            period_score = 15
            period_match = "exact"
        elif target_period and price.period[:4] == target_period[:4]:
            period_score = 7
            period_match = "year"
        market_prices.append({
            "price_id": price.id,
            "material_name": material.name,
            "material_category": material.category,
            "spec": price.spec,
            "unit": price.unit or material.default_unit,
            "price": price.price,
            "period": price.period,
            "region": price_region.name,
            "region_id": price.region_id,
            "trust_level": price.trust_level,
            "source_type": price.source_type,
            "price_basis": price.price_basis,
            "notes": price.notes,
            "region_match": region_match,
            "period_match": period_match,
            "relevance": round(relevance / 100, 4),
            "_rank": relevance + region_score + period_score,
        })
    market_prices.sort(key=lambda value: (value["_rank"], value["price_id"]), reverse=True)
    for value in market_prices:
        value.pop("_rank", None)
    return {
        "project": project_context,
        "work_items": work_items,
        "required_terms": required_terms,
        "material_requirements": material_requirements,
        "resource_requirements": resource_requirements,
        "core_object_groups": sorted(_core_object_groups(target_text)),
        "core_material_families": sorted(_core_material_family_groups(target_text)),
        "required_cost_categories": required_cost_categories(
            boq.get("name", ""),
            boq.get("feature", ""),
            source_costs or boq.get("source_costs") or {},
        ),
        "candidates": safe_candidates,
        "historical_project_references": historical_references,
        "similar_project_results": list(boq.get("similar_project_results") or [])[:3],
        "market_prices": market_prices[:max(1, int(price_limit))],
    }


def validate_ai_generated_quota(
    boq: dict,
    generated_quota: dict,
    market_prices: Iterable[dict] = (),
) -> dict:
    """Normalize AI results and separate actionable estimates from unusable data.

    Missing project features, unit conversions, assumptions, and weak price
    evidence lower confidence instead of discarding the whole result. Only
    invalid structure, empty components, impossible quantities, or invalid
    categories remain hard errors.
    """
    errors: list[str] = []
    warnings: list[str] = []
    boq_name = _text(boq.get("name"))
    boq_feature = _text(boq.get("feature"))
    boq_unit = _text(boq.get("unit"))
    if len(_match_text(boq_name)) < 2:
        errors.append("清单工程名称不足，不能生成补充定额")
    if len(_match_text(boq_feature)) < 3:
        warnings.append("项目特征及工作内容不足，补充组成需按工程名称和造价常识人工复核")
    if not boq_unit:
        warnings.append("清单单位为空，补充定额的单位和含量必须人工确认")

    quota = dict(generated_quota or {})
    quota["major"] = _text(quota.get("major") or boq.get("major")) or "其他"
    quota["code"] = _text(quota.get("code")) or "AI补充定额"
    quota["name"] = _text(quota.get("name"))
    quota["feature"] = _text(quota.get("feature")) or boq_feature or boq_name
    quota["unit"] = _text(quota.get("unit")) or boq_unit
    quota["source"] = "AI补充定额/市场估算"
    quota["category"] = "AI补充定额/市场估算"
    quota["notes"] = _text(quota.get("notes"))
    quota["unit_conversion"] = _text(
        quota.get("unit_conversion") or quota.get("unitConversion")
    )
    allow_specification_adjustment = bool(quota.get("allow_specification_adjustment"))
    specification_errors: list[str] = []
    specification_adjusted = False
    source_conversion_text = f"{boq_name} {boq_feature}"
    inferred_quota_conversion = _infer_unit_conversion_basis(
        source_conversion_text,
        boq_unit,
        quota["unit"],
    )
    if not quota["unit_conversion"] and inferred_quota_conversion:
        quota["unit_conversion"] = inferred_quota_conversion
        warnings.append(f"已根据清单明确尺寸自动生成单位换算：{inferred_quota_conversion}")
    raw_assumptions = quota.get("assumptions")
    quota["assumptions"] = [
        _text(value)
        for value in (raw_assumptions if isinstance(raw_assumptions, list) else [])
        if _text(value)
    ]
    if not quota["name"]:
        errors.append("补充定额名称为空")
    quota_unit_mismatch = bool(
        boq_unit
        and quota["unit"]
        and _canonical_unit(quota["unit"]) != _canonical_unit(boq_unit)
    )
    if quota_unit_mismatch:
        conversion = quota["unit_conversion"]
        if conversion and _unit_conversion_basis_is_actionable(
            conversion,
            category="",
            component_name=quota["name"],
            boq_unit=boq_unit,
            component_unit=quota["unit"],
        ):
            warnings.append(
                f"补充定额单位“{quota['unit']}”与清单单位“{boq_unit}”不同，"
                f"已记录换算口径：{conversion}"
            )
        else:
            original_unit = quota["unit"]
            quota["unit"] = boq_unit
            quota["unit_conversion"] = (
                f"项目级补充定额统一按每1{boq_unit}形成综合单价；"
                f"AI原输出定额单位为{original_unit}，各人材机按其单位含量分别计价"
            )
            warnings.append(
                f"补充定额单位“{original_unit}”与清单单位“{boq_unit}”不同，"
                "已统一为清单单位口径；各组成含量仍需按独立计量关系复核"
            )
            quota_unit_mismatch = False
    # Validate the AI quota against its complete generated evidence. AI often
    # puts an explicit grade/material only on a component (for example
    # ``预拌混凝土C25``), not on the quota headline.
    raw_components = quota.get("components") or quota.get("items") or []
    generated_component_text = " ".join(
        " ".join(
            _text(value.get(key))
            for key in ("code", "name", "feature", "unit", "note")
            if _text(value.get(key))
        )
        for value in raw_components
        if isinstance(value, dict)
    )
    generated_evidence_text = " ".join(
        value for value in (
            quota["name"], quota["feature"], generated_component_text,
        ) if value
    )
    logic_ok, _, logic_reasons = _candidate_logic(
        f"{boq_name} {boq_feature}",
        generated_evidence_text,
        boq_unit,
        quota["unit"],
    )
    if not logic_ok:
        scope_errors = [
            reason for reason in logic_reasons
            if "专用部位" in reason or "工程部位冲突" in reason
        ]
        if scope_errors:
            errors.extend(scope_errors)
        else:
            warnings.extend(logic_reasons)
    scope_errors = quota_reference_scope_errors(boq_name, boq_feature, quota)
    if scope_errors:
        errors.extend(scope_errors)
    generated_specification = specification_relation(
        f"{boq_name} {boq_feature}",
        generated_evidence_text,
        allow_quantity_conversion=True,
        require_target_specs=True,
        # A generic AI quota may omit a target layer thickness while its
        # component still identifies the correct material. This is a warning
        # case; explicit grade/diameter/model conflicts remain hard errors.
        allow_missing_convertible_specs=True,
    )
    if generated_specification.get("errors"):
        specification_errors.extend(generated_specification["errors"])
        # Explicit grade/model/diameter conflicts are deterministic errors.
        # ``allow_specification_adjustment`` must never turn C25/C30, DN or
        # material-grade conflicts into payable warnings. Convertible
        # thickness differences are returned in ``conversions`` below.
        errors.extend(generated_specification["errors"])
    if generated_specification.get("conversions"):
        specification_adjusted = True
        quota["specification_adjustments"] = list(
            generated_specification.get("conversions") or []
        )
        warnings.extend(generated_specification.get("warnings") or [])
        if quota_unit_mismatch and not quota["unit_conversion"]:
            message = "AI补充定额与清单存在可换算规格差异，但缺少规格含量换算公式"
            specification_errors.append(message)
            warnings.append(message + "；已按组成单位含量口径保留并标记复核")
    if specification_adjusted:
        quota.setdefault("specification_adjustments", [])
        quota["assumptions"] = list(dict.fromkeys([
            *(quota.get("assumptions") or []),
            "AI提取规格已按原清单规格进行换算；含量、价格和适用范围需人工复核",
        ]))
    target_sizes = _size_specs(f"{boq_name} {boq_feature}")
    generated_sizes = _size_specs(generated_evidence_text)
    target_specs = _technical_specs(f"{boq_name} {boq_feature}")
    generated_specs = _technical_specs(generated_evidence_text)
    if generated_sizes and not target_sizes:
        warnings.append("补充定额增加了清单未明确的规格尺寸，请确认是否为合理市场假设")
    elif generated_sizes.difference(target_sizes):
        warnings.append("补充定额含清单未提供的规格尺寸，请确认换算依据")
    if generated_specs.difference(target_specs):
        warnings.append("补充定额含清单未明确的材料等级、管径或厚度假设，需人工确认")
    quota_relation = fuzz.WRatio(
        _match_text(boq_name),
        _match_text(f"{quota['name']} {quota['feature']}"),
    )
    target_crafts = _text_terms(f"{boq_name} {boq_feature}", CRAFT_TERMS)
    target_materials = _clause_material_terms(f"{boq_name} {boq_feature}")
    target_locations = _location_groups(f"{boq_name} {boq_feature}")
    quota_text = f"{quota['name']} {quota['feature']}"
    supported_quota_terms = bool(
        target_crafts.intersection(_text_terms(quota_text, CRAFT_TERMS))
        or target_materials.intersection(_clause_material_terms(quota_text))
        or target_locations.intersection(_location_groups(quota_text))
    )
    if quota_relation < 45 and not supported_quota_terms:
        warnings.append("补充定额名称和工作内容与清单工程名称关联度较低，需重点复核")

    # A market assumption for steel grade/thickness is allowed when the BOQ
    # is measured by tonne, but it must remain visible and traceable. It is
    # never a substitute for the design fact, and it must not hide an object,
    # scope, unit-conversion, or calculation conflict.
    steel_object = any(term in _match_text(f"{boq_name} {boq_feature}") for term in (
        "预埋铁件", "预埋件", "铁件", "钢板", "钢构件", "钢支架", "钢件",
    ))
    steel_dimensions = _technical_specs(f"{boq_name} {boq_feature}") | _size_specs(
        f"{boq_name} {boq_feature}"
    )
    if steel_object and not steel_dimensions:
        warnings.append(
            "清单涉及铁件/钢构件，但未提供钢材牌号、规格、尺寸、厚度或重量；"
            "已允许按项目地区/最新月份的市场常规假设估价，但必须标注假设、价格来源和人工复核"
        )
        assumptions = quota.get("assumptions") or []
        if not isinstance(assumptions, list) or not any(_text(value) for value in assumptions):
            quota["assumptions"] = [
                "清单未明确钢材牌号、规格和单件重量，按项目地区同期常规市场组成暂估",
                "材料规格、损耗和适用范围属于市场假设，需在施工图或询价资料齐全后复核",
            ]
            warnings.append("铁件/钢构件缺少设计规格，已自动补充市场暂估假设并标记复核")

    evidence_by_id = {
        int(value["price_id"]): value
        for value in market_prices
        if value.get("price_id") is not None
    }
    if not isinstance(raw_components, list) or not raw_components:
        errors.append("补充定额没有人工、材料、机械或分包组成明细")
        raw_components = []
    generated_spec_conversion_required = bool(
        allow_specification_adjustment
        and generated_specification.get("conversions")
    )
    components = []
    reliable_count = 0
    estimated_count = 0
    allowed_source_ids = set(evidence_by_id)
    for index, raw in enumerate(raw_components[:40], start=1):
        if not isinstance(raw, dict):
            errors.append(f"第{index}个组成不是有效字段结构")
            continue
        raw = dict(raw)
        category = AI_COMPOSITION_CATEGORY_ALIASES.get(
            _text(raw.get("cat") or raw.get("category") or raw.get("type")),
            _text(raw.get("cat") or raw.get("category") or raw.get("type")),
        )
        name = _text(raw.get("name"))
        unit = _text(raw.get("unit"))
        own_unit_conversion = _text(
            raw.get("unitConversion") or raw.get("unit_conversion")
        )
        unit_conversion = own_unit_conversion
        component_feature = _text(raw.get("feature"))
        if generated_spec_conversion_required and category in MATERIAL_COMPONENT_CATEGORIES:
            # Generated quotas can be accepted as controlled substitutes when
            # their explicit thickness differs, but the material quantity must
            # carry that ratio into the calculation. A pre-existing non-1
            # factor is treated as already applied to avoid double conversion.
            existing_spec_factor = _number(
                raw.get("specConversionFactor")
                or raw.get("spec_conversion_factor"),
                1.0,
            )
            if abs(existing_spec_factor - 1.0) < 0.0000001:
                raw = apply_specification_quantity_conversion(
                    raw,
                    f"{boq_name} {boq_feature}",
                    f"{quota['name']} {quota['feature']}",
                )
            raw["specConversionApplied"] = True
        component_unit_mismatch = bool(
            boq_unit
            and unit
            and _canonical_unit(boq_unit) != _canonical_unit(unit)
        )
        component_blocked_by_conversion = False
        component_conversion_source = _component_conversion_source(
            boq_name,
            boq_feature,
            name,
            component_feature,
            boq_unit,
            unit,
        )
        if component_unit_mismatch and not unit_conversion:
            inferred_component_conversion = _infer_unit_conversion_basis(
                component_conversion_source,
                boq_unit,
                unit,
            )
            if inferred_component_conversion:
                unit_conversion = inferred_component_conversion
                warnings.append(
                    f"{name or f'第{index}个组成'}已根据清单明确尺寸自动生成单位换算："
                    f"{inferred_component_conversion}"
                )
        if component_unit_mismatch and not unit_conversion:
            # A quota-level formula may explicitly describe how all relevant
            # components are converted. Keep the inherited basis visible on
            # the component so the exported calculation remains auditable.
            unit_conversion = quota["unit_conversion"]
        qty = _number(
            raw.get("qty")
            if raw.get("qty") not in (None, "")
            else raw.get("quantity") if raw.get("quantity") not in (None, "")
            else raw.get("content")
        )
        loss = _percent(raw.get("loss") if "loss" in raw else raw.get("loss_rate"))
        no_tax_price = _number(
            raw.get("noTaxPrice")
            if raw.get("noTaxPrice") not in (None, "")
            else raw.get("no_tax_price") if raw.get("no_tax_price") not in (None, "")
            else raw.get("unitPrice") if raw.get("unitPrice") not in (None, "")
            else raw.get("price")
        )
        tax_rate = _percent(raw.get("taxRate") if "taxRate" in raw else raw.get("tax_rate"))
        tax_price = _number(
            raw.get("taxPrice")
            if raw.get("taxPrice") not in (None, "")
            else raw.get("tax_price")
        )
        basis = _text(
            raw.get("calculation_basis") or raw.get("calculationBasis")
            or raw.get("basis") or raw.get("price_basis")
        )
        source_ids = raw.get("source_evidence_ids") or raw.get("sourceEvidenceIds") or []
        if not isinstance(source_ids, list):
            source_ids = []
        valid_source_ids = []
        for value in source_ids:
            try:
                price_id = int(value)
            except (TypeError, ValueError):
                continue
            if price_id in allowed_source_ids and price_id not in valid_source_ids:
                valid_source_ids.append(price_id)
        if category not in AI_COMPOSITION_CATEGORIES:
            errors.append(f"第{index}个组成费用类别“{category}”不符合定额库格式")
        if not name:
            errors.append(f"第{index}个组成名称为空")
        if not unit:
            errors.append(f"第{index}个组成单位为空")
        if component_unit_mismatch:
            # Only material dimensions require a physical conversion formula.
            # Labour, machinery and subcontract components are naturally
            # expressed as consumption per BOQ unit (工日/m2, 台班/m3, etc.).
            physical_conversion_required = category in MATERIAL_COMPONENT_CATEGORIES
            conversion_is_actionable = bool(
                unit_conversion
                and _unit_conversion_basis_is_actionable(
                    unit_conversion,
                    category=category,
                    component_name=name,
                    boq_unit=boq_unit,
                    component_unit=unit,
                )
            )
            if physical_conversion_required and not conversion_is_actionable and qty > 0:
                # In quota pricing, component qty is itself a consumption per
                # one BOQ unit (for example 0.15m3/m2 or 1.03m2/m2).  A missing
                # prose formula lowers evidence quality but must not discard a
                # positive, calculable AI result.  Explicit dimensions below
                # still override this ratio, and object/spec conflicts remain
                # hard errors.
                unit_conversion = (
                    f"计量关系：每1{boq_unit or '清单单位'}计取{qty:g}{unit}；"
                    "该值为单位消耗量，需结合清单尺寸、施工工艺或定额消耗水平复核"
                )
                basis = basis or unit_conversion
                warnings.append(
                    f"{name or f'第{index}个组成'}缺少物理换算公式，"
                    f"已按AI单位消耗量{qty:g}{unit}/{boq_unit or '清单单位'}形成可计算暂估"
                )
            elif category not in MATERIAL_COMPONENT_CATEGORIES and qty > 0:
                # 人工、机械和专业分包通常直接按清单单位估算消耗量，
                # 不应因缺少一段文字说明而整条拒绝；保留可审计的复核提示。
                basis = basis or (
                    f"按每1{boq_unit or '清单单位'}计取{qty:g}{unit}，"
                    "依据清单工作内容估算，需人工复核"
                )
                warnings.append(
                    f"{name or f'第{index}个组成'}按清单单位计取，含量依据已自动补充，需人工复核"
                )
            elif unit_conversion and not own_unit_conversion:
                warnings.append(
                    f"{name or f'第{index}个组成'}沿用补充定额级单位换算说明：{unit_conversion}"
                )
        inferred_quantity = _infer_area_volume_quantity(
            component_conversion_source,
            boq_unit,
            unit,
        )
        declared_quantity = _declared_conversion_quantity(
            unit_conversion,
            boq_unit,
            unit,
        ) if component_unit_mismatch else None
        if declared_quantity is not None:
            if qty <= 0 or abs(qty - declared_quantity) > max(0.0000001, declared_quantity * 0.005):
                warnings.append(
                    f"{name or f'第{index}个组成'}含量{qty:g}与其明确换算公式不一致，"
                    f"已按公式校正为{declared_quantity:g}"
                )
                qty = declared_quantity
                basis = "；".join(
                    value for value in (
                        basis,
                        f"含量校正：采用换算公式中的{declared_quantity:g}{unit}/{boq_unit}",
                    )
                    if value
                )
            # A component-specific formula is stronger evidence than a broad
            # BOQ clause with another layer thickness.
            inferred_quantity = declared_quantity
        if (
            inferred_quantity is not None
            and category in MATERIAL_COMPONENT_CATEGORIES
            and component_unit_mismatch
            and (
                qty <= 0
                or abs(qty - inferred_quantity) > max(0.000001, inferred_quantity * 0.05)
            )
        ):
            if qty <= 0:
                warnings.append(
                    f"{name or f'第{index}个组成'}未提供有效含量，已按清单明确厚度"
                    f"自动计算为{inferred_quantity:g}"
                )
            else:
                warnings.append(
                    f"{name or f'第{index}个组成'}含量{qty:g}与清单明确厚度换算值"
                    f"{inferred_quantity:g}不一致，已按面积/体积关系校正"
                )
            qty = inferred_quantity
            basis = "；".join(
                value for value in (
                    basis,
                    f"含量校正：按明确厚度换算为{inferred_quantity:g}{unit}/{boq_unit}",
                )
                if value
            )
        if not component_feature:
            warnings.append(f"第{index}个组成“{name or '未命名'}”没有对应的清单工作内容")
        if qty <= 0:
            errors.append(f"第{index}个组成“{name or '未命名'}”含量必须大于0")
        component_blocked_by_loss = False
        if loss < 0 or loss > 100:
            errors.append(f"第{index}个组成“{name or '未命名'}”损耗率超出0%至100%")
        elif loss >= 100:
            component_blocked_by_loss = True
            warnings.append(
                f"第{index}个组成“{name or '未命名'}”损耗率为{loss:g}%，"
                "超过造价常规安全线，禁止计价，需人工复核"
            )
        if no_tax_price <= 0:
            errors.append(f"第{index}个组成“{name or '未命名'}”缺少有效除税单价")
        if tax_rate < 0 or tax_rate > 100:
            errors.append(f"第{index}个组成“{name or '未命名'}”税率超出0%至100%")
        if len(_match_text(basis)) < 6:
            warnings.append(f"第{index}个组成“{name or '未命名'}”缺少可复核的价格计算依据")

        component_text = f"{name} {component_feature}"
        component_sizes = _size_specs(component_text)
        if component_sizes.difference(target_sizes):
            warnings.append(f"第{index}个组成“{name or '未命名'}”增加了清单未提供的规格尺寸")
        component_specs = _technical_specs(component_text)
        if component_specs.difference(target_specs):
            if steel_object:
                warnings.append(f"第{index}个组成“{name or '未命名'}”采用清单未明确的钢材等级、尺寸或厚度市场假设")
            else:
                warnings.append(f"第{index}个组成“{name or '未命名'}”增加了清单未明确的等级、管径或厚度假设")
        target_scopes = _text_terms(f"{boq_name} {boq_feature}", SPECIAL_SCOPE_TERMS)
        component_scopes = _text_terms(component_text, SPECIAL_SCOPE_TERMS)
        unexpected_scopes = component_scopes.difference(target_scopes)
        if unexpected_scopes:
            errors.append(
                f"第{index}个组成“{name or '未命名'}”含清单未说明的专用部位："
                + "/".join(sorted(unexpected_scopes))
            )
        component_locations = _location_groups(component_text)
        if target_locations and component_locations and not target_locations.intersection(component_locations):
            warnings.append(f"第{index}个组成“{name or '未命名'}”施工部位与清单冲突")
        component_materials = _clause_material_terms(component_text)
        derived_material_basis = _derived_material_supported_by_work(
            f"{boq_name} {boq_feature}",
            component_text,
            component_materials,
        )
        if category in MATERIAL_COMPONENT_CATEGORIES:
            material_allowed, material_reason = _material_match_allowed(
                f"{boq_name} {boq_feature}",
                component_text,
            )
            if material_allowed and not target_materials and component_materials and not derived_material_basis:
                material_allowed = False
                material_reason = (
                    "清单未明确该材料，且当前工序没有可核验的材料推导依据："
                    + "/".join(sorted(component_materials))
                )
            if not material_allowed:
                if derived_material_basis:
                    warnings.append(
                        f"第{index}个{category}“{name or '未命名'}”由清单明确工序“"
                        f"{derived_material_basis}”推导，材料规格与含量属于AI市场假设，需人工复核"
                    )
                    quota["assumptions"] = list(dict.fromkeys([
                        *(quota.get("assumptions") or []),
                        f"{name or '未命名'}依据清单工序“{derived_material_basis}”补充，规格和含量需人工复核",
                    ]))
                else:
                    errors.append(
                        f"第{index}个{category}“{name or '未命名'}”材料对象不符合清单："
                        f"{material_reason}"
                    )
        component_crafts = _text_terms(component_text, CRAFT_TERMS)
        relation_score = fuzz.WRatio(
            _match_text(component_text),
            _match_text(f"{boq_name} {boq_feature}"),
        )
        supported_component_terms = bool(
            target_crafts.intersection(component_crafts)
            or target_materials.intersection(component_materials)
            or target_locations.intersection(component_locations)
        )
        if relation_score < 35 and not supported_component_terms:
            warnings.append(f"第{index}个组成“{name or '未命名'}”无法清晰对应清单名称或工作内容")
        expected_tax_price = no_tax_price * (1 + tax_rate / 100)
        if tax_price <= 0:
            tax_price = expected_tax_price
        elif expected_tax_price and abs(tax_price - expected_tax_price) / expected_tax_price > 0.03:
            warnings.append(f"{name}含税单价与税率不一致，已按除税单价和税率重算")
            tax_price = expected_tax_price

        component_level = "red"
        compatible_evidence = []
        compatible_source_ids = []
        for price_id in valid_source_ids:
            evidence = evidence_by_id[price_id]
            name_score = fuzz.WRatio(_match_text(name), _match_text(
                f"{evidence.get('material_name', '')} {evidence.get('spec', '')}"
            ))
            unit_matches = (
                not evidence.get("unit")
                or _canonical_unit(unit) == _canonical_unit(evidence.get("unit"))
            )
            if name_score >= 65 and unit_matches:
                compatible_evidence.append(evidence)
                compatible_source_ids.append(price_id)
        evidence_labels = [
            f"#{price_id} {evidence_by_id[price_id].get('material_name', '')} "
            f"{evidence_by_id[price_id].get('spec', '')} "
            f"{evidence_by_id[price_id].get('region', '')} "
            f"{evidence_by_id[price_id].get('period', '')}"
            for price_id in compatible_source_ids
        ]
        if compatible_evidence:
            reference = compatible_evidence[0]
            reference_price = _number(reference.get("price"))
            basis_text = _text(reference.get("price_basis")).lower()
            if "含税" in basis_text or "tax_inclusive" in basis_text:
                comparable = tax_price
            else:
                comparable = no_tax_price
            deviation = abs(comparable - reference_price) / reference_price if reference_price else 1
            basis_known = any(value in basis_text for value in (
                "含税", "除税", "tax_inclusive", "tax_exclusive", "pretax", "no_tax",
            ))
            scope_reliable = (
                reference.get("region_match") in {"selected", "exact", "city", "province"}
                and reference.get("period_match") in {"exact", "year"}
            )
            if deviation <= 0.15 and basis_known and scope_reliable:
                component_level = "green"
                reliable_count += 1
            elif deviation <= 0.35:
                component_level = "yellow"
                warnings.append(f"{name}为相似确认价格，地区、计价期、税价口径或偏差需复核")
            else:
                warnings.append(f"{name}与本地确认价格偏差{deviation:.0%}，可靠性不足")
        elif valid_source_ids:
            estimated_count += 1
            warnings.append(f"{name}引用的信息价名称、规格或单位不相容")
        else:
            estimated_count += 1
            warnings.append(f"{name}无本地已确认价格佐证，属于AI市场估算")

        specification_factor = _number(
            raw.get("specConversionFactor")
            or raw.get("spec_conversion_factor"),
            1.0,
        )
        if specification_factor <= 0:
            specification_factor = 1.0
        no_tax_total = qty * specification_factor * (1 + loss / 100) * no_tax_price
        tax_total = qty * specification_factor * (1 + loss / 100) * tax_price
        if component_blocked_by_conversion or component_blocked_by_loss:
            no_tax_total = 0.0
            tax_total = 0.0
        components.append({
            "cat": category,
            "code": _text(raw.get("code")),
            "name": name,
            "feature": _text(raw.get("feature")),
            "unit": unit,
            "unitConversion": unit_conversion,
            "requiresUnitConversion": _component_requires_physical_conversion(
                category,
                boq_unit,
                unit,
            ),
            "qty": round(qty, 8),
            "specConversionFactor": round(specification_factor, 8),
            "specConversionBasis": _text(
                raw.get("specConversionBasis") or raw.get("spec_conversion_basis")
            ),
            "specConversionApplied": bool(raw.get("specConversionApplied")),
            "loss": round(loss, 8),
            "noTaxPrice": round(no_tax_price, 8),
            "taxRate": round(tax_rate, 8),
            "taxPrice": round(tax_price, 8),
            "noTaxTotal": round(no_tax_total, 8),
            "taxTotal": round(tax_total, 8),
            "note": _text(raw.get("note")),
            "calculationBasis": basis,
            "sourceEvidenceIds": compatible_source_ids,
            "sourceEvidence": evidence_labels,
            "evidenceLevel": component_level,
            "blockedByUnitConversion": component_blocked_by_conversion,
            "blockedByLossRate": component_blocked_by_loss,
            "included": not (component_blocked_by_conversion or component_blocked_by_loss),
        })

    required_key_terms = extract_quota_key_terms(boq_name, boq_feature)
    generated_key_terms = extract_quota_key_terms(
        f"{quota['name']} {quota['feature']} {generated_component_text}",
    )
    if required_key_terms["strict"] and required_key_terms["crafts"]:
        generated_craft_groups = _semantic_groups(
            f"{quota['name']} {quota['feature']} {generated_component_text}",
            CRAFT_EQUIVALENT_GROUPS,
        )
        missing_crafts = {
            term for term in set(required_key_terms["crafts"])
            if term not in set(generated_key_terms["crafts"])
            and not (
                _semantic_groups(term, CRAFT_EQUIVALENT_GROUPS)
                & generated_craft_groups
            )
        }
        if missing_crafts:
            errors.append(
                "AI补充定额缺少清单核心工艺词：" + "、".join(sorted(missing_crafts))
            )
    if required_key_terms["road_scope_groups"] and generated_key_terms["road_scope_groups"]:
        conflicting_groups = (
            set(required_key_terms["road_scope_groups"])
            - set(generated_key_terms["road_scope_groups"])
        )
        if conflicting_groups:
            errors.append(
                "AI补充定额关键道路部位冲突：" + "、".join(sorted(conflicting_groups))
            )

    component_anchor_text = " ".join(
        f"{value.get('name', '')} {value.get('feature', '')}"
        for value in components
    )
    target_anchors = _cjk_bigrams(f"{boq_name} {boq_feature}")
    generated_anchors = _cjk_bigrams(
        f"{quota['name']} {quota['feature']} {component_anchor_text}"
    )
    shared_anchors = target_anchors.intersection(generated_anchors)
    semantic_anchor = bool(
        _core_object_groups(f"{boq_name} {boq_feature}").intersection(
            _core_object_groups(f"{quota['name']} {quota['feature']} {component_anchor_text}")
        )
        or target_materials.intersection(
            _clause_material_terms(f"{quota['name']} {quota['feature']} {component_anchor_text}")
        )
        or target_crafts.intersection(
            _text_terms(f"{quota['name']} {quota['feature']} {component_anchor_text}", CRAFT_TERMS)
        )
        or _semantic_groups(
            f"{boq_name} {boq_feature}", CRAFT_EQUIVALENT_GROUPS,
        ).intersection(_semantic_groups(
            f"{quota['name']} {quota['feature']} {component_anchor_text}",
            CRAFT_EQUIVALENT_GROUPS,
        ))
    )
    if target_anchors and not semantic_anchor and (
        not shared_anchors or (len(shared_anchors) < 2 and quota_relation < 50)
    ):
        errors.append("补充定额未保留清单工程对象核心词，疑似把原清单对象替换成了其他工程")

    quota["components"] = deduplicate_quota_components(components)
    quota.pop("items", None)
    quota["no_tax_price"] = round(sum(
        value["noTaxTotal"] for value in components if value.get("included", True)
    ), 8)
    quota["tax_price"] = round(sum(
        value["taxTotal"] for value in components if value.get("included", True)
    ), 8)
    confidence = min(max(_number(quota.get("confidence")), 0), 1)
    if specification_adjusted:
        confidence = min(confidence, 0.55)
    if errors:
        evidence_level = "red"
    elif components and reliable_count == len(components) and confidence >= 0.80:
        evidence_level = "green"
    elif confidence >= 0.55 and all(value["evidenceLevel"] != "red" for value in components):
        evidence_level = "yellow"
    else:
        evidence_level = "red"
    quota["confidence"] = confidence
    quota["evidence_level"] = evidence_level
    quota["evidence_summary"] = (
        f"{reliable_count}/{len(components)}个分量有相容的本地已确认价格佐证；"
        f"{estimated_count}个分量为AI市场估算"
    )
    # Component-level unit/loss issues no longer discard otherwise usable
    # components. Hard row errors still block the generated quota; warnings
    # and excluded components remain visible for review and traceability.
    actionable = bool(components) and not errors
    status = (
        "accepted" if actionable and evidence_level == "green" and not warnings
        else "accepted_with_warning" if actionable
        else "rejected"
    )
    return {
        "valid": actionable,
        "actionable": actionable,
        "quota": quota,
        "errors": list(dict.fromkeys(errors)),
        "specification_errors": list(dict.fromkeys(specification_errors)),
        "warnings": list(dict.fromkeys(warnings)),
        "evidence_level": evidence_level,
        "status": status,
    }


REFERENCE_ONLY_TERMS = (
    "详见图纸", "见图纸", "设计图纸", "施工图纸", "设计说明",
    "满足设计", "满足规范", "按图施工", "按设计施工", "自行考虑",
    "综合考虑", "未尽事宜",
)


def _is_reference_only_clause(value: str) -> bool:
    normalized = _match_text(value)
    return not normalized or any(term in normalized for term in REFERENCE_ONLY_TERMS)


def split_boq_work_items(item_name: str, feature: str = "") -> list[str]:
    """把清单特征按编号/换行拆成独立工作内容，过滤标题和图纸说明。"""
    values = [_text(item_name)]
    values.extend(re.split(
        r"[\r\n；;]+|(?=\s*(?:\d+\s*[、．)]|\d+\.(?!\d)|\(\s*\d+\s*\)|（\s*\d+\s*）|[①②③④⑤⑥⑦⑧⑨⑩]))",
        _text(feature),
    ))
    result = []
    seen = set()
    for value in values:
        raw_value = value.strip()
        is_item_name = _match_text(raw_value) == _match_text(item_name)
        is_heading = bool(raw_value) and raw_value.endswith(("：", ":"))
        value = re.sub(
            r"^\s*(?:\d+\s*[、．)]|\d+\.(?!\d)|\(\s*\d+\s*\)|（\s*\d+\s*）|[①②③④⑤⑥⑦⑧⑨⑩])\s*",
            "",
            value,
        ).strip(" ：:，,；;")
        value = re.sub(
            r"[，,；;]?\s*(?:具体)?(?:详见|见)(?:设计|施工)?(?:图纸|说明).*$",
            "",
            value,
        ).strip(" ：:，,；;")
        if not is_item_name:
            value = re.sub(
                r"^(?:基层(?:种类)?[:：])?(?:钢筋)?混凝土结构层\s*(?:[，,、；;]\s*|$)",
                "",
                value,
            ).strip(" ：:，,；;")
        if _is_reference_only_clause(value):
            continue
        # Numbered feature blocks often start with a descriptive heading such as
        # “混凝土路面：”; it is context, not an additional cost clause.
        if is_heading:
            continue
        if len(_match_text(value)) < 3:
            continue
        key = _match_text(value)
        if key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def boq_cost_clauses(item_name: str, feature: str = "") -> list[str]:
    """返回清单项目特征中的每一条可计价工作内容，不依赖固定关键词。"""
    result = []
    for clause in split_boq_work_items(item_name, feature):
        if _match_text(clause) == _match_text(item_name):
            continue
        result.append(clause)
    return result


def _material_term_is_covered(term: str, component_text: str) -> bool:
    """Treat explicit compound materials and approved substitutes as coverage."""
    target = _match_text(term).lower()
    normalized = _match_text(component_text).lower()
    if not target or target in normalized:
        return True
    if target == "水泥" and "混凝土" in normalized:
        return True
    if target == "沥青" and any(value in normalized for value in ("玛蹄脂", "油膏", "密封膏", "密封胶")):
        return True
    for family, component_terms in MATERIAL_DECOMPOSITION_GROUPS.items():
        if target == _match_text(family).lower() and any(
            _match_text(value).lower() in normalized
            for value in component_terms
        ):
            return True
    for aliases in MATERIAL_EQUIVALENT_GROUPS.values():
        normalized_aliases = {_match_text(value).lower() for value in aliases}
        if target in normalized_aliases and any(value in normalized for value in normalized_aliases):
            return True
    return False


def material_coverage_gaps(item_name: str, feature: str, compositions: Iterable[dict]) -> list[str]:
    """找出清单明确出现、但当前工料机名称和特征中没有体现的材料词。"""
    target_text = f"{_text(item_name)} {_text(feature)}"
    component_text = " ".join(
        " ".join(
            _text(detail.get(key))
            for key in ("cat", "category", "code", "name", "quotaCode", "quotaName", "unit")
            if _text(detail.get(key))
        )
        for detail in compositions
        if detail.get("included", True)
    )
    target_terms = _clause_material_terms(target_text)
    # 如“地砖”已出现，不再重复报告其子词“砖”。
    specific_terms = _specific_terms(target_terms)
    normalized_components = _match_text(component_text).lower()
    return sorted(
        (
            term for term in specific_terms
            if not _material_term_is_covered(term, normalized_components)
        ),
        key=lambda value: (-len(value), value),
    )


def uncovered_cost_clauses(
    item_name: str,
    feature: str,
    compositions: Iterable[dict],
    matched_source_clauses: Iterable[str] = (),
) -> list[str]:
    """按规格、材料和工艺检查清单成本工作内容是否已被当前定额组覆盖。"""
    details = list(compositions)
    aggregate = " ".join(
        " ".join(_text(value) for value in (
            detail.get("quotaCode"), detail.get("quotaName"), detail.get("code"),
            detail.get("name"), detail.get("cat"), detail.get("category"), detail.get("unit"),
            detail.get("feature"), detail.get("note"),
        ))
        for detail in details
        if detail.get("included", True)
    )
    aggregate_crafts = _text_terms(aggregate, CRAFT_TERMS)
    aggregate_materials = _clause_material_terms(aggregate)
    source_clauses = [_text(value) for value in matched_source_clauses]
    uncovered = []
    for clause in boq_cost_clauses(item_name, feature):
        # A recorded source clause is provenance only, not proof of actual
        # coverage. Always validate the current composition evidence below.
        clause_crafts = _text_terms(clause, CRAFT_TERMS)
        clause_materials = _clause_material_terms(clause)
        evidence_checks = []
        if clause_crafts:
            evidence_checks.append(bool(clause_crafts.intersection(aggregate_crafts)))
        if clause_materials:
            evidence_checks.append(bool(clause_materials.intersection(aggregate_materials)))
        if _technical_specs(clause) or _size_specs(clause):
            specification = specification_relation(
                clause,
                aggregate,
                allow_quantity_conversion=True,
                require_target_specs=True,
            )
            evidence_checks.append(not specification.get("errors"))
        covered = all(evidence_checks) if evidence_checks else fuzz.WRatio(
            _match_text(clause), _match_text(aggregate),
        ) >= 78
        if not covered:
            uncovered.append(clause)
    return uncovered


def build_quota_audit_report(boq: dict, current: dict) -> dict:
    """Run the fixed local audit checks used before AI quota review."""
    boq_name = _text(boq.get("name"))
    boq_feature = _text(boq.get("feature"))
    boq_unit = _text(boq.get("unit"))
    target_text = f"{boq_name} {boq_feature}".strip()
    components = [dict(value) for value in current.get("compositions") or [] if isinstance(value, dict)]
    quota_matches = [dict(value) for value in current.get("quota_matches") or [] if isinstance(value, dict)]
    matched_clauses = [
        str(clause)
        for match in quota_matches
        for clause in (match.get("matched_source_clauses") or [match.get("source_clause")])
        if clause and str(clause) != "主体工序"
    ]
    uncovered_clauses = uncovered_cost_clauses(
        boq_name,
        boq_feature,
        components,
        matched_clauses,
    )
    missing_materials = material_coverage_gaps(
        boq_name,
        boq_feature,
        components,
    )
    source_costs = current.get("source_costs") or {
        "labor": current.get("imported_labor"),
        "material": current.get("imported_material"),
        "machinery": current.get("imported_machinery"),
        "management": current.get("imported_management"),
        "profit": current.get("imported_profit"),
    }
    category_coverage = cost_category_coverage(
        boq_name,
        boq_feature,
        components,
        source_costs=source_costs,
        fee_details=current.get("fee_details") or [],
    )
    issues = []

    for clause in uncovered_clauses:
        issues.append({
            "severity": "high" if len(_match_text(clause)) >= 4 else "medium",
            "category": "人材机遗漏",
            "message": f"工作内容未完整覆盖：{clause}",
            "affected_component": clause,
        })
    for material in missing_materials:
        issues.append({
            "severity": "high",
            "category": "人材机遗漏",
            "message": f"清单明确材料未进入组成：{material}",
            "affected_component": material,
        })
    for category in category_coverage.get("missing") or []:
        issues.append({
            "severity": "high",
            "category": "人材机管利遗漏",
            "message": (
                f"费用类别未形成完整组成：{category}。"
                f"依据：{category_coverage.get('reasons', {}).get(category) or '未提供'}"
            ),
            "affected_component": category,
        })
    for category in category_coverage.get("reference_only") or []:
        issues.append({
            "severity": "high",
            "category": "人材机管利遗漏",
            "message": (
                f"{category}仅保留原表费用参考，未找到可确认的本地定额组成，"
                "不能按已匹配定额直接采用"
            ),
            "affected_component": category,
        })

    duplicate_quota_counts: dict[object, int] = {}
    for match in quota_matches:
        key = match.get("quota_id") or f"generated:{match.get('quota_code')}:{match.get('quota_name')}"
        duplicate_quota_counts[key] = duplicate_quota_counts.get(key, 0) + 1
    for key, count in duplicate_quota_counts.items():
        if count > 1:
            issues.append({
                "severity": "high",
                "category": "重复匹配",
                "message": f"同一份定额或补充定额被重复计入 {count} 次",
                "affected_component": str(key),
            })

    duplicate_component_counts: dict[tuple, int] = {}
    for component in components:
        if not component.get("included", True):
            continue
        key = (
            component.get("quotaId"),
            _text(component.get("quotaCode")),
            _text(component.get("quotaName")),
            _text(component.get("cat") or component.get("category")),
            _text(component.get("code")),
            _text(component.get("name")),
            _text(component.get("unit")),
            round(_number(component.get("qty")), 8),
            round(_percent(component.get("loss") if "loss" in component else component.get("loss_rate")), 8),
            round(_number(component.get("noTaxPrice") if "noTaxPrice" in component else component.get("no_tax_price")), 8),
            round(_number(component.get("taxRate") if "taxRate" in component else component.get("tax_rate")), 8),
            round(_number(component.get("taxPrice") if "taxPrice" in component else component.get("tax_price")), 8),
        )
        duplicate_component_counts[key] = duplicate_component_counts.get(key, 0) + 1
    for key, count in duplicate_component_counts.items():
        if count > 1:
            issues.append({
                "severity": "high",
                "category": "重复匹配",
                "message": f"相同人材机分量重复计入 {count} 次：{key[6]}",
                "affected_component": str(key[6]),
            })

    component_text = " ".join(
        " ".join(value for value in (
            _text(component.get("quotaCode")),
            _text(component.get("quotaName")),
            _text(component.get("code")),
            _text(component.get("name")),
            _text(component.get("unit")),
        ) if value)
        for component in components
    )
    for match in quota_matches:
        score = _number(match.get("score"))
        if score < 0.55:
            issues.append({
                "severity": "high" if score < 0.45 else "medium",
                "category": "乱匹配",
                "message": (
                    f"匹配置信度仅 {score:.0%}，存在低质量替代："
                    f"{match.get('quota_code') or ''} {match.get('quota_name') or ''}"
                ),
                "affected_component": match.get("quota_name") or match.get("quota_code") or "",
            })
        candidate_name = " ".join(value for value in (
            _text(match.get("quota_code")),
            _text(match.get("quota_name")),
            _text(match.get("quota_feature")),
        ) if value)
        candidate_details = [
            value for value in components
            if value.get("quotaId") == match.get("quota_id")
            or (
                not match.get("quota_id")
                and _text(value.get("quotaName")) == _text(match.get("quota_name"))
            )
        ]
        candidate_spec_text = " ".join(
            " ".join(
                _text(value.get(key))
                for key in ("name", "feature", "code", "unit")
                if _text(value.get(key))
            )
            for value in candidate_details
        )
        specification = specification_relation(
            target_text,
            f"{candidate_name} {candidate_spec_text}",
            allow_quantity_conversion=True,
            require_target_specs=True,
        )
        for reason in specification.get("errors") or []:
            issues.append({
                "severity": "high",
                "category": "规格冲突",
                "message": reason,
                "affected_component": candidate_name,
            })
        if specification.get("conversions"):
            expected_factor = 1.0
            for conversion in specification["conversions"]:
                expected_factor *= float(conversion.get("ratio") or 1.0)
            material_details = [
                value for value in candidate_details
                if _text(value.get("cat") or value.get("category")) in MATERIAL_COMPONENT_CATEGORIES
            ]
            conversion_factors = [
                _number(value.get("specConversionFactor"), 1.0)
                for value in material_details
                if value.get("included", True)
            ]
            if material_details and not any(
                abs(value - expected_factor) <= max(0.0001, expected_factor * 0.001)
                for value in conversion_factors
            ):
                issues.append({
                    "severity": "high",
                    "category": "单位换算",
                    "message": (
                        "清单与候选定额的明确厚度不同，但材料含量未按规格比例换算："
                        f"预计系数 {expected_factor:.6g}"
                    ),
                    "affected_component": candidate_name,
                })
        if candidate_name and boq_unit:
            logic_ok, _, logic_reasons = _candidate_logic(
                target_text,
                f"{candidate_name} {component_text}",
                boq_unit,
                _text(match.get("quota_unit")),
            )
            if not logic_ok:
                issues.append({
                    "severity": "high",
                    "category": "乱匹配",
                    "message": "；".join(logic_reasons[:3]),
                    "affected_component": candidate_name,
                })

    category_totals: dict[str, float] = {}
    other_total = 0.0
    for component in components:
        if not component.get("included", True):
            continue
        category = _text(component.get("cat") or component.get("category"))
        qty = _number(component.get("qty"))
        loss = _percent(component.get("loss") if "loss" in component else component.get("loss_rate"))
        no_tax_price = _number(
            component.get("noTaxPrice")
            if "noTaxPrice" in component
            else component.get("no_tax_price")
        )
        tax_rate = _percent(
            component.get("taxRate")
            if "taxRate" in component
            else component.get("tax_rate")
        )
        tax_price = _number(
            component.get("taxPrice")
            if "taxPrice" in component
            else component.get("tax_price")
        )
        no_tax_total = _number(
            component.get("noTaxTotal")
            if "noTaxTotal" in component
            else component.get("no_tax_total")
        )
        tax_total = _number(
            component.get("taxTotal")
            if "taxTotal" in component
            else component.get("tax_total")
        )
        expected_no_total = qty * (1 + loss / 100) * no_tax_price
        expected_tax_total = qty * (1 + loss / 100) * tax_price
        expected_tax_price = no_tax_price * (1 + tax_rate / 100)
        name = _text(component.get("name")) or _text(component.get("quotaName"))
        if qty <= 0 or no_tax_price <= 0:
            issues.append({
                "severity": "high",
                "category": "计算错误",
                "message": f"{name} 的含量或除税单价必须大于0",
                "affected_component": name,
            })
        if not (0 <= loss <= 100):
            issues.append({
                "severity": "high",
                "category": "计算错误",
                "message": f"{name} 的损耗率超出0%-100%",
                "affected_component": name,
                "actual_calculation": f"{loss}%",
            })
        if not (0 <= tax_rate <= 100):
            issues.append({
                "severity": "high",
                "category": "计算错误",
                "message": f"{name} 的税率超出0%-100%",
                "affected_component": name,
                "actual_calculation": f"{tax_rate}%",
            })
        if no_tax_total and abs(no_tax_total - expected_no_total) > max(0.01, abs(expected_no_total) * 0.001):
            issues.append({
                "severity": "high",
                "category": "计算错误",
                "message": f"{name} 的除税合计与含量/损耗/单价计算不一致",
                "affected_component": name,
                "expected_calculation": f"{expected_no_total:.4f}",
                "actual_calculation": f"{no_tax_total:.4f}",
            })
        if tax_total and abs(tax_total - expected_tax_total) > max(0.01, abs(expected_tax_total) * 0.001):
            issues.append({
                "severity": "high",
                "category": "计算错误",
                "message": f"{name} 的含税合计与含量/损耗/单价计算不一致",
                "affected_component": name,
                "expected_calculation": f"{expected_tax_total:.4f}",
                "actual_calculation": f"{tax_total:.4f}",
            })
        if tax_price and abs(tax_price - expected_tax_price) > max(0.01, abs(expected_tax_price) * 0.001):
            issues.append({
                "severity": "medium",
                "category": "计算错误",
                "message": f"{name} 的含税单价与除税单价/税率不一致",
                "affected_component": name,
                "expected_calculation": f"{expected_tax_price:.4f}",
                "actual_calculation": f"{tax_price:.4f}",
            })
        if category in AI_COMPOSITION_CATEGORIES:
            category_totals[category] = category_totals.get(category, 0.0) + (
                no_tax_total if no_tax_total else expected_no_total
            )
        else:
            other_total += no_tax_total if no_tax_total else expected_no_total

    labor = category_totals.get("人工费", 0.0)
    material = (
        category_totals.get("材料费", 0.0)
        + category_totals.get("辅材费", 0.0)
        + category_totals.get("主材费", 0.0)
    )
    machinery = category_totals.get("机械费", 0.0)
    subcontract = category_totals.get("专业分包", 0.0)
    management = _number(current.get("management"))
    profit = _number(current.get("profit"))
    comprehensive = _number(current.get("comprehensive_price"))
    expected_comprehensive = labor + material + machinery + subcontract + other_total + management + profit
    if comprehensive and abs(comprehensive - expected_comprehensive) > max(0.01, abs(expected_comprehensive) * 0.001):
        issues.append({
            "severity": "high",
            "category": "计算错误",
            "message": "综合单价与当前人材机、管理费和利润汇总不一致",
            "expected_calculation": f"{expected_comprehensive:.4f}",
            "actual_calculation": f"{comprehensive:.4f}",
        })

    # A mathematically consistent sum can still be commercially implausible.
    # Treat the imported price as evidence, not as an automatic target: a
    # large deviation triggers review and never changes the calculated price.
    reference_price = _number(
        current.get("source_comprehensive_price")
        or current.get("imported_comprehensive_price")
    )
    if reference_price <= 0:
        historical = current.get("historical_reference") or {}
        reference_price = _number(historical.get("comprehensive_price"))
    if comprehensive <= 0 and expected_comprehensive > 0:
        issues.append({
            "severity": "high",
            "category": "综合价合理性",
            "message": "已形成有效人材机组成，但综合单价为空或为0",
            "expected_calculation": f"{expected_comprehensive:.4f}",
            "actual_calculation": f"{comprehensive:.4f}",
        })
    if reference_price > 0 and comprehensive > 0:
        deviation = abs(comprehensive - reference_price) / reference_price
        if deviation > 0.50:
            severity = "high"
        elif deviation > 0.30:
            severity = "medium"
        else:
            severity = ""
        if severity:
            issues.append({
                "severity": severity,
                "category": "综合价合理性",
                "message": (
                    f"套定额综合单价与原表/历史案例参考价偏差 {deviation:.0%}，"
                    "需要核对地区、计价期、含税口径、工程对象和单位换算"
                ),
                "expected_calculation": f"参考价 {reference_price:.4f}",
                "actual_calculation": f"套定额价 {comprehensive:.4f}",
            })
    quantity = _number(boq.get("quantity"))
    total_price = _number(current.get("total_price"))
    if quantity > 0 and comprehensive > 0 and total_price > 0:
        expected_total = comprehensive * quantity
        if abs(total_price - expected_total) > max(0.01, abs(expected_total) * 0.001):
            issues.append({
                "severity": "high",
                "category": "综合价合理性",
                "message": "清单合价与综合单价×工程量不一致",
                "expected_calculation": f"{comprehensive:.4f} × {quantity:.6g} = {expected_total:.4f}",
                "actual_calculation": f"{total_price:.4f}",
            })
    direct_cost = labor + material + machinery + subcontract + other_total
    fee_total = management + profit
    if direct_cost > 0 and fee_total / direct_cost > 0.50:
        issues.append({
            "severity": "medium",
            "category": "综合价合理性",
            "message": "管理费和利润合计占直接费超过50%，可能存在取费基数或重复计费问题",
            "expected_calculation": f"直接费 {direct_cost:.4f}",
            "actual_calculation": f"管理费+利润 {fee_total:.4f}（占 {fee_total / direct_cost:.0%}）",
        })
    if comprehensive > 0 and direct_cost > 0 and comprehensive / direct_cost > 5:
        issues.append({
            "severity": "medium",
            "category": "综合价合理性",
            "message": "综合单价明显高于直接费，可能存在重复套项、单位含量错误或费用口径混用",
            "expected_calculation": f"直接费 {direct_cost:.4f}",
            "actual_calculation": f"综合单价 {comprehensive:.4f}",
        })

    issues = list({json.dumps(value, ensure_ascii=False, sort_keys=True, default=str): value for value in issues}.values())
    high_issues = [value for value in issues if value.get("severity") == "high"]
    medium_issues = [value for value in issues if value.get("severity") == "medium"]
    needs_correction = bool(high_issues or medium_issues or current.get("evidence_level") in {"red", "yellow"})
    return {
        "work_items": split_boq_work_items(boq_name, boq_feature),
        "coverage": {
            "uncovered_work_items": uncovered_clauses,
            "missing_materials": missing_materials,
            "cost_categories": category_coverage,
        },
        "category_coverage": category_coverage,
        "duplicate_quotas": list(duplicate_quota_counts),
        "duplicate_components": [key[6] for key, count in duplicate_component_counts.items() if count > 1],
        "issues": issues,
        "needs_correction": needs_correction,
    }


def _quota_covers_clause(item: QuotaItem, clause: str) -> bool:
    """Return true when one quota's own compositions already cover a BOQ clause."""
    if not clause:
        return True
    reference = quota_reference_dict(item)
    components = reference.get("compositions") or []
    if not uncovered_cost_clauses(
        "",
        clause,
        components,
    ):
        return True

    target_materials = _clause_material_terms(clause)
    target_crafts = _text_terms(clause, CRAFT_TERMS)
    if not target_materials:
        return False
    candidate_text = " ".join(value for value in (
        _text(item.name),
        *(
            " ".join(
                _text(component.get(key))
                for key in ("name", "code", "feature", "unit")
                if _text(component.get(key))
            )
            for component in components
        ),
    ) if value)
    candidate_materials = _clause_material_terms(candidate_text)
    if not target_materials.intersection(candidate_materials):
        return False
    generic_cover_crafts = {"制作", "安装", "绑扎"}
    if target_crafts and not target_crafts.intersection(generic_cover_crafts):
        return False
    specification = specification_relation(
        clause,
        candidate_text,
        allow_quantity_conversion=True,
        require_target_specs=True,
    )
    if specification.get("errors"):
        return False
    return True


def _quota_equivalent_clause_cover(item: QuotaItem, clause: str) -> str:
    """Return a review note when an existing component is a logical process substitute."""
    if _clause_material_terms(clause):
        return ""
    target_groups = _semantic_groups(clause, CRAFT_EQUIVALENT_GROUPS)
    if not target_groups:
        return ""
    target_locations = _location_groups(clause)
    item_locations = _location_groups(_text(item.name))
    if target_locations and item_locations and not target_locations.intersection(item_locations):
        return ""
    for detail in quota_reference_dict(item).get("compositions") or []:
        component_text = " ".join(
            _text(detail.get(key))
            for key in ("cat", "code", "name", "unit")
            if _text(detail.get(key))
        )
        shared_groups = target_groups.intersection(
            _semantic_groups(component_text, CRAFT_EQUIVALENT_GROUPS)
        )
        if shared_groups:
            label = "/".join(sorted(shared_groups))
            component_name = _text(detail.get("name")) or "定额组成"
            return f"{label}按“{component_name}”工艺替代，需人工确认适用性"
    return ""


def find_quota_composition_matches(
    session: Session,
    item_name: str,
    feature: str = "",
    unit: str = "",
    major: str = "",
    candidate_index: list[dict] | None = None,
    item_code: str = "",
    max_matches: int = 5,
    minimum_score: float = PRIMARY_MATCH_THRESHOLD,
    supplemental_minimum_score: float | None = None,
    reference_price: float | None = None,
    source_costs: dict | None = None,
    fallback_candidate_index: list[dict] | None = None,
) -> list[dict]:
    """Match each BOQ feature clause independently and retain logical low-score substitutes."""
    source_costs = source_costs or {}
    requirements = required_cost_categories(item_name, feature, source_costs)
    required = requirements["required"]
    index = candidate_index if candidate_index is not None else build_quota_match_index(session, major)
    # A specified major is a hard recall boundary. Cross-major search must be
    # an explicit caller decision, otherwise a similar-looking quota can
    # silently enter pricing from another discipline.
    fallback_index = fallback_candidate_index or index
    fallback_entries = {value["item"].id: value for value in fallback_index}
    clauses = boq_cost_clauses(item_name, feature)
    search_inputs = [(item_name, feature, "主体工序", item_code, False)]
    for clause in clauses:
        search_inputs.append((clause, clause, f"工作内容：{clause}", "", False))

    # A fallback is allowed to rescue a missing local quota, but it must not
    # turn a shared word such as "道路" or "安装" into an adopted price.
    # Keep the user's primary threshold configurable while enforcing a safety
    # floor for candidates that will enter composition and price calculation.
    fallback_threshold = max(40.0, float(minimum_score))
    direct_threshold = max(
        float(minimum_score),
        float(supplemental_minimum_score) if supplemental_minimum_score is not None else float(minimum_score) + SUPPLEMENTAL_MATCH_OFFSET,
    )
    effective_max = max(int(max_matches), len(search_inputs))
    results = []
    results_by_id = {}
    selected_names = []

    def add_match(raw_match, source_clause, fallback=False, fallback_reason=""):
        item = raw_match["item"]
        item_id = item.id
        existing = results_by_id.get(item_id)
        if existing is not None:
            clauses = existing.setdefault("matched_source_clauses", [])
            if source_clause not in clauses:
                clauses.append(source_clause)
                existing["source_clause"] = "\n".join(clauses)
                existing["reasons"] = list(dict.fromkeys([
                    *existing.get("reasons", []),
                    source_clause,
                ]))
            if fallback:
                existing["evidence_level"] = "red"
                existing["reasons"] = list(dict.fromkeys([
                    *existing.get("reasons", []),
                    fallback_reason or "低相似度替补，材料或含量需人工调整",
                ]))
            if raw_match.get("components") and existing.get("component_only"):
                component_keys = {
                    (
                        _text(value.get("cat")), _text(value.get("code")),
                        _text(value.get("name")), _text(value.get("unit")),
                    )
                    for value in existing.get("components") or []
                }
                for value in raw_match["components"]:
                    key = (
                        _text(value.get("cat")), _text(value.get("code")),
                        _text(value.get("name")), _text(value.get("unit")),
                    )
                    if key not in component_keys:
                        existing.setdefault("components", []).append(dict(value))
                        component_keys.add(key)
            return existing

        match = dict(raw_match)
        required_category = _text(raw_match.get("required_category"))
        role = (
            f"补充{required_category}"
            if required_category
            else (
                "补充材料"
                if raw_match.get("components")
                else ("主体" if not results else "补充工序")
            )
        )
        if fallback:
            role = f"{role}（低相似度）"
        match.update({
            "role": role,
            "source_clause": source_clause,
            "matched_source_clauses": [source_clause],
            "evidence_level": (
                "red"
                if fallback
                else ("green" if float(match["score"]) >= 0.75 else "yellow")
            ),
        })
        if fallback:
            match["reasons"] = list(dict.fromkeys([
                *match.get("reasons", []),
                fallback_reason or "低相似度替补，材料或含量需人工调整",
            ]))
        if raw_match.get("components"):
            match["components"] = [dict(value) for value in raw_match["components"]]
            match["component_only"] = True
        results.append(match)
        results_by_id[item_id] = match
        selected_names.append(_match_text(item.name))
        return match

    def scope_text(item):
        entry = fallback_entries.get(item.id)
        reference = (entry or {}).get("reference") or quota_reference_dict(item)
        composition_text = " ".join(
            " ".join(
                _text(value.get(key))
                for key in ("cat", "code", "name", "unit")
                if _text(value.get(key))
            )
            for value in reference.get("compositions") or []
        )
        return " ".join(value for value in (
            _text(item.code), _text(item.name), _text(item.unit), composition_text,
        ) if value)

    def acceptable_fallback(clause, match):
        if float(match["score"]) < fallback_threshold / 100:
            return False
        item = match["item"]
        candidate = scope_text(item)
        target_text = clause
        shared = (
            _text_terms(target_text, CRAFT_TERMS).intersection(_text_terms(candidate, CRAFT_TERMS))
            or _clause_material_terms(target_text).intersection(_clause_material_terms(candidate))
            or _location_groups(target_text).intersection(_location_groups(candidate))
            or _size_specs(target_text).intersection(_size_specs(candidate))
        )
        return bool(shared) and (
            match["name_score"] >= 0.55
            or match["context_score"] >= 0.65
            or match["component_name_score"] >= 0.62
        )

    def component_fallback(clause):
        target = _match_text(clause)
        target_materials = _clause_material_terms(clause)
        target_material_groups = _semantic_groups(clause, MATERIAL_EQUIVALENT_GROUPS)
        target_crafts = _text_terms(clause, CRAFT_TERMS)
        target_craft_groups = _semantic_groups(clause, CRAFT_EQUIVALENT_GROUPS)
        target_locations = _location_groups(clause)
        best = None
        for entry in fallback_index:
            item = entry["item"]
            entry = fallback_entries.get(item.id)
            reference_compositions = (
                (entry or {}).get("reference") or quota_reference_dict(item)
            ).get("compositions") or []
            item_locations = _location_groups(_text(item.name))
            item_scope_text = " ".join(value for value in (
                _text(item.code), _text(item.name), _text(item.feature),
                *(
                    " ".join(_text(component.get(key)) for key in ("name", "feature", "note"))
                    for component in reference_compositions
                ),
            ) if value)
            scope_conflict = _unsupported_special_scope(
                clause,
                item_scope_text,
            )
            if scope_conflict:
                # Component-only fallback must obey the same location guard as
                # a full quota match. Never borrow a component from an
                # exterior/interior/roof quota for an unspecified BOQ scope.
                continue
            for component in reference_compositions:
                category = _text(component.get("cat"))
                if category not in QUOTA_DETAIL_CATEGORIES:
                    continue
                component_name_raw = _text(component.get("name")) or _text(component.get("code"))
                component_code_raw = _text(component.get("code"))
                component_unit_raw = _text(component.get("unit"))
                component_name = _match_text(component_name_raw)
                component_text = _match_text(
                    f"{category} {component_code_raw} {component_name_raw} {component_unit_raw}"
                )
                if _unsupported_special_scope(clause, component_text):
                    continue
                if _explicit_material_conflict(clause, component_text):
                    continue
                if _earthwork_logic_conflict(clause, component_text):
                    continue
                component_materials = _clause_material_terms(component_text)
                component_material_groups = _semantic_groups(
                    component_text, MATERIAL_EQUIVALENT_GROUPS,
                )
                component_crafts = _text_terms(component_text, CRAFT_TERMS)
                component_craft_groups = _semantic_groups(
                    component_text, CRAFT_EQUIVALENT_GROUPS,
                )
                common_materials = target_materials.intersection(component_materials)
                common_material_groups = target_material_groups.intersection(
                    component_material_groups
                )
                common_crafts = target_crafts.intersection(component_crafts)
                common_craft_groups = target_craft_groups.intersection(component_craft_groups)

                if target_materials:
                    if category not in MATERIAL_COMPONENT_CATEGORIES:
                        continue
                    if not common_materials and not common_material_groups:
                        continue
                    # 具体密封/薄膜类材料不能被“混凝土、砂浆”等泛材料词抢走。
                    if target_material_groups and not common_material_groups:
                        continue
                elif target_crafts:
                    if not common_crafts and not common_craft_groups:
                        continue
                    if (
                        target_locations and item_locations
                        and not target_locations.intersection(item_locations)
                    ):
                        continue
                else:
                    continue

                name_score = max(
                    fuzz.ratio(target, component_name),
                    fuzz.WRatio(target, component_name),
                )
                context_score = fuzz.WRatio(target, component_text)
                name_materials = _clause_material_terms(component_name_raw)
                name_material_groups = _semantic_groups(
                    component_name_raw, MATERIAL_EQUIVALENT_GROUPS,
                )
                name_crafts = _text_terms(component_name_raw, CRAFT_TERMS)
                name_craft_groups = _semantic_groups(
                    component_name_raw, CRAFT_EQUIVALENT_GROUPS,
                )
                semantic_bonus = 0.0
                if target_materials:
                    if target_materials.intersection(name_materials):
                        semantic_bonus += 34
                    elif target_material_groups.intersection(name_material_groups):
                        semantic_bonus += 30
                    elif common_materials:
                        semantic_bonus += 16
                    else:
                        semantic_bonus += 12
                    for group in common_material_groups:
                        for priority, term in enumerate(
                            MATERIAL_REPLACEMENT_PRIORITY.get(group, ())
                        ):
                            if term in component_name_raw:
                                semantic_bonus += max(6, 18 - priority * 3)
                                break
                    semantic_bonus += 10
                else:
                    if target_crafts.intersection(name_crafts):
                        semantic_bonus += 28
                    elif target_craft_groups.intersection(name_craft_groups):
                        semantic_bonus += 24
                    elif common_crafts:
                        semantic_bonus += 14
                    else:
                        semantic_bonus += 10
                    if target_locations.intersection(item_locations):
                        semantic_bonus += 8
                unit_bonus = 0.0
                if _canonical_unit(unit) and _canonical_unit(component_unit_raw):
                    if _canonical_unit(unit) == _canonical_unit(component_unit_raw):
                        unit_bonus = 6
                    else:
                        # A composition may legitimately use m3/kg/工日 for one m2
                        # BOQ item; keep it eligible and require quantity review.
                        unit_bonus = -2
                component_score = min(
                    100.0,
                    name_score * 0.56 + context_score * 0.22 + semantic_bonus + unit_bonus,
                )
                if component_score < 28:
                    continue

                selected_component = dict(component)
                reasons = [
                    f"组成类别:{category or '未分类'}",
                    f"组成明细相关:{component_name_raw}（{component_score:.0f}%）",
                ]
                if unit_bonus > 0:
                    reasons.append(f"组成单位一致:{component_unit_raw}")
                elif unit_bonus < 0:
                    reasons.append(
                        f"组成单位:{component_unit_raw or '未提供'}，需按清单单位复核含量"
                    )
                if common_material_groups:
                    reasons.append(
                        f"材料同类替补:{'/'.join(sorted(common_material_groups))}"
                    )
                if common_craft_groups:
                    reasons.append(
                        f"工艺同类替补:{'/'.join(sorted(common_craft_groups))}"
                    )
                if target_materials:
                    matched_materials = set(common_materials)
                    if not matched_materials and common_material_groups:
                        matched_materials = {
                            term for term in target_materials
                            if _semantic_groups(term, MATERIAL_EQUIVALENT_GROUPS).intersection(
                                common_material_groups
                            )
                        }
                    material = max(matched_materials or target_materials, key=len)
                    reference_name = component_name_raw or "材料组成"
                    if material not in reference_name:
                        selected_component["name"] = f"{material}（以{reference_name}替补）"
                    selected_component["feature"] = (
                        f"参考组成：{reference_name}；"
                        f"对应清单特征：{clause}"
                    )
                    selected_component["note"] = "；".join(filter(None, (
                        _text(component.get("note")),
                        "低相似度材料替补，材料品种、规格、含量、计量口径和单价需人工调整",
                    )))
                candidate = {
                    "item": item,
                    "score": component_score / 100,
                    "name_score": round(name_score / 100, 4),
                    "context_score": round(context_score / 100, 4),
                    "component_name_score": round(component_score / 100, 4),
                    "reasons": reasons,
                    "components": [selected_component],
                }
                if best is None or candidate["score"] > best["score"]:
                    best = candidate
        return best

    def prefer_component_only(clause, match, component_match=None):
        """Prefer a single material composition when a whole quota only matches by a common word."""
        if not _clause_material_terms(clause):
            return False
        candidate_text = " ".join(value for value in (
            _text(match["item"].name),
            _text(match["item"].feature),
        ) if value)
        target_crafts = _text_terms(clause, CRAFT_TERMS)
        candidate_crafts = _text_terms(candidate_text, CRAFT_TERMS)
        craft_shared = target_crafts.intersection(candidate_crafts)
        craft_shared = craft_shared or _semantic_groups(
            clause, CRAFT_EQUIVALENT_GROUPS,
        ).intersection(_semantic_groups(candidate_text, CRAFT_EQUIVALENT_GROUPS))
        target_core = _core_object_groups(clause)
        candidate_core = _core_object_groups(candidate_text)
        if target_core.intersection(candidate_core) or craft_shared:
            return False
        # A component substitute may rescue a missing quota, but it must not
        # replace a complete quota merely because one material name scores
        # slightly higher. Require a clearly weak full-quota result and a
        # meaningful score advantage for the component evidence.
        full_score = float(match.get("score") or 0)
        component_score = float(component_match.get("score") or 0) if component_match else 0.0
        return (
            full_score < 0.58
            and component_score >= full_score + 0.08
            and float(match.get("name_score") or 0) < 0.68
        )

    # When the BOQ explicitly requires machinery, the primary candidate must
    # contain machinery. This prevents a labor-only quota from winning on name
    # similarity and leaving the machine cost behind.
    source_machine_present = _number(source_costs.get("machinery")) > 0
    primary_required_category = (
        "机械费"
        if source_machine_present or (required.get("机械费") and len(clauses) <= 1)
        else None
    )
    primary = find_quota_match_details(
        session,
        item_name,
        feature,
        unit,
        major=major,
        candidate_index=index,
        item_code=item_code,
        minimum_score=minimum_score,
        reference_price=reference_price,
        required_category=primary_required_category,
        source_costs=source_costs,
    )
    if primary:
        add_match(primary, "主体工序")

    for name, context, source_clause, code, _ in search_inputs[1:]:
        clause_requires_mechanical = (
            source_machine_present
            and infer_mechanical_requirement(name, context)["required"]
        )
        for existing in list(results):
            # Do not let a main quota silently absorb a later numbered feature.
            # Each clause must obtain its own fresh candidate evaluation; the
            # same quota may still be returned if it independently passes that
            # evaluation and is then deduplicated by quota id.
            if source_clause in (existing.get("matched_source_clauses") or []) and _quota_covers_clause(existing["item"], name):
                add_match(
                    {
                        "item": existing["item"],
                        "score": existing["score"],
                        "name_score": existing["name_score"],
                        "context_score": existing["context_score"],
                        "component_name_score": existing["component_name_score"],
                        "reasons": existing["reasons"],
                    },
                    source_clause,
                )
                break
            equivalent_reason = _quota_equivalent_clause_cover(existing["item"], name)
            if source_clause in (existing.get("matched_source_clauses") or []) and equivalent_reason:
                add_match(
                    {
                        "item": existing["item"],
                        "score": existing["score"],
                        "name_score": existing["name_score"],
                        "context_score": existing["context_score"],
                        "component_name_score": existing["component_name_score"],
                        "reasons": existing["reasons"],
                    },
                    source_clause,
                    fallback=True,
                    fallback_reason=equivalent_reason,
                )
                break
        else:
            match = find_quota_match_details(
                session,
                name,
                context,
                unit,
                major=major,
                candidate_index=index,
                item_code=code,
                minimum_score=minimum_score,
                reference_price=reference_price,
                required_category="机械费" if clause_requires_mechanical else None,
                source_costs=source_costs,
            )
            mechanical_already_covered = False
            if match is not None and clause_requires_mechanical:
                target_profiles = mechanical_dedup_keys(name, context)
                existing_profiles = set()
                for existing in results:
                    existing_profiles.update(
                        mechanical_dedup_keys(
                            existing["item"].name,
                            existing["item"].feature,
                        )
                    )
                if target_profiles and target_profiles.intersection(existing_profiles):
                    match = None
                    mechanical_already_covered = True
            if mechanical_already_covered:
                continue
            if match and match["item"].id in results_by_id:
                match = None
            if match and float(match["score"]) < direct_threshold / 100:
                match = None
            component_match = component_fallback(name)
            if component_match is not None and clause_requires_mechanical:
                target_profiles = mechanical_dedup_keys(name, context)
                existing_profiles = set()
                for existing in results:
                    existing_profiles.update(
                        mechanical_dedup_keys(
                            existing["item"].name,
                            existing["item"].feature,
                        )
                    )
                if target_profiles and target_profiles.intersection(existing_profiles):
                    # A component fallback can otherwise add a second
                    # mechanical quota after the primary quota already covers
                    # the same construction profile.
                    component_match = None
            if clause_requires_mechanical and component_match is not None:
                component_categories = {
                    _text(value.get("cat") or value.get("category"))
                    for value in component_match.get("components") or []
                }
                if "机械费" not in component_categories:
                    component_match = None
            component_is_usable = bool(
                component_match
                and float(component_match["score"]) >= fallback_threshold / 100
            )
            if match and component_is_usable and prefer_component_only(name, match, component_match):
                add_match(
                    component_match,
                    source_clause,
                    fallback=True,
                    fallback_reason="优先引用材料组成，整条候选定额的工序与清单不完全一致；材料、含量和计量口径需人工调整",
                )
            elif match:
                add_match(match, source_clause)
            else:
                if component_is_usable:
                    add_match(component_match, source_clause, fallback=True)
                else:
                    fallback = find_quota_match_details(
                        session,
                        name,
                        context,
                        unit,
                        major="",
                        candidate_index=fallback_index,
                        item_code=code,
                        minimum_score=fallback_threshold,
                        reference_price=reference_price,
                        required_category="机械费" if clause_requires_mechanical else None,
                        source_costs=source_costs,
                    )
                    if fallback and acceptable_fallback(name, fallback):
                        add_match(fallback, source_clause, fallback=True)
        if len(results) >= effective_max:
            break

    def result_has_category(group: str) -> bool:
        allowed = COST_CATEGORY_GROUPS[group]
        for match in results:
            details = match.get("components")
            if details is None:
                details = quota_reference_dict(match["item"]).get("compositions") or []
            if any(_text(value.get("cat") or value.get("category")) in allowed for value in details):
                return True
        return False

    def machine_result_covers_manual_coordination() -> bool:
        """Check whether a selected machine quota already includes manual work."""
        for selected in results:
            details = selected.get("components")
            if details is None:
                details = quota_reference_dict(selected["item"]).get("compositions") or []
            has_machine = any(
                _text(value.get("cat") or value.get("category")) == "机械费"
                for value in details
            )
            if not has_machine:
                continue
            if quota_covers_labor_coordination(
                selected["item"].name,
                selected["item"].feature,
            ):
                return True
            if any(
                quota_covers_labor_coordination(
                    value.get("name", ""), value.get("feature", "")
                )
                for value in details
            ):
                return True
        return False

    def category_match(group: str):
        allowed = COST_CATEGORY_GROUPS[group]
        labor_is_covered = group == "人工费" and machine_result_covers_manual_coordination()
        # Search the full row first, then each numbered work-content clause.
        candidates_to_try = [(item_name, feature, "费用类别补充：" + group)]
        candidates_to_try.extend(
            (clause, clause, f"费用类别补充：{group} / {clause}")
            for clause in boq_cost_clauses(item_name, feature)
        )
        for target_name, target_feature, source_clause in candidates_to_try:
            if source_clause != "费用类别补充：" + group:
                clause_text = f"{target_name} {target_feature}"
                if group == "机械费":
                    # For an imported machine-cost column, the full row is
                    # sufficient evidence even if the feature text omits the
                    # word "mechanical". Numbered clauses are searched only
                    # when they explicitly identify a machine process, which
                    # avoids adding the same machine cost for every clause.
                    if not infer_mechanical_requirement(
                        target_name, target_feature
                    )["required"] and _number(source_costs.get("machinery")) <= 0:
                        continue
                    if (
                        _number(source_costs.get("machinery")) > 0
                        and source_clause != "费用类别补充：" + group
                    ):
                        continue
                if group == "材料费" and not (
                    _clause_material_terms(clause_text) or _technical_specs(clause_text)
                ):
                    continue
            match = find_quota_match_details(
                session,
                target_name,
                target_feature,
                unit,
                major=major,
                candidate_index=index,
                item_code=item_code if target_name == item_name else "",
                minimum_score=fallback_threshold,
                reference_price=reference_price,
                required_category=allowed,
                source_costs=source_costs,
            )
            if match is None:
                continue
            if labor_is_covered and (
                is_full_manual_excavation_quota(match["item"].name, match["item"].feature)
                or not quota_covers_labor_coordination(
                    match["item"].name, match["item"].feature
                )
            ):
                # Do not add a complete manual excavation operation merely to
                # satisfy the source labor column when the selected machine
                # quota already includes manual clearing/edge trimming.
                continue
            minimum_category_score = 0.50 if group == "机械费" else 0.55
            if float(match.get("score") or 0) < minimum_category_score:
                continue
            reference = quota_reference_dict(match["item"])
            components = [
                dict(value)
                for value in reference.get("compositions") or []
                if _text(value.get("cat") or value.get("category")) in allowed
            ]
            if not components:
                continue
            match["components"] = components
            match["required_category"] = group
            match["component_only"] = True
            match["reasons"] = list(dict.fromkeys([
                *match.get("reasons", []),
                f"补齐费用类别：{group}",
            ]))
            return match, source_clause
        return None, ""

    # A complete unit-price analysis is assembled from category evidence. A
    # combined quota may satisfy several groups; only the missing group is
    # added from a supplemental quota to avoid double counting.
    source_required_groups = {
        group
        for group, key in (("人工费", "labor"), ("材料费", "material"), ("机械费", "machinery"))
        if required.get(group) or _number(source_costs.get(key)) > 0
    }
    for group in ("人工费", "材料费", "机械费"):
        if group not in source_required_groups or result_has_category(group):
            continue
        supplemental, source_clause = category_match(group)
        if supplemental is not None:
            add_match(supplemental, source_clause, fallback=True,
                      fallback_reason=f"补齐清单要求的{group}组成，需复核含量与适用范围")
    return results


def _composition_dict(composition: QuotaComposition) -> dict:
    return {
        "cat": composition.category,
        "code": composition.code,
        "name": composition.name,
        "feature": composition.feature,
        "unit": composition.unit,
        "qty": composition.qty,
        "loss": f"{composition.loss_rate:g}%" if composition.loss_rate else "",
        "noTaxPrice": composition.no_tax_price,
        "taxRate": f"{composition.tax_rate:g}%" if composition.tax_rate else "",
        "taxPrice": composition.tax_price,
        "noTaxTotal": composition.no_tax_total,
        "taxTotal": composition.tax_total,
        "note": composition.note,
    }


def quota_to_dict(item: QuotaItem) -> dict:
    """输出为 run(5).py / 装饰成本定额.json 兼容结构。"""
    return {
        "code": item.code,
        "name": item.name,
        "feature": item.feature,
        "unit": item.unit,
        "taxPrice": item.tax_price,
        "noTaxPrice": item.no_tax_price,
        "note": item.category,
        "source": item.source,
        "items": [_composition_dict(comp) for comp in item.compositions],
    }


def search_quota_items(
    session: Session,
    *,
    major: str = "",
    keyword: str = "",
    category: str = "",
    limit: int = 2000,
) -> list[QuotaItem]:
    query = session.query(QuotaItem).options(selectinload(QuotaItem.compositions))
    if major:
        query = query.filter(QuotaItem.major.in_(quota_major_variants(major)))
    if keyword:
        key = keyword.strip()
        query = query.filter(or_(
            QuotaItem.name.contains(key),
            QuotaItem.code.contains(key),
            QuotaItem.feature.contains(key),
            QuotaItem.notes.contains(key),
        ))
    if category:
        query = query.filter(QuotaItem.category == category)
    return query.order_by(QuotaItem.major, QuotaItem.name, QuotaItem.id).limit(limit).all()


def list_quota_majors(session: Session) -> list[str]:
    values = [
        row[0]
        for row in session.query(QuotaItem.major)
        .distinct()
        .order_by(QuotaItem.major)
        .all()
        if row[0]
    ]
    for major in DEFAULT_MAJORS:
        if major not in values:
            values.append(major)
    return values


def list_quota_categories(session: Session, major: str = "") -> list[str]:
    query = session.query(QuotaItem.category)
    if major:
        query = query.filter(QuotaItem.major.in_(quota_major_variants(major)))
    values = [
        row[0]
        for row in query.distinct().order_by(QuotaItem.category).all()
        if row[0]
    ]
    return values


def quota_composition_category_counts(session: Session, major: str = "") -> dict[str, int]:
    query = (
        session.query(QuotaComposition.category, func.count(QuotaComposition.id))
        .join(QuotaItem, QuotaComposition.quota_item_id == QuotaItem.id)
    )
    if major:
        query = query.filter(QuotaItem.major == major)
    return {
        category: int(count)
        for category, count in query.group_by(QuotaComposition.category).all()
        if category
    }


def get_quota_item(session: Session, quota_id: int) -> QuotaItem | None:
    return session.query(QuotaItem).filter(QuotaItem.id == quota_id).first()


def delete_quota_item(session: Session, quota_id: int, audit: bool = True) -> bool:
    item = session.query(QuotaItem).filter(QuotaItem.id == quota_id).first()
    if item is None:
        return False
    name = item.name
    session.delete(item)
    if audit:
        write_audit(session, "delete", "quota_item", item.id, f"删除定额: {name}")
    return True


def _save_compositions(
    item: QuotaItem,
    compositions: Iterable[dict],
) -> list[QuotaComposition]:
    item.compositions = []
    result = []
    for index, values in enumerate(compositions):
        composition = QuotaComposition(
            sort_order=index,
            category=_text(values.get("cat") or values.get("category")),
            code=_text(values.get("code")),
            name=_text(values.get("name")),
            feature=_text(values.get("feature")),
            unit=_text(values.get("unit")),
            qty=_number(values.get("qty")),
            loss_rate=_percent(values.get("loss")),
            no_tax_price=_number(values.get("noTaxPrice") or values.get("no_tax_price")),
            tax_rate=_percent(values.get("taxRate") or values.get("tax_rate")),
            tax_price=_number(values.get("taxPrice") or values.get("tax_price")),
            no_tax_total=_number(values.get("noTaxTotal") or values.get("no_tax_total")),
            tax_total=_number(values.get("taxTotal") or values.get("tax_total")),
            note=_text(values.get("note")),
        )
        item.compositions.append(composition)
        result.append(composition)
    return result


def save_quota_item(
    session: Session,
    *,
    major: str,
    code: str,
    name: str,
    source_key: str | None = None,
    feature: str = "",
    unit: str = "",
    tax_price: float = 0.0,
    no_tax_price: float = 0.0,
    category: str = "",
    source: str = "",
    notes: str = "",
    compositions: Iterable[dict] | None = None,
    quota_id: int | None = None,
    audit: bool = True,
) -> QuotaItem:
    if quota_id:
        item = session.query(QuotaItem).filter(QuotaItem.id == quota_id).first()
    else:
        item = None
    if item is None:
        source_key = source_key or _source_key(
            "quota",
            major,
            code,
            name,
            feature,
            unit,
        )
        item = session.query(QuotaItem).filter(QuotaItem.source_key == source_key).first()
    if item is None:
        item = QuotaItem(
            source_key=source_key or _source_key("quota", major, code, name, feature, unit),
        )
        session.add(item)
    item.major = _text(major) or "装修"
    item.code = _text(code)
    item.name = _text(name)
    item.feature = _text(feature)
    item.unit = _text(unit)
    item.tax_price = _number(tax_price)
    item.no_tax_price = _number(no_tax_price)
    item.category = _text(category)
    item.source = _text(source) or item.source
    item.notes = _text(notes)
    if compositions is not None:
        _save_compositions(item, compositions)
    session.flush()
    if audit:
        write_audit(
            session,
            "update" if quota_id else "create",
            "quota_item",
            item.id,
            f"保存定额: {item.name}",
        )
    return item


def preserve_imported_cost_references(
    session: Session,
    rows: Iterable[dict],
    *,
    source_file: str = "",
    major: str = "导入表人材机",
    audit: bool = True,
) -> dict:
    """Extract priced labor/material/machine data from an imported BOQ.

    Imported cost columns are commonly engineering-quantity totals rather
    than unit prices. The saved records are therefore explicitly marked as
    source references and converted to a unit basis. They remain separate from
    formal quota data so users can filter, review, and reuse them without
    losing the original evidence.
    """
    source_file = _text(source_file)
    resolved_major = _text(major) or "导入表人材机"
    category_fields = (
        ("人工费", "labor", "人工费"),
        ("材料费", "material", "材料费"),
        ("机械费", "machinery", "机械费"),
    )
    imported = 0
    updated = 0
    skipped = 0
    component_count = 0
    category_counts = {category: 0 for category, _, _ in category_fields}

    for row in rows or ():
        row = dict(row or {})
        name = _text(row.get("name"))
        unit = _text(row.get("unit"))
        if not name:
            skipped += 1
            continue
        quantity = _number(row.get("quantity"))
        source_comprehensive = _number(
            row.get("imported_comprehensive_price")
            or row.get("comprehensive_price")
        )
        source_total = _number(
            row.get("imported_total_price")
            or row.get("total_price")
        )
        raw_costs = {}
        for _, field, _ in category_fields:
            raw = _number(row.get(f"imported_{field}"))
            if raw <= 0:
                raw = _number(row.get(field))
            raw_costs[field] = raw if raw > 0 else 0.0
        component_sum = sum(raw_costs.values())
        unit_comprehensive = (
            source_comprehensive
            if source_comprehensive > 0
            else source_total / quantity
            if source_total > 0 and quantity > 0
            else 0.0
        )
        total_basis = False
        conversion_basis = ""
        if quantity > 0 and abs(quantity - 1.0) > 0.000001 and component_sum > 0:
            denominator = max(component_sum, source_total, 0.000001)
            if source_total > 0 and abs(component_sum - source_total) / denominator <= 0.25:
                total_basis = True
                conversion_basis = f"组成合计 {component_sum:g} ÷ 工程量 {quantity:g}"
            elif unit_comprehensive > 0:
                denominator = max(component_sum / quantity, unit_comprehensive, 0.000001)
                if abs(component_sum / quantity - unit_comprehensive) / denominator <= 0.35:
                    total_basis = True
                    conversion_basis = (
                        f"组成合计 {component_sum:g} ÷ 工程量 {quantity:g}"
                        f" ≈ 单位综合单价 {unit_comprehensive:g}"
                    )
        components = []
        for category, field, label in category_fields:
            raw = raw_costs[field]
            if raw <= 0:
                continue
            unit_price = raw / quantity if total_basis and quantity > 0 else raw
            components.append({
                "cat": category,
                "code": f"SOURCE-{field.upper()}",
                "name": f"{name} / 原表{label}",
                "feature": _text(row.get("feature")),
                "unit": unit,
                "qty": 1.0,
                "noTaxPrice": unit_price,
                "taxPrice": unit_price,
                "note": (
                    "导入表费用参考；"
                    + (
                        f"原始金额按{conversion_basis}折算为单位价格，税价口径需复核"
                        if total_basis
                        else "按原表单位金额保存，税价口径需复核"
                    )
                ),
            })
            category_counts[category] += 1
        if not components:
            skipped += 1
            continue

        source_identity = "|".join((
            source_file,
            _text(row.get("source_tab")),
            _text(row.get("seq")),
            _text(row.get("code")),
            name,
            _text(row.get("feature")),
            unit,
        ))
        digest = hashlib.sha1(source_identity.encode("utf-8")).hexdigest()[:16]
        code = _text(row.get("code")) or f"SOURCE-{digest}"
        feature = _text(row.get("feature"))
        source_key = _source_key(
            "imported-cost", resolved_major, f"{code}:{digest}", name, feature, unit,
        )
        existing = session.query(QuotaItem).filter(
            QuotaItem.source_key == source_key
        ).first()
        if existing is None:
            imported += 1
        else:
            updated += 1
        save_quota_item(
            session,
            quota_id=existing.id if existing else None,
            source_key=source_key,
            major=resolved_major,
            code=code,
            name=name,
            feature=feature,
            unit=unit,
            tax_price=unit_comprehensive,
            no_tax_price=unit_comprehensive,
            category="导入表人材机参考",
            source=source_file,
            notes=(
                f"来源工作表：{_text(row.get('source_tab')) or '未标明'}；"
                f"清单编码：{_text(row.get('code')) or '未标明'}；"
                f"原始工程量：{quantity:g}；"
                f"单位价格口径：{conversion_basis or '原表已为单位金额'}；"
                "仅保存导入表人材机费用证据，不替代正式定额。"
            ),
            compositions=components,
            audit=False,
        )
        component_count += len(components)

    session.flush()
    if audit and (imported or updated):
        write_audit(
            session,
            "import",
            "quota_item",
            detail=(
                f"导入表人材机参考提取: 新增 {imported} 条，更新 {updated} 条，"
                f"组成 {component_count} 条，跳过 {skipped} 行"
            ),
        )
    return {
        "success": True,
        "major": resolved_major,
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "components": component_count,
        "categories": category_counts,
        "total": imported + updated,
    }


def save_ai_generated_quota(
    session: Session,
    generated: dict,
    *,
    boq: dict | None = None,
    project: dict | None = None,
    audit: bool = True,
) -> QuotaItem:
    """Persist a reviewed AI estimate as reusable, clearly marked AI quota data."""
    generated = dict(generated or {})
    boq = dict(boq or {})
    project = dict(project or {})
    name = _text(generated.get("name")) or _text(boq.get("name"))
    unit = _text(generated.get("unit")) or _text(boq.get("unit"))
    code = _text(generated.get("code")) or "AI-GENERATED"
    major = _text(generated.get("major")) or _text(boq.get("major")) or "其他"
    feature = _text(generated.get("feature")) or _text(boq.get("feature"))
    city = _text(project.get("city"))
    province = _text(project.get("province"))
    period = _text(project.get("pricing_date") or project.get("price_year"))
    region_key = "|".join((province, city, period))
    source_key = _source_key("ai", major, code, name, feature, unit) + "|" + hashlib.sha1(
        region_key.encode("utf-8")
    ).hexdigest()[:12]
    metadata = {
        "record_type": "ai_quota",
        "status": "待人工复核",
        "project_id": project.get("project_id"),
        "project_name": project.get("name"),
        "province": province,
        "city": city,
        "district": _text(project.get("district")),
        "pricing_date": period,
        "price_year": _text(project.get("price_year")),
        "pricing_stage": _text(project.get("stage")),
        "source_type": _text(generated.get("source_type")) or "ai_generated_market_estimate",
        "evidence_level": _text(generated.get("evidence_level")) or "red",
        "confidence": _number(generated.get("confidence")),
        "assumptions": generated.get("assumptions") or [],
        "unit_conversion": _text(generated.get("unit_conversion")),
        "evidence_summary": _text(generated.get("evidence_summary")),
        "source_evidence": generated.get("source_evidence") or generated.get("sourceEvidence") or [],
        "boq_name": _text(boq.get("name")),
        "boq_code": _text(boq.get("code")),
        "boq_feature": _text(boq.get("feature")),
    }
    notes = json.dumps(metadata, ensure_ascii=False, default=str)
    return save_quota_item(
        session,
        major=major,
        code=code,
        name=name,
        source_key=source_key,
        feature=feature,
        unit=unit,
        tax_price=_number(generated.get("tax_price") or generated.get("taxPrice")),
        no_tax_price=_number(generated.get("no_tax_price") or generated.get("noTaxPrice")),
        category="AI定额数据",
        source="AI生成/联网检索市场估算",
        notes=notes,
        compositions=generated.get("components") or generated.get("items") or [],
        audit=audit,
    )


def import_quota_excel(
    filepath: str | Path,
    session: Session,
    *,
    major: str | None = None,
    audit: bool = True,
) -> dict:
    """Import valid quota details or grouped quota compositions from Excel."""
    path = Path(filepath)
    resolved_major = _text(major) or infer_quota_major(path) or "装修"
    imported = 0
    updated = 0
    skipped = 0
    rejection_reasons: dict[str, int] = {}

    header_aliases = {
        "category": ("类别", "费用类别", "组成类别"),
        "code": ("定额编码", "编码", "项目编码"),
        "name": ("定额名称", "工料机名称", "材料名称", "项目名称", "工程名称", "名称"),
        "feature": ("工作内容", "项目特征及工作内容", "项目特征", "特征描述"),
        "unit": ("计量单位", "单位"),
        "qty": ("单位含量", "消耗量", "含量"),
        "loss": ("损耗率%", "损耗率", "损耗"),
        "no_tax_price": ("除税单价", "不含税单价", "税前单价"),
        "tax_rate": ("税率%", "税率"),
        "tax_price": ("含税单价",),
        "no_tax_total": ("除税合价", "不含税合价", "税前合价"),
        "tax_total": ("含税合价",),
    }

    def normalize_header(value) -> str:
        return re.sub(r"[\s　:：()（）]+", "", _text(value)).lower()

    normalized_aliases = {
        field: tuple(normalize_header(alias) for alias in aliases)
        for field, aliases in header_aliases.items()
    }

    def header_indexes(rows: list[tuple], row_index: int, depth: int) -> dict[str, int | None]:
        width = max((len(row) for row in rows[row_index:row_index + depth]), default=0)
        columns = []
        for column in range(width):
            parts = []
            for row in rows[row_index:row_index + depth]:
                value = normalize_header(row[column] if column < len(row) else "")
                if value and value not in parts:
                    parts.append(value)
            columns.append("".join(parts))
        indexes: dict[str, int | None] = {field: None for field in header_aliases}
        for field, aliases in normalized_aliases.items():
            candidates = []
            for column, value in enumerate(columns):
                for alias in aliases:
                    if value == alias:
                        candidates.append((3, -len(value), column))
                    elif alias and alias in value:
                        candidates.append((2, -len(value), column))
            if candidates:
                indexes[field] = max(candidates)[2]
        return indexes

    def find_tables(sheets) -> list[tuple[str, list[tuple], int, int, dict]]:
        found = []
        for sheet_name, sheet_rows in sheets:
            rows = [tuple(row) for row in sheet_rows]
            best = None
            for row_index in range(min(len(rows), 50)):
                for depth in (1, 2, 3):
                    indexes = header_indexes(rows, row_index, depth)
                    if indexes["category"] is None or indexes["name"] is None or indexes["unit"] is None:
                        continue
                    if indexes["no_tax_price"] is None and indexes["tax_price"] is None:
                        continue
                    score = (
                        8
                        + (2 if indexes["code"] is not None else 0)
                        + (2 if indexes["feature"] is not None else 0)
                        + (3 if indexes["qty"] is not None else 0)
                        + (1 if indexes["loss"] is not None else 0)
                        + (1 if indexes["tax_rate"] is not None else 0)
                        + (1 if indexes["no_tax_total"] is not None else 0)
                        + (1 if indexes["tax_total"] is not None else 0)
                    )
                    candidate = (score, -row_index, -depth, row_index, depth, indexes)
                    if best is None or candidate[:3] > best[:3]:
                        best = candidate
            if best is not None:
                found.append((sheet_name, rows, best[3], best[4], best[5]))
        return found

    def reject(reason: str):
        nonlocal skipped
        skipped += 1
        rejection_reasons[reason] = rejection_reasons.get(reason, 0) + 1

    def cell(values, indexes, field: str) -> str:
        index = indexes.get(field)
        return _text(values[index]) if index is not None and index < len(values) else ""

    def normalize_category(value) -> str:
        category = _text(value)
        return AI_COMPOSITION_CATEGORY_ALIASES.get(category, category)

    def invalid_unit_reason(unit) -> str:
        value = _text(unit)
        normalized = re.sub(r"\s+", "", value)
        if not normalized or normalized in {"单位", "计量单位"}:
            return "计量单位为空或无效"
        if looks_like_engineering_code(normalized) or re.fullmatch(r"\d+(?:\.\d+)?", normalized):
            return "计量单位被识别为编码或数字"
        if len(normalized) > 20 or "\n" in value or any(mark in value for mark in ("：", ":", "；", ";")):
            return "计量单位包含项目描述，疑似列错位"
        return ""

    def invalid_detail_reason(category, code, name, unit, qty, no_tax_price, tax_price, *, require_qty: bool) -> str:
        if category not in QUOTA_DETAIL_CATEGORIES:
            return "类别不是有效人材机类别"
        if _text(name) in set(QUOTA_DETAIL_CATEGORIES):
            return "定额名称只是费用类别标题"
        reason = invalid_engineering_name_reason(name, code)
        if reason:
            return reason.replace("工程名称", "定额名称")
        reason = invalid_unit_reason(unit)
        if reason:
            return reason
        if require_qty and _number(qty) <= 0:
            return "单位含量为空、为0或不是有效正数"
        if _number(no_tax_price) <= 0 and _number(tax_price) <= 0:
            return "除税单价和含税单价均为空或非正数"
        return ""

    def save_detail(category, code, name, unit, no_tax_price, tax_price, sheet_name):
        nonlocal imported, updated
        source_key = _source_key(
            "excel-detail", resolved_major, f"{category}:{code}", name, "", unit,
        )
        existing = session.query(QuotaItem).filter(QuotaItem.source_key == source_key).first()
        imported += int(existing is None)
        updated += int(existing is not None)
        save_quota_item(
            session,
            quota_id=existing.id if existing else None,
            source_key=source_key,
            major=resolved_major,
            code=code,
            name=name,
            feature="",
            unit=unit,
            tax_price=tax_price,
            no_tax_price=no_tax_price,
            category="明细",
            source=str(filepath),
            notes=f"明细表导入；来源工作表：{sheet_name}",
            compositions=[{
                "cat": category, "code": code, "name": name, "feature": "", "unit": unit,
                "qty": 1, "noTaxPrice": no_tax_price, "taxPrice": tax_price,
                "noTaxTotal": no_tax_price, "taxTotal": tax_price,
            }],
            audit=False,
        )

    try:
        sheets = read_excel_sheets(path)
        tables = find_tables(sheets)
        if not tables:
            return {
                "success": False, "major": resolved_major, "imported": 0, "updated": 0,
                "skipped": 0, "total": 0,
                "error": "未找到包含类别、定额名称、单位和单价的有效定额表头。",
            }

        for sheet_name, rows, header_row, header_depth, indexes in tables:
            data_rows = rows[header_row + header_depth:]
            detail_only = indexes["qty"] is None and indexes["loss"] is None
            if detail_only:
                for values in data_rows:
                    if not any(value not in (None, "") for value in values):
                        continue
                    category = normalize_category(cell(values, indexes, "category"))
                    code = cell(values, indexes, "code")
                    name = cell(values, indexes, "name")
                    unit = cell(values, indexes, "unit")
                    no_tax_price = _number(cell(values, indexes, "no_tax_price"))
                    tax_price = _number(cell(values, indexes, "tax_price"))
                    reason = invalid_detail_reason(
                        category, code, name, unit, 1, no_tax_price, tax_price,
                        require_qty=False,
                    )
                    if reason:
                        reject(reason)
                        continue
                    save_detail(category, code, name, unit, no_tax_price, tax_price, sheet_name)
                continue

            current: dict | None = None

            def flush_current():
                nonlocal current, imported, updated
                if current is None:
                    return
                if not current["compositions"]:
                    reject("定额主项没有有效人材机明细")
                    current = None
                    return
                current["no_tax_price"] = sum(
                    _number(item.get("noTaxTotal"))
                    or _number(item.get("qty")) * (1 + _number(item.get("loss"))) * _number(item.get("noTaxPrice"))
                    for item in current["compositions"]
                )
                current["tax_price"] = sum(
                    _number(item.get("taxTotal"))
                    or _number(item.get("qty")) * (1 + _number(item.get("loss"))) * _number(item.get("taxPrice"))
                    for item in current["compositions"]
                )
                source_key = _source_key(
                    "excel", resolved_major, current["code"], current["name"],
                    current["feature"], current["unit"],
                )
                existing = session.query(QuotaItem).filter(QuotaItem.source_key == source_key).first()
                imported += int(existing is None)
                updated += int(existing is not None)
                save_quota_item(
                    session,
                    quota_id=existing.id if existing else None,
                    source_key=source_key,
                    major=resolved_major,
                    code=current["code"],
                    name=current["name"],
                    feature=current["feature"],
                    unit=current["unit"],
                    tax_price=current["tax_price"],
                    no_tax_price=current["no_tax_price"],
                    category="组合定额",
                    source=str(filepath),
                    notes=f"组合定额导入；来源工作表：{sheet_name}",
                    compositions=current["compositions"],
                    audit=False,
                )
                current = None

            for values in data_rows:
                if not any(value not in (None, "") for value in values):
                    continue
                raw_category = cell(values, indexes, "category")
                category = normalize_category(raw_category)
                code = cell(values, indexes, "code")
                name = cell(values, indexes, "name")
                unit = cell(values, indexes, "unit")
                feature = cell(values, indexes, "feature")
                if not raw_category and name:
                    flush_current()
                    reason = invalid_engineering_name_reason(name, code)
                    if reason:
                        reject(reason.replace("工程名称", "定额名称"))
                        continue
                    reason = invalid_unit_reason(unit)
                    if reason:
                        reject(reason)
                        continue
                    current = {
                        "code": code, "name": name, "feature": feature, "unit": unit,
                        "tax_price": 0.0, "no_tax_price": 0.0, "compositions": [],
                    }
                    continue
                if current is None:
                    reject("人材机明细缺少有效定额主项")
                    continue
                qty = _number(cell(values, indexes, "qty"))
                loss = _percent(cell(values, indexes, "loss"))
                no_tax_price = _number(cell(values, indexes, "no_tax_price"))
                tax_price = _number(cell(values, indexes, "tax_price"))
                reason = invalid_detail_reason(
                    category, code, name, unit, qty, no_tax_price, tax_price,
                    require_qty=True,
                )
                if reason:
                    reject(reason)
                    continue
                current["compositions"].append({
                    "cat": category,
                    "code": code,
                    "name": name,
                    "feature": feature,
                    "unit": unit,
                    "qty": qty,
                    "loss": loss,
                    "noTaxPrice": no_tax_price,
                    "taxRate": _percent(cell(values, indexes, "tax_rate")),
                    "taxPrice": tax_price,
                    "noTaxTotal": _number(cell(values, indexes, "no_tax_total")),
                    "taxTotal": _number(cell(values, indexes, "tax_total")),
                    "note": "",
                })
            flush_current()

        session.flush()
        if audit:
            write_audit(
                session,
                "import",
                "quota_item",
                detail=f"定额库导入: 新增 {imported} 条，更新 {updated} 条，拦截 {skipped} 行",
            )
        return {
            "success": True,
            "major": resolved_major,
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "total": imported + updated,
            "rejection_reasons": rejection_reasons,
            "sheets": len(tables),
        }
    except Exception as error:
        session.rollback()
        return {
            "success": False,
            "major": resolved_major,
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "total": imported + updated,
            "rejection_reasons": rejection_reasons,
            "error": str(error),
        }


def import_quota_json(
    filepath: str | Path,
    session: Session,
    *,
    major: str = "装修",
    audit: bool = True,
) -> dict:
    """导入 run(5).py 兼容的装饰成本定额.json。"""
    with open(filepath, "r", encoding="utf-8") as file:
        payload = json.load(file)

    if isinstance(payload, dict):
        items_by_major = payload
    elif isinstance(payload, list):
        items_by_major = {major: payload}
    else:
        return {"success": False, "imported": 0, "updated": 0, "skipped": 0, "error": "定额 JSON 格式不正确"}

    imported = 0
    updated = 0
    skipped = 0
    for current_major, records in items_by_major.items():
        if not isinstance(records, list):
            continue
        for record in records:
            if not isinstance(record, dict) or not _text(record.get("name")):
                skipped += 1
                continue
            code = _text(record.get("code"))
            name = _text(record.get("name"))
            feature = _text(record.get("feature"))
            unit = _text(record.get("unit"))
            source_key = _source_key("json", current_major, code, name, feature, unit)
            existing = session.query(QuotaItem).filter(QuotaItem.source_key == source_key).first()
            if existing is None:
                imported += 1
            else:
                updated += 1
            save_quota_item(
                session,
                quota_id=existing.id if existing else None,
                source_key=source_key,
                major=current_major,
                code=code,
                name=name,
                feature=feature,
                unit=unit,
                tax_price=_number(record.get("taxPrice")),
                no_tax_price=_number(record.get("noTaxPrice")),
                category=_text(record.get("note") or record.get("remark")),
                source=str(filepath),
                notes="",
                compositions=record.get("items") or [],
                audit=False,
            )
            if (imported + updated) % 200 == 0:
                session.flush()
    session.flush()
    if audit:
        write_audit(
            session,
            "import",
            "quota_item",
            detail=f"定额 JSON 导入: 新增 {imported} 条，更新 {updated} 条，跳过 {skipped} 条",
        )
    return {
        "success": True,
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "total": imported + updated,
    }


def export_quota_json(items: Iterable[QuotaItem], filepath: str | Path):
    payload = [quota_to_dict(item) for item in items]
    with open(filepath, "w", encoding="utf-8", newline="\n") as file:
        json.dump(payload, file, ensure_ascii=False, indent=1)


def export_quota_excel(items: Iterable[QuotaItem], filepath: str | Path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "成本组价明细"
    worksheet.append(list(QUOTA_EXCEL_HEADERS))
    fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for item in items:
        for composition in item.compositions:
            if composition.category not in QUOTA_DETAIL_CATEGORIES:
                continue
            worksheet.append([
                composition.category,
                composition.code,
                composition.name,
                composition.unit,
                composition.no_tax_price,
                composition.tax_price,
            ])
    for column in worksheet.columns:
        letter = column[0].column_letter
        worksheet.column_dimensions[letter].width = 16
    worksheet.column_dimensions["C"].width = 42
    worksheet.column_dimensions["D"].width = 12
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(filepath)
