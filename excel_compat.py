"""统一兼容 xlsx、xls 以及部分软件导出的文本表格。"""
from __future__ import annotations

import csv
import io
import zipfile
from pathlib import Path


def _read_text_table(data: bytes):
    encodings = ("utf-8-sig", "gb18030", "utf-16", "utf-16-le", "utf-16-be")
    for encoding in encodings:
        try:
            text = data.decode(encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
        if not text.strip():
            continue
        delimiter = "\t" if "\t" in text else ","
        rows = [tuple(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
        if rows and any(len(row) > 1 for row in rows):
            return [("文本表格", rows)]
    return None


def read_excel_sheets(filepath: str | Path):
    """按文件签名读取，避免扩展名错误触发底层编码异常。"""
    path = Path(filepath)
    if path.name.startswith("~$"):
        raise ValueError("当前选择的是 Excel 临时锁定文件（~$），请关闭文件后选择原始文件。")
    data = path.read_bytes()
    if data.startswith(b"PK\x03\x04"):
        from openpyxl import load_workbook

        try:
            book = load_workbook(str(path), read_only=True, data_only=True)
        except (UnicodeError, zipfile.BadZipFile, OSError) as error:
            raise ValueError("Excel 文件损坏或内部编码异常，请在 Excel 中另存为新的 .xlsx 后重新导入。") from error
        try:
            return [(sheet.title, [tuple(row) for row in sheet.iter_rows(values_only=True)]) for sheet in book.worksheets]
        finally:
            book.close()
    if data.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        # Some vendor-exported BIFF files have malformed style records.  The
        # calamine reader is more tolerant and keeps the same row-oriented
        # shape as the openpyxl/xlrd branches.
        try:
            from python_calamine import load_workbook

            book = load_workbook(str(path))
            sheets = []
            for sheet_name in book.sheet_names:
                sheet = book.get_sheet_by_name(sheet_name)
                # Calamine raises on a few zero-sized sheets in malformed
                # legacy files (for example an empty "漏项" tab).
                if not getattr(sheet, "height", 0) or not getattr(sheet, "width", 0):
                    sheets.append((sheet_name, []))
                    continue
                sheets.append((sheet_name, [tuple(row) for row in sheet.iter_rows()]))
            return sheets
        except Exception:
            pass

        import xlrd

        try:
            book = xlrd.open_workbook(str(path), on_demand=True)
        except (UnicodeError, xlrd.biffh.XLRDError, OSError) as error:
            raise ValueError("无法读取旧版 Excel 文件，请在 Excel 中另存为 .xlsx 后重新导入。") from error
        try:
            return [(sheet.name, [tuple(cell.value for cell in sheet.row(row_index)) for row_index in range(sheet.nrows)]) for sheet in book.sheets()]
        finally:
            book.release_resources()
    text_sheets = _read_text_table(data)
    if text_sheets:
        return text_sheets
    raise ValueError("无法识别文件格式，请选择完整的 .xlsx 或 .xls 文件。")
