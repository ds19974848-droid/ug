"""工具函数：Excel导入导出、通用辅助"""
import os
from datetime import datetime
from pathlib import Path
from typing import Optional
from sqlalchemy.orm import Session
from .db import get_session, write_audit
from .models import Material, MaterialPrice, MaterialAlias, Region
from .models import Project, ProjectItem, OfficialSource, SubscriptionTask, PriceHistory
from .material_service import build_material_identity_cache, find_or_create_material, infer_price_basis


def export_prices_to_excel(session: Session, filepath: str, region_id: int = None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    query = session.query(
        Material.name.label("material_name"), Material.category, MaterialPrice.spec,
        MaterialPrice.unit, Region.name.label("region_name"), MaterialPrice.period,
        MaterialPrice.price, MaterialPrice.trust_level, MaterialPrice.source_type,
        MaterialPrice.price_basis, MaterialPrice.is_confirmed,
    ).join(MaterialPrice).join(Region)
    if region_id:
        query = query.filter(MaterialPrice.region_id == region_id)
    query = query.filter(MaterialPrice.is_withdrawn.is_(False))
    data = query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "材料价格库"
    headers = ["材料名称", "分类", "规格", "单位", "地区", "期数", "价格", "可信等级", "来源类型", "价格依据", "已确认"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in data:
        ws.append([row.material_name, row.category, row.spec, row.unit,
                    row.region_name, row.period, row.price, row.trust_level,
                    row.source_type or row.trust_level, row.price_basis or "原文发布",
                    "是" if row.is_confirmed else "否"])
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    wb.save(filepath)


def import_prices_from_excel(filepath: str, session: Session) -> dict:
    from openpyxl import load_workbook
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    imported = 0
    skipped = 0
    errors = []
    aliases = {
        "name": ("材料名称", "材料", "名称", "资源名称", "物料名称", "品名", "品种名称"),
        "region": ("地区", "城市", "地市", "项目地区"),
        "period": ("期数", "期间", "月份", "发布期", "发布日期"),
        "price": ("价格", "单价", "信息价", "含税信息价", "除税信息价", "综合价", "综合单价"),
        "spec": ("规格", "型号", "规格型号", "材质规格"),
        "unit": ("单位", "计量单位"),
        "trust": ("可信等级", "数据来源", "来源类型"),
        "basis": ("价格依据", "税价类型", "价格基准"),
    }
    if not headers:
        workbook.close()
        return {"success": False, "imported": 0, "error": "Excel 没有数据"}
    header_index = {str(value).strip(): index for index, value in enumerate(headers) if value is not None}
    mapping = {
        key: next((header_index[label] for label in labels if label in header_index), None)
        for key, labels in aliases.items()
    }
    missing = [key for key in ("name", "region", "period", "price") if mapping[key] is None]
    if missing:
        workbook.close()
        labels = {"name": "材料名称", "region": "地区", "period": "期数", "price": "价格"}
        return {"success": False, "imported": 0, "error": "缺少必需列: " + ", ".join(labels[key] for key in missing)}

    def mapped_value(row, key, default=""):
        column = mapping.get(key)
        cell = row[column] if column is not None and column < len(row) else default
        return default if cell is None else cell

    material_cache = build_material_identity_cache(session)
    price_header = str(headers[mapping["price"]]).strip()
    for idx, row in enumerate(rows, start=2):
        try:
            material_name = str(mapped_value(row, "name")).strip()
            region_name = str(mapped_value(row, "region")).strip()
            period = str(mapped_value(row, "period")).strip()
            price = float(str(mapped_value(row, "price")).replace(",", "").replace("￥", "").replace("¥", ""))
            spec = str(mapped_value(row, "spec")).strip()
            unit = str(mapped_value(row, "unit")).strip()
            raw_trust = str(mapped_value(row, "trust", "官方信息价")).strip()
            trust_level = "official_manual" if "官方" in raw_trust else "market_reference" if "市场" in raw_trust else raw_trust
            source_type = "official" if "官方" in raw_trust else "market_reference" if "市场" in raw_trust else "manual"
            price_basis = infer_price_basis(price_header, mapped_value(row, "basis"), raw_trust)
            region = session.query(Region).filter(Region.name == region_name).first()
            if not region:
                skipped += 1
                continue
            material, _ = find_or_create_material(
                session, material_name, unit=unit, spec=spec, cache=material_cache,
            )
            existing = session.query(MaterialPrice).filter(
                MaterialPrice.material_id == material.id,
                MaterialPrice.region_id == region.id,
                MaterialPrice.period == period,
                MaterialPrice.spec == spec,
            ).first()
            if existing is None:
                existing = next(
                    (
                        pending for pending in session.new
                        if isinstance(pending, MaterialPrice)
                        and pending.material_id == material.id
                        and pending.region_id == region.id
                        and pending.period == period
                        and pending.spec == spec
                        and not pending.is_withdrawn
                    ),
                    None,
                )
            if existing:
                existing.price = price
                existing.unit = unit
                existing.spec = spec
                existing.trust_level = trust_level
                existing.source_type = source_type
                existing.price_basis = price_basis
            else:
                mp = MaterialPrice(
                    material_id=material.id, region_id=region.id,
                    period=period, price=price, unit=unit, spec=spec,
                    trust_level=trust_level, source_type=source_type,
                    price_basis=price_basis, is_confirmed=True,
                )
                session.add(mp)
            session.add(PriceHistory(
                material_id=material.id,
                region_id=region.id,
                period=period,
                price=price,
                unit=unit,
                spec=spec,
                trust_level=trust_level,
                source_type=source_type,
                price_basis=price_basis,
                notes=f"Excel导入: {Path(filepath).name}",
            ))
            imported += 1
        except Exception as e:
            errors.append(f"行 {idx}: {e}")
            skipped += 1
    workbook.close()
    session.commit()
    write_audit(session, "import", "material_price", detail=f"Excel导入: {imported}条")
    session.commit()
    return {"success": True, "imported": imported, "skipped": skipped, "errors": errors}


def export_project_to_excel(project_id: int, filepath: str) -> bool:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    session = get_session()
    try:
        project = session.query(Project).filter(Project.id == project_id).first()
        if not project:
            return False
        items = session.query(ProjectItem).filter(
            ProjectItem.project_id == project_id
        ).order_by(ProjectItem.seq_no).all()
        wb = Workbook()
        ws = wb.active
        ws.title = project.name[:31]
        headers = [
            "序号", "清单编号", "清单名称", "规格", "单位", "工程量",
            "人工费单价", "材料费单价", "机械费单价", "未拆分直接费单价",
            "管理费单价", "利润单价", "综合单价", "合价", "备注",
        ]
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
        for item in items:
            ws.append([
                item.seq_no, item.item_code, item.item_name, item.spec, item.unit, item.quantity,
                item.labor_unit_price, item.material_unit_price, item.machinery_unit_price,
                item.unallocated_unit_price, item.management_unit_price, item.profit_unit_price,
                item.unit_price, item.total_price, item.notes,
            ])
        ws.append([])
        ws.append(["总造价", project.total_amount])
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
        wb.save(filepath)
        write_audit(session, "export", "project", project_id, detail=f"导出报价: {project.name}")
        session.commit()
        return True
    except Exception:
        session.rollback()
        return False
    finally:
        session.close()


def is_official_domain(url: str) -> bool:
    from urllib.parse import urlparse

    host = (urlparse(url).hostname or "").lower().rstrip(".")
    trusted_official_hosts = {
        "202.61.90.35",
        "sceci.net",
        "www.sceci.net",
    }
    return host in trusted_official_hosts or host == "gov.cn" or host.endswith(".gov.cn")


def detect_anomaly(price: float, previous_price: Optional[float] = None) -> tuple:
    from .config import config
    reasons = []
    if price < config.PRICE_MIN_SANE:
        reasons.append(f"价格过低 ({price} < {config.PRICE_MIN_SANE})")
    if price > config.PRICE_MAX_SANE:
        reasons.append(f"价格过高 ({price} > {config.PRICE_MAX_SANE})")
    if previous_price and previous_price > 0:
        change = abs(price - previous_price) / previous_price
        if change > config.PRICE_CHANGE_THRESHOLD:
            direction = "涨" if price > previous_price else "跌"
            reasons.append(f"价格{direction}幅异常 ({change:.1%} > {config.PRICE_CHANGE_THRESHOLD:.0%})")
    return (len(reasons) > 0, "; ".join(reasons))
