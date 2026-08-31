"""Extract construction bill rows from common Glodon Excel exports.

The visible matching table intentionally keeps the application's existing
header.  The extractor also returns source code, quantity, and source tab so
matching and audit information are not lost when those fields are not visible
in the compact table.
"""

from __future__ import annotations

import math
import os
import re
import tempfile
import zipfile
from pathlib import Path

import openpyxl


F1_PREFIX = "F.1"
F2_PREFIX = "F.2"
MEASURE_LABEL = "单价措施项目清单"
CODE_HEADER = "项目编码"
SUBTOTAL = "小计"
FENBU_KEY = "分部分项工程清单与计价表"
TECH_KEY = "施工技术措施项目清单与计价"
CALC_KEY = "综合单价计算表"


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell(row, index: int):
    return row[index] if index is not None and index < len(row) else None


def _is_sequence(value) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.0+)?", _text(value)))


NON_QUOTA_COST_MARKERS = (
    "专业工程暂估价", "计日工", "总承包服务费", "预算包干费", "工程优质费",
    "概算幅度差", "暂列金额", "规费", "税金", "其他项目清单",
    "井点降水工程", "排水降水工程", "施工排水降水", "大型机械设备进出场及安拆",
    "脚手架工程", "模板工程", "垂直运输工程", "安全文明施工费", "夜间施工",
    "二次搬运", "冬雨季施工", "已完工程及设备保护",
)

NON_ITEM_ROW_MARKERS = (
    "分部分项工程", "施工技术措施项目", "单价措施项目", "措施项目清单",
    "单位工程汇总", "单项工程汇总", "工程项目汇总", "费用汇总",
    "合计", "小计", "本页小计", "累计", "总计",
)
_REGION_CODE_PREFIXES = "京津冀晋蒙辽吉黑沪苏浙皖闽赣鲁豫鄂湘粤桂琼川贵云渝藏陕甘青宁新港澳"


def _normal(value) -> str:
    return re.sub(r"\s+", "", _text(value))


def is_non_quota_cost_name(value) -> bool:
    name = _normal(value)
    return bool(name) and any(marker in name for marker in NON_QUOTA_COST_MARKERS)


def has_engineering_quantity(value) -> bool:
    text = _text(value).replace(",", "")
    if not text or not re.fullmatch(r"[+]?\d+(?:\.\d+)?", text):
        return False
    try:
        return float(text) > 0
    except ValueError:
        return False


def looks_like_engineering_code(value) -> bool:
    """Return True for identifiers that cannot be treated as item names."""
    text = _normal(value).upper().replace("（", "(").replace("）", ")")
    if not text:
        return False
    if re.fullmatch(rf"[{_REGION_CODE_PREFIXES}]?\d{{7,20}}", text):
        return True
    if re.fullmatch(r"[A-Z]{1,8}[-_/]?[A-Z0-9]*(?:[-_/][A-Z0-9]+)+", text):
        return True
    if re.fullmatch(r"[A-Z]{0,5}\d(?:[0-9._/-]{5,}\d)", text):
        return True
    return False


def invalid_engineering_name_reason(name, code="") -> str:
    """Reject codes, headings and numeric labels masquerading as names."""
    value = _text(name)
    normalized = _normal(value)
    if not normalized:
        return "工程名称为空"
    if code and _normal(code).upper() == normalized.upper():
        return "工程名称与项目编码完全相同"
    if looks_like_engineering_code(value):
        return "工程名称实际为项目/定额编码"
    if re.fullmatch(r"[（(]?\d+(?:\.\d+)?[）)]?", normalized):
        return "工程名称仅为序号或数字"
    if normalized in {"项目名称", "工程名称", "清单名称", "定额名称", "名称"}:
        return "表头行不能作为工程清单"
    if any(marker in normalized for marker in NON_ITEM_ROW_MARKERS):
        return "汇总、章节或措施标题行不属于可套定额清单"
    if is_non_quota_cost_name(normalized):
        return "费用汇总或措施费用行不属于分部分项清单"
    if not re.search(r"[A-Z\u4e00-\u9fff]", normalized, flags=re.IGNORECASE):
        return "工程名称没有可识别的工程对象"
    return ""


def normalized_boq_feature(name, feature="", code="") -> str:
    """Keep meaningful feature text; otherwise use a valid engineering name."""
    value = _text(feature)
    if (
        not value
        or looks_like_engineering_code(value)
        or (code and _normal(value).upper() == _normal(code).upper())
    ):
        return _text(name)
    return value


def boq_item_rejection_reason(
    *, name, feature="", unit="", quantity=None, code="", require_quantity=True,
) -> str:
    reason = invalid_engineering_name_reason(name, code)
    if reason:
        return reason
    if not _text(unit) or looks_like_engineering_code(unit):
        return "计量单位为空或无效"
    if require_quantity and not has_engineering_quantity(quantity):
        return "工程量为空、为0或不是有效正数"
    # Feature may legitimately be absent in some BOQs. In that case the valid
    # engineering name is the matching description; a code-like feature is
    # ignored instead of being shown as work content.
    effective_feature = normalized_boq_feature(name, feature, code)
    if invalid_engineering_name_reason(effective_feature, code):
        return "项目特征及工作内容没有可识别的工程含义"
    return ""


def filter_valid_boq_items(items: list[dict]) -> tuple[list[dict], dict[str, int]]:
    accepted = []
    rejected: dict[str, int] = {}
    for raw in items:
        item = dict(raw or {})
        reason = boq_item_rejection_reason(
            name=item.get("name"), feature=item.get("feature"),
            unit=item.get("unit"), quantity=item.get("quantity"),
            code=item.get("code"), require_quantity=True,
        )
        if reason:
            rejected[reason] = rejected.get(reason, 0) + 1
            continue
        item["feature"] = normalized_boq_feature(
            item.get("name"), item.get("feature"), item.get("code"),
        )
        accepted.append(item)
    return accepted, rejected


def _append_feature(item: dict | None, value) -> None:
    value = _text(value)
    if not item or not value:
        return
    current = _text(item.get("feature"))
    if value not in current.splitlines():
        item["feature"] = "\n".join(part for part in (current, value) if part)


def _repair_styles(source: str, target: str):
    """Repair the malformed cellStyles block found in some exported files."""
    with zipfile.ZipFile(source) as source_zip, zipfile.ZipFile(
        target, "w", zipfile.ZIP_DEFLATED
    ) as target_zip:
        for info in source_zip.infolist():
            data = source_zip.read(info.filename)
            if info.filename == "xl/styles.xml":
                styles = data.decode("utf-8")
                styles = re.sub(
                    r"<cellStyles\b[^>]*>.*?</cellStyles>",
                    '<cellStyles count="0" />',
                    styles,
                    flags=re.S,
                )
                data = styles.encode("utf-8")
            target_zip.writestr(info, data)


def _load_workbook_safe(path: str):
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True), None
    except Exception:
        file_handle, repaired = tempfile.mkstemp(
            prefix="dashuo-bill-list-", suffix=".xlsx"
        )
        os.close(file_handle)
        _repair_styles(path, repaired)
        try:
            workbook = openpyxl.load_workbook(
                repaired, read_only=True, data_only=True
            )
        except Exception:
            Path(repaired).unlink(missing_ok=True)
            raise
        return workbook, repaired


def _empty_costs():
    return (None, None, None, None, None)


def _parse_f2_cost_maps(workbook, sheet_names):
    maps = []
    for sheet_name in sheet_names:
        costs = {}
        current_code = None
        current_costs = None

        def save_current():
            if current_code:
                costs[current_code] = current_costs or _empty_costs()

        for row in workbook[sheet_name].iter_rows(values_only=True):
            values = [_normal(value) for value in row]
            header_index = next(
                (index for index, value in enumerate(values) if value == CODE_HEADER),
                None,
            )
            if header_index is not None:
                save_current()
                current_code = _text(_cell(row, header_index + 1)).split(")")[-1].strip()
                current_costs = None
                continue
            if current_code and _normal(_cell(row, 0)) == SUBTOTAL:
                current_costs = tuple(_cell(row, index) for index in (10, 12, 13, 14, 15))
        save_current()
        maps.append(costs)
    return maps


def _lookup_f2_cost(maps, sheet_index: int, code: str):
    if sheet_index < len(maps) and code in maps[sheet_index]:
        return maps[sheet_index][code]
    candidates = [costs[code] for costs in maps if code in costs]
    return candidates[0] if len(candidates) == 1 else _empty_costs()


def _f1_columns(workbook, sheet_name):
    """Find F.1 columns from the actual two-row header, not fixed offsets."""
    rows = list(workbook[sheet_name].iter_rows(values_only=True))[:12]
    max_column = max((len(row) for row in rows), default=0)
    # Rows 3-5 are the merged header block. Do not scan data rows, otherwise
    # words inside feature descriptions can be mistaken for cost columns.
    header_rows = rows[2:5]
    labels = {column: " ".join(_normal(_cell(row, column)) for row in header_rows) for column in range(max_column)}
    exact = {}
    for row in header_rows:
        for column, value in enumerate(row):
            normalized = _normal(value)
            if normalized:
                exact.setdefault(normalized, column)
    def find(*tokens):
        for token in tokens:
            normalized = _normal(token)
            if normalized in exact:
                return exact[normalized]
        return next((column for column, text in labels.items() if all(_normal(token) in text for token in tokens)), None)
    return {
        "seq": find("序号"),
        "code": find("项目编码"),
        "name": find("项目名称"),
        "feature": find("项目特征") or find("特征描述"),
        "unit": find("计量", "单位") or find("单位"),
        "quantity": find("工程量"),
        "comprehensive": find("综合单价"),
        "total": find("合价"),
        "main_material": find("主材"),
        "aux_material": find("辅材"),
        "loss": find("损耗"),
        "labor": find("人工"),
        "special_material": find("美缝") or find("涂料"),
        "transport": find("二次搬运") or find("加工费") or find("运费"),
    }


def _extract_format_a(workbook):
    f1_names = [name for name in workbook.sheetnames if name.startswith(F1_PREFIX)]
    f2_names = [name for name in workbook.sheetnames if name.startswith(F2_PREFIX)]
    f2_maps = _parse_f2_cost_maps(workbook, f2_names)
    items = []
    for sheet_index, sheet_name in enumerate(f1_names):
        in_measure = False
        previous_item = None
        columns = _f1_columns(workbook, sheet_name)
        # Older exports and hand-built workbooks can omit header labels.
        columns = {key: (value if value is not None else fallback) for (key, value), fallback in zip(
            columns.items(), (0, 1, 2, 3, 4, 5, 6, 7, 10, 11, 12, 13, 14, 15)
        )}
        for row in workbook[sheet_name].iter_rows(values_only=True):
            # The sheet title itself contains the measures label; only the
            # item-name columns indicate that the measures section has begun.
            row_text = _text(_cell(row, 2)) + _text(_cell(row, 3))
            if MEASURE_LABEL in row_text:
                in_measure = True
            sequence = _cell(row, columns["seq"])
            code = _text(_cell(row, columns["code"]))
            name_col = _text(_cell(row, columns["name"]))
            alternate_name = _text(_cell(row, columns["feature"]))
            name = name_col or alternate_name
            if not _is_sequence(sequence) and not code and not name:
                _append_feature(previous_item, _cell(row, columns["feature"]))
                continue
            if not _is_sequence(sequence) or not code or not name:
                continue
            if is_non_quota_cost_name(name):
                continue
            if not has_engineering_quantity(_cell(row, columns["quantity"])):
                continue
            costs = (
                _empty_costs()
                if in_measure
                else _lookup_f2_cost(f2_maps, sheet_index, code)
            )
            breakdown = {
                "主材": _cell(row, columns["main_material"]),
                "辅材": _cell(row, columns["aux_material"]),
                "损耗": _cell(row, columns["loss"]),
                "人工": _cell(row, columns["labor"]),
                "美缝/勾缝/涂料": _cell(row, columns["special_material"]),
                "二次搬运/成品保护/加工/运费": _cell(row, columns["transport"]),
            }
            breakdown_text = "；".join(f"{key}：{_text(value)}" for key, value in breakdown.items() if value not in (None, ""))
            item = {
                    "seq": _text(sequence) or str(len(items) + 1),
                    "code": code,
                    "name": name,
                    "feature": _text(_cell(row, columns["feature"])),
                    "unit": _text(_cell(row, columns["unit"])),
                    "quantity": _text(_cell(row, columns["quantity"])),
                    "analysis": "报价组成：" + breakdown_text if breakdown_text else "",
                    "comprehensive_price": _cell(row, columns["comprehensive"]),
                    "total_price": _cell(row, columns["total"]),
                    "costs": costs,
                    "breakdown": breakdown,
                    "source_tab": sheet_name,
                }
            if costs == _empty_costs() and breakdown_text:
                material = sum(float(value or 0) for key, value in breakdown.items() if key != "人工" and isinstance(value, (int, float)))
                labor = breakdown.get("人工") if isinstance(breakdown.get("人工"), (int, float)) else None
                item["costs"] = (labor, material or None, 0, None, None)
            items.append(item)
            previous_item = item
    return items


def _calc_sheet_for(list_name: str, calc_names):
    list_prefix = list_name.split("表10", 1)[0]
    is_tech = TECH_KEY in list_name
    candidates = [
        name for name in calc_names if not list_prefix or name.startswith(list_prefix)
    ]
    if not candidates:
        candidates = list(calc_names)
    if len(candidates) == 1:
        return candidates[0]
    for name in candidates:
        name_is_tech = "技术措施" in name
        if name_is_tech == is_tech:
            return name
    return candidates[0] if candidates else None


def _parse_calc_costs(workbook, sheet_name):
    if not sheet_name:
        return {}
    costs = {}
    for row in workbook[sheet_name].iter_rows(values_only=True):
        if _is_sequence(_cell(row, 0)) and _text(_cell(row, 1)):
            costs[_text(_cell(row, 1))] = tuple(
                _cell(row, index) for index in range(5, 10)
            )
    return costs


def _extract_format_b(workbook):
    list_names = [
        name
        for name in workbook.sheetnames
        if FENBU_KEY in name or TECH_KEY in name
    ]
    calc_names = [name for name in workbook.sheetnames if CALC_KEY in name]
    items = []
    for sheet_name in list_names:
        calc_name = _calc_sheet_for(sheet_name, calc_names)
        cost_map = _parse_calc_costs(workbook, calc_name)
        previous_item = None
        for row in workbook[sheet_name].iter_rows(values_only=True):
            sequence = _cell(row, 0)
            code = _text(_cell(row, 1))
            name = _text(_cell(row, 2))
            if not _is_sequence(sequence) and not code and not name:
                _append_feature(previous_item, _cell(row, 3))
                continue
            if not _is_sequence(sequence) or not code or not name:
                continue
            if is_non_quota_cost_name(name):
                continue
            if not has_engineering_quantity(_cell(row, 5)):
                continue
            item = {
                    "seq": _text(sequence) or str(len(items) + 1),
                    "code": code,
                    "name": name,
                    "feature": _text(_cell(row, 3)),
                    "unit": _text(_cell(row, 4)),
                    "quantity": _text(_cell(row, 5)),
                    "analysis": "",
                    "comprehensive_price": _cell(row, 7),
                    "total_price": _cell(row, 6),
                    "costs": cost_map.get(code, _empty_costs()),
                    "source_tab": sheet_name,
                }
            items.append(item)
            previous_item = item
    return items


def extract_specialized_bill_list(path: str):
    """Return parsed rows for a supported workbook, or ``None`` for generic Excel.

    The workbook is opened read-only and closed before this function returns.
    A detected format with zero item rows is returned as an empty result so the
    caller can report a useful format-specific error instead of silently using
    a less accurate generic importer.
    """
    suffix = Path(path).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        return None
    workbook, repaired_path = _load_workbook_safe(path)
    try:
        format_a = any(name.startswith(F1_PREFIX) for name in workbook.sheetnames)
        format_b = any(
            FENBU_KEY in name or TECH_KEY in name for name in workbook.sheetnames
        )
        if format_a:
            items, rejected = filter_valid_boq_items(_extract_format_a(workbook))
            return {
                "format": "广联达 F.1/F.2",
                "items": items,
                "rejected": sum(rejected.values()),
                "rejection_reasons": rejected,
                "source_tabs": [name for name in workbook.sheetnames if name.startswith(F1_PREFIX)],
            }
        if format_b:
            items, rejected = filter_valid_boq_items(_extract_format_b(workbook))
            return {
                "format": "广联达表10.2.2-16/17",
                "items": items,
                "rejected": sum(rejected.values()),
                "rejection_reasons": rejected,
                "source_tabs": [
                    name
                    for name in workbook.sheetnames
                    if FENBU_KEY in name or TECH_KEY in name
                ],
            }
        return None
    finally:
        workbook.close()
        if repaired_path:
            Path(repaired_path).unlink(missing_ok=True)
