"""固定工程清单库的导入、导出与价格匹配。"""
from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from rapidfuzz import fuzz, process
from sqlalchemy import or_
from sqlalchemy.orm import Session

from .db import write_audit
from .models import CostItem
from .text_utils import normalize_text


FIELD_MAP = {
    "数据ID": "source_key",
    "数据来源": "data_source",
    "原始序号": "original_seq",
    "项目名称": "item_name",
    "项目特征": "features",
    "单位": "unit",
    "综合单价": "comprehensive_price",
    "含税价": "tax_inclusive_price",
    "人工费": "labor_cost",
    "材料费": "material_cost",
    "机械费": "machinery_cost",
    "管理费": "management_cost",
    "利润": "profit",
    "完整分类": "full_category",
    "原始类别": "original_category",
    "价格分析": "price_analysis",
    "AI分析": "ai_analysis",
    "备注": "notes",
}
NUMERIC_FIELDS = {
    "original_seq", "comprehensive_price", "tax_inclusive_price", "labor_cost",
    "material_cost", "machinery_cost", "management_cost", "profit",
}


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _number(value, default=0.0):
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def import_cost_items_from_excel(
    filepath: str | Path,
    session: Session,
    *,
    sheet_name: str = "全部数据",
    audit: bool = True,
) -> dict:
    """幂等导入固定清单；相同数据 ID 会更新而不会重复。"""
    from openpyxl import load_workbook
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return {"success": False, "imported": 0, "updated": 0, "skipped": 0, "error": "Excel 没有数据"}
        header_indexes = {
            FIELD_MAP[_text(header)]: index
            for index, header in enumerate(headers)
            if _text(header) in FIELD_MAP
        }
        required = {"item_name", "comprehensive_price"}
        if not required.issubset(header_indexes):
            missing = "、".join(sorted(required - set(header_indexes)))
            return {"success": False, "imported": 0, "updated": 0, "skipped": 0, "error": f"缺少必要字段: {missing}"}

        existing = {
            item.source_key: item
            for item in session.query(CostItem).all()
        }
        imported = 0
        updated = 0
        skipped = 0
        for row_number, row in enumerate(rows, start=2):
            values = {}
            for field, index in header_indexes.items():
                value = row[index] if index < len(row) else None
                if field in NUMERIC_FIELDS:
                    values[field] = int(_number(value)) if field == "original_seq" else _number(value)
                else:
                    values[field] = _text(value)
            if not values.get("item_name"):
                skipped += 1
                continue
            source_key = values.get("source_key") or f"import-{uuid4().hex}"
            values["source_key"] = source_key
            item = existing.get(source_key)
            if item is None:
                item = CostItem(**values)
                session.add(item)
                existing[source_key] = item
                imported += 1
            else:
                for field, value in values.items():
                    setattr(item, field, value)
                updated += 1
            if (imported + updated) % 500 == 0:
                session.flush()
        session.flush()
        if audit:
            write_audit(
                session,
                "import",
                "cost_item",
                detail=f"固定清单导入: 新增 {imported} 条，更新 {updated} 条，跳过 {skipped} 条",
            )
        return {
            "success": True,
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "total": imported + updated,
        }
    finally:
        workbook.close()


def export_cost_items_to_excel(items: list[CostItem], filepath: str | Path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "固定工程清单库"
    headers = list(FIELD_MAP)
    fields = [FIELD_MAP[header] for header in headers]
    worksheet.append(headers)
    fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for item in items:
        worksheet.append([getattr(item, field) for field in fields])
    for column in worksheet.columns:
        letter = column[0].column_letter
        worksheet.column_dimensions[letter].width = 18
    worksheet.column_dimensions["E"].width = 42
    workbook.save(filepath)


def find_cost_item_reference(
    session: Session,
    item_name: str,
    unit: str = "",
    category: str = "",
    *,
    min_score: float = 0.45,
    contains_limit: int = 200,
    candidate_limit: int = 1500,
) -> tuple[CostItem | None, float]:
    """优先精确/包含匹配，最后在候选集中做中文模糊匹配。

    返回 (CostItem or None, score_float)，score 范围 0.0-1.0。
    """
    keyword = _text(item_name)
    if not keyword:
        return None, 0.0
    keyword_norm = normalize_text(keyword)
    unit = _text(unit)
    unit_norm = normalize_text(unit) if unit else ""

    query = session.query(CostItem)
    if category:
        query = query.filter(CostItem.full_category.contains(category))

    # 1) 精确匹配（库中原始文本相等）
    exact = query.filter(CostItem.item_name == keyword).order_by(CostItem.id).first()
    if exact:
        if not unit or exact.unit == unit:
            return exact, 1.0
        # unit 不匹配：仍然返回，但降低置信度（便于人工复核）
        return exact, 0.9

    # 2) 包含匹配（先用 DB 限定候选，然后在 Python 侧用规范化+快速相似度排序）
    contains_q = query.filter(
        or_(CostItem.item_name.contains(keyword), CostItem.features.contains(keyword))
    ).limit(int(contains_limit))
    contains = contains_q.all()
    if contains:
        # 若指定 unit，优先筛选 unit 相同的候选
        if unit:
            same_unit = [item for item in contains if item.unit == unit]
            if same_unit:
                contains = same_unit
        best = None
        best_score = -1.0
        for item in contains:
            score = fuzz.WRatio(keyword_norm, normalize_text(item.item_name))
            if score > best_score:
                best_score = score
                best = item
        if best is not None:
            final_score = float(best_score) / 100.0
            return best, final_score

    # 3) 全库近似匹配（受 candidate_limit 限制）
    candidates = query.order_by(CostItem.id).limit(int(candidate_limit)).all()
    if not candidates:
        return None, 0.0

    # 若指定 unit，则把同单位候选放前面（提高命中概率）
    if unit:
        same_unit = [c for c in candidates if c.unit == unit]
        if same_unit:
            prioritized = same_unit + [c for c in candidates if c.unit != unit]
            candidates = prioritized

    # 构建 id->normalized name 映射供 rapidfuzz 使用
    choices = {candidate.id: normalize_text(candidate.item_name) for candidate in candidates}
    try:
        match = process.extractOne(keyword_norm, choices, scorer=fuzz.WRatio)
    except Exception:
        match = None

    if not match:
        return None, 0.0

    # process.extractOne 返回 (best_value, score, key) when choices is dict
    try:
        _, score, matched_key = match
    except Exception:
        # 兼容不同 rapidfuzz 版本的返回结构
        if isinstance(match, tuple) and len(match) >= 2:
            score = match[1]
            matched_key = match[-1]
        else:
            return None, 0.0

    matched_score = float(score) / 100.0
    # 如果分数太低则视为未命中（但仍把分数返回用于诊断）
    if matched_score < float(min_score):
        return None, matched_score

    matched_item = next((c for c in candidates if c.id == matched_key), None)
    return matched_item, matched_score
