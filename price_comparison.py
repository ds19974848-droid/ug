"""Excel 表1/表2逐行对比与结果导出。"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
from typing import Any

from .excel_compat import read_excel_sheets


DEFAULT_COLUMN_LABELS = ["序号", "定额编号", "单项名称", "单位", "工程量", "综合单价", "合计"]
NUMERIC_LABELS = ("工程量", "综合单价", "合计")


@dataclass
class SourceTable:
    path: str
    file_name: str
    sheet_name: str
    rows: list[list[Any]]
    column_labels: list[str]
    numeric_columns: dict[str, int]

    @property
    def max_row(self) -> int:
        return len(self.rows)

    @property
    def max_column(self) -> int:
        return len(self.column_labels)


@dataclass
class ComparisonRow:
    excel_row: int
    values_a: list[Any]
    values_b: list[Any]
    equal: bool
    diff_fields: list[str]
    numeric_diffs: dict[str, float | None]
    diff_note: str


@dataclass
class ComparisonResult:
    table_a: SourceTable
    table_b: SourceTable
    column_labels: list[str]
    numeric_columns: dict[str, int]
    rows: list[ComparisonRow]
    diff_cells: int

    @property
    def equal_rows(self) -> int:
        return sum(row.equal for row in self.rows)

    @property
    def different_rows(self) -> int:
        return len(self.rows) - self.equal_rows


def display_value(value: Any) -> str:
    """Convert Excel values to stable human-readable text."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _numeric_value(value: Any) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = display_value(value).replace(",", "").replace("￥", "").replace("¥", "").strip()
    if not text or text in {"-", "—", "/"}:
        return None
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def numeric_value(value: Any) -> float | None:
    parsed = _numeric_value(value)
    return float(parsed) if parsed is not None else None


def _same_value(left: Any, right: Any) -> bool:
    left_text = display_value(left)
    right_text = display_value(right)
    if not left_text and not right_text:
        return True
    left_number = _numeric_value(left)
    right_number = _numeric_value(right)
    if left_number is not None and right_number is not None:
        return left_number == right_number
    return left_text == right_text


def _trim_rows(rows: list[list[Any]]) -> list[list[Any]]:
    last_row = 0
    last_column = 0
    for row_index, row in enumerate(rows, start=1):
        non_empty = [index for index, value in enumerate(row, start=1) if display_value(value)]
        if non_empty:
            last_row = row_index
            last_column = max(last_column, max(non_empty))
    if not last_row or not last_column:
        return []
    return [list(row[:last_column]) + [None] * max(0, last_column - len(row)) for row in rows[:last_row]]


def _find_column_labels(rows: list[list[Any]], max_column: int) -> list[str]:
    labels = [f"列{index + 1}" for index in range(max_column)]
    best_row: list[Any] | None = None
    best_score = 0
    tokens = set(DEFAULT_COLUMN_LABELS)
    for row in rows[: min(len(rows), 60)]:
        score = sum(1 for value in row if display_value(value) in tokens)
        if score > best_score:
            best_score = score
            best_row = row
    if best_row is not None and best_score >= 2:
        for index in range(max_column):
            value = display_value(best_row[index] if index < len(best_row) else "")
            if value:
                labels[index] = value
    elif max_column == len(DEFAULT_COLUMN_LABELS):
        labels = DEFAULT_COLUMN_LABELS.copy()
    return labels


def _numeric_columns(labels: list[str]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, label in enumerate(labels):
        for numeric_label in NUMERIC_LABELS:
            if numeric_label in label and numeric_label not in result:
                result[numeric_label] = index
    if len(labels) == 7:
        for label, index in zip(NUMERIC_LABELS, (4, 5, 6)):
            result.setdefault(label, index)
    return result


def load_comparison_table(filepath: str | Path, sheet_name: str = "") -> SourceTable:
    """Read the first worksheet from xlsx or xls and preserve its row positions."""
    path = Path(filepath)
    sheets = read_excel_sheets(path)
    active_sheet_name, source_rows = next(
        ((name, data) for name, data in sheets if not sheet_name or name == sheet_name),
        sheets[0] if sheets else ("", []),
    )
    rows = [list(row) for row in source_rows]

    rows = _trim_rows(rows)
    if not rows:
        raise ValueError("Excel 没有可对比的数据")
    labels = _find_column_labels(rows, max(len(row) for row in rows))
    return SourceTable(
        path=str(path),
        file_name=path.name,
        sheet_name=active_sheet_name,
        rows=rows,
        column_labels=labels,
        numeric_columns=_numeric_columns(labels),
    )


def compare_tables(table_a: SourceTable, table_b: SourceTable) -> ComparisonResult:
    max_rows = max(table_a.max_row, table_b.max_row)
    max_columns = max(table_a.max_column, table_b.max_column)
    labels = list(table_a.column_labels)
    if len(labels) < max_columns:
        labels.extend(table_b.column_labels[len(labels):max_columns])
    if len(labels) < max_columns:
        labels.extend(f"列{index + 1}" for index in range(len(labels), max_columns))

    numeric_columns = _numeric_columns(labels)
    rows: list[ComparisonRow] = []
    diff_cells = 0
    for row_index in range(max_rows):
        values_a = list(table_a.rows[row_index]) if row_index < table_a.max_row else []
        values_b = list(table_b.rows[row_index]) if row_index < table_b.max_row else []
        values_a += [None] * (max_columns - len(values_a))
        values_b += [None] * (max_columns - len(values_b))
        diff_fields = [
            labels[column]
            for column in range(max_columns)
            if not _same_value(values_a[column], values_b[column])
        ]
        diff_cells += len(diff_fields)
        numeric_diffs: dict[str, float | None] = {}
        for label, column in numeric_columns.items():
            left = _numeric_value(values_a[column]) if column < len(values_a) else None
            right = _numeric_value(values_b[column]) if column < len(values_b) else None
            numeric_diffs[label] = float(right - left) if left is not None and right is not None else None
        if not diff_fields:
            note = f"{_column_range(max_columns)} 各字段一致"
        else:
            details = []
            for field in diff_fields:
                column = labels.index(field)
                details.append(f"{field}：A={display_value(values_a[column]) or '空'}，B={display_value(values_b[column]) or '空'}")
            note = "；".join(details)
        rows.append(ComparisonRow(row_index + 1, values_a, values_b, not diff_fields, diff_fields, numeric_diffs, note))
    return ComparisonResult(table_a, table_b, labels, numeric_columns, rows, diff_cells)


def _excel_column_name(column_number: int) -> str:
    result = ""
    while column_number:
        column_number, remainder = divmod(column_number - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _column_range(column_count: int) -> str:
    return f"A:{_excel_column_name(column_count)}"


def _format_number(value: float | None) -> str:
    if value is None:
        return ""
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.6f}".rstrip("0").rstrip(".")


def export_comparison(result: ComparisonResult, filepath: str | Path) -> None:
    """Export the comparison using the requested A/B side-by-side layout."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    raw_columns = len(result.column_labels)
    result_start = 2 + raw_columns * 2
    difference_names = list(NUMERIC_LABELS)
    total_columns = result_start + 1 + len(difference_names) + 1
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "表3-1逐行对比"

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    worksheet.cell(1, 1, "【表3-1】工程施工费预算表｜原始数据并排逐行对比")
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1 + raw_columns)
    worksheet.cell(2, 1, f"原表A：{result.table_a.file_name}")
    worksheet.merge_cells(start_row=2, start_column=2 + raw_columns, end_row=2, end_column=1 + raw_columns * 2)
    worksheet.cell(2, 2 + raw_columns, f"原表B：{result.table_b.file_name}")
    worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=1 + raw_columns)
    worksheet.cell(3, 1, f"来源：{result.table_a.path}")
    worksheet.merge_cells(start_row=3, start_column=2 + raw_columns, end_row=3, end_column=1 + raw_columns * 2)
    worksheet.cell(3, 2 + raw_columns, f"来源：{result.table_b.path}")
    worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_columns)
    worksheet.cell(
        4,
        1,
        f"对比范围：A1:{_excel_column_name(raw_columns)}{len(result.rows)}；"
        f"按相同 Excel 行号逐字段核对；数值差额均为“原表B - 原表A”。 "
        f"逐行结论：共对比 {len(result.rows)} 个 Excel 行，完全一致 {result.equal_rows} 行，"
        f"存在差异 {result.different_rows} 行，差异单元格 {result.diff_cells} 个。",
    )

    source_a_start = 2
    source_b_start = source_a_start + raw_columns
    worksheet.cell(6, 1, "定位")
    worksheet.merge_cells(start_row=6, start_column=source_a_start, end_row=6, end_column=source_a_start + raw_columns - 1)
    worksheet.cell(6, source_a_start, f"原表A 原始数据（{_column_range(raw_columns)}）")
    worksheet.merge_cells(start_row=6, start_column=source_b_start, end_row=6, end_column=source_b_start + raw_columns - 1)
    worksheet.cell(6, source_b_start, f"原表B 原始数据（{_column_range(raw_columns)}）")
    worksheet.merge_cells(start_row=6, start_column=result_start, end_row=6, end_column=result_start + 1)
    worksheet.cell(6, result_start, "逐行对比结果")
    difference_start = result_start + 2
    worksheet.merge_cells(start_row=6, start_column=difference_start, end_row=6, end_column=difference_start + len(difference_names) - 1)
    worksheet.cell(6, difference_start, "数值差额（原表B - 原表A）")
    worksheet.cell(6, total_columns, "差异说明")

    headers = ["Excel行号"]
    headers.extend(f"A_{label}" for label in result.column_labels)
    headers.extend(f"B_{label}" for label in result.column_labels)
    headers.extend(["对比结果", "差异字段"])
    headers.extend(f"{label}差额" for label in difference_names)
    headers.append("逐字段说明")
    for column, value in enumerate(headers, start=1):
        worksheet.cell(7, column, value)

    for row_number, row in enumerate(result.rows, start=8):
        worksheet.cell(row_number, 1, row.excel_row)
        for column, value in enumerate(row.values_a, start=source_a_start):
            worksheet.cell(row_number, column, value)
        for column, value in enumerate(row.values_b, start=source_b_start):
            worksheet.cell(row_number, column, value)
        worksheet.cell(row_number, result_start, "完全一致" if row.equal else "存在差异")
        worksheet.cell(row_number, result_start + 1, "无" if row.equal else "、".join(row.diff_fields))
        for offset, label in enumerate(difference_names):
            worksheet.cell(row_number, difference_start + offset, row.numeric_diffs.get(label))
        worksheet.cell(row_number, total_columns, row.diff_note)

    dark_blue = PatternFill("solid", fgColor="1D4ED8")
    light_blue = PatternFill("solid", fgColor="DBEAFE")
    light_orange = PatternFill("solid", fgColor="FFF7ED")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in worksheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=total_columns):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
    worksheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    worksheet["A1"].fill = dark_blue
    for row in (6, 7):
        for cell in worksheet[row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = dark_blue
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in worksheet[4]:
        cell.fill = light_blue
    for row_number, comparison_row in enumerate(result.rows, start=8):
        for column in range(1, total_columns + 1):
            cell = worksheet.cell(row_number, column)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if not comparison_row.equal:
                cell.fill = light_orange
        for column in range(difference_start, difference_start + len(difference_names)):
            worksheet.cell(row_number, column).number_format = "#,##0.00;[Red]-#,##0.00"

    worksheet.freeze_panes = "A8"
    worksheet.auto_filter.ref = f"A7:{get_column_letter(total_columns)}{max(7, 7 + len(result.rows))}"
    widths = [10] + [16] * (raw_columns * 2) + [14, 24, 14, 14, 14, 42]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.row_dimensions[1].height = 26
    worksheet.row_dimensions[4].height = 42
    worksheet.row_dimensions[6].height = 28
    worksheet.row_dimensions[7].height = 34
    workbook.save(str(filepath))


def export_database_comparison(
    rows: list[dict[str, Any]],
    filepath: str | Path,
    *,
    title: str,
    region: str,
    period: str,
    source_name: str,
) -> None:
    """Export project/imported-table rows against the confirmed local price database."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    headers = [
        "序号", "清单编号", "清单名称", "规格", "单位", "工程量",
        "原清单单价", "数据库最新价", "价格差额", "变化率", "预计造价变化",
        "数据库来源", "匹配度", "状态", "说明",
    ]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "价格库对比"
    total_columns = len(headers)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    worksheet.cell(1, 1, title)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    worksheet.cell(2, 1, f"地区：{region or '未指定'}    期数：{period or '最新'}    对比来源：{source_name}")
    worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_columns)
    matched = sum(row.get("status") == "已匹配" for row in rows)
    estimated = sum(row.get("estimated_change") or 0 for row in rows)
    worksheet.cell(3, 1, f"共 {len(rows)} 行，已匹配 {matched} 行，未匹配 {len(rows) - matched} 行，预计造价变化：{estimated:,.2f} 元")
    for column, header in enumerate(headers, start=1):
        worksheet.cell(5, column, header)
    for row_number, row in enumerate(rows, start=6):
        values = [
            row.get("seq_no", row_number - 5), row.get("item_code", ""), row.get("item_name", ""),
            row.get("spec", ""), row.get("unit", ""), row.get("quantity"), row.get("old_price"),
            row.get("new_price"), row.get("price_diff"), row.get("change_rate"), row.get("estimated_change"),
            row.get("source", ""), row.get("confidence"), row.get("status", ""), row.get("note", ""),
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row_number, column, value)

    dark_blue = PatternFill("solid", fgColor="1D4ED8")
    light_blue = PatternFill("solid", fgColor="DBEAFE")
    light_orange = PatternFill("solid", fgColor="FFF7ED")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in worksheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=total_columns):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    worksheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    worksheet["A1"].fill = dark_blue
    for cell in worksheet[2]:
        cell.fill = light_blue
    for cell in worksheet[3]:
        cell.fill = light_blue
    for cell in worksheet[5]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = dark_blue
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_number, row in enumerate(rows, start=6):
        for column in range(1, total_columns + 1):
            cell = worksheet.cell(row_number, column)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row.get("status") != "已匹配":
                cell.fill = light_orange
        for column in (6, 7, 8, 9, 11):
            worksheet.cell(row_number, column).number_format = "#,##0.00;[Red]-#,##0.00"
        worksheet.cell(row_number, 10).number_format = "0.00%;[Red]-0.00%"
        worksheet.cell(row_number, 13).number_format = "0%"
    worksheet.freeze_panes = "A6"
    worksheet.auto_filter.ref = f"A5:{get_column_letter(total_columns)}{max(5, 5 + len(rows))}"
    widths = [8, 18, 30, 22, 10, 12, 14, 14, 14, 12, 16, 42, 10, 12, 52]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.row_dimensions[1].height = 26
    worksheet.row_dimensions[3].height = 28
    worksheet.row_dimensions[5].height = 36
    workbook.save(str(filepath))
