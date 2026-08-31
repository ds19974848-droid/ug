"""官方信息价批量订阅与待入库处理。"""
from __future__ import annotations

import csv
import base64
import hashlib
import json
import re
import time
import threading
import zipfile
from collections import deque
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Callable, Optional
from urllib.parse import unquote, urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from .api_fetcher import API_HANDLERS, fetch_api_source
from .config import config
from .db import get_session, write_audit
from .models import (
    Material,
    MaterialPrice,
    OfficialSource,
    PriceHistory,
    Region,
    SourceDocument,
    SubscriptionTask,
)
from .material_service import build_material_identity_cache, find_or_create_material, infer_price_basis
from .utils import detect_anomaly, is_official_domain
from .source_display import display_source_name


ATTACHMENT_EXTENSIONS = {".xlsx", ".xls", ".csv", ".pdf", ".zip"}
CONTENT_KEYWORDS = (
    "人工、材料、机械信息价", "人工材料机械信息价", "材料信息价", "信息价",
    "材料价格", "价格信息", "建材价格", "工料机价格",
)
DISCOVERY_KEYWORDS = CONTENT_KEYWORDS + ("工程造价信息", "造价信息", "文件下载")
DOWNLOAD_KEYWORDS = ("信息价附件", "价格表", "材料价格", "下载信息价", "信息价")
PAGINATION_KEYWORDS = ("下一页", "下页", "后一页", "尾页", "末页", "更多")
NON_PRICE_MARKERS = (
    "操作手册", "办事指南", "中标候选人", "处罚决定", "招标文件", "申请表",
    "点击", "上传", "下载", "页面描述", "本页面", "建设单位", "项目负责人",
)
NON_MATERIAL_PRICE_MARKERS = (
    "工程造价指数", "造价指数", "价格指数", "指数调整", "建筑面积",
    "项目总造价", "工程总价", "合价", "总价", "费率", "税率",
    "百分比", "单方造价", "造价指标", "投资估算", "概算金额",
)
CITY_HEADERS = ("城市", "地区", "区域", "地市", "所属城市", "适用地区")
SPECIALTY_KEYWORDS = {
    "建筑与装饰工程": ("建筑", "房建", "土建", "装饰", "装修", "幕墙", "建筑材料"),
    "安装工程": ("安装", "电气", "给排水", "消防", "暖通", "通风"),
    "市政工程": ("市政", "道路", "排水", "管网", "沥青", "城市道路"),
    "园林绿化工程": ("园林", "绿化", "苗木", "景观"),
    "城市轨道交通工程": ("轨道交通", "地铁", "轨道", "车站", "区间"),
    "公路工程": ("公路", "路基", "路面", "交通工程"),
    "桥梁与隧道工程": ("桥梁", "桥涵", "隧道", "预应力", "支座"),
    "水利水电工程": ("水利", "水电", "水工", "堤防", "泵站"),
}
COMMON_UNITS = {
    "t", "kg", "g", "m", "m2", "m3", "㎡", "m²", "立方米", "平方米", "米",
    "个", "只", "套", "台", "件", "块", "根", "樘", "组", "副", "座", "孔",
    "延米", "吨", "千克", "升", "l", "工日", "台班",
}
NAME_HEADERS = ("材料名称", "名称", "材料", "品种名称", "品名", "项目名称")
SPEC_HEADERS = ("规格型号", "规格", "型号", "材质规格")
UNIT_HEADERS = ("单位", "计量单位")
PRICE_HEADERS = (
    "信息价", "除税信息价", "含税信息价", "含税价", "除税价",
    "市场价", "综合价", "综合单价", "单价", "价格",
)
JSON_FIELD_ALIASES = {
    "name": (
        "材料名称", "名称", "品名", "品种名称", "项目名称", "materialname",
        "productname", "cailiaoname", "clmc", "pm", "mc",
    ),
    "spec": (
        "规格型号", "规格", "型号", "材质规格", "specification", "spec",
        "model", "ggxh", "gg", "xh",
    ),
    "unit": ("计量单位", "单位", "unitname", "unit", "jldw", "dw"),
    "city": ("城市", "地区", "区域", "地市", "所属城市", "regionname", "cityname"),
    "price": (
        "除税信息价", "含税信息价", "信息价", "综合单价", "综合价", "市场价",
        "含税价", "除税价", "不含税价", "单价", "价格", "informationprice",
        "taxprice", "unitprice", "price", "xxj", "hsj", "bhsj", "csj", "dj",
    ),
}


def friendly_subscription_error(error: Exception | str) -> str:
    text = str(error)
    lowered = text.lower()
    if "nameresolutionerror" in lowered or "getaddrinfo failed" in lowered or "failed to resolve" in lowered:
        return "官网地址已失效或域名无法解析，请在来源配置中更换为现行官方栏目网址"
    if "timed out" in lowered or "timeout" in lowered:
        return "官网响应超时，请稍后重试"
    if "401" in text:
        return "官网要求登录（HTTP 401），请使用软件浏览器登录后再验证"
    if "403" in text:
        return "官网拒绝自动访问（HTTP 403），需要登录授权或单独适配该网站"
    if "404" in text:
        return "官网页面不存在（HTTP 404），栏目网址可能已经调整"
    if "connection" in lowered and ("refused" in lowered or "failed" in lowered):
        return "暂时无法连接官网，请检查网络或稍后重试"
    return text[:240]


def ensure_working_city_sources(session: Session, city: str) -> list[OfficialSource]:
    """Ensure configured direct-crawl sources exist for a city."""
    from .official_sources_config import (
        CITY_SOURCES,
        get_city_audit_status,
        get_public_market_source_config,
        get_zaojiahome_province,
    )

    city = (city or "").strip()
    audit = get_city_audit_status(city)
    region = session.query(Region).filter(Region.name == city).first()
    province = (region.province if region else "") or get_zaojiahome_province(city)
    all_configs = [
        source
        for source in CITY_SOURCES.get(city, [])
        if source.get("data_status") == "working"
        and source.get("strategy") == "api"
    ]
    official_configs = [source for source in all_configs if source.get("is_official", True)]
    market_configs = [source for source in all_configs if source.get("source_class") == "market_reference"]
    if not market_configs:
        generic_market = get_public_market_source_config(city, province)
        if generic_market:
            market_configs = [generic_market]
    # Official data is step 1. Public market reference is always step 2 when
    # available, so cities without an official API can use the same workflow.
    configs = list(official_configs)
    if market_configs:
        configs.extend(market_configs[:1])
    if not configs:
        return []

    if region is None:
        region = Region(name=city, province=get_zaojiahome_province(city), is_active=True)
        session.add(region)
        session.flush()
    elif not region.province:
        region.province = get_zaojiahome_province(city)
    existing = session.query(OfficialSource).filter(
        OfficialSource.region_id == region.id,
    ).all()
    selected = []
    for config in configs:
        configured_url = config.get("url", "").strip().rstrip("/")
        host = urlparse(config.get("url", "")).netloc.lower()
        source = next(
            (
                item for item in existing
                if item.url.strip().rstrip("/") == configured_url
                or (host and urlparse(item.url).netloc.lower() == host)
            ),
            None,
        )
        if source is None:
            source = OfficialSource(region_id=region.id)
            session.add(source)
            existing.append(source)
        source.name = config.get("name", f"{city}官方来源")
        source.url = config["url"]
        source.source_type = config.get("strategy", "api")
        source.is_official = config.get("is_official", True)
        source.is_active = True
        source.notes = json.dumps({
            "description": config.get("note", ""),
            "source_class": config.get("source_class", "official" if source.is_official else "market_reference"),
            "province": config.get("province", get_zaojiahome_province(city)),
            "catalog_url": config.get("url", ""),
        }, ensure_ascii=False)
        selected.append(source)
    session.commit()
    return selected


_SUBSCRIPTION_RUN_LOCK = threading.Lock()


class SubscriptionEngine:
    def __init__(
        self,
        period: str = "",
        specialty: str = "",
        keywords: Optional[list[str]] = None,
        progress_callback: Optional[Callable[[dict], None]] = None,
    ):
        self.session: Optional[Session] = None
        self.period = self._normalize_period(period)
        self.specialty = specialty.strip()
        self.keywords = [value.strip() for value in (keywords or []) if value.strip()][:30]
        self.progress_callback = progress_callback
        self._price_cache: dict[tuple[int, int, str, str, str], MaterialPrice | None] = {}
        self._batch_total = 0
        self._batch_index = 0
        self._batch_started = 0.0
        self._phase = "准备"
        self._current_stats: dict = {}

    def __enter__(self):
        self.session = get_session()
        self._material_cache = build_material_identity_cache(self.session)
        return self

    def __exit__(self, *args):
        if self.session:
            self.session.close()

    def _begin_source_stats(self):
        self._current_stats = {
            "candidate_count": 0,
            "accepted_count": 0,
            "stored_count": 0,
            "rejected_count": 0,
            "rejection_reasons": {},
        }

    def _record_rejection(self, reason: str):
        stats = self._current_stats
        stats["rejected_count"] = int(stats.get("rejected_count", 0)) + 1
        reasons = stats.setdefault("rejection_reasons", {})
        key = str(reason or "未通过数据校验")[:120]
        reasons[key] = int(reasons.get(key, 0)) + 1

    def _set_phase(self, source: OfficialSource, phase: str, percent: int, message: str, **counts):
        self._phase = phase
        self._current_stats["phase"] = phase
        self._emit_progress(source, percent, message, **counts)

    def _sync_task_stats(self, task: SubscriptionTask):
        stats = self._current_stats or {}
        task.parsed_count = int(stats.get("candidate_count", 0))
        task.stored_count = int(stats.get("stored_count", 0))
        task.rejected_count = int(stats.get("rejected_count", 0))
        task.phase = str(stats.get("phase") or self._phase or "")

    def _stats_snapshot(self) -> dict:
        stats = self._current_stats or {}
        reasons = stats.get("rejection_reasons", {})
        return {
            "parsed_prices": int(stats.get("candidate_count", 0)),
            "accepted_prices": int(stats.get("accepted_count", 0)),
            "stored_prices": int(stats.get("stored_count", 0)),
            "rejected_prices": int(stats.get("rejected_count", 0)),
            "rejection_reasons": dict(sorted(reasons.items(), key=lambda item: item[1], reverse=True)[:8]),
        }

    def _validated_source_config(self, source: OfficialSource):
        """Resolve an official source against the verified city catalog."""
        from .official_sources_config import get_city_audit_status, get_city_source_config

        region = self.session.query(Region).filter(Region.id == source.region_id).first()
        city = region.name if region else ""
        source_config = get_city_source_config(city, source.url)
        audit = get_city_audit_status(city)
        return city, source_config, audit

    @staticmethod
    def _is_discovered_source(source: OfficialSource) -> bool:
        if source.source_type == "ai_discovered":
            return True
        try:
            metadata = json.loads(source.notes or "{}")
        except (TypeError, json.JSONDecodeError):
            return False
        return bool(isinstance(metadata, dict) and metadata.get("ai_discovery_completed"))

    def _save_api_prices(self, task, source, prices, document):
        count = 0
        records = []
        seen_records = set()
        variant_counts = {}
        for price_record in prices:
            self._current_stats["candidate_count"] = int(self._current_stats.get("candidate_count", 0)) + 1
            name = price_record.get("name", "").strip()
            if not name:
                self._record_rejection("材料名称为空")
                continue
            spec = price_record.get("spec", "").strip()
            unit = price_record.get("unit", "").strip()
            record_city = str(
                price_record.get("city")
                or price_record.get("region")
                or price_record.get("city_name")
                or ""
            ).strip()
            if record_city and not self._city_matches(record_city, self._expected_city(source)):
                self._record_rejection("接口记录城市与当前来源城市不一致")
                continue
            category = str(price_record.get("category", "") or "").strip()
            code = str(price_record.get("code", "") or "").strip()
            price = price_record.get("price")
            identity = (name, spec, unit, category, code)
            duplicate_key = (identity, str(price), str(price_record.get("raw_price", "")))
            if duplicate_key in seen_records:
                continue
            seen_records.add(duplicate_key)
            variant_index = variant_counts.get(identity, 0)
            variant_counts[identity] = variant_index + 1
            source_key = "api:" + "|".join((code, category, unit))
            if variant_index:
                source_key += f"|variant:{variant_index}"
            records.append((price_record, source_key[:300]))
        for record_index, (price_record, source_key) in enumerate(records, 1):
            name = price_record.get("name", "").strip()
            spec = price_record.get("spec", "").strip()
            unit = price_record.get("unit", "").strip()
            price = price_record.get("price")
            if price is None or price <= 0:
                self._record_rejection("价格为空或非正数")
                continue
            category = price_record.get("category", "")
            notes = (
                f"API:{display_source_name(source.name, source.url, source.source_type)}|CAT:{category}|CODE:{price_record.get('code', '')}"
                f"|RAW:{price_record.get('raw_price', price)}"
            )
            count += self._save_price_record(
                name,
                spec,
                unit,
                price,
                task,
                source,
                document,
                category=category,
                record_notes=notes,
                source_key=source_key,
                price_basis=price_record.get("price_basis", ""),
                count_candidate=False,
            )
            if record_index % 500 == 0:
                self.session.flush()
        self.session.flush()
        return count

    def run_sources(self, source_ids: list[int]) -> dict:
        if not _SUBSCRIPTION_RUN_LOCK.acquire(blocking=False):
            message = "已有信息价抓取任务正在运行，请等待其完成后再抓取当前网址"
            return {
                "total": len(source_ids), "success": 0, "empty": 0,
                "failed": len(source_ids), "new_prices": 0,
                "errors": [message], "details": [], "busy": True,
            }
        try:
            return self._run_sources_unlocked(source_ids)
        finally:
            _SUBSCRIPTION_RUN_LOCK.release()

    def _run_sources_unlocked(self, source_ids: list[int]) -> dict:
        result = {
            "total": len(source_ids), "success": 0, "empty": 0,
            "failed": 0, "new_prices": 0, "parsed_prices": 0,
            "stored_prices": 0, "rejected_prices": 0,
            "errors": [], "details": [],
        }
        self._batch_total = max(1, len(source_ids))
        self._batch_started = time.monotonic()
        # Group sources by city. Step 1 is the official source; step 2 is the
        # public market reference fallback. Stop after the first source that
        # produces valid records so source classes are never mixed.
        source_rows = []
        for source_id in source_ids:
            source = self.session.query(OfficialSource).filter(OfficialSource.id == source_id).first()
            if source:
                source_rows.append(source)
        by_region: dict[int, list[OfficialSource]] = {}
        for source in source_rows:
            by_region.setdefault(source.region_id, []).append(source)
        city_groups = list(by_region.values())
        self._batch_total = max(1, len(city_groups))
        for group_index, group in enumerate(city_groups):
            self._batch_index = group_index
            ordered = sorted(group, key=lambda source: (not source.is_official, source.id))
            group_result = None
            for candidate_index, source in enumerate(ordered):
                source_result = self.run_single_source(source.id)
                result["details"].append(source_result)
                valid_records = bool(
                    source_result.get("stored_prices")
                    or source_result.get("accepted_prices")
                    or source_result.get("new_prices")
                )
                group_result = source_result
                if valid_records:
                    if candidate_index > 0:
                        source_result["fallback_used"] = True
                        source_result["fallback_reason"] = "第一步官方来源未形成可用价格，已切换第二步公开市场参考价"
                    break
            source_result = group_result or {"success": False, "error": "来源不存在"}
            if source_result.get("success"):
                result["success"] += 1
                count = source_result.get("new_prices", 0)
                result["new_prices"] += count
                result["parsed_prices"] += int(source_result.get("parsed_prices") or 0)
                result["stored_prices"] += int(source_result.get("stored_prices") or count or 0)
                result["rejected_prices"] += int(source_result.get("rejected_prices") or 0)
                if not (source_result.get("stored_prices") or source_result.get("accepted_prices") or count):
                    result["empty"] += 1
                    reason = source_result.get("failure_reason") or source_result.get("message")
                    if reason:
                        result["errors"].append(str(reason)[:300])
            else:
                result["failed"] += 1
                result["errors"].append(source_result.get("error", "未知错误"))
        return result

    def run_all_active(self) -> dict:
        sources = self.session.query(OfficialSource).filter(
            OfficialSource.is_active.is_(True),
        ).order_by(OfficialSource.region_id, OfficialSource.is_official.desc(), OfficialSource.id).all()
        by_city: dict[str, list[tuple[tuple[int, int, int], OfficialSource]]] = {}
        for source in sources:
            city, source_config, _audit = self._validated_source_config(source)
            if not city or not source.url:
                continue
            metadata = {}
            try:
                metadata = json.loads(source.notes or "{}")
                if not isinstance(metadata, dict):
                    metadata = {}
            except (TypeError, json.JSONDecodeError):
                pass
            status = source_config.get("data_status", "") if source_config else metadata.get("validation_status", "")
            source_class = source_config.get("source_class", "") if source_config else metadata.get("source_class", "")
            if status in {"wrong_content", "no_public_source"}:
                continue
            # User-entered URLs are intentionally excluded from unattended
            # batch updates. They can require an active account and are run
            # explicitly from the selected-city flow.
            if source.source_type == "user_url":
                continue
            rank = (
                0 if source.is_official and status == "working" and source.source_type == "api" else
                1 if source.is_official and status == "working" else
                2 if source.is_official else
                3 if source_class == "market_reference" else 4,
                0 if source.source_type == "api" else 1 if source.source_type == "crawl" else 2,
                source.id,
            )
            by_city.setdefault(city, []).append((rank, source))
        source_ids = []
        for items in by_city.values():
            if not items:
                continue
            ordered = [source for _, source in sorted(items, key=lambda item: item[0])]
            primary = ordered[0]
            source_ids.append(primary.id)
            if primary.is_official:
                market_fallback = None
                for candidate in ordered[1:]:
                    try:
                        metadata = json.loads(candidate.notes or "{}")
                    except (TypeError, json.JSONDecodeError):
                        metadata = {}
                    if (
                        not candidate.is_official
                        and isinstance(metadata, dict)
                        and metadata.get("source_class") == "market_reference"
                    ):
                        market_fallback = candidate
                        break
                if market_fallback is not None:
                    source_ids.append(market_fallback.id)
        return self.run_sources(source_ids)

    def run_single_source(self, source_id: int) -> dict:
        source = self.session.query(OfficialSource).filter(OfficialSource.id == source_id).first()
        if not source:
            return {"success": False, "error": "来源不存在"}
        if self._batch_total == 0:
            self._batch_total = 1
            self._batch_index = 0
            self._batch_started = time.monotonic()
        task = None
        self._begin_source_stats()
        api_fallback_reason = ""
        try:
            self._set_phase(source, "source_discovery", 1, "识别城市、来源和期数")
            task = self._create_task(source)
            source_url = source.url or ""
            api_handler_available = any(pattern in source_url for pattern in API_HANDLERS)
            source_config = None
            if source.is_official:
                city, source_config, _audit = self._validated_source_config(source)
            else:
                city = source.region_obj.name if source.region_obj else ""

            # The dedicated adapter is an optimization, not a hard gate. A
            # stale adapter, empty response, or changed endpoint must fall
            # through to the same generic pipeline instead of ending as an
            # unexplained zero-result task.
            if api_handler_available:
                source_class = ""
                try:
                    source_class = json.loads(source.notes or "{}").get("source_class", "")
                except (TypeError, json.JSONDecodeError):
                    source_class = ""
                step_label = "第二步公开市场参考价" if source_class == "market_reference" else "第一步官方专用接口"
                self._set_phase(source, "api_fetch", 8, f"{step_label}：正在调用来源")
                try:
                    api_result = fetch_api_source(
                        source_url,
                        self.session,
                        task_id=task.id,
                        region_id=source.region_id,
                        source_id=source.id,
                        period_filter=self.period,
                        progress_callback=lambda progress, message: self._emit_progress(source, progress, message),
                    )
                    api_fallback_reason = str(api_result.get("error") or "").strip()
                    api_prices = api_result.get("prices") or []
                    if not api_fallback_reason and api_prices:
                        actual_period = self._normalize_period(api_result.get("period", "")) or task.period
                        if self.period and actual_period and actual_period != self.period:
                            api_fallback_reason = f"接口返回期数 {actual_period} 与选择期数 {self.period} 不一致"
                            api_prices = []
                        if api_fallback_reason:
                            task.message = self._append_task_message(task.message, f"专用接口未完成：{api_fallback_reason}；已自动转入通用抓取")
                            api_prices = []
                    if not api_fallback_reason and api_prices:
                        task.period = actual_period
                        document_payload = api_result.get("document") or {}
                        content = document_payload.get("content", b"")
                        if content:
                            document = self._store_document(
                                task,
                                source,
                                document_payload.get("file_name", f"{actual_period}_官方信息价.xls"),
                                document_payload.get("url", source_url),
                                content,
                            )
                            document.period = actual_period
                            count = self._save_api_prices(task, source, api_prices, document)
                            if count or self._current_stats.get("accepted_count", 0):
                                cleanup = self._prune_price_storage({source.region_id})
                                document.is_parsed = True
                                result_message = api_result.get("message", "").strip()
                                document.parse_result = result_message or f"接口候选 {len(api_prices)} 条，入库 {count} 条"
                                task.status = "success"
                                task.result_count = count
                                task.progress = 100
                                task.message = document.parse_result
                                self._sync_task_stats(task)
                                task.finished_at = datetime.utcnow()
                                source.last_check_at = datetime.utcnow()
                                source.last_result = "success" if count else "empty"
                                self.session.commit()
                                self._set_phase(source, "complete", 100, "抓取完成")
                                result = {
                                    "success": True,
                                    "new_prices": count,
                                    "task_id": task.id,
                                    "source_id": source.id,
                                    "region_id": source.region_id,
                                    "city": city,
                                    "source_name": display_source_name(source.name, source.url, source.source_type),
                                    "period": actual_period,
                                    "phase": "store" if count else "validate",
                                    "message": result_message or (
                                        "接口数据有效且已存在，未产生新增记录"
                                        if not count and self._current_stats.get("accepted_count", 0)
                                        else ""
                                    ),
                                    "storage_cleanup": cleanup,
                                }
                                result.update(self._stats_snapshot())
                                result["failure_stage"] = "" if self._current_stats.get("accepted_count", 0) else "validate"
                                result["failure_reason"] = "；".join(result["rejection_reasons"].keys()) if not self._current_stats.get("accepted_count", 0) else ""
                                result["next_action"] = "核对期数、单位或来源文件" if not self._current_stats.get("accepted_count", 0) else ""
                                return result
                        else:
                            api_fallback_reason = "专用接口返回了价格，但没有返回可追溯原始文件"
                    elif not api_fallback_reason:
                        api_fallback_reason = "专用接口返回空数据"
                except Exception as error:
                    api_fallback_reason = friendly_subscription_error(error)
                if api_fallback_reason:
                    task.message = self._append_task_message(task.message, f"专用接口未完成：{api_fallback_reason}；已自动转入通用抓取")

            self._set_phase(source, "document_fetch", 15, "正在访问来源并寻找公开文件或网页数据")
            count = self._run_task(task, source)
            cleanup = self._prune_price_storage({source.region_id})
            task.status = "success"
            task.result_count = count
            task.progress = 100
            self._sync_task_stats(task)
            has_valid_records = bool(self._current_stats.get("accepted_count"))
            task.failure_stage = "" if count or has_valid_records else ("validate" if self._current_stats.get("rejected_count") else "document_fetch")
            task.failure_reason = "" if count or has_valid_records else (
                "；".join(self._stats_snapshot().get("rejection_reasons", {}).keys())
                if self._current_stats.get("rejected_count")
                else "未发现可解析的价格表或附件"
            )
            task.next_action = "" if count or has_valid_records else "打开来源页面核对具体期刊，或完成浏览器登录后重试"
            if cleanup["removed_prices"] or cleanup["removed_regions"]:
                task.message = (
                    f"{task.message or '抓取完成'}；已按保留策略清理 {cleanup['removed_prices']} 条旧价格，"
                    f"当前保留 {cleanup['kept_regions']} 个城市"
                )
            task.finished_at = datetime.utcnow()
            source.last_check_at = datetime.utcnow()
            source.last_result = "success" if count else "empty"
            self.session.commit()
            self._set_phase(source, "complete" if count or has_valid_records else ("validate" if self._current_stats.get("rejected_count") else "failed"), 100, "抓取完成" if count or has_valid_records else "抓取结束，未形成可入库价格")
            result = {
                "success": True, "new_prices": count, "parsed_prices": count, "task_id": task.id,
                "region_id": source.region_id,
                "city": city,
                "period": task.period,
                "source_id": source.id, "source_name": display_source_name(source.name, source.url, source.source_type),
                "message": task.message or "",
                "storage_cleanup": cleanup,
            }
            result.update(self._stats_snapshot())
            result["failure_stage"] = "" if count or has_valid_records else ("validate" if self._current_stats.get("rejected_count") else "document_fetch")
            result["failure_reason"] = (
                "；".join(result["rejection_reasons"].keys())
                if result.get("failure_stage") == "validate"
                else ("未发现可解析的价格表或附件" if not count else "")
            )
            result["next_action"] = (
                "打开来源页面核对具体期刊，或完成浏览器登录后重试"
                if not count and not has_valid_records else ""
            )
            return result
        except Exception as error:
            self.session.rollback()
            self._price_cache.clear()
            if task is not None:
                task = self.session.query(SubscriptionTask).filter(SubscriptionTask.id == task.id).first()
            source = self.session.query(OfficialSource).filter(OfficialSource.id == source_id).first()
            if task is not None:
                task.status = "failed"
                task.message = friendly_subscription_error(error)
                self._sync_task_stats(task)
                task.failure_stage = self._phase or "failed"
                task.failure_reason = task.message
                task.next_action = "检查来源网址、登录状态和目标期数后重试"
                task.finished_at = datetime.utcnow()
            if source is not None:
                source.last_check_at = datetime.utcnow()
                source.last_result = "failed"
            self.session.commit()
            if source is not None:
                self._emit_progress(source, 100, "该来源抓取失败")
            result = {
                "success": False, "error": friendly_subscription_error(error), "parsed_prices": 0,
                "task_id": task.id if task else None,
                "source_id": source_id,
                "source_name": display_source_name(source.name, source.url, source.source_type) if source else "",
            }
            result.update(self._stats_snapshot())
            result["failure_stage"] = self._phase or "failed"
            result["failure_reason"] = result["error"]
            result["next_action"] = "检查来源网址、登录状态和目标期数后重试"
            return result

    def ingest_local_file(self, source_id: int, filepath: str | Path) -> dict:
        source = self.session.query(OfficialSource).filter(OfficialSource.id == source_id).first()
        if not source:
            return {"success": False, "error": "来源不存在"}
        task = self._create_task(source)
        try:
            path = Path(filepath)
            content = path.read_bytes()
            task.progress = 30
            document = self._store_document(task, source, path.name, str(path), content)
            count = self._parse_document(content, path.suffix.lower(), task, source, document)
            cleanup = self._prune_price_storage({source.region_id})
            document.is_parsed = True
            document.parse_result = f"解析 {count} 条新增或更新价格"
            task.status = "success"
            task.progress = 100
            task.result_count = count
            task.finished_at = datetime.utcnow()
            self.session.commit()
            return {"success": True, "new_prices": count, "task_id": task.id, "storage_cleanup": cleanup}
        except Exception as error:
            self.session.rollback()
            task = self.session.query(SubscriptionTask).filter(SubscriptionTask.id == task.id).first()
            if task:
                task.status = "failed"
                task.message = str(error)[:1000]
                task.finished_at = datetime.utcnow()
                self.session.commit()
            return {"success": False, "error": str(error)}

    def _create_task(self, source: OfficialSource) -> SubscriptionTask:
        task = SubscriptionTask(
            source_id=source.id,
            region_id=source.region_id,
            period=self.period or datetime.now().strftime("%Y-%m"),
            status="running",
            progress=1,
            started_at=datetime.utcnow(),
        )
        self.session.add(task)
        self.session.flush()
        self.session.commit()
        return task

    def _run_task(self, task: SubscriptionTask, source: OfficialSource) -> int:
        if source.is_official and not is_official_domain(source.url):
            raise ValueError(f"来源“{source.name}”未通过官方域名校验")
        task.progress = 5
        self.session.commit()
        self._set_phase(source, "access_check", 5, "正在检查来源访问状态")
        # A user-entered URL may require a member session. Do the browser
        # capture first; a preliminary requests.get() used to fail with 401/
        # 403 and abort the task before the saved browser cookies were used.
        browser_attempted = False
        if source.source_type in {"user_url", "ai_discovered", "browser_login"}:
            from .browser_login import has_valid_cookies

            if has_valid_cookies(source.id):
                browser_attempted = True
                self._set_phase(source, "login_check", 18, "正在使用已保存登录状态加载来源")
                browser_count = self._run_browser_capture(task, source)
                if browser_count > 0:
                    return browser_count
        try:
            response = self._request(source.url, source)
        except requests.HTTPError as error:
            status = getattr(error.response, "status_code", None)
            if status in (401, 403):
                raise RuntimeError(
                    "来源要求登录或拒绝自动访问。请使用‘浏览器登录’完成授权后重试；"
                    "普通浏览器的登录状态不会自动共享给软件。"
                ) from error
            raise
        self._set_phase(source, "source_discovery", 10, "来源已连接，正在识别页面、附件和数据接口")
        content_type = response.headers.get("Content-Type", "").lower()
        response_name = self._response_filename(response, source.url)
        extension = self._response_extension(response, source.url, response_name)
        if "json" in content_type or extension == ".json":
            name = response_name
            document = self._store_document(task, source, name, source.url, response.content)
            self._set_phase(source, "validate", 45, "正在解析并校验 JSON 价格记录", documents=1)
            count = self._parse_json(response.content, task, source, document)
            document.is_parsed = True
            document.parse_result = f"JSON接口解析 {count} 条新增或更新价格"
            self._set_phase(source, "store", 90, "JSON 价格记录已写入来源城市和期数", documents=1)
            return count
        if extension in ATTACHMENT_EXTENSIONS or self._is_document_content_type(content_type):
            name = response_name
            document = self._store_document(task, source, name, source.url, response.content)
            self._set_phase(source, "validate", 45, "正在解析并校验价格文件", documents=1)
            count = self._parse_document(response.content, extension or Path(name).suffix.lower(), task, source, document)
            document.is_parsed = True
            document.parse_result = f"解析 {count} 条新增或更新价格"
            self._set_phase(source, "store", 90, "价格文件解析完成，正在保存结果", documents=1)
            return count

        root_html = self._response_text(response)
        if self._looks_like_login_page(root_html):
            self._set_phase(source, "login_check", 14, "来源页面要求登录，正在检查已保存授权")
            from .browser_login import has_valid_cookies

            if not browser_attempted and has_valid_cookies(source.id):
                browser_attempted = True
                total_count = self._run_browser_capture(task, source)
                if total_count > 0:
                    return total_count
            raise RuntimeError(
                "来源页面是登录页或会员页，当前没有可用的数据访问授权。"
                "请点击‘浏览器登录’完成登录后重试。"
            )
        queue = deque([(source.url, root_html, self._page_is_relevant(source.url, root_html))])
        visited_pages = set()
        visited_documents = set()
        total_count = 0
        while queue and len(visited_pages) < config.MAX_SOURCE_PAGES:
            page_url, html, page_is_relevant = queue.popleft()
            if page_url in visited_pages:
                continue
            visited_pages.add(page_url)
            html_bytes = html.encode("utf-8", errors="ignore")
            html_name = self._safe_name(Path(urlparse(page_url).path).name or "index.html", ".html")
            page_document = self._store_document(task, source, html_name, page_url, html_bytes)
            parsed = self._parse_html_tables(html, task, source, page_document)
            total_count += parsed
            page_document.is_parsed = True
            page_document.parse_result = f"网页表格解析 {parsed} 条新增或更新价格"

            attachment_urls, content_urls = self._discover_links(
                html, page_url, source.url, page_is_relevant,
            )
            for document_url in attachment_urls:
                if len(visited_documents) >= config.MAX_SOURCE_DOCUMENTS:
                    break
                if document_url in visited_documents:
                    continue
                visited_documents.add(document_url)
                try:
                    document_response = self._request(document_url, source)
                    name = self._response_filename(document_response, document_url)
                    extension = self._response_extension(document_response, document_url, name)
                    document = self._store_document(
                        task, source, name, document_url, document_response.content,
                    )
                    parsed = self._parse_document(
                        document_response.content, extension, task, source, document,
                    )
                    document.is_parsed = True
                    document.parse_result = f"解析 {parsed} 条新增或更新价格"
                    total_count += parsed
                    progress = min(90, 15 + len(visited_pages) * 3 + len(visited_documents))
                    self._emit_progress(
                        source,
                        progress,
                        "正在扫描官网附件",
                        pages=len(visited_pages),
                        documents=len(visited_documents),
                    )
                except Exception as error:
                    if isinstance(error, SQLAlchemyError):
                        raise
                    task.message = (task.message + f"\n附件失败 {document_url}: {error}").strip()[-4000:]

            for content_url in content_urls:
                if content_url in visited_pages or len(visited_pages) + len(queue) >= config.MAX_SOURCE_PAGES:
                    continue
                try:
                    content_response = self._request(content_url, source)
                    content_type = content_response.headers.get("Content-Type", "").lower()
                    name = self._response_filename(content_response, content_url)
                    extension = self._response_extension(content_response, content_url, name)
                    if extension in ATTACHMENT_EXTENSIONS or self._is_document_content_type(content_type):
                        if content_url not in visited_documents and len(visited_documents) < config.MAX_SOURCE_DOCUMENTS:
                            visited_documents.add(content_url)
                            document = self._store_document(
                                task, source, name, content_url, content_response.content,
                            )
                            parsed = self._parse_document(
                                content_response.content, extension, task, source, document,
                            )
                            document.is_parsed = True
                            document.parse_result = f"解析 {parsed} 条新增或更新价格"
                            total_count += parsed
                            progress = min(90, 15 + len(visited_pages) * 3 + len(visited_documents))
                            self._emit_progress(
                                source,
                                progress,
                                "正在扫描官网附件",
                                pages=len(visited_pages),
                                documents=len(visited_documents),
                            )
                    elif "html" in content_type or not content_type:
                        content_html = self._response_text(content_response)
                        queue.append((
                            content_url,
                            content_html,
                            page_is_relevant or self._page_is_relevant(content_url, content_html),
                        ))
                except Exception as error:
                    if isinstance(error, SQLAlchemyError):
                        raise
                    task.message = (task.message + f"\n栏目失败 {content_url}: {error}").strip()[-4000:]
            task.progress = min(90, 15 + len(visited_pages) * 3 + len(visited_documents))
            self.session.commit()
            self._emit_progress(
                source,
                task.progress,
                "正在扫描官网栏目",
                pages=len(visited_pages),
                documents=len(visited_documents),
            )
        discovered_metadata = {}
        if self._is_discovered_source(source):
            try:
                discovered_metadata = json.loads(source.notes or "{}")
                if not isinstance(discovered_metadata, dict):
                    discovered_metadata = {}
            except (TypeError, json.JSONDecodeError):
                discovered_metadata = {}
        should_try_browser = (
            self._looks_like_dynamic_page(source.url, root_html)
            or bool(discovered_metadata.get("api_candidates"))
            or source.source_type == "ai_discovered"
        )
        if total_count == 0 and not browser_attempted and should_try_browser:
            self._emit_progress(
                source,
                max(task.progress, 35),
                "静态页面无数据，正在启动浏览器加载动态官网",
                pages=len(visited_pages),
                documents=len(visited_documents),
            )
            total_count += self._run_browser_capture(task, source)
        return total_count

    @staticmethod
    def _looks_like_dynamic_page(url: str, html: str) -> bool:
        lowered = (html or "").lower()
        return bool(urlparse(url).fragment) or (
            "<script" in lowered
            and any(marker in lowered for marker in ("static/js/", "chunk-vendors", "webpack", "id=app", 'id="app"'))
            and "<table" not in lowered
        )

    def _run_browser_capture(self, task: SubscriptionTask, source: OfficialSource) -> int:
        from .browser_capture import capture_dynamic_page
        from .browser_login import get_source_cookie_header

        self._set_phase(source, "browser_dynamic", max(20, task.progress or 20), "正在加载动态页面并监听数据接口")
        capture = capture_dynamic_page(
            source.url,
            timeout=max(30, config.SUBSCRIPTION_TIMEOUT),
            cookie_header=get_source_cookie_header(source.id) or "",
        )
        if not capture.get("success"):
            task.message = self._append_task_message(
                task.message,
                f"动态官网浏览器加载失败：{capture.get('error', '未知原因')}",
            )
            return 0

        count = 0
        rendered_html = capture.get("html", "")
        if rendered_html:
            rendered_document = self._store_document(
                task,
                source,
                "browser_rendered.html",
                capture.get("url") or source.url,
                rendered_html.encode("utf-8", errors="ignore"),
            )
            parsed = self._parse_html_tables(rendered_html, task, source, rendered_document)
            rendered_document.is_parsed = True
            rendered_document.parse_result = f"浏览器渲染表格解析 {parsed} 条新增或更新价格"
            count += parsed
            attachments, _ = self._discover_links(
                rendered_html,
                capture.get("url") or source.url,
                source.url,
                True,
            )
            count += self._parse_dynamic_attachments(attachments, task, source)

        parsed_resources = 0
        for index, resource in enumerate(capture.get("resources", []), 1):
            url = str(resource.get("url", ""))
            if not url or not self._resource_url_is_allowed(url, source.url):
                continue
            body = resource.get("body", "")
            if not body:
                continue
            try:
                content = base64.b64decode(body) if resource.get("base64_encoded") else str(body).encode("utf-8")
            except (ValueError, TypeError):
                continue
            extension = self._captured_resource_extension(url, resource.get("mime_type", ""))
            if extension not in ATTACHMENT_EXTENSIONS and extension != ".json":
                continue
            name = self._safe_name(Path(urlparse(url).path).name or f"browser_response_{index}", extension)
            document = self._store_document(task, source, name, url, content)
            try:
                if extension == ".json":
                    parsed = self._parse_json(content, task, source, document)
                else:
                    parsed = self._parse_document(content, extension, task, source, document)
                document.is_parsed = True
                document.parse_result = f"浏览器接口解析 {parsed} 条新增或更新价格"
                count += parsed
                parsed_resources += 1
            except Exception as error:
                document.parse_result = f"浏览器接口解析失败：{friendly_subscription_error(error)}"

        if count == 0:
            task.message = self._append_task_message(
                task.message,
                f"浏览器已加载动态官网并检查 {parsed_resources} 个数据接口，"
                "但未发现同时包含材料名称、单位和价格的可核验记录；可能需要具体期刊页、网站登录授权或专用接口适配",
            )
        self._set_phase(
            source,
            "validate" if count == 0 else "store",
            90,
            "动态页面未发现可入库价格" if count == 0 else "动态页面价格已解析并准备入库",
            pages=1,
            documents=parsed_resources,
        )
        self._emit_progress(
            source,
            90,
            "动态官网解析完成",
            pages=1,
            documents=parsed_resources,
        )
        return count

    def _parse_dynamic_attachments(
        self,
        urls: list[str],
        task: SubscriptionTask,
        source: OfficialSource,
    ) -> int:
        count = 0
        for url in urls[:config.MAX_SOURCE_DOCUMENTS]:
            if not self._resource_url_is_allowed(url, source.url):
                continue
            try:
                response = self._request(url, source)
                name = self._response_filename(response, url)
                extension = self._response_extension(response, url, name)
                document = self._store_document(task, source, name, url, response.content)
                parsed = self._parse_document(response.content, extension, task, source, document)
                document.is_parsed = True
                document.parse_result = f"动态页面附件解析 {parsed} 条新增或更新价格"
                count += parsed
            except Exception as error:
                task.message = self._append_task_message(
                    task.message,
                    f"动态页面附件失败 {url}: {friendly_subscription_error(error)}",
                )
        return count

    @staticmethod
    def _resource_url_is_allowed(url: str, root_url: str) -> bool:
        resource_host = (urlparse(url).hostname or "").lower()
        root_host = (urlparse(root_url).hostname or "").lower()
        return resource_host == root_host or (
            is_official_domain(url) and is_official_domain(root_url)
        )

    @staticmethod
    def _captured_resource_extension(url: str, mime_type: str) -> str:
        extension = Path(urlparse(url).path).suffix.lower()
        if extension in ATTACHMENT_EXTENSIONS:
            return extension
        mime_type = (mime_type or "").lower()
        if "json" in mime_type:
            return ".json"
        if "spreadsheet" in mime_type or "openxmlformats" in mime_type:
            return ".xlsx"
        if "excel" in mime_type or "ms-excel" in mime_type:
            return ".xls"
        if "csv" in mime_type:
            return ".csv"
        if "pdf" in mime_type:
            return ".pdf"
        return ""

    @staticmethod
    def _append_task_message(existing: str, message: str) -> str:
        return "\n".join(value for value in (existing.strip(), message.strip()) if value)[-4000:]

    def _emit_progress(
        self,
        source: OfficialSource,
        source_percent: int,
        stage: str,
        pages: int = 0,
        documents: int = 0,
    ):
        if self.progress_callback is None:
            return
        source_percent = max(0, min(100, int(source_percent)))
        total_sources = max(1, self._batch_total)
        overall = ((self._batch_index + source_percent / 100) / total_sources) * 100
        elapsed = max(0.0, time.monotonic() - self._batch_started)
        eta_seconds = None
        if elapsed >= 1 and overall >= 3:
            eta_seconds = max(0, round(elapsed * (100 - overall) / overall))
        try:
            self.progress_callback({
                "percent": round(overall),
                "source_percent": source_percent,
                "source_index": self._batch_index + 1,
                "source_total": total_sources,
                "source_name": display_source_name(source.name, source.url, source.source_type),
                "phase": self._phase,
                "stage": stage,
                "pages": pages,
                "documents": documents,
                "elapsed_seconds": round(elapsed),
                "eta_seconds": eta_seconds,
            })
        except Exception:
            pass

    def _request(self, url: str, source: OfficialSource):
        from .source_credentials import source_auth_headers
        headers = {
            "User-Agent": config.USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/pdf,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        }
        headers.update(source_auth_headers(source.id))
        response = requests.get(
            url,
            headers=headers,
            timeout=config.SUBSCRIPTION_TIMEOUT,
        )
        response.raise_for_status()
        return response

    @staticmethod
    def _response_text(response) -> str:
        response.encoding = response.apparent_encoding or response.encoding or "utf-8"
        return response.text

    @staticmethod
    def _is_document_content_type(content_type: str) -> bool:
        return any(token in content_type for token in ("pdf", "spreadsheet", "excel", "csv"))

    def _discover_links(
        self,
        html: str,
        page_url: str,
        root_url: str,
        page_is_relevant: bool = False,
    ) -> tuple[list[str], list[str]]:
        root_host = urlparse(root_url).netloc.lower()
        soup = BeautifulSoup(html, "lxml")
        attachments: list[tuple[int, str]] = []
        content_pages: list[tuple[int, str]] = []
        candidates = []
        for anchor in soup.find_all("a", href=True):
            candidates.append((anchor.get("href", ""), anchor.get_text(" ", strip=True), anchor.get("title", "")))
        for element in soup.find_all(attrs={"onclick": True}):
            onclick = element.get("onclick", "")
            matches = re.findall(
                r"(?:window\.open|location(?:\.href|\.assign|\.replace))\s*\(?'?\s*['\"]([^'\"]+)",
                onclick,
                re.I,
            )
            for match in matches:
                candidates.append((match, element.get_text(" ", strip=True), element.get("title", "")))
        for element in soup.find_all(attrs=True):
            for attr_name in ("data-url", "data-href", "data-download", "data-file"):
                value = element.get(attr_name)
                if value:
                    candidates.append((str(value), element.get_text(" ", strip=True), element.get("title", "")))
        for form in soup.find_all("form", action=True):
            candidates.append((form.get("action", ""), form.get_text(" ", strip=True), form.get("title", "")))
        for element in soup.find_all(["iframe", "embed"], src=True):
            candidates.append((element.get("src", ""), element.get("title", ""), ""))
        for script in soup.find_all("script"):
            script_text = script.get_text(" ", strip=True)
            for match in re.findall(
                r"['\"]((?:https?://|/)[^'\"]+(?:download|file|attach|excel|xlsx|xls|pdf|price|info)[^'\"]*)['\"]",
                script_text,
                re.I,
            ):
                candidates.append((match, "脚本发现的下载或数据接口", ""))
        for href, anchor_text, title in candidates:
            href = href.strip()
            if not href or href.startswith(("javascript:", "mailto:", "#")):
                continue
            full_url = urljoin(page_url, href)
            candidate_host = urlparse(full_url).netloc.lower()
            if candidate_host != root_host and not (
                is_official_domain(full_url) and is_official_domain(root_url)
            ):
                continue
            extension = Path(urlparse(full_url).path).suffix.lower()
            blob = f"{anchor_text} {title} {unquote(full_url)}".lower()
            score = self._link_score(blob)
            download_hint = any(keyword in blob for keyword in DOWNLOAD_KEYWORDS)
            pagination = page_is_relevant and (
                any(keyword in blob for keyword in PAGINATION_KEYWORDS)
                or bool(re.search(r"(?:page|pageno|index)[_=/.-]?\d+", blob, re.I))
            )
            if (extension in ATTACHMENT_EXTENSIONS or download_hint) and (score > 0 or download_hint):
                attachments.append((score + 100, full_url))
            elif score > 0 or pagination:
                content_pages.append((score, full_url))
        return self._ranked_unique(attachments), self._ranked_unique(content_pages)

    def _page_is_relevant(self, url: str, html: str) -> bool:
        soup = BeautifulSoup(html, "lxml")
        title = soup.title.get_text(" ", strip=True) if soup.title else ""
        sample = soup.get_text(" ", strip=True)[:12000]
        return self._link_score(f"{unquote(url)} {title} {sample}".lower()) > 0

    @staticmethod
    def _looks_like_login_page(html: str) -> bool:
        soup = BeautifulSoup(html or "", "lxml")
        title = soup.title.get_text(" ", strip=True) if soup.title else ""
        text = soup.get_text(" ", strip=True)[:16000]
        compact = re.sub(r"\s+", "", f"{title} {text}").lower()
        if any(marker in compact for marker in ("请输入验证码", "会员登录", "请先登录", "登录后查看", "账号登录", "密码登录")):
            return True
        password_inputs = soup.select("input[type='password'], input[name*='password'], input[name*='pwd']")
        return bool(password_inputs) and not any(keyword.replace(" ", "") in compact for keyword in CONTENT_KEYWORDS)

    def _link_score(self, text: str) -> int:
        score = sum(12 for keyword in CONTENT_KEYWORDS if keyword.lower() in text)
        score += sum(3 for keyword in DISCOVERY_KEYWORDS if keyword.lower() in text)
        if self.specialty and self.specialty not in {"全部", "全部专业"}:
            terms = SPECIALTY_KEYWORDS.get(self.specialty, (self.specialty,))
            score += sum(10 for term in terms if term.lower() in text)
        score += sum(4 for keyword in self.keywords if keyword.lower() in text)
        period = self._infer_period(text)
        if period:
            score += 20
            if self.period and period == self.period:
                score += 40
            score += max(0, int(period.replace("-", "")) - 200000) // 100
        return score

    @staticmethod
    def _ranked_unique(items: list[tuple[int, str]]) -> list[str]:
        seen = set()
        result = []
        for _, url in sorted(items, key=lambda item: item[0], reverse=True):
            if url not in seen:
                seen.add(url)
                result.append(url)
        return result

    def _response_filename(self, response, url: str) -> str:
        disposition = response.headers.get("Content-Disposition", "")
        match = re.search(r"filename\*?=(?:UTF-8''|\")?([^\";]+)", disposition, re.I)
        if match:
            return self._safe_name(unquote(match.group(1).strip()), "")
        name = unquote(Path(urlparse(url).path).name) or "source_document"
        content_type = response.headers.get("Content-Type", "").lower()
        fallback = ".pdf" if "pdf" in content_type else ".xlsx" if ("excel" in content_type or "spreadsheet" in content_type) else ""
        return self._safe_name(name, fallback)

    @staticmethod
    def _response_extension(response, url: str, name: str = "") -> str:
        """Infer download type even when the URL is an extensionless endpoint."""
        candidates = (Path(name).suffix, Path(urlparse(url).path).suffix)
        for candidate in candidates:
            if candidate and candidate.lower() in ATTACHMENT_EXTENSIONS:
                return candidate.lower()
        content_type = (response.headers.get("Content-Type", "") or "").lower()
        if "json" in content_type:
            return ".json"
        if "spreadsheet" in content_type or "openxmlformats" in content_type:
            return ".xlsx"
        if "excel" in content_type or "ms-excel" in content_type:
            return ".xls"
        if "csv" in content_type:
            return ".csv"
        if "pdf" in content_type:
            return ".pdf"
        body = response.content[:8]
        if body.startswith(b"PK"):
            return ".xlsx"
        if body.startswith(b"%PDF"):
            return ".pdf"
        if body.startswith(b"\xd0\xcf\x11\xe0"):
            return ".xls"
        return ""

    @staticmethod
    def _safe_name(name: str, fallback_extension: str) -> str:
        name = re.sub(r"[<>:\"/\\|?*\x00-\x1f]", "_", name).strip(" .") or "source_document"
        if fallback_extension and not Path(name).suffix:
            name += fallback_extension
        return name[:180]

    def _store_document(
        self,
        task: SubscriptionTask,
        source: OfficialSource,
        name: str,
        url: str,
        content: bytes,
    ) -> SourceDocument:
        digest = hashlib.sha256(content).hexdigest()[:12]
        folder = config.SOURCE_DIR / str(source.id) / task.period
        folder.mkdir(parents=True, exist_ok=True)
        safe_name = self._safe_name(name, "")
        file_path = folder / f"{digest}_{safe_name}"
        if not file_path.exists():
            file_path.write_bytes(content)
        document = SourceDocument(
            task_id=task.id,
            source_id=source.id,
            file_name=safe_name,
            file_path=str(file_path),
            file_type=Path(safe_name).suffix.lower().lstrip(".") or "html",
            file_size=len(content),
            period=self._infer_period(f"{name} {url}") or task.period,
            url=url,
        )
        self.session.add(document)
        self.session.flush()
        return document

    def _parse_document(
        self,
        content: bytes,
        extension: str,
        task: SubscriptionTask,
        source: OfficialSource,
        document: SourceDocument,
    ) -> int:
        extension = extension.lower()
        if extension not in ATTACHMENT_EXTENSIONS and extension != ".json":
            if content.startswith(b"PK"):
                extension = ".xlsx"
            elif content.startswith(b"%PDF"):
                extension = ".pdf"
            elif content.startswith(b"\xd0\xcf\x11\xe0"):
                extension = ".xls"
        inferred_period = self._infer_period(f"{document.file_name} {document.url}")
        if inferred_period:
            document.period = inferred_period
        if extension == ".xlsx":
            return self._parse_excel(content, task, source, document)
        if extension == ".xls":
            return self._parse_xls(content, task, source, document)
        if extension == ".csv":
            return self._parse_csv(content, task, source, document)
        if extension == ".pdf":
            return self._parse_pdf(content, task, source, document)
        if extension == ".zip":
            return self._parse_zip(content, task, source, document)
        text = content.decode("utf-8", errors="ignore")
        document.period = self._infer_period(text[:20000]) or document.period
        return self._parse_html_tables(text, task, source, document)

    def _parse_zip(
        self,
        content: bytes,
        task: SubscriptionTask,
        source: OfficialSource,
        document: SourceDocument,
    ) -> int:
        count = 0
        total_size = 0
        with zipfile.ZipFile(BytesIO(content)) as archive:
            members = [item for item in archive.infolist() if not item.is_dir()][:200]
            for member in members:
                extension = Path(member.filename).suffix.lower()
                if extension not in ATTACHMENT_EXTENSIONS - {".zip"}:
                    continue
                if member.file_size > 50 * 1024 * 1024:
                    continue
                total_size += member.file_size
                if total_size > 500 * 1024 * 1024:
                    break
                original_period = document.period
                document.period = self._infer_period(member.filename) or original_period
                try:
                    count += self._parse_document(
                        archive.read(member), extension, task, source, document,
                    )
                finally:
                    document.period = original_period
        return count

    def _parse_html_tables(
        self,
        html: str,
        task: SubscriptionTask,
        source: OfficialSource,
        document: SourceDocument,
    ) -> int:
        soup = BeautifulSoup(html, "lxml")
        count = 0
        for table in soup.find_all("table"):
            rows = [
                [cell.get_text(" ", strip=True) for cell in row.find_all(["td", "th"])]
                for row in table.find_all("tr")
            ]
            count += self._parse_grid(rows, task, source, document)
        return count

    def _parse_excel(
        self,
        content: bytes,
        task: SubscriptionTask,
        source: OfficialSource,
        document: SourceDocument,
    ) -> int:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as error:
            raise RuntimeError(f"Excel 解析失败（目前支持标准 xlsx 文件）: {error}") from error
        count = 0
        original_period = document.period
        try:
            for worksheet in workbook.worksheets:
                rows = [["" if value is None else str(value).strip() for value in row] for row in worksheet.iter_rows(values_only=True)]
                period_text = " ".join(" ".join(row) for row in rows[:20])
                document.period = self._infer_period(period_text) or original_period
                try:
                    count += self._parse_grid(rows, task, source, document)
                finally:
                    document.period = original_period
        finally:
            workbook.close()
        return count

    def _parse_xls(
        self,
        content: bytes,
        task: SubscriptionTask,
        source: OfficialSource,
        document: SourceDocument,
    ) -> int:
        try:
            import xlrd

            workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        except Exception as error:
            raise RuntimeError(f"XLS 解析失败: {error}") from error
        count = 0
        original_period = document.period
        try:
            for sheet in workbook.sheets():
                rows = [
                    ["" if value is None else str(value).strip() for value in sheet.row_values(row_index)]
                    for row_index in range(sheet.nrows)
                ]
                period_text = " ".join(" ".join(row) for row in rows[:20])
                document.period = self._infer_period(period_text) or original_period
                try:
                    count += self._parse_grid(rows, task, source, document)
                finally:
                    document.period = original_period
        finally:
            workbook.release_resources()
        return count

    def _parse_csv(
        self,
        content: bytes,
        task: SubscriptionTask,
        source: OfficialSource,
        document: SourceDocument,
    ) -> int:
        text = content.decode("utf-8-sig", errors="replace")
        return self._parse_grid(list(csv.reader(StringIO(text))), task, source, document)

    def _parse_json(
        self,
        content: bytes,
        task: SubscriptionTask,
        source: OfficialSource,
        document: SourceDocument,
    ) -> int:
        text = content.decode("utf-8-sig", errors="replace")
        payload = json.loads(text)
        if isinstance(payload, str):
            payload = json.loads(payload)
        document.period = self._infer_period(text[:100000]) or document.period
        count = 0
        for records in self._json_record_lists(payload):
            for raw_record in records:
                record = self._json_price_record(raw_record)
                if record:
                    name, spec, unit, price, record_city = record
                    if record_city and not self._city_matches(record_city, self._expected_city(source)):
                        self._record_rejection("接口记录城市与当前来源城市不一致")
                        continue
                    count += self._save_price_record(name, spec, unit, price, task, source, document)
        return count

    @classmethod
    def _json_record_lists(cls, value) -> list[list[dict]]:
        result = []
        stack = [value]
        seen_ids = set()
        while stack:
            current = stack.pop()
            current_id = id(current)
            if current_id in seen_ids:
                continue
            seen_ids.add(current_id)
            if isinstance(current, list):
                dictionaries = [item for item in current if isinstance(item, dict)]
                if dictionaries:
                    result.append(dictionaries)
                stack.extend(current)
            elif isinstance(current, dict):
                stack.extend(current.values())
        return result

    @classmethod
    def _json_price_record(cls, record: dict):
        flattened = {}
        stack = [("", record)]
        while stack:
            prefix, value = stack.pop()
            for key, item in value.items():
                compound_key = f"{prefix}{key}" if not prefix else f"{prefix}.{key}"
                if isinstance(item, dict):
                    stack.append((compound_key, item))
                elif not isinstance(item, (list, tuple)):
                    flattened[compound_key] = item

        selected = {}
        for field, aliases in JSON_FIELD_ALIASES.items():
            for key, value in flattened.items():
                normalized = cls._normalize_json_key(key)
                if any(normalized == alias or normalized.endswith(alias) for alias in aliases):
                    if value not in (None, ""):
                        selected[field] = value
                        break
        if not {"name", "unit", "price"}.issubset(selected):
            return None
        price = cls._to_price(selected["price"])
        name = str(selected["name"]).strip()
        unit = str(selected["unit"]).strip()
        if price is None or not name or not unit:
            return None
        return (
            name,
            str(selected.get("spec", "")).strip(),
            unit,
            price,
            str(selected.get("city", "")).strip(),
        )

    @staticmethod
    def _normalize_json_key(value: str) -> str:
        return re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", value.lower())

    def _parse_pdf(
        self,
        content: bytes,
        task: SubscriptionTask,
        source: OfficialSource,
        document: SourceDocument,
    ) -> int:
        try:
            import fitz

            pdf = fitz.open(stream=content, filetype="pdf")
            table_count = 0
            lines = []
            for page in pdf:
                try:
                    finder = page.find_tables()
                    for table in finder.tables:
                        rows = [["" if cell is None else str(cell).strip() for cell in row] for row in table.extract()]
                        table_count += self._parse_grid(rows, task, source, document)
                except Exception:
                    pass
                lines.extend(page.get_text().splitlines())
            pdf.close()
        except Exception as error:
            raise RuntimeError(f"PDF 解析失败: {error}") from error
        full_text = " ".join(lines)
        document.period = self._infer_period(" ".join(lines[:200])) or document.period
        if table_count:
            return table_count
        if not self._has_price_context(full_text[:30000]):
            return 0
        count = 0
        for line in lines:
            line = line.strip()
            if not line or len(line) < 5:
                continue
            unit_match = re.search(
                r"(?<![A-Za-z0-9/])(m2|m3|m²|m³|㎡|t|kg|吨|千克|米|个|只|套|台|件|块|根|樘|组|座|工日|台班)(?![A-Za-z0-9/])",
                line,
                re.I,
            )
            if unit_match is None:
                continue
            numbers = re.findall(r"(?<![A-Za-z])\d+(?:,\d{3})*(?:\.\d+)?", line)
            if not numbers:
                continue
            price = self._to_price(numbers[-1])
            name = line[:unit_match.start()].strip(" :-")
            if name and price is not None:
                count += self._save_price_record(
                    name, "", unit_match.group(0), price, task, source, document,
                )
        return count

    def _parse_grid(
        self,
        rows: list[list[str]],
        task: SubscriptionTask,
        source: OfficialSource,
        document: SourceDocument,
    ) -> int:
        if not rows:
            return 0
        grid_text = " ".join(" ".join(str(cell) for cell in row) for row in rows[:80])
        if any(marker in grid_text for marker in NON_MATERIAL_PRICE_MARKERS):
            # Keep genuine material rows in mixed tables, but reject tables
            # whose visible context is clearly an index/summary table.
            material_signal = any(
                marker in grid_text for marker in ("材料名称", "材料价格", "建材", "人工", "机械", "工料机")
            )
            if not material_signal:
                return 0
        header_index = None
        mapping = None
        for index, row in enumerate(rows[:30]):
            candidate = self._header_mapping(row)
            if candidate and {"name", "unit", "price"}.issubset(candidate):
                header_index = index
                mapping = candidate
                break
        count = 0
        if mapping is not None:
            for row in rows[header_index + 1:]:
                record = self._mapped_record(row, mapping, self._expected_city(source))
                if record:
                    count += self._save_price_record(*record, task, source, document)
            return count
        if not self._has_price_context(grid_text):
            return 0
        for row in rows:
            cells = [str(value).strip() for value in row if str(value).strip()]
            if len(cells) < 4:
                continue
            unit_index = next((index for index, value in enumerate(cells[1:5], 1) if value.lower() in COMMON_UNITS), None)
            if unit_index is None:
                continue
            price = next((self._to_price(value) for value in cells[unit_index + 1:] if self._to_price(value) is not None), None)
            if price is not None:
                spec = cells[1] if unit_index > 1 else ""
                count += self._save_price_record(cells[0], spec, cells[unit_index], price, task, source, document)
        return count

    @staticmethod
    def _header_mapping(row: list[str]) -> dict[str, int]:
        mapping = {}
        for index, value in enumerate(row):
            header = str(value).replace("\n", "").replace(" ", "").strip()
            if not header:
                continue
            if "name" not in mapping and any(token in header for token in NAME_HEADERS):
                mapping["name"] = index
            if "spec" not in mapping and any(token in header for token in SPEC_HEADERS):
                mapping["spec"] = index
            if "unit" not in mapping and any(token in header for token in UNIT_HEADERS):
                mapping["unit"] = index
            if "price" not in mapping and any(token in header for token in PRICE_HEADERS):
                mapping["price"] = index
            if "city" not in mapping and any(token in header for token in CITY_HEADERS):
                mapping["city"] = index
        return mapping

    @staticmethod
    def _city_matches(value: str, expected_city: str) -> bool:
        value = re.sub(r"\s+", "", str(value or ""))
        expected_city = re.sub(r"\s+", "", str(expected_city or ""))
        if not value or not expected_city:
            return True
        aliases = {expected_city, expected_city.removesuffix("市")}
        return any(alias and (alias in value or value in alias) for alias in aliases)

    def _expected_city(self, source: OfficialSource) -> str:
        region = self.session.query(Region).filter(Region.id == source.region_id).first()
        return region.name if region else ""

    def _mapped_record(self, row: list[str], mapping: dict[str, int], expected_city: str = ""):
        def value(key):
            index = mapping.get(key)
            return str(row[index]).strip() if index is not None and index < len(row) and row[index] is not None else ""

        name = value("name")
        price = self._to_price(value("price"))
        unit = value("unit")
        city_value = value("city")
        if city_value and not self._city_matches(city_value, expected_city):
            self._record_rejection("表格行城市与当前来源城市不一致")
            return None
        if not name or not unit or price is None or name in NAME_HEADERS:
            return None
        return name, value("spec"), unit, price

    @staticmethod
    def _has_price_context(text: str) -> bool:
        compact = re.sub(r"\s+", "", text or "")
        return any(keyword.replace(" ", "") in compact for keyword in CONTENT_KEYWORDS)

    @staticmethod
    def _is_plausible_price_record(name: str, spec: str, unit: str, price: float) -> bool:
        combined = f"{name} {spec}".strip()
        if not (2 <= len(name) <= 160) or not unit or len(unit) > 20 or price <= 0:
            return False
        normalized_unit = re.sub(r"\s+", "", unit).lower()
        if normalized_unit in {"项", "式", "%", "费率"} or "元/" in normalized_unit:
            return False
        if re.fullmatch(r"[\d\W_]+", name):
            return False
        if any(marker in combined for marker in NON_PRICE_MARKERS):
            return False
        if any(marker in combined for marker in NON_MATERIAL_PRICE_MARKERS):
            return False
        if "《" in combined or "标准》" in combined:
            return False
        if re.search(r"(?:GB|GB/T|JGJ|DGJ|DB\d|T/)\s*[-/\d]", combined, re.I):
            return False
        if 1900 <= price <= 2100 and re.search(r"标准|规范|规程|图集|GB|JGJ|DGJ", combined, re.I):
            return False
        return True

    @staticmethod
    def _to_price(value) -> float | None:
        text = str(value).replace(",", "").replace("￥", "").replace("¥", "").strip()
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return None
        try:
            return float(match.group())
        except ValueError:
            return None

    def _save_price_record(
        self,
        name: str,
        spec: str,
        unit: str,
        price: float,
        task: SubscriptionTask,
        source: OfficialSource,
        document: SourceDocument,
        category: str = "",
        record_notes: str = "",
        source_key: str = "",
        price_basis: str = "",
        count_candidate: bool = True,
    ) -> int:
        if count_candidate:
            self._current_stats["candidate_count"] = int(self._current_stats.get("candidate_count", 0)) + 1
        name = re.sub(r"\s+", " ", name).strip()
        spec = re.sub(r"\s+", " ", spec).strip()
        unit = unit.strip()
        if not self._is_plausible_price_record(name, spec, unit, price):
            self._record_rejection("名称、单位或价格未通过价格记录校验")
            return 0
        region = self.session.query(Region).filter(Region.id == source.region_id).first()
        if not region:
            self._record_rejection("来源未绑定城市")
            return 0
        record_period = document.period or task.period
        target_period = self._normalize_period(self.period)
        normalized_record_period = self._normalize_period(record_period)
        if target_period and normalized_record_period and target_period != normalized_record_period:
            self._record_rejection(f"期数不一致（目标 {target_period}，记录 {normalized_record_period}）")
            return 0
        record_period = normalized_record_period or record_period
        self._current_stats["accepted_count"] = int(self._current_stats.get("accepted_count", 0)) + 1
        material, _ = find_or_create_material(
            self.session,
            name,
            unit=unit,
            spec=spec,
            category=category,
            cache=self._material_cache,
        )
        is_public_market_reference = (
            not source.is_official
            and "wlg.zaojiahome.com" in (source.url or "").lower()
        )
        price_source_type = (
            "official"
            if source.is_official
            else (
                "market_reference"
                if is_public_market_reference or source.source_type != "user_url"
                else "user_url"
            )
        )
        source_key = (source_key or "")[:300]
        price_key = (material.id, region.id, record_period, spec, source_key)
        if price_key in self._price_cache:
            existing = self._price_cache[price_key]
        else:
            existing = self.session.query(MaterialPrice).filter(
                MaterialPrice.material_id == material.id,
                MaterialPrice.region_id == region.id,
                MaterialPrice.period == record_period,
                MaterialPrice.spec == spec,
                MaterialPrice.source_key == source_key,
            ).first()
            # Rows written before source_key was introduced can be upgraded in place
            # for the first matching API row instead of being duplicated.
            if existing is None and source_key:
                existing = self.session.query(MaterialPrice).filter(
                    MaterialPrice.material_id == material.id,
                    MaterialPrice.region_id == region.id,
                    MaterialPrice.period == record_period,
                    MaterialPrice.spec == spec,
                    MaterialPrice.source_key == "",
                ).first()
            self._price_cache[price_key] = existing
        if existing:
            if abs(existing.price - price) < 0.000001 and existing.unit == unit:
                return 0
            previous = self.session.query(MaterialPrice).filter(
                MaterialPrice.material_id == material.id,
                MaterialPrice.region_id == region.id,
                MaterialPrice.spec == spec,
                MaterialPrice.id != existing.id,
            ).order_by(MaterialPrice.period.desc()).first()
            is_anomaly, _ = detect_anomaly(price, previous.price if previous else None)
            if is_anomaly:
                self._record_rejection("价格相对历史值异常")
                return 0
            existing.price = price
            existing.unit = unit
            existing.source_key = source_key
            existing.source_doc_id = document.id
            existing.is_confirmed = True
            existing.is_anomaly = False
            existing.anomaly_reason = ""
            existing.source_type = price_source_type
            existing.price_basis = infer_price_basis(price_basis, source.name, source.notes, record_notes)
            existing.is_withdrawn = False
            if record_notes:
                existing.notes = record_notes
            existing.updated_at = datetime.utcnow()
            action = "update"
        else:
            previous = self.session.query(MaterialPrice).filter(
                MaterialPrice.material_id == material.id,
                MaterialPrice.region_id == region.id,
                MaterialPrice.spec == spec,
            ).order_by(MaterialPrice.period.desc()).first()
            is_anomaly, anomaly_reason = detect_anomaly(price, previous.price if previous else None)
            if is_anomaly:
                self._record_rejection("价格相对历史值异常")
                return 0
            existing = MaterialPrice(
                material_id=material.id,
                region_id=region.id,
                period=record_period,
                price=price,
                unit=unit,
                spec=spec,
                source_key=source_key,
                trust_level="official" if source.is_official else "market_reference",
                source_type=price_source_type,
                price_basis=infer_price_basis(price_basis, source.name, source.notes, record_notes),
                source_doc_id=document.id,
                is_confirmed=True,
                is_anomaly=False,
                anomaly_reason="",
                notes=record_notes,
            )
            self.session.add(existing)
            self._price_cache[price_key] = existing
            action = "create"
        self.session.add(PriceHistory(
            material_id=material.id,
            region_id=region.id,
            period=record_period,
            price=price,
            unit=unit,
            spec=spec,
            trust_level="official" if source.is_official else "market_reference",
            source_type=price_source_type,
            price_basis=infer_price_basis(price_basis, source.name, source.notes, record_notes),
            source_doc_id=document.id,
            notes=record_notes or f"订阅来源: {display_source_name(source.name, source.url, source.source_type)}",
        ))
        write_audit(
            self.session,
            action,
            "material_price",
            detail=f"{'官方订阅' if source.is_official else '市场参考'}: {region.name} {name} {spec} {record_period} {price}",
        )
        self._current_stats["stored_count"] = int(self._current_stats.get("stored_count", 0)) + 1
        return 1

    def _prune_price_storage(self, protected_region_ids: set[int] | None = None) -> dict:
        """Keep the newest period for at most five cities, protecting current fetches."""
        if not config.KEEP_LATEST_PRICE_PERIOD_ONLY:
            return {"removed_prices": 0, "removed_history": 0,
                    "removed_regions": 0, "kept_regions": 0}
        managed_types = {"official", "market_reference", "user_url"}
        prices = self.session.query(MaterialPrice).filter(
            MaterialPrice.is_withdrawn.is_(False),
            MaterialPrice.source_doc_id.is_not(None),
        ).all()
        by_region: dict[int, list[MaterialPrice]] = {}
        for price in prices:
            if (price.source_type or "official") in managed_types:
                by_region.setdefault(price.region_id, []).append(price)

        def period_key(value: str, created_at=None) -> int:
            match = re.search(r"(20\d{2})[-/.年_](\d{1,2})", str(value or ""))
            if match:
                return int(match.group(1)) * 100 + int(match.group(2))
            if created_at:
                return int(created_at.strftime("%Y%m"))
            return 0

        latest_by_region = {
            region_id: max((period_key(item.period, item.created_at) for item in rows), default=0)
            for region_id, rows in by_region.items()
        }
        ranked_regions = sorted(
            latest_by_region,
            key=lambda region_id: (
                latest_by_region[region_id],
                max((item.updated_at or item.created_at for item in by_region[region_id]), default=datetime.min),
            ),
            reverse=True,
        )
        max_regions = max(1, int(config.MAX_PRICE_REGIONS))
        protected = {
            int(region_id)
            for region_id in (protected_region_ids or set())
            if region_id in latest_by_region
        }
        # The current fetch must remain visible even when its period is older
        # than the five cities already stored. Fill the remaining slots by
        # recency, preserving the global city-count limit.
        keep_regions = set(list(protected)[:max_regions])
        for region_id in ranked_regions:
            if len(keep_regions) >= max_regions:
                break
            keep_regions.add(region_id)
        removed_prices = []
        for region_id, rows in by_region.items():
            latest_key = latest_by_region[region_id]
            for price in rows:
                if region_id not in keep_regions or period_key(price.period, price.created_at) != latest_key:
                    removed_prices.append(price)
        removed_doc_ids = {price.source_doc_id for price in removed_prices if price.source_doc_id}
        for price in removed_prices:
            self.session.delete(price)

        removed_histories = []
        for history in self.session.query(PriceHistory).filter(
            PriceHistory.source_doc_id.is_not(None),
        ).all():
            if (history.source_type or "official") not in managed_types:
                continue
            if (
                history.region_id not in keep_regions
                or period_key(history.period, history.created_at) != latest_by_region.get(history.region_id, 0)
            ):
                removed_histories.append(history)
        removed_doc_ids.update(history.source_doc_id for history in removed_histories if history.source_doc_id)
        for history in removed_histories:
            self.session.delete(history)
        self.session.flush()

        for doc_id in removed_doc_ids:
            document = self.session.query(SourceDocument).filter(SourceDocument.id == doc_id).first()
            if not document:
                continue
            still_used = (
                self.session.query(MaterialPrice.id).filter(MaterialPrice.source_doc_id == doc_id).first()
                or self.session.query(PriceHistory.id).filter(PriceHistory.source_doc_id == doc_id).first()
            )
            if still_used:
                continue
            try:
                path = Path(document.file_path)
                source_root = config.SOURCE_DIR.resolve()
                if path.exists() and source_root in path.resolve().parents:
                    path.unlink()
            except (OSError, RuntimeError):
                pass
            self.session.delete(document)
        self.session.flush()
        return {
            "removed_prices": len(removed_prices),
            "removed_history": len(removed_histories),
            "removed_regions": len(set(by_region) - keep_regions),
            "kept_regions": len(keep_regions),
        }

    @staticmethod
    def _normalize_period(period: str) -> str:
        if not period:
            return ""
        match = re.search(r"(20\d{2})\s*[年./-]\s*(\d{1,2})", period)
        return f"{match.group(1)}-{int(match.group(2)):02d}" if match else period.strip()

    @classmethod
    def _infer_period(cls, text: str) -> str:
        matches = re.findall(r"(20\d{2})\s*[年./_-]\s*(\d{1,2})\s*月?", text or "")
        valid = [f"{year}-{int(month):02d}" for year, month in matches if 1 <= int(month) <= 12]
        return max(valid) if valid else ""


def prune_retrieved_price_storage() -> dict:
    """Apply the bounded storage policy during startup or maintenance."""
    with SubscriptionEngine() as engine:
        result = engine._prune_price_storage()
        engine.session.commit()
        return result


def confirm_pending_prices(pending_ids: list[int]) -> dict:
    session = get_session()
    try:
        confirmed = 0
        for pending_id in pending_ids:
            price = session.query(MaterialPrice).filter(MaterialPrice.id == pending_id).first()
            if price and not price.is_confirmed:
                price.is_confirmed = True
                price.is_anomaly = False
                price.anomaly_reason = ""
                confirmed += 1
        write_audit(session, "confirm", "material_price", detail=f"确认入库 {confirmed} 条")
        session.commit()
        return {"success": True, "confirmed": confirmed}
    except Exception as error:
        session.rollback()
        return {"success": False, "error": str(error)}
    finally:
        session.close()


def reject_pending_prices(pending_ids: list[int]) -> dict:
    session = get_session()
    try:
        rejected = 0
        for pending_id in pending_ids:
            price = session.query(MaterialPrice).filter(MaterialPrice.id == pending_id).first()
            if price and not price.is_confirmed:
                session.delete(price)
                rejected += 1
        write_audit(session, "delete", "material_price", detail=f"驳回待入库 {rejected} 条")
        session.commit()
        return {"success": True, "rejected": rejected}
    except Exception as error:
        session.rollback()
        return {"success": False, "error": str(error)}
    finally:
        session.close()
